import contextlib
import base64
import html
import http.server
import io
import json
import math
import threading
import time
from reportlab.lib.utils import ImageReader
import os
import re
import sqlite3
import hashlib
import secrets
import uuid
import zipfile
from calendar import monthrange
from datetime import datetime, timedelta, date
from functools import partial
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional
import urllib.parse

from dotenv import load_dotenv
from textwrap import dedent
import pandas as pd
from PIL import Image, ImageOps, ImageEnhance
from pypdf import PdfReader
import pytesseract

from openpyxl import load_workbook

from backup_utils import ensure_monthly_backup, get_backup_status


import streamlit as st
from streamlit.components.v1 import html as st_components_html
from collections import OrderedDict
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    from storage_paths import get_storage_dir
except ModuleNotFoundError:  # pragma: no cover - defensive for bundled test imports
    import importlib.util

    _storage_module_path = Path(__file__).resolve().parent / "storage_paths.py"
    spec = importlib.util.spec_from_file_location("storage_paths", _storage_module_path)
    module = importlib.util.module_from_spec(spec)
    loader = spec.loader
    if loader is None:  # pragma: no cover - should not happen
        raise
    loader.exec_module(module)
    get_storage_dir = module.get_storage_dir

# ---------- Config ----------
load_dotenv()
DEFAULT_BASE_DIR = get_storage_dir()
BASE_DIR = Path(os.getenv("APP_STORAGE_DIR", DEFAULT_BASE_DIR))
BASE_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = os.getenv("DB_PATH", str(BASE_DIR / "ps_crm.db"))
PROJECT_ROOT = Path(__file__).resolve().parent
DATE_FMT = "%d-%m-%Y"
CURRENCY_SYMBOL = os.getenv("APP_CURRENCY_SYMBOL", "৳")
BACKUP_DIR = BASE_DIR / "backups"
BACKUP_RETENTION_COUNT = int(os.getenv("PS_CRM_BACKUP_RETENTION", "12"))
BACKUP_MIRROR_DIR = os.getenv("PS_CRM_BACKUP_MIRROR_DIR")
BACKUP_MIRROR_PATH = (
    Path(BACKUP_MIRROR_DIR).expanduser() if BACKUP_MIRROR_DIR else None
)

UPLOADS_DIR = BASE_DIR / "uploads"
DELIVERY_ORDER_DIR = UPLOADS_DIR / "delivery_orders"
SERVICE_DOCS_DIR = UPLOADS_DIR / "service_documents"
MAINTENANCE_DOCS_DIR = UPLOADS_DIR / "maintenance_documents"
CUSTOMER_DOCS_DIR = UPLOADS_DIR / "customer_documents"
OPERATIONS_OTHER_DIR = UPLOADS_DIR / "operations_other_documents"
SERVICE_BILL_DIR = UPLOADS_DIR / "service_bills"
REPORT_DOCS_DIR = UPLOADS_DIR / "report_documents"
QUOTATION_RECEIPT_DIR = UPLOADS_DIR / "quotation_receipts"
QUOTATION_DOCS_DIR = UPLOADS_DIR / "quotation_documents"
DELIVERY_RECEIPT_DIR = UPLOADS_DIR / "delivery_receipts"
DOCUMENT_UPLOAD_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif"}
QUOTATION_EDITOR_PORT = int(os.getenv("QUOTATION_EDITOR_PORT", "8502"))

DEFAULT_QUOTATION_VALID_DAYS = 30

REQUIRED_CUSTOMER_FIELDS = {
    "name": "Name",
    "phone": "Phone",
    "address": "Address",
}

SERVICE_STATUS_OPTIONS = ["In progress", "Completed", "Haven't started"]
DEFAULT_SERVICE_STATUS = SERVICE_STATUS_OPTIONS[0]
GENERATOR_CONDITION_OPTIONS = ["Mint", "Good", "Bad"]

REPORT_PERIOD_OPTIONS = OrderedDict(
    [
        ("daily", "Daily"),
        ("weekly", "Weekly"),
        ("monthly", "Monthly"),
    ]
)

SERVICE_REPORT_FIELDS = OrderedDict(
    [
        (
            "customer_name",
            {
                "label": "Customer Name",
                "type": "text",
                "help": "Who received the service work.",
            },
        ),
        (
            "reported_complaints",
            {
                "label": "Reported Complaints",
                "type": "text",
                "help": "Issues raised by the customer.",
            },
        ),
        (
            "product_details",
            {
                "label": "Product Details",
                "type": "text",
                "help": "Model, serial, or generator description.",
            },
        ),
        (
            "details_remarks",
            {
                "label": "Remarks",
                "type": "text",
                "help": "Notes or actions taken for this row.",
            },
        ),
        (
            "remarks_history",
            {
                "label": "Remarks history",
                "type": "text",
                "help": "Previous remarks captured for this row.",
            },
        ),
        (
            "phone",
            {
                "label": "Phone",
                "type": "text",
                "help": "Customer phone or contact number.",
            },
        ),
        ("qty", {"label": "Qty", "type": "number", "step": 1.0}),
        (
            "progress_status",
            {
                "label": "Progress",
                "type": "select",
                "options": ["Ongoing", "Done", "Rejected"],
                "help": "Defaults to Ongoing for new uploads. Choose Done or Rejected when updating.",
            },
        ),
        (
            "payment_status",
            {
                "label": "Payment Status",
                "type": "select",
                "options": ["Pending", "Paid"],
                "help": "Track whether the service has been paid.",
            },
        ),
        (
            "quotation_tk",
            {
                "label": "Quotation Tk",
                "type": "number",
                "format": "%.2f",
                "step": 100.0,
                "help": "Quoted amount in Taka.",
            },
        ),
        (
            "bill_price_tk",
            {
                "label": "Bill Price Tk",
                "type": "number",
                "format": "%.2f",
                "step": 100.0,
                "help": "Final billed amount in Taka.",
            },
        ),
        (
            "work_done_date",
            {
                "label": "Work Done Date",
                "type": "date",
                "format": "DD-MM-YYYY",
                "help": "When the work was completed.",
            },
        ),
        (
            "donation_cost",
            {
                "label": "Donation Cost",
                "type": "number",
                "format": "%.2f",
                "step": 100.0,
                "help": "Any donation or complimentary cost.",
            },
        ),
    ]
)

FOLLOW_UP_REPORT_FIELDS = OrderedDict(
    [
        (
            "follow_up_date",
            {"label": "Date", "type": "date", "format": "DD-MM-YYYY"},
        ),
        ("client_name", {"label": "Customer Name", "type": "text"}),
        ("address", {"label": "Address", "type": "text"}),
        ("contact", {"label": "Phone", "type": "text"}),
        ("product_detail", {"label": "Product Detail", "type": "text"}),
        ("qty", {"label": "Qty", "type": "number", "step": 1.0}),
        (
            "notes",
            {
                "label": "Remarks",
                "type": "text",
                "help": "Follow-up notes or updates for this row.",
            },
        ),
        (
            "remarks_history",
            {
                "label": "Remarks history",
                "type": "text",
                "help": "Previous remarks captured for this row.",
            },
        ),
        ("person_in_charge", {"label": "Person In Charge", "type": "text"}),
        (
            "progress_status",
            {
                "label": "Progress",
                "type": "select",
                "options": ["Ongoing", "Done", "Rejected"],
                "help": "Defaults to Ongoing for new uploads. Choose Done or Rejected when updating.",
            },
        ),
        (
            "reminder_date",
            {
                "label": "Reminder date",
                "type": "date",
                "format": "DD-MM-YYYY",
                "help": "Optional reminder date for this follow-up.",
            },
        ),
    ]
)

REPORT_GRID_FIELDS = SERVICE_REPORT_FIELDS

REPORT_TEMPLATE_LABELS = OrderedDict(
    [
        ("service", "Service report"),
        ("sales", "Sales report"),
        ("follow_up", "Follow-up report"),
    ]
)

REPORT_TEMPLATE_FIELDS = OrderedDict(
    [
        ("service", SERVICE_REPORT_FIELDS),
        ("sales", SERVICE_REPORT_FIELDS),
        ("follow_up", FOLLOW_UP_REPORT_FIELDS),
    ]
)

REPORT_TEMPLATE_SUMMARY_FIELDS = {
    "service": {
        "tasks": "reported_complaints",
        "remarks": "details_remarks",
        "research": "product_details",
    },
    "sales": {
        "tasks": "reported_complaints",
        "remarks": "details_remarks",
        "research": "product_details",
    },
    "follow_up": {
        "tasks": "notes",
        "remarks": "notes",
        "research": "product_detail",
    },
}

REPORT_TEMPLATE_DISPLAY_COLUMNS = {
    key: [config["label"] for config in fields.values()]
    for key, fields in REPORT_TEMPLATE_FIELDS.items()
}

REPORT_GRID_DISPLAY_COLUMNS = REPORT_TEMPLATE_DISPLAY_COLUMNS["service"]
ALL_REPORT_DISPLAY_COLUMNS: list[str] = []
for template_columns in REPORT_TEMPLATE_DISPLAY_COLUMNS.values():
    for column in template_columns:
        if column not in ALL_REPORT_DISPLAY_COLUMNS:
            ALL_REPORT_DISPLAY_COLUMNS.append(column)

_quotation_editor_server: Optional[http.server.ThreadingHTTPServer] = None
_quotation_editor_thread: Optional[threading.Thread] = None


def _normalize_report_template(value: Optional[str]) -> str:
    normalized = (clean_text(value) or "").lower().replace(" ", "_")
    if normalized in REPORT_TEMPLATE_FIELDS:
        return normalized
    return "service"


def _get_report_grid_fields(template_key: Optional[str] = None) -> OrderedDict:
    normalized = _normalize_report_template(template_key)
    return REPORT_TEMPLATE_FIELDS.get(normalized, REPORT_GRID_FIELDS)


def _get_report_display_columns(template_key: Optional[str] = None) -> list[str]:
    normalized = _normalize_report_template(template_key)
    return REPORT_TEMPLATE_DISPLAY_COLUMNS.get(normalized, REPORT_GRID_DISPLAY_COLUMNS)


def _default_report_grid_row(template_key: Optional[str] = None) -> dict[str, object]:
    row: dict[str, object] = {}
    for key, config in _get_report_grid_fields(template_key).items():
        if key == "progress_status":
            row[key] = "Ongoing"
        elif config["type"] == "number":
            row[key] = None
        else:
            row[key] = ""
    return row


def _coerce_grid_number(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.replace(",", "")
        try:
            return float(cleaned)
        except ValueError:
            return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_grid_rows(
    rows: Iterable[dict],
    *,
    template_key: Optional[str] = None,
) -> list[dict[str, object]]:
    normalized: list[dict[str, object]] = []
    if not rows:
        return normalized
    fields = _get_report_grid_fields(template_key)
    for raw in rows:
        if not isinstance(raw, dict):
            continue
        entry: dict[str, object] = {}
        for key, config in fields.items():
            value = raw.get(key)
            if config["type"] in {"text", "select"}:
                entry[key] = clean_text(value)
            elif config["type"] == "number":
                entry[key] = _coerce_grid_number(value)
            elif config["type"] == "date":
                entry[key] = to_iso_date(value)
            else:
                entry[key] = value
        if any(val not in (None, "") for val in entry.values()):
            normalized.append(entry)
    return normalized


def parse_report_grid_payload(
    value: Optional[str],
    *,
    template_key: Optional[str] = None,
) -> list[dict[str, object]]:
    text = clean_text(value)
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except (TypeError, ValueError):
        return []
    if isinstance(parsed, list):
        return _normalize_grid_rows(parsed, template_key=template_key)
    return []


def prepare_report_grid_payload(
    rows: Iterable[dict], *, template_key: Optional[str] = None
) -> Optional[str]:
    normalized = _normalize_grid_rows(rows, template_key=template_key)
    if not normalized:
        return None
    return json.dumps(normalized, ensure_ascii=False)


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower())

REPORT_COLUMN_ALIASES = {
    "customer_name": ["customer", "client", "client_name", "company", "name"],
    "reported_complaints": ["complaints", "issues", "issue", "problem"],
    "product_details": [
        "product",
        "product_detail",
        "product_details",
        "product_info",
        "products_sold",
        "products",
        "model",
        "serial",
        "generator",
    ],
    "details_remarks": ["remarks", "remark", "notes", "note", "comments", "comment"],
    "remarks_history": ["remarks_history", "remarks history", "history"],
    "phone": ["phone", "contact", "mobile", "contact_no", "contact number", "whatsapp"],
    "qty": ["qty", "quantity", "pcs", "pieces", "units"],
    "progress_status": ["progress", "status"],
    "payment_status": ["payment", "payment_status", "payment status"],
    "quotation_tk": [
        "quotation",
        "quotation_tk",
        "quotation amount",
        "quotation value",
        "quote",
    ],
    "work_done_date": ["work_done_date", "work done date", "completion_date"],
    "donation_cost": ["donation", "donation_cost", "complimentary cost"],
    "follow_up_date": ["follow_up_date", "follow-up date", "follow up date", "date"],
    "client_name": ["client_name", "customer_name", "customer", "client"],
    "address": ["address", "location"],
    "contact": ["phone", "contact", "mobile"],
    "product_detail": ["product_detail", "product detail", "product details", "product"],
    "notes": ["remarks", "notes", "comment", "comments"],
    "person_in_charge": ["person_in_charge", "person in charge", "incharge", "responsible"],
    "reminder_date": ["reminder_date", "reminder date", "reminder"],
}


def _build_report_header_map(
    *, template_key: Optional[str] = None
) -> dict[str, str]:
    header_map: dict[str, str] = {}
    for key, config in _get_report_grid_fields(template_key).items():
        header_map[_normalize_header(key)] = key
        header_map[_normalize_header(config["label"])] = key
        for alias in REPORT_COLUMN_ALIASES.get(key, []):
            header_map[_normalize_header(alias)] = key
    return header_map


def _suggest_report_column_mapping(
    columns: Iterable[str], *, template_key: Optional[str] = None
) -> dict[str, str]:
    """Suggest a mapping from target fields to uploaded columns."""

    header_map = _build_report_header_map(template_key=template_key)

    suggestions: dict[str, str] = {}
    for col in columns:
        normalized = _normalize_header(str(col))
        target = header_map.get(normalized)
        if target and target not in suggestions:
            suggestions[target] = col
    return suggestions


def _load_report_grid_dataframe(file_bytes: bytes, filename: str) -> Optional[pd.DataFrame]:
    """Convert uploaded file content into a DataFrame for report imports."""

    name = filename.lower()
    if not file_bytes:
        return None
    buffer = io.BytesIO(file_bytes)
    try:
        if name.endswith(".csv"):
            df = pd.read_csv(buffer)
        else:
            df = pd.read_excel(buffer)
    except Exception:
        return None
    if df is None or df.empty:
        return df
    cleaned_columns: list[str] = []
    seen_headers: dict[str, int] = {}
    for idx, col in enumerate(df.columns, start=1):
        if pd.isna(col):
            cleaned = ""
        else:
            cleaned = str(col).strip()
        if not cleaned or cleaned.lower().startswith("unnamed"):
            cleaned = f"Column {idx}"
        base_header = cleaned
        seen_count = seen_headers.get(base_header, 0)
        if seen_count:
            cleaned = f"{base_header} ({seen_count + 1})"
        seen_headers[base_header] = seen_count + 1
        cleaned_columns.append(cleaned)
    df.columns = cleaned_columns
    return df


def _import_report_grid_from_dataframe(
    raw_df: Optional[pd.DataFrame],
    column_mapping: Optional[dict[str, str]] = None,
    *,
    template_key: Optional[str] = None,
) -> list[dict[str, object]]:
    """Parse a DataFrame into report grid rows with optional custom mapping."""

    imported_rows: list[dict[str, object]] = []
    if raw_df is None or raw_df.empty:
        return imported_rows

    fields = _get_report_grid_fields(template_key)
    header_map = _build_report_header_map(template_key=template_key)

    resolved_mapping: dict[str, str] = {}
    if column_mapping:
        for source, target in column_mapping.items():
            if source in raw_df.columns and target in fields:
                resolved_mapping[source] = target

    if not resolved_mapping:
        for col in raw_df.columns:
            normalized = _normalize_header(str(col))
            target = header_map.get(normalized)
            if target:
                resolved_mapping[col] = target

    if not resolved_mapping:
        return imported_rows

    for _, row in raw_df.iterrows():
        entry = _default_report_grid_row(template_key)
        for source, target in resolved_mapping.items():
            value = row.get(source)
            config = fields.get(target, {})
            if config.get("type") == "number":
                entry[target] = _coerce_grid_number(value)
            elif config.get("type") == "date":
                entry[target] = to_iso_date(value)
            else:
                entry[target] = clean_text(value)
        if any(val not in (None, "") for val in entry.values()):
            imported_rows.append(entry)

    return imported_rows


def _import_report_grid_from_file(
    uploaded_file, column_mapping: Optional[dict[str, str]] = None
) -> list[dict[str, object]]:
    """Parse an uploaded spreadsheet into report grid rows.

    Supports CSV and Excel formats. Column headers are matched against
    ``REPORT_GRID_FIELDS`` labels and keys (case-insensitive). A custom
    ``column_mapping`` can be supplied to map uploaded columns to fields.
    """

    if uploaded_file is None:
        return []

    file_bytes = uploaded_file.getvalue()
    dataframe = _load_report_grid_dataframe(file_bytes, uploaded_file.name)
    return _import_report_grid_from_dataframe(dataframe, column_mapping)


def format_report_grid_rows_for_display(
    rows: Iterable[dict],
    *,
    empty_ok: bool = False,
    template_key: Optional[str] = None,
) -> pd.DataFrame:
    normalized = _normalize_grid_rows(rows, template_key=template_key)
    display_columns = _get_report_display_columns(template_key)
    if not normalized and not empty_ok:
        return pd.DataFrame(columns=display_columns)
    if not normalized:
        normalized = []
    formatted: list[dict[str, object]] = []
    fields = _get_report_grid_fields(template_key)
    for entry in normalized:
        display_row: dict[str, object] = {}
        for key, config in fields.items():
            label = config["label"]
            value = entry.get(key)
            if config["type"] == "text":
                display_row[label] = clean_text(value) or ""
            elif config["type"] == "number":
                display_row[label] = value if value is not None else None
            elif config["type"] == "date":
                iso = to_iso_date(value)
                if iso:
                    try:
                        parsed = datetime.strptime(iso, "%Y-%m-%d")
                        display_row[label] = parsed.strftime(DATE_FMT)
                    except ValueError:
                        display_row[label] = iso
                else:
                    display_row[label] = ""
            else:
                display_row[label] = value
        formatted.append(display_row)
    if not formatted:
        return pd.DataFrame(columns=display_columns)
    df = pd.DataFrame(formatted)
    return df.reindex(columns=display_columns)


def _grid_rows_for_editor(
    rows: Iterable[dict], *, template_key: Optional[str] = None
) -> list[dict[str, object]]:
    """Coerce stored report rows into a format suitable for the data editor."""

    normalized = _normalize_grid_rows(rows, template_key=template_key)
    source_rows: list[dict[str, object]]
    if normalized:
        source_rows = normalized
    else:
        source_rows = [dict(entry) for entry in rows or []]  # type: ignore[arg-type]
    if not source_rows:
        return []

    editor_rows: list[dict[str, object]] = []
    fields = _get_report_grid_fields(template_key)
    for entry in source_rows:
        editor_entry: dict[str, object] = {}
        for key, config in fields.items():
            value = entry.get(key)
            if config["type"] == "text":
                editor_entry[key] = clean_text(value) or ""
            elif config["type"] == "number":
                editor_entry[key] = _coerce_grid_number(value)
            elif config["type"] == "date":
                iso = to_iso_date(value)
                if iso:
                    try:
                        editor_entry[key] = datetime.strptime(
                            iso, "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        editor_entry[key] = None
                else:
                    editor_entry[key] = None
            else:
                editor_entry[key] = value
        editor_rows.append(editor_entry)
    return editor_rows


def _grid_rows_from_editor(
    df: Optional[pd.DataFrame], *, template_key: Optional[str] = None
) -> list[dict[str, object]]:
    """Normalize rows captured from the Streamlit data editor widget."""

    if df is None or not isinstance(df, pd.DataFrame):
        return []
    try:
        records = df.to_dict("records")
    except Exception:
        return []
    return _normalize_grid_rows(records, template_key=template_key)


def _build_report_column_config(
    fields: Mapping[str, dict[str, object]]
) -> dict[str, object]:
    column_config: dict[str, object] = {}
    for key, config in fields.items():
        label = config.get("label", key)
        help_text = config.get("help")
        if config.get("type") == "number":
            column_config[key] = st.column_config.NumberColumn(
                label,
                help=help_text,
                format=config.get("format"),
                step=config.get("step"),
            )
        elif config.get("type") == "date":
            column_config[key] = st.column_config.DateColumn(
                label,
                help=help_text,
                format=config.get("format", "DD-MM-YYYY"),
            )
        elif config.get("type") == "select":
            column_config[key] = st.column_config.SelectboxColumn(
                label,
                options=config.get("options") or [],
                help=help_text,
            )
        else:
            column_config[key] = st.column_config.TextColumn(label, help=help_text)
    return column_config


def _summarize_grid_column(rows: Iterable[dict[str, object]], key: str) -> Optional[str]:
    """Combine a grid column into a legacy text summary for backwards compatibility."""

    values: list[str] = []
    seen: set[str] = set()
    for row in rows or []:
        text = clean_text(row.get(key))
        if not text:
            continue
        if text in seen:
            continue
        seen.add(text)
        values.append(text)
    if not values:
        return None
    return "\n".join(values)


def _report_remarks_field(template_key: Optional[str]) -> str:
    normalized = _normalize_report_template(template_key)
    if normalized == "follow_up":
        return "notes"
    return "details_remarks"


def _append_report_remarks_history(
    rows: list[dict[str, object]],
    previous_rows: list[dict[str, object]],
    *,
    template_key: Optional[str] = None,
) -> list[dict[str, object]]:
    remarks_key = _report_remarks_field(template_key)
    history_key = "remarks_history"
    now_stamp = datetime.now().strftime("%d-%m-%Y %H:%M")
    updated_rows: list[dict[str, object]] = []
    for idx, row in enumerate(rows):
        updated = dict(row)
        previous = previous_rows[idx] if idx < len(previous_rows) else {}
        previous_history = clean_text(previous.get(history_key)) or ""
        old_remark = clean_text(previous.get(remarks_key)) or ""
        new_remark = clean_text(updated.get(remarks_key)) or ""
        history_lines: list[str] = []
        if previous_history:
            history_lines.append(previous_history)
        if old_remark and new_remark and new_remark != old_remark:
            if not previous_history:
                history_lines.append(f"Previous: {old_remark}")
            history_lines.append(f"{now_stamp} - {new_remark}")
        updated[history_key] = "\n".join(history_lines) if history_lines else previous_history
        updated_rows.append(updated)
    return updated_rows

NOTIFICATION_BUFFER_KEY = "runtime_notifications"
MAX_RUNTIME_NOTIFICATIONS = 40
ACTIVITY_FEED_LIMIT = 25

NOTIFICATION_EVENT_LABELS = {
    "customer_created": "Customer added",
    "customer_updated": "Customer updated",
    "customer_deleted": "Customer removed",
    "service_created": "Service created",
    "service_updated": "Service updated",
    "maintenance_created": "Maintenance created",
    "maintenance_updated": "Maintenance updated",
    "quotation_created": "Quotation created",
    "quotation_updated": "Quotation updated",
    "quotation_deleted": "Quotation deleted",
    "warranty_updated": "Warranty updated",
    "report_submitted": "Report submitted",
    "report_updated": "Report updated",
    "report_deleted": "Report deleted",
    "delivery_order_created": "Delivery order saved",
    "delivery_order_updated": "Delivery order updated",
    "delivery_order_deleted": "Delivery order deleted",
    "work_done_created": "Work done saved",
    "work_done_updated": "Work done updated",
    "work_done_deleted": "Work done deleted",
}

LEAD_REMARK_TAG = "Lead / Chasing"

def customer_complete_clause(alias: str = "") -> str:
    prefix = f"{alias}." if alias else ""
    return " AND ".join(
        [
            f"TRIM(COALESCE({prefix}name, '')) <> ''",
            f"TRIM(COALESCE({prefix}phone, '')) <> ''",
            f"TRIM(COALESCE({prefix}address, '')) <> ''",
        ]
    )


def customer_incomplete_clause(alias: str = "") -> str:
    return f"NOT ({customer_complete_clause(alias)})"


def _is_blank_field(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not clean_text(value)
    if isinstance(value, (list, dict, tuple, set)):
        return len(value) == 0
    return False


def _has_quotation_items(items: Optional[list[dict[str, object]]]) -> bool:
    if not items:
        return False
    for entry in items:
        if not isinstance(entry, dict):
            continue
        if clean_text(entry.get("description")):
            return True
        quantity = _coerce_float(entry.get("quantity"), 0.0)
        rate = _coerce_float(entry.get("rate"), 0.0)
        total_price = _coerce_float(entry.get("total_price"), 0.0)
        if quantity > 0 or rate > 0 or total_price > 0:
            return True
    return False


def _is_lead_customer(remarks: Optional[str]) -> bool:
    if not remarks:
        return False
    normalized = clean_text(remarks).lower()
    return "lead / chasing" in normalized or "lead: chasing" in normalized


def _strip_lead_tag(remarks: Optional[str]) -> Optional[str]:
    if not remarks:
        return None
    parts = [part.strip() for part in remarks.split("|")]
    cleaned_parts = []
    for part in parts:
        if not part:
            continue
        if _is_lead_customer(part):
            continue
        cleaned_parts.append(part)
    cleaned = " | ".join(cleaned_parts).strip()
    return cleaned or None


def _promote_lead_customer(
    conn,
    *,
    name: Optional[str],
    company: Optional[str],
    phone: Optional[str],
) -> bool:
    customer_name = clean_text(name)
    company_name = clean_text(company)
    phone_number = clean_text(phone)

    if not any([customer_name, company_name, phone_number]):
        return False

    cursor = conn.cursor()

    def _fetch_existing(query: str, params: tuple[object, ...]):
        return cursor.execute(query, params).fetchone()

    existing = None
    if phone_number:
        existing = _fetch_existing(
            """
            SELECT customer_id, remarks
            FROM customers
            WHERE TRIM(IFNULL(phone, '')) = ?
            ORDER BY customer_id DESC
            LIMIT 1
            """,
            (phone_number,),
        )
    if existing is None and company_name:
        existing = _fetch_existing(
            """
            SELECT customer_id, remarks
            FROM customers
            WHERE LOWER(IFNULL(company_name, '')) = LOWER(?)
            ORDER BY customer_id DESC
            LIMIT 1
            """,
            (company_name,),
        )
    if existing is None and customer_name:
        existing = _fetch_existing(
            """
            SELECT customer_id, remarks
            FROM customers
            WHERE LOWER(IFNULL(name, '')) = LOWER(?)
            ORDER BY customer_id DESC
            LIMIT 1
            """,
            (customer_name,),
        )

    if not existing:
        return False

    customer_id, remarks = existing
    if not _is_lead_customer(remarks):
        return False

    cleaned = _strip_lead_tag(remarks)
    cursor.execute(
        "UPDATE customers SET remarks=? WHERE customer_id=?",
        (cleaned, customer_id),
    )
    conn.commit()
    return True

# ---------- Schema ----------
ADMIN_DEFAULT_PASSWORD = "PANNAPS123"

SCHEMA_SQL = """
PRAGMA foreign_keys = ON;
CREATE TABLE IF NOT EXISTS users (
    user_id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE,
    pass_hash TEXT,
    phone TEXT,
    email TEXT,
    title TEXT,
    role TEXT DEFAULT 'staff',
    created_at TEXT DEFAULT (datetime('now'))
);
CREATE TABLE IF NOT EXISTS user_sessions (
    token TEXT PRIMARY KEY,
    user_id INTEGER NOT NULL,
    created_at TEXT DEFAULT (datetime('now')),
    last_seen TEXT DEFAULT (datetime('now')),
    expires_at TEXT,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_user_sessions_user_id ON user_sessions(user_id);
CREATE INDEX IF NOT EXISTS idx_user_sessions_expires ON user_sessions(expires_at);
CREATE TABLE IF NOT EXISTS customers (
    customer_id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    company_name TEXT,
    phone TEXT,
    address TEXT,
    delivery_address TEXT,
    remarks TEXT,
    purchase_date TEXT,
    product_info TEXT,
    delivery_order_code TEXT,
    sales_person TEXT,
    amount_spent REAL,
    created_by INTEGER,
    attachment_path TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    dup_flag INTEGER DEFAULT 0,
    FOREIGN KEY(created_by) REFERENCES users(user_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS customer_groups (
    group_id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    created_by INTEGER,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(created_by) REFERENCES users(user_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS customer_group_members (
    group_id INTEGER NOT NULL,
    customer_id INTEGER NOT NULL,
    created_at TEXT DEFAULT (datetime('now')),
    PRIMARY KEY (group_id, customer_id),
    FOREIGN KEY(group_id) REFERENCES customer_groups(group_id) ON DELETE CASCADE,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_customer_groups_name ON customer_groups(name);
CREATE INDEX IF NOT EXISTS idx_customer_group_members_customer ON customer_group_members(customer_id);
CREATE TABLE IF NOT EXISTS products (
    product_id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    model TEXT,
    serial TEXT,
    dup_flag INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS orders (
    order_id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER,
    order_date TEXT,
    delivery_date TEXT,
    notes TEXT,
    dup_flag INTEGER DEFAULT 0,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS order_items (
    order_item_id INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id INTEGER,
    product_id INTEGER,
    quantity INTEGER DEFAULT 1,
    FOREIGN KEY(order_id) REFERENCES orders(order_id) ON DELETE CASCADE,
    FOREIGN KEY(product_id) REFERENCES products(product_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS warranties (
    warranty_id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER,
    product_id INTEGER,
    serial TEXT,
    issue_date TEXT,
    expiry_date TEXT,
    status TEXT DEFAULT 'active',
    remarks TEXT,
    dup_flag INTEGER DEFAULT 0,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE SET NULL,
    FOREIGN KEY(product_id) REFERENCES products(product_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS delivery_orders (
    do_number TEXT PRIMARY KEY,
    customer_id INTEGER,
    order_id INTEGER,
    description TEXT,
    sales_person TEXT,
    remarks TEXT,
    file_path TEXT,
    record_type TEXT DEFAULT 'delivery_order',
    status TEXT DEFAULT 'pending',
    payment_receipt_path TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    deleted_at TEXT,
    deleted_by INTEGER,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE SET NULL,
    FOREIGN KEY(order_id) REFERENCES orders(order_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS services (
    service_id INTEGER PRIMARY KEY AUTOINCREMENT,
    do_number TEXT,
    customer_id INTEGER,
    service_date TEXT,
    service_start_date TEXT,
    service_end_date TEXT,
    description TEXT,
    status TEXT DEFAULT 'In progress',
    remarks TEXT,
    service_product_info TEXT,
    condition_status TEXT,
    condition_remarks TEXT,
    bill_amount REAL,
    bill_document_path TEXT,
    report_id INTEGER,
    report_row_index INTEGER,
    payment_status TEXT DEFAULT 'pending',
    payment_receipt_path TEXT,
    created_by INTEGER,
    updated_at TEXT DEFAULT (datetime('now')),
    deleted_at TEXT,
    deleted_by INTEGER,
    FOREIGN KEY(do_number) REFERENCES delivery_orders(do_number) ON DELETE SET NULL,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS maintenance_records (
    maintenance_id INTEGER PRIMARY KEY AUTOINCREMENT,
    do_number TEXT,
    customer_id INTEGER,
    maintenance_date TEXT,
    maintenance_start_date TEXT,
    maintenance_end_date TEXT,
    description TEXT,
    status TEXT DEFAULT 'In progress',
    remarks TEXT,
    maintenance_product_info TEXT,
    total_amount REAL,
    payment_status TEXT DEFAULT 'pending',
    payment_receipt_path TEXT,
    created_by INTEGER,
    updated_at TEXT DEFAULT (datetime('now')),
    deleted_at TEXT,
    deleted_by INTEGER,
    FOREIGN KEY(do_number) REFERENCES delivery_orders(do_number) ON DELETE SET NULL,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS service_documents (
    document_id INTEGER PRIMARY KEY AUTOINCREMENT,
    service_id INTEGER,
    file_path TEXT,
    original_name TEXT,
    uploaded_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(service_id) REFERENCES services(service_id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS maintenance_documents (
    document_id INTEGER PRIMARY KEY AUTOINCREMENT,
    maintenance_id INTEGER,
    file_path TEXT,
    original_name TEXT,
    uploaded_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(maintenance_id) REFERENCES maintenance_records(maintenance_id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS customer_notes (
    note_id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER NOT NULL,
    note TEXT,
    remind_on TEXT,
    is_done INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_customer_notes_customer ON customer_notes(customer_id);
CREATE INDEX IF NOT EXISTS idx_customer_notes_remind ON customer_notes(remind_on, is_done);
CREATE TABLE IF NOT EXISTS customer_documents (
    document_id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER NOT NULL,
    doc_type TEXT NOT NULL,
    file_path TEXT,
    original_name TEXT,
    uploaded_by INTEGER,
    uploaded_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    updated_by INTEGER,
    deleted_at TEXT,
    deleted_by INTEGER,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE CASCADE,
    FOREIGN KEY(uploaded_by) REFERENCES users(user_id) ON DELETE SET NULL
);
CREATE INDEX IF NOT EXISTS idx_customer_documents_customer ON customer_documents(customer_id);
CREATE TABLE IF NOT EXISTS operations_other_documents (
    document_id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER NOT NULL,
    description TEXT,
    items_payload TEXT,
    file_path TEXT,
    original_name TEXT,
    uploaded_by INTEGER,
    uploaded_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    deleted_at TEXT,
    deleted_by INTEGER,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE CASCADE,
    FOREIGN KEY(uploaded_by) REFERENCES users(user_id) ON DELETE SET NULL
);
CREATE INDEX IF NOT EXISTS idx_operations_other_documents_customer ON operations_other_documents(customer_id);
CREATE TABLE IF NOT EXISTS import_history (
    import_id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER,
    product_id INTEGER,
    order_id INTEGER,
    order_item_id INTEGER,
    warranty_id INTEGER,
    do_number TEXT,
    import_tag TEXT,
    imported_at TEXT DEFAULT (datetime('now')),
    original_date TEXT,
    customer_name TEXT,
    address TEXT,
    delivery_address TEXT,
    phone TEXT,
    product_label TEXT,
    notes TEXT,
    amount_spent REAL,
    quantity INTEGER DEFAULT 1,
    imported_by INTEGER,
    deleted_at TEXT,
    deleted_by INTEGER,
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE SET NULL,
    FOREIGN KEY(product_id) REFERENCES products(product_id) ON DELETE SET NULL,
    FOREIGN KEY(order_id) REFERENCES orders(order_id) ON DELETE SET NULL,
    FOREIGN KEY(order_item_id) REFERENCES order_items(order_item_id) ON DELETE SET NULL,
    FOREIGN KEY(warranty_id) REFERENCES warranties(warranty_id) ON DELETE SET NULL,
    FOREIGN KEY(imported_by) REFERENCES users(user_id) ON DELETE SET NULL,
    FOREIGN KEY(deleted_by) REFERENCES users(user_id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS work_reports (
    report_id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    period_type TEXT NOT NULL,
    period_start TEXT NOT NULL,
    period_end TEXT NOT NULL,
    tasks TEXT,
    remarks TEXT,
    research TEXT,
    grid_payload TEXT,
    attachment_path TEXT,
    import_file_path TEXT,
    report_template TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_work_reports_user_period ON work_reports(user_id, period_type, period_start);
CREATE INDEX IF NOT EXISTS idx_work_reports_period ON work_reports(period_type, period_start, period_end);
CREATE UNIQUE INDEX IF NOT EXISTS uniq_work_reports_user_period ON work_reports(user_id, period_type, period_start);
CREATE TABLE IF NOT EXISTS quotations (
    quotation_id INTEGER PRIMARY KEY AUTOINCREMENT,
    reference TEXT,
    quote_date TEXT,
    customer_name TEXT,
    customer_company TEXT,
    customer_address TEXT,
    customer_district TEXT,
    customer_contact TEXT,
    attention_name TEXT,
    attention_title TEXT,
    subject TEXT,
    salutation TEXT,
    introduction TEXT,
    closing TEXT,
    quote_type TEXT,
    total_amount REAL,
    discount_pct REAL,
    status TEXT DEFAULT 'pending',
    payment_receipt_path TEXT,
    follow_up_status TEXT,
    follow_up_notes TEXT,
    follow_up_date TEXT,
    reminder_label TEXT,
    letter_template TEXT,
    salesperson_name TEXT,
    salesperson_title TEXT,
    salesperson_contact TEXT,
    salesperson_email TEXT,
    document_path TEXT,
    items_payload TEXT,
    remarks_internal TEXT,
    created_by INTEGER,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    deleted_at TEXT,
    deleted_by INTEGER,
    FOREIGN KEY(created_by) REFERENCES users(user_id) ON DELETE SET NULL
);
CREATE INDEX IF NOT EXISTS idx_quotations_status ON quotations(status);
CREATE INDEX IF NOT EXISTS idx_quotations_owner ON quotations(created_by);
CREATE TABLE IF NOT EXISTS needs (
    need_id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER,
    product TEXT,
    unit TEXT,
    notes TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS activity_log (
    activity_id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    event_type TEXT,
    entity_type TEXT,
    entity_id INTEGER,
    description TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE SET NULL
);
CREATE INDEX IF NOT EXISTS idx_activity_log_created ON activity_log(created_at DESC);
CREATE INDEX IF NOT EXISTS idx_activity_log_entity ON activity_log(entity_type, entity_id);
CREATE TABLE IF NOT EXISTS staff_admin_messages (
    message_id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    message TEXT NOT NULL,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE SET NULL
);
CREATE INDEX IF NOT EXISTS idx_staff_admin_messages_created_at ON staff_admin_messages(created_at DESC);
CREATE TRIGGER IF NOT EXISTS prevent_admin_delete
BEFORE DELETE ON users
WHEN LOWER(OLD.role) = 'admin'
BEGIN
    SELECT RAISE(ABORT, 'Cannot delete admin user');
END;
"""

# ---------- Helpers ----------
def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=30)
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.execute("PRAGMA busy_timeout = 30000;")
    conn.execute("PRAGMA journal_mode = WAL;")
    return conn

def init_schema(conn):
    ensure_upload_dirs()
    conn.executescript(SCHEMA_SQL)
    for attempt in range(3):
        try:
            ensure_schema_upgrades(conn)
            break
        except sqlite3.OperationalError as exc:
            if "locked" not in str(exc).lower() or attempt == 2:
                raise
            time.sleep(0.5 * (attempt + 1))
    conn.commit()
    # bootstrap admin if empty
    cur = conn.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        admin_user = os.getenv("ADMIN_USER", "admin")
        admin_pass = os.getenv("ADMIN_PASS", ADMIN_DEFAULT_PASSWORD)
        h = hashlib.sha256(admin_pass.encode("utf-8")).hexdigest()
        conn.execute("INSERT INTO users (username, pass_hash, role) VALUES (?, ?, 'admin')", (admin_user, h))
        conn.commit()


def ensure_schema_upgrades(conn):
    def has_column(table: str, column: str) -> bool:
        cur = conn.execute(f"PRAGMA table_info({table})")
        return any(str(row[1]) == column for row in cur.fetchall())

    def add_column(table: str, column: str, definition: str) -> None:
        if not has_column(table, column):
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

    def ensure_trigger(name: str, sql: str) -> None:
        cur = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='trigger' AND name=?",
            (name,),
        )
        if cur.fetchone() is None:
            conn.execute(sql)

    add_column("customers", "company_name", "TEXT")
    add_column("customers", "delivery_address", "TEXT")
    add_column("customers", "remarks", "TEXT")
    add_column("customers", "purchase_date", "TEXT")
    add_column("customers", "product_info", "TEXT")
    add_column("customers", "delivery_order_code", "TEXT")
    add_column("customers", "attachment_path", "TEXT")
    add_column("customers", "sales_person", "TEXT")
    add_column("customers", "amount_spent", "REAL")
    add_column("customers", "created_by", "INTEGER")
    add_column("users", "phone", "TEXT")
    add_column("users", "email", "TEXT")
    add_column("users", "title", "TEXT")
    add_column("services", "status", "TEXT DEFAULT 'In progress'")
    add_column("services", "service_start_date", "TEXT")
    add_column("services", "service_end_date", "TEXT")
    add_column("services", "service_product_info", "TEXT")
    add_column("services", "condition_status", "TEXT")
    add_column("services", "condition_remarks", "TEXT")
    add_column("services", "bill_amount", "REAL")
    add_column("services", "bill_document_path", "TEXT")
    add_column("services", "created_by", "INTEGER")
    add_column("services", "report_id", "INTEGER")
    add_column("services", "report_row_index", "INTEGER")
    add_column("maintenance_records", "status", "TEXT DEFAULT 'In progress'")
    add_column("services", "payment_status", "TEXT DEFAULT 'pending'")
    add_column("services", "payment_receipt_path", "TEXT")
    add_column("maintenance_records", "payment_status", "TEXT DEFAULT 'pending'")
    add_column("maintenance_records", "payment_receipt_path", "TEXT")
    add_column("services", "deleted_at", "TEXT")
    add_column("services", "deleted_by", "INTEGER")
    add_column("maintenance_records", "deleted_at", "TEXT")
    add_column("maintenance_records", "deleted_by", "INTEGER")
    add_column("quotations", "payment_receipt_path", "TEXT")
    add_column("quotations", "items_payload", "TEXT")
    add_column("maintenance_records", "maintenance_start_date", "TEXT")
    add_column("maintenance_records", "maintenance_end_date", "TEXT")
    add_column("maintenance_records", "maintenance_product_info", "TEXT")
    add_column("maintenance_records", "total_amount", "REAL")

    ensure_trigger(
        "prevent_admin_delete",
        """
        CREATE TRIGGER prevent_admin_delete
        BEFORE DELETE ON users
        WHEN LOWER(OLD.role) = 'admin'
        BEGIN
            SELECT RAISE(ABORT, 'Cannot delete admin user');
        END;
        """,
    )

    admin_hash = hashlib.sha256(ADMIN_DEFAULT_PASSWORD.encode("utf-8")).hexdigest()
    conn.execute(
        "UPDATE users SET pass_hash=? WHERE LOWER(username)='admin'",
        (admin_hash,),
    )
    add_column("maintenance_records", "created_by", "INTEGER")
    add_column("quotations", "document_path", "TEXT")
    add_column("warranties", "remarks", "TEXT")
    add_column("delivery_orders", "remarks", "TEXT")
    add_column("delivery_orders", "items_payload", "TEXT")
    add_column("delivery_orders", "total_amount", "REAL")
    add_column("delivery_orders", "created_by", "INTEGER")
    add_column("delivery_orders", "record_type", "TEXT DEFAULT 'delivery_order'")
    add_column("delivery_orders", "status", "TEXT DEFAULT 'pending'")
    add_column("delivery_orders", "payment_receipt_path", "TEXT")
    add_column("delivery_orders", "updated_at", "TEXT DEFAULT (datetime('now'))")
    add_column("delivery_orders", "deleted_at", "TEXT")
    add_column("delivery_orders", "deleted_by", "INTEGER")
    add_column("operations_other_documents", "updated_at", "TEXT DEFAULT (datetime('now'))")
    add_column("operations_other_documents", "updated_by", "INTEGER")
    add_column("operations_other_documents", "deleted_at", "TEXT")
    add_column("operations_other_documents", "deleted_by", "INTEGER")
    add_column("customer_documents", "updated_at", "TEXT DEFAULT (datetime('now'))")
    add_column("customer_documents", "updated_by", "INTEGER")
    add_column("customer_documents", "deleted_at", "TEXT")
    add_column("customer_documents", "deleted_by", "INTEGER")
    add_column("import_history", "amount_spent", "REAL")
    add_column("import_history", "imported_by", "INTEGER")
    add_column("import_history", "delivery_address", "TEXT")
    add_column("import_history", "quantity", "INTEGER DEFAULT 1")
    add_column("import_history", "deleted_by", "INTEGER")
    add_column("work_reports", "grid_payload", "TEXT")
    add_column("work_reports", "attachment_path", "TEXT")
    add_column("work_reports", "import_file_path", "TEXT")
    add_column("work_reports", "report_template", "TEXT")
    add_column("service_documents", "uploaded_by", "INTEGER")
    add_column("maintenance_documents", "uploaded_by", "INTEGER")
    add_column("quotations", "salesperson_email", "TEXT")
    add_column("quotations", "deleted_at", "TEXT")
    add_column("quotations", "deleted_by", "INTEGER")

    # Remove stored email data for privacy; the app no longer collects it.
    if has_column("customers", "email"):
        conn.execute("UPDATE customers SET email=NULL WHERE email IS NOT NULL")

    cur = conn.execute(
        """
        SELECT report_id, user_id, LOWER(COALESCE(period_type, '')), COALESCE(period_start, '')
        FROM work_reports
        ORDER BY user_id, LOWER(COALESCE(period_type, '')), COALESCE(period_start, ''), report_id DESC
        """
    )
    seen_keys: set[tuple[int, str, str]] = set()
    duplicates: list[int] = []
    for report_id, user_id, period_type, period_start in cur.fetchall():
        key = (int(user_id), period_type or "", period_start or "")
        if key in seen_keys:
            duplicates.append(int(report_id))
        else:
            seen_keys.add(key)
    for report_id in duplicates:
        conn.execute("DELETE FROM work_reports WHERE report_id=?", (int(report_id),))

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS customer_notes (
            note_id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            note TEXT,
            remind_on TEXT,
            is_done INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_customer_notes_customer ON customer_notes(customer_id)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_customer_notes_remind ON customer_notes(remind_on, is_done)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_customers_created_by ON customers(created_by)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_import_history_imported_by ON import_history(imported_by)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dashboard_remarks (
            remark_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            note TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dashboard_remarks_created_at ON dashboard_remarks(created_at DESC)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dashboard_remarks_user ON dashboard_remarks(user_id)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS staff_admin_messages (
            message_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            message TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_staff_admin_messages_created_at ON staff_admin_messages(created_at DESC)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS operations_other_documents (
            document_id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            description TEXT,
            items_payload TEXT,
            file_path TEXT,
            original_name TEXT,
            uploaded_by INTEGER,
            uploaded_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            updated_by INTEGER,
            deleted_at TEXT,
            deleted_by INTEGER,
            FOREIGN KEY(customer_id) REFERENCES customers(customer_id) ON DELETE CASCADE,
            FOREIGN KEY(uploaded_by) REFERENCES users(user_id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_operations_other_documents_customer ON operations_other_documents(customer_id)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS work_reports (
            report_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            period_type TEXT NOT NULL,
            period_start TEXT NOT NULL,
            period_end TEXT NOT NULL,
            tasks TEXT,
            remarks TEXT,
            research TEXT,
            attachment_path TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_work_reports_user_period ON work_reports(user_id, period_type, period_start)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_work_reports_period ON work_reports(period_type, period_start, period_end)"
    )
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS uniq_work_reports_user_period ON work_reports(user_id, period_type, period_start)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS activity_log (
            activity_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            event_type TEXT,
            entity_type TEXT,
            entity_id INTEGER,
            description TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_activity_log_created ON activity_log(created_at DESC)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_activity_log_entity ON activity_log(entity_type, entity_id)"
    )


def get_current_user() -> dict:
    return st.session_state.get("user") or {}


def current_user_id() -> Optional[int]:
    user = get_current_user()
    try:
        return int(user.get("user_id")) if user.get("user_id") is not None else None
    except (TypeError, ValueError):
        return None


def current_user_is_admin() -> bool:
    return get_current_user().get("role") == "admin"


def customer_scope_filter(alias: str = "") -> tuple[str, tuple[object, ...]]:
    user = get_current_user()
    if not user or user.get("role") == "admin":
        return "", ()
    user_id = current_user_id()
    if user_id is None:
        return "1=0", ()
    prefix = f"{alias}." if alias else ""
    return f"{prefix}created_by = ?", (user_id,)


def accessible_customer_ids(conn) -> Optional[set[int]]:
    if current_user_is_admin():
        return None
    user_id = current_user_id()
    if user_id is None:
        return set()
    df = df_query(conn, "SELECT customer_id FROM customers WHERE created_by=?", (user_id,))
    if df.empty:
        return set()
    ids: set[int] = set()
    for value in df["customer_id"].dropna().tolist():
        try:
            ids.add(int(value))
        except (TypeError, ValueError):
            continue
    return ids


def filter_delivery_orders_for_view(
    do_df: pd.DataFrame,
    allowed_customers: Optional[set[int]],
    *,
    record_types: Optional[set[str]] = None,
) -> pd.DataFrame:
    """Limit delivery order rows to records the current user can access.

    Admins (``allowed_customers`` is ``None``) see all rows. Staff can always
    see records linked to customers they own **or** records they personally
    created, even when the customer link is missing.
    """

    if do_df is None or do_df.empty or allowed_customers is None:
        return do_df

    viewer_id = current_user_id()

    def _allowed(row: pd.Series) -> bool:
        if record_types is not None:
            row_type = clean_text(row.get("record_type")) or "delivery_order"
            if row_type not in record_types:
                return False
        cust_id = row.get("customer_id")
        creator_id = row.get("created_by")

        try:
            if pd.notna(cust_id) and int(cust_id) in allowed_customers:
                return True
        except Exception:
            pass

        if viewer_id is not None:
            try:
                if pd.notna(creator_id) and int(creator_id) == int(viewer_id):
                    return True
            except Exception:
                pass

        return False

    return do_df[do_df.apply(_allowed, axis=1)]


def df_query(conn, q, params=()):
    return pd.read_sql_query(q, conn, params=params)

def fmt_dates(df: pd.DataFrame, cols):
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce").dt.strftime(DATE_FMT)
    return df


def add_months(base: date, months: int) -> date:
    """Return ``base`` shifted by ``months`` while clamping the day to the target month."""

    if not isinstance(base, date):
        raise TypeError("base must be a date instance")
    try:
        months = int(months)
    except (TypeError, ValueError):
        raise TypeError("months must be an integer") from None

    month_index = base.month - 1 + months
    year = base.year + month_index // 12
    month = month_index % 12 + 1
    day = min(base.day, monthrange(year, month)[1])
    return date(year, month, day)


def month_bucket_counts(
    conn,
    table: str,
    date_column: str,
    *,
    where: Optional[str] = None,
    params: Optional[Iterable[object]] = None,
) -> tuple[int, int]:
    """Return the current and previous month counts for ``table.date_column``."""

    params = tuple(params or ())
    criteria = [f"{date_column} IS NOT NULL"]
    if where:
        criteria.append(f"({where})")
    where_clause = " AND ".join(criteria)
    query = dedent(
        f"""
        SELECT
            SUM(CASE WHEN strftime('%Y-%m', {date_column}) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) AS current_month,
            SUM(CASE WHEN strftime('%Y-%m', {date_column}) = strftime('%Y-%m', date('now', '-1 month')) THEN 1 ELSE 0 END) AS previous_month
        FROM {table}
        WHERE {where_clause}
        """
    )
    cur = conn.execute(query, params)
    row = cur.fetchone()
    if not row:
        return 0, 0
    current, previous = row
    current_count = int(current or 0)
    previous_count = int(previous or 0)
    return current_count, previous_count


def format_metric_delta(current: int, previous: int) -> str:
    """Format a delta label comparing the current value to the previous month."""

    diff = int(current) - int(previous)
    if diff == 0:
        return "On par with last month"
    if previous == 0:
        return f"+{current} (new this month)"
    pct = (diff / previous) * 100
    return f"{diff:+d} ({pct:+.1f}%) vs last month"


def sales_scope_filter(alias: str = "d") -> tuple[str, tuple[object, ...]]:
    if current_user_is_admin():
        return "", ()
    user_id = current_user_id()
    if user_id is None:
        return "1=0", ()
    prefix = f"{alias}." if alias else ""
    return f"({prefix}created_by = ? OR c.created_by = ?)", (user_id, user_id)


def fetch_sales_metrics(conn, scope_clause: str, scope_params: tuple[object, ...]) -> dict[str, float]:
    is_admin = current_user_is_admin()
    user_id = current_user_id()

    def _sum_delivery(where_clause: str, params: tuple[object, ...]) -> float:
        filters = [
            "d.deleted_at IS NULL",
            "COALESCE(d.record_type, 'delivery_order') IN ('delivery_order', 'work_done')",
            "d.status = 'paid'",
            "d.total_amount IS NOT NULL",
        ]
        if scope_clause:
            filters.append(scope_clause)
        query = dedent(
            f"""
            SELECT COALESCE(SUM(d.total_amount), 0) AS total
            FROM delivery_orders d
            LEFT JOIN customers c ON c.customer_id = d.customer_id
            WHERE {" AND ".join(filters)} AND {where_clause}
            """
        )
        row = conn.execute(query, params).fetchone()
        return float(row[0] or 0)

    def _sum_quotations(where_clause: str) -> float:
        filters = [
            "q.deleted_at IS NULL",
            "q.status = 'paid'",
            "q.total_amount IS NOT NULL",
        ]
        params: list[object] = []
        if not is_admin and user_id is not None:
            filters.append("q.created_by = ?")
            params.append(user_id)
        query = dedent(
            f"""
            SELECT COALESCE(SUM(q.total_amount), 0) AS total
            FROM quotations q
            WHERE {" AND ".join(filters)} AND {where_clause}
            """
        )
        row = conn.execute(query, tuple(params)).fetchone()
        return float(row[0] or 0)

    def _sum_services(where_clause: str) -> float:
        filters = [
            "s.deleted_at IS NULL",
            "s.payment_status = 'paid'",
            "s.bill_amount IS NOT NULL",
        ]
        params: list[object] = []
        if not is_admin and user_id is not None:
            filters.append("(s.created_by = ? OR c.created_by = ?)")
            params.extend([user_id, user_id])
        query = dedent(
            f"""
            SELECT COALESCE(SUM(s.bill_amount), 0) AS total
            FROM services s
            LEFT JOIN customers c ON c.customer_id = s.customer_id
            WHERE {" AND ".join(filters)} AND {where_clause}
            """
        )
        row = conn.execute(query, tuple(params)).fetchone()
        return float(row[0] or 0)

    def _sum_maintenance(where_clause: str) -> float:
        filters = [
            "m.deleted_at IS NULL",
            "m.payment_status = 'paid'",
            "m.total_amount IS NOT NULL",
        ]
        params: list[object] = []
        if not is_admin and user_id is not None:
            filters.append("(m.created_by = ? OR c.created_by = ?)")
            params.extend([user_id, user_id])
        query = dedent(
            f"""
            SELECT COALESCE(SUM(m.total_amount), 0) AS total
            FROM maintenance_records m
            LEFT JOIN customers c ON c.customer_id = m.customer_id
            WHERE {" AND ".join(filters)} AND {where_clause}
            """
        )
        row = conn.execute(query, tuple(params)).fetchone()
        return float(row[0] or 0)

    def _sum_all(
        delivery_clause: str,
        quote_clause: str,
        service_clause: str,
        maintenance_clause: str,
    ) -> float:
        delivery_total = _sum_delivery(delivery_clause, scope_params)
        quote_total = _sum_quotations(quote_clause)
        service_total = _sum_services(service_clause)
        maintenance_total = _sum_maintenance(maintenance_clause)
        return delivery_total + quote_total + service_total + maintenance_total

    return {
        "daily": _sum_all(
            "date(d.created_at) = date('now')",
            "date(q.created_at) = date('now')",
            "date(s.updated_at) = date('now')",
            "date(m.updated_at) = date('now')",
        ),
        "weekly": _sum_all(
            "date(d.created_at) >= date('now', '-6 day')",
            "date(q.created_at) >= date('now', '-6 day')",
            "date(s.updated_at) >= date('now', '-6 day')",
            "date(m.updated_at) >= date('now', '-6 day')",
        ),
        "monthly": _sum_all(
            "strftime('%Y-%m', d.created_at) = strftime('%Y-%m', 'now')",
            "strftime('%Y-%m', q.created_at) = strftime('%Y-%m', 'now')",
            "strftime('%Y-%m', s.updated_at) = strftime('%Y-%m', 'now')",
            "strftime('%Y-%m', m.updated_at) = strftime('%Y-%m', 'now')",
        ),
    }


def format_sales_amount(value: float) -> str:
    return format_money(value) or f"{_coerce_float(value, 0.0):,.2f}"


def upcoming_warranty_projection(conn, months_ahead: int = 6) -> pd.DataFrame:
    """Return a month-by-month projection of expiring active warranties."""

    try:
        months = int(months_ahead)
    except (TypeError, ValueError):
        months = 6
    months = max(1, min(months, 24))

    today = date.today()
    start_month = date(today.year, today.month, 1)
    last_bucket = add_months(start_month, months - 1)
    last_day = monthrange(last_bucket.year, last_bucket.month)[1]
    range_end = date(last_bucket.year, last_bucket.month, last_day)

    scope_clause, scope_params = customer_scope_filter("c")
    projection = df_query(
        conn,
        dedent(
            """
            SELECT strftime('%Y-%m', w.expiry_date) AS month_bucket,
                   COUNT(*) AS total
            FROM warranties w
            LEFT JOIN customers c ON c.customer_id = w.customer_id
            WHERE w.status='active'
              AND w.expiry_date IS NOT NULL
              AND date(w.expiry_date) BETWEEN date(?) AND date(?)
              {scope_filter}
            GROUP BY month_bucket
            ORDER BY month_bucket
            """
        ).format(scope_filter=f" AND {scope_clause}" if scope_clause else ""),
        params=(start_month.isoformat(), range_end.isoformat(), *scope_params),
    )

    records: list[dict[str, object]] = []
    current_bucket = start_month
    while len(records) < months:
        bucket_key = current_bucket.strftime("%Y-%m")
        label = current_bucket.strftime("%b %Y")
        matching = projection[projection["month_bucket"] == bucket_key]
        if matching.empty:
            count = 0
        else:
            count = int(matching.iloc[0]["total"] or 0)
        records.append({"Month": label, "Expiring warranties": count})
        current_bucket = add_months(current_bucket, 1)

    return pd.DataFrame(records)


def upcoming_warranty_breakdown(
    conn, days_ahead: int = 60, group_by: str = "sales_person"
) -> pd.DataFrame:
    """Summarise upcoming expiries grouped by a chosen dimension."""

    try:
        days = int(days_ahead)
    except (TypeError, ValueError):
        days = 60
    days = max(1, min(days, 365))

    grouping_options = {
        "sales_person": (
            "COALESCE(NULLIF(TRIM(c.sales_person), ''), 'Unassigned')",
            "Sales person",
        ),
        "customer": (
            "COALESCE(NULLIF(TRIM(c.name), ''), '(Unknown customer)')",
            "Customer",
        ),
        "product": (
            "COALESCE(NULLIF(TRIM(COALESCE(p.name, '') || CASE WHEN p.model IS NULL OR TRIM(p.model) = '' THEN '' ELSE ' ' || p.model END), ''), '(Unspecified product)')",
            "Product",
        ),
    }

    normalized_group = (group_by or "sales_person").lower()
    group_expr, column_label = grouping_options.get(
        normalized_group, grouping_options["sales_person"]
    )

    today = date.today()
    range_end = today + timedelta(days=days)

    scope_clause, scope_params = customer_scope_filter("c")
    scope_filter = f" AND {scope_clause}" if scope_clause else ""
    breakdown = df_query(
        conn,
        dedent(
            f"""
            SELECT {group_expr} AS bucket,
                   COUNT(*) AS total
            FROM warranties w
            LEFT JOIN customers c ON c.customer_id = w.customer_id
            LEFT JOIN products p ON p.product_id = w.product_id
            WHERE w.status='active'
              AND w.expiry_date IS NOT NULL
              AND date(w.expiry_date) BETWEEN date(?) AND date(?)
              AND p.name IS NOT NULL
              AND TRIM(p.name) != ''
              {scope_filter}
            GROUP BY bucket
            ORDER BY total DESC, bucket ASC
            """
        ),
        params=(today.isoformat(), range_end.isoformat(), *scope_params),
    )

    if breakdown.empty:
        return pd.DataFrame(columns=[column_label, "Expiring warranties"])

    renamed = breakdown.rename(
        columns={"bucket": column_label, "total": "Expiring warranties"}
    )
    renamed["Expiring warranties"] = renamed["Expiring warranties"].astype(int)
    return renamed


def clean_text(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    value = str(value).strip()
    return value or None


def _extract_text_from_quotation_upload(upload) -> tuple[str, list[str]]:
    """Return extracted text and warnings from an uploaded quotation file."""

    warnings: list[str] = []
    text_content = ""
    suffix = Path(upload.name).suffix.lower()
    file_bytes = upload.getvalue()

    def _ensure_ocr_engine_available() -> bool:
        global _OCR_ENGINE_AVAILABLE
        if _OCR_ENGINE_AVAILABLE is True:
            return True
        if _OCR_ENGINE_AVAILABLE is False:
            return False
        try:
            pytesseract.get_tesseract_version()
            _OCR_ENGINE_AVAILABLE = True
            return True
        except pytesseract.TesseractNotFoundError:
            warnings.append(
                "OCR engine unavailable. Install Tesseract OCR and ensure it is on your PATH."
            )
            _OCR_ENGINE_AVAILABLE = False
            return False
        except Exception as exc:  # pragma: no cover - defensive for OCR init failures
            warnings.append(f"OCR failed to initialize: {exc}")
            _OCR_ENGINE_AVAILABLE = False
            return False

    def _ocr_image(image: Image.Image) -> str:
        if not _ensure_ocr_engine_available():
            return ""

        try:
            grayscale = ImageOps.grayscale(image)
            boosted = ImageOps.autocontrast(grayscale)
            boosted = ImageEnhance.Contrast(boosted).enhance(1.8)
            boosted = ImageEnhance.Sharpness(boosted).enhance(1.2)
            primary_text = pytesseract.image_to_string(boosted)

            if len(primary_text.strip()) >= 12:
                return primary_text

            inverted = ImageOps.invert(grayscale)
            inverted = ImageEnhance.Contrast(inverted).enhance(1.6)
            alt_text = pytesseract.image_to_string(inverted)
            return alt_text if len(alt_text.strip()) > len(primary_text.strip()) else primary_text
        except pytesseract.TesseractNotFoundError:
            _OCR_ENGINE_AVAILABLE = False
            warnings.append(
                "OCR engine unavailable. Install Tesseract OCR and ensure it is on your PATH."
            )
            return ""
        except Exception as exc:  # pragma: no cover - defensive against OCR failures
            warnings.append(f"OCR failed: {exc}")
            return ""

    if suffix == ".pdf":
        pages: list[str] = []
        try:
            reader = PdfReader(io.BytesIO(file_bytes))
            for page in reader.pages:
                try:
                    text_block = page.extract_text() or ""
                except Exception:
                    text_block = ""
                image_ocr: list[str] = []
                for image_file in getattr(page, "images", []):
                    try:
                        pil_image = Image.open(io.BytesIO(image_file.data))
                        image_ocr.append(_ocr_image(pil_image))
                    except Exception:
                        continue
                if image_ocr:
                    text_block = "\n".join([text_block, *image_ocr]).strip()
                pages.append(text_block)
            text_content = "\n".join(pages)
            if not text_content.strip():
                warnings.append("No readable text found in the uploaded PDF.")
        except Exception as exc:  # pragma: no cover - defensive against damaged uploads
            warnings.append(f"Could not read PDF: {exc}")
        if len(text_content.strip()) < 80:
            try:
                from pdf2image import convert_from_bytes

                last_page = max(1, min(2, len(pages) or 2))
                raster_pages = convert_from_bytes(
                    file_bytes, dpi=300, first_page=1, last_page=last_page
                )
                extra_ocr = [_ocr_image(image) for image in raster_pages]
                combined = "\n".join([text_content, *extra_ocr]).strip()
                if combined:
                    text_content = combined
            except Exception as exc:  # pragma: no cover - optional dependency
                warnings.append(
                    "Add pdf2image with poppler to strengthen OCR for scanned PDFs"
                )
                if str(exc).strip():
                    warnings.append(f"OCR raster fallback failed: {exc}")
    else:
        try:
            image = Image.open(io.BytesIO(file_bytes))
        except Exception as exc:  # pragma: no cover - defensive against damaged uploads
            warnings.append(f"Could not open the uploaded image: {exc}")
            return "", warnings

        text_content = _ocr_image(image)

    return text_content, warnings


OCR_UPLOAD_SUFFIXES = {
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".gif",
    ".tif",
    ".tiff",
}


_OCR_ENGINE_AVAILABLE: Optional[bool] = None


def _ocr_uploads_enabled() -> bool:
    return bool(st.session_state.get("ocr_uploads_enabled", True))


def _run_upload_ocr(upload, *, key_prefix: str) -> tuple[str, list[str]]:
    if upload is None or not _ocr_uploads_enabled():
        return "", []
    suffix = Path(upload.name).suffix.lower() if getattr(upload, "name", None) else ""
    if suffix not in OCR_UPLOAD_SUFFIXES:
        return "", []
    token = f"{getattr(upload, 'name', '')}:{getattr(upload, 'size', '')}"
    token_key = f"{key_prefix}_ocr_token"
    text_key = f"{key_prefix}_ocr_text"
    warnings_key = f"{key_prefix}_ocr_warnings"
    if st.session_state.get(token_key) != token:
        text_content, warnings = _extract_text_from_quotation_upload(upload)
        st.session_state[token_key] = token
        st.session_state[text_key] = text_content
        st.session_state[warnings_key] = warnings
    text_content = clean_text(st.session_state.get(text_key)) or ""
    warnings = st.session_state.get(warnings_key, [])
    return text_content, warnings


def _show_ocr_warnings_once(warnings: list[str]) -> None:
    for warning in warnings:
        warning_key = f"ocr_warning_seen_{hash(warning)}"
        if st.session_state.get(warning_key):
            continue
        st.session_state[warning_key] = True
        st.caption(f"OCR notice: {warning}")


def _render_upload_ocr_preview(
    upload,
    *,
    key_prefix: str,
    label: str = "Auto-detected text (OCR)",
    show_preview: bool = False,
) -> tuple[str, list[str]]:
    text_content, warnings = _run_upload_ocr(upload, key_prefix=key_prefix)
    if warnings:
        _show_ocr_warnings_once(warnings)
    if not show_preview:
        return text_content, warnings
    if not text_content and not warnings:
        return text_content, warnings
    with st.expander(label):
        if warnings:
            for warning in warnings:
                st.warning(warning, icon="⚠️")
        if text_content:
            st.text_area(
                "Extracted text",
                value=text_content,
                height=200,
                key=f"{key_prefix}_ocr_text_area",
            )
    return text_content, warnings


def _items_blank(items: list[dict[str, object]], *, fields: tuple[str, ...]) -> bool:
    if not items:
        return True
    for item in items:
        for field in fields:
            value = item.get(field)
            if value not in (None, "", 0, 0.0):
                return False
    return True


def _apply_ocr_autofill(
    *,
    upload,
    ocr_key_prefix: str,
    doc_type: str,
    details_key_prefix: str,
) -> None:
    text_content, warnings = _run_upload_ocr(upload, key_prefix=ocr_key_prefix)
    if warnings:
        _show_ocr_warnings_once(warnings)
    if not text_content:
        return
    if doc_type == "Quotation":
        metadata = _extract_quotation_metadata(text_content)
        ref_key = f"{details_key_prefix}_quotation_reference"
        if clean_text(metadata.get("quotation_reference")) and not clean_text(
            st.session_state.get(ref_key)
        ):
            st.session_state[ref_key] = metadata.get("quotation_reference")
        date_key = f"{details_key_prefix}_quotation_date"
        parsed_date = metadata.get("quotation_date")
        if parsed_date and not st.session_state.get(date_key):
            st.session_state[date_key] = parsed_date
        items_key = f"{details_key_prefix}_quotation_items"
        detected_items = metadata.get("_detected_items") or []
        if detected_items:
            existing_items = st.session_state.get(items_key, [])
            if _items_blank(existing_items, fields=("description", "quantity", "rate")):
                mapped_items = []
                for item in detected_items:
                    mapped_items.append(
                        {
                            "description": item.get("description") or "",
                            "quantity": _coerce_float(item.get("quantity"), 1.0),
                            "rate": _coerce_float(item.get("rate"), 0.0),
                        }
                    )
                st.session_state[items_key] = mapped_items
    elif doc_type in ("Delivery order", "Work done", "Service", "Maintenance"):
        lines = [line.strip() for line in text_content.splitlines() if line.strip()]
        detected_items = _parse_line_items_from_text(lines)
        if doc_type in ("Delivery order", "Work done"):
            items_key = f"{details_key_prefix}_delivery_items"
            existing_items = st.session_state.get(items_key, [])
            if detected_items and _items_blank(
                existing_items, fields=("description", "quantity", "unit_price")
            ):
                mapped_items = []
                for item in detected_items:
                    mapped_items.append(
                        {
                            "description": item.get("description") or "",
                            "quantity": _coerce_float(item.get("quantity"), 1.0),
                            "unit_price": _coerce_float(item.get("rate"), 0.0),
                        }
                    )
                st.session_state[items_key] = mapped_items
        elif doc_type == "Service":
            items_key = f"{details_key_prefix}_service_items"
            existing_items = st.session_state.get(items_key, [])
            if detected_items and _items_blank(
                existing_items, fields=("description", "quantity", "unit_price")
            ):
                mapped_items = []
                for item in detected_items:
                    mapped_items.append(
                        {
                            "description": item.get("description") or "",
                            "quantity": _coerce_float(item.get("quantity"), 1.0),
                            "unit_price": _coerce_float(item.get("rate"), 0.0),
                        }
                    )
                st.session_state[items_key] = mapped_items
            date_key = f"{details_key_prefix}_service_date"
            parsed_date = _parse_date_from_text(text_content)
            if parsed_date and not st.session_state.get(date_key):
                st.session_state[date_key] = parsed_date
        else:
            items_key = f"{details_key_prefix}_maintenance_items"
            existing_items = st.session_state.get(items_key, [])
            if detected_items and _items_blank(
                existing_items, fields=("description", "quantity", "unit_price")
            ):
                mapped_items = []
                for item in detected_items:
                    mapped_items.append(
                        {
                            "description": item.get("description") or "",
                            "quantity": _coerce_float(item.get("quantity"), 1.0),
                            "unit_price": _coerce_float(item.get("rate"), 0.0),
                        }
                    )
                st.session_state[items_key] = mapped_items
            date_key = f"{details_key_prefix}_maintenance_date"
            parsed_date = _parse_date_from_text(text_content)
            if parsed_date and not st.session_state.get(date_key):
                st.session_state[date_key] = parsed_date


def _parse_date_from_text(value: str) -> Optional[date]:
    cleaned = clean_text(value)
    if not cleaned:
        return None
    cleaned = cleaned.replace(".", "-").replace("/", "-")
    candidates = re.findall(r"\d{1,2}-\d{1,2}-\d{2,4}", cleaned)
    if not candidates:
        return None
    for candidate in candidates:
        for fmt in ["%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d"]:
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    return None


def _parse_line_items_from_text(lines: list[str]) -> list[dict[str, object]]:
    """Extract probable line items from OCR'd inquiry text."""

    items: list[dict[str, object]] = []
    qty_pattern = re.compile(r"(?P<qty>\d+(?:\.\d+)?)\s*(pcs|nos|units|unit|qty|set|sets|pairs)?", re.IGNORECASE)
    rate_pattern = re.compile(r"(?:@|x|\b)\s*(?P<rate>\d{2,}(?:[\d,.]*\d)?)")
    model_pattern = re.compile(r"model[:\s]*([A-Za-z0-9-_.]+)", re.IGNORECASE)

    for raw_line in lines:
        line = raw_line.strip()
        if len(line.split()) < 2:
            continue

        qty_match = qty_pattern.search(line)
        model_match = model_pattern.search(line)
        rate_match = rate_pattern.search(line)

        qty_value = float(qty_match.group("qty")) if qty_match else 1.0
        rate_value = 0.0
        if rate_match:
            try:
                rate_value = float(rate_match.group("rate").replace(",", ""))
            except ValueError:
                rate_value = 0.0

        model_value = clean_text(model_match.group(1)) if model_match else None

        # Remove obvious numeric tokens when building description for clarity.
        cleaned_line = re.sub(r"\s{2,}", " ", line)
        description = cleaned_line
        if qty_match:
            description = description.replace(qty_match.group(0), "").strip(", -")
        if rate_match:
            description = description.replace(rate_match.group(0), "").strip(", -")
        if model_value:
            description = description.replace(model_value, "").strip(", -")
        description = description.strip() or cleaned_line

        if not description:
            continue

        items.append(
            {
                "description": description,
                "model": model_value or "",
                "quantity": qty_value,
                "rate": rate_value,
                "discount": 0.0,
            }
        )

    return items


def _extract_quotation_metadata(text: str) -> dict[str, object]:
    """Detect useful fields and probable items from uploaded quotation text."""

    if not text:
        return {}

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    normalized = "\n".join(lines)
    updates: dict[str, object] = {}

    def _match(patterns: Iterable[str]) -> Optional[str]:
        for pattern in patterns:
            match = re.search(pattern, normalized, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        return None

    def _header_value(prefix: str) -> Optional[str]:
        for line in lines:
            if line.lower().startswith(prefix.lower()):
                match = re.match(
                    rf"^{re.escape(prefix)}[:\s,\-]*?(.*)$",
                    line,
                    flags=re.IGNORECASE,
                )
                if not match:
                    continue
                candidate = match.group(1).strip()
                if candidate:
                    return candidate
        return None

    def _collect_after(keyword: str, max_lines: int = 6) -> list[str]:
        for idx, line in enumerate(lines):
            if line.lower().startswith(keyword.lower()):
                collected: list[str] = []
                for entry in lines[idx + 1 : idx + 1 + max_lines]:
                    if not entry.strip():
                        break
                    if re.match(
                        r"^(subject|attn\.?|attention|dear|ref\.?|reference|date|phone|tel|mobile|email)\b",
                        entry.strip(),
                        flags=re.IGNORECASE,
                    ):
                        break
                    collected.append(entry)
                return collected
        return []

    def _parse_amount(value: str) -> Optional[float]:
        match = re.search(r"([1-9]\d{1,2}(?:,\d{3})*(?:\.\d+)?)", value)
        if not match:
            return None
        try:
            return float(match.group(1).replace(",", ""))
        except ValueError:
            return None

    def _extract_price_schedule() -> list[dict[str, object]]:
        """Capture the generator line item from the uploaded quotation."""

        description_lines: list[str] = []
        price_hint: Optional[float] = None
        quantity_hint: float = 1.0

        for idx, line in enumerate(lines):
            if re.search(r"brand\s+new\s+diesel\s+generating\s+set", line, re.IGNORECASE):
                # Collect nearby descriptive lines until we hit an obvious footer.
                block: list[str] = [line]
                for entry in lines[idx + 1 : min(len(lines), idx + 12)]:
                    if re.search(r"total\s+amount", entry, re.IGNORECASE):
                        break
                    block.append(entry)
                description_lines = [entry for entry in block if entry.strip()]
                joined_block = " ".join(description_lines)
                price_hint = _parse_amount(joined_block)
                qty_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:set|pcs|piece|unit)s?", joined_block, re.IGNORECASE)
                if qty_match:
                    try:
                        quantity_hint = float(qty_match.group(1))
                    except ValueError:
                        quantity_hint = 1.0
                break

        if not description_lines:
            return []

        cleaned_description = "; ".join(description_lines)
        items: list[dict[str, object]] = [
            {
                "description": cleaned_description,
                "model": "",
                "quantity": quantity_hint,
                "rate": price_hint or 0.0,
                "discount": 0.0,
            }
        ]
        return items

    reference = _header_value("ref") or _match(
        [
            r"quotation\s*(?:no\.|number|#)[:#\s]*([\w/-]+)",
            r"quote\s*(?:no\.|number|#)[:#\s]*([\w/-]+)",
            r"reference[:#\s]*([\w/-]+)",
            r"ref[:#\s]*([\w/-]+)",
            r"(?:inquiry|enquiry)\s*(?:no\.|number|#)[:#\s]*([\w/-]+)",
        ]
    )
    if reference:
        updates["quotation_reference"] = reference

    subject_line = _match([r"subject[:\s]*([^\n]+)", r"scope[:\s]*([^\n]+)"])
    if subject_line:
        updates["quotation_subject"] = subject_line

    attention = _header_value("attention") or _match(
        [
            r"attn\.?[:\s]*([^\n]+)",
            r"attention[:\s]*([^\n]+)",
            r"dear\s+([^\n,]+)",
            r"contact[:\s]*([^\n,]+)",
        ]
    )
    if attention:
        updates["quotation_attention_name"] = attention
        updates.setdefault("quotation_customer_contact_name", attention)

    company = _header_value("to") or _match(
        [
            r"company[:\s]*([^\n]+)",
            r"to[:\s]*([^\n]+)",
            r"for[:\s]*([^\n]+)",
            r"customer[:\s]*([^\n]+)",
        ]
    )
    phone = _match([r"(?:phone|tel|mobile)[:\s]*([+\d][\d\s\-()]+)", r"(?:cell|contact)[:\s]*([+\d][\d\s\-()]+)"])
    if phone:
        updates["quotation_customer_contact"] = phone

    email = _match([r"email[:\s]*([^\n\s]+@[^\n\s]+)"])
    if email:
        updates["quotation_customer_email"] = email
        existing = updates.get("quotation_customer_contact")
        updates["quotation_customer_contact"] = (
            f"{existing}, {email}" if existing else email
        )

    address_block = _collect_after("to")
    address = _match([r"address[:\s]*([^\n]+(?:\n[^\n]+){0,3})", r"(?:office|site)\s*address[:\s]*([^\n]+)"])
    address_lines = []
    if address_block:
        if not company:
            company = address_block[0]
            address_lines = address_block[1:]
        else:
            if address_block and clean_text(address_block[0]) == clean_text(company):
                address_lines = address_block[1:]
            else:
                address_lines = address_block
    if company:
        updates["quotation_company_name"] = company
    if address_lines:
        updates["quotation_customer_address"] = "\n".join(address_lines)
    elif address:
        updates["quotation_customer_address"] = address.replace("\n", " ")

    detected_date = _header_value("date") or _match(
        [
            r"date[:\s]*([\d./\-]+)",
            r"valid\s*(?:until|till|up to)[:\s]*([\d./\-]+)",
        ]
    )
    parsed_date = _parse_date_from_text(detected_date) if detected_date else None
    if parsed_date:
        updates["quotation_date"] = parsed_date

    total_matches = re.findall(
        r"(?:grand\s*total|total\s*amount|sub\s*total|total)[^\d]{0,10}([\d,.]+)",
        normalized,
        flags=re.IGNORECASE,
    )
    parsed_totals: list[float] = []
    for raw in total_matches:
        cleaned = raw.replace(",", "")
        try:
            parsed_totals.append(float(cleaned))
        except ValueError:
            continue
    if parsed_totals:
        best_total = max(parsed_totals)
        updates["quotation_detected_total"] = best_total

    detected_items = _extract_price_schedule() or _parse_line_items_from_text(lines)
    if detected_items:
        updates["_detected_items"] = detected_items

    return updates


def _parse_sqlite_timestamp(value: Optional[str]) -> Optional[datetime]:
    text = clean_text(value)
    if not text:
        return None
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def format_time_ago(value: Optional[str]) -> str:
    timestamp = _parse_sqlite_timestamp(value)
    if not timestamp:
        return clean_text(value) or ""
    if timestamp.tzinfo is not None:
        timestamp = timestamp.replace(tzinfo=None)
    seconds = max(int((datetime.utcnow() - timestamp).total_seconds()), 0)
    if seconds < 5:
        return "just now"
    if seconds < 60:
        return f"{seconds}s ago"
    minutes = seconds // 60
    if minutes < 60:
        return f"{minutes}m ago"
    hours = minutes // 60
    if hours < 24:
        return f"{hours}h ago"
    days = hours // 24
    if days < 7:
        return f"{days}d ago"
    weeks = days // 7
    if weeks < 5:
        return f"{weeks}w ago"
    months = days // 30
    if months < 12:
        return f"{months}mo ago"
    years = days // 365
    return f"{years}y ago"


def _notification_store() -> list[dict[str, object]]:
    buffer = st.session_state.get(NOTIFICATION_BUFFER_KEY)
    if not isinstance(buffer, list):
        buffer = []
    st.session_state[NOTIFICATION_BUFFER_KEY] = buffer
    return buffer


def get_runtime_notifications() -> list[dict[str, object]]:
    return list(_notification_store())


def push_runtime_notification(
    title: Optional[str],
    message: Optional[str],
    *,
    severity: str = "info",
    details: Optional[Iterable[str]] = None,
) -> None:
    if not title and not message:
        return
    entry = {
        "title": clean_text(title) or "Notification",
        "message": clean_text(message) or "",
        "severity": (clean_text(severity) or "info").lower(),
        "timestamp": datetime.utcnow().isoformat(timespec="seconds"),
        "details": [
            clean_text(item) for item in (details or []) if clean_text(item)
        ],
    }
    buffer = _notification_store()
    buffer.append(entry)
    if len(buffer) > MAX_RUNTIME_NOTIFICATIONS:
        del buffer[0 : len(buffer) - MAX_RUNTIME_NOTIFICATIONS]
    st.session_state[NOTIFICATION_BUFFER_KEY] = buffer


def _build_staff_alerts(conn, *, user_id: Optional[int]) -> list[dict[str, object]]:
    alerts: list[dict[str, object]] = []
    if user_id is None:
        return alerts

    today_iso = date.today().isoformat()
    follow_ups = df_query(
        conn,
        dedent(
            """
            SELECT reference, follow_up_date, reminder_label
            FROM quotations
            WHERE created_by = ?
              AND follow_up_date IS NOT NULL
              AND deleted_at IS NULL
              AND LOWER(IFNULL(status, 'pending')) <> 'paid'
            ORDER BY date(follow_up_date) ASC
            LIMIT 12
            """
        ),
        (user_id,),
    )
    if not follow_ups.empty:
        for _, row in follow_ups.iterrows():
            follow_date_val = clean_text(row.get("follow_up_date"))
            follow_dt = pd.to_datetime(follow_date_val, errors="coerce")
            severity = "info"
            if pd.notna(follow_dt) and follow_dt.date() <= date.today():
                severity = "warning"
            follow_label = format_period_range(follow_date_val, follow_date_val) or (follow_date_val or "(date pending)")
            alerts.append(
                {
                    "title": clean_text(row.get("reference")) or "Quotation follow-up",
                    "message": clean_text(row.get("reminder_label"))
                    or f"Follow-up scheduled for {follow_label}",
                    "severity": severity,
                }
            )

    follow_up_reports = df_query(
        conn,
        dedent(
            """
            SELECT report_id, grid_payload
            FROM work_reports
            WHERE user_id=?
              AND LOWER(COALESCE(report_template, '')) = 'follow_up'
              AND grid_payload IS NOT NULL
              AND grid_payload != ''
            ORDER BY datetime(updated_at) DESC
            LIMIT 50
            """
        ),
        (user_id,),
    )
    follow_up_reminders: list[tuple[date, dict[str, object]]] = []
    if not follow_up_reports.empty:
        for _, row in follow_up_reports.iterrows():
            grid_rows = parse_report_grid_payload(
                row.get("grid_payload"), template_key="follow_up"
            )
            for entry in grid_rows:
                reminder_iso = to_iso_date(entry.get("reminder_date"))
                if not reminder_iso:
                    continue
                reminder_dt = pd.to_datetime(reminder_iso, errors="coerce")
                if pd.isna(reminder_dt):
                    continue
                reminder_date = reminder_dt.date()
                customer_label = (
                    clean_text(entry.get("client_name"))
                    or clean_text(entry.get("product_detail"))
                    or clean_text(entry.get("contact"))
                    or "Follow-up"
                )
                message = clean_text(entry.get("notes")) or "Follow-up reminder"
                follow_up_reminders.append(
                    (
                        reminder_date,
                        {
                            "title": customer_label,
                            "message": f"{message} (due {format_period_range(reminder_iso, reminder_iso)})",
                            "severity": "warning"
                            if reminder_date <= date.today()
                            else "info",
                        },
                    )
                )
    if follow_up_reminders:
        follow_up_reminders.sort(key=lambda item: item[0])
        for _, reminder in follow_up_reminders[:12]:
            alerts.append(reminder)

    return alerts


def _fetch_entity_activity(
    conn,
    entity_types: Iterable[str],
    *,
    user_filter: Optional[int] = None,
    limit: int = 30,
) -> pd.DataFrame:
    types = [clean_text(t) for t in entity_types if clean_text(t)]
    if not types:
        return pd.DataFrame()

    filters = [f"a.entity_type IN ({','.join('?' for _ in types)})"]
    params: list[object] = list(types)
    try:
        resolved_limit = max(1, min(int(limit), 200))
    except (TypeError, ValueError):
        resolved_limit = 30

    if user_filter is not None:
        filters.append("a.user_id = ?")
        params.append(int(user_filter))

    params.append(resolved_limit)

    where_clause = " AND ".join(filters)

    return df_query(
        conn,
        dedent(
            f"""
            SELECT a.activity_id,
                   a.entity_type,
                   a.event_type,
                   a.description,
                   a.created_at,
                   a.user_id,
                   COALESCE(u.username, 'Team member') AS actor
            FROM activity_log a
            LEFT JOIN users u ON u.user_id = a.user_id
            WHERE {where_clause}
            ORDER BY datetime(a.created_at) DESC, a.activity_id DESC
            LIMIT ?
            """
        ),
        tuple(params),
    )


def log_activity(
    conn,
    *,
    event_type: Optional[str],
    description: Optional[str],
    entity_type: Optional[str] = None,
    entity_id: Optional[int] = None,
    user_id: Optional[int] = None,
) -> None:
    event_key = clean_text(event_type)
    description_text = clean_text(description)
    if not event_key and not description_text:
        return
    actor_id = user_id if user_id is not None else current_user_id()
    actor_label = clean_text(get_current_user().get("username")) if get_current_user() else None
    label = NOTIFICATION_EVENT_LABELS.get(
        event_key, (event_key or "Activity").replace("_", " ").title()
    )
    should_notify = event_key not in {"login", "logout"}
    try:
        conn.execute(
            """
            INSERT INTO activity_log (user_id, event_type, entity_type, entity_id, description, created_at)
            VALUES (?, ?, ?, ?, ?, datetime('now'))
            """,
            (
                actor_id,
                event_key,
                clean_text(entity_type),
                entity_id,
                description_text or description or "",
            ),
        )
        conn.commit()
    except sqlite3.Error:
        with contextlib.suppress(Exception):
            conn.rollback()
    else:
        if should_notify:
            message = description_text or description or label or "Activity logged"
            details: list[str] = []
            if entity_type and entity_id:
                details.append(
                    f"{clean_text(entity_type).title()} #{entity_id}" if clean_text(entity_type) else f"Record #{entity_id}"
                )
            push_runtime_notification(
                label or "Activity logged",
                f"{actor_label or 'Team member'}: {message}",
                severity="info",
                details=details,
            )


def fetch_activity_feed(conn, limit: int = ACTIVITY_FEED_LIMIT) -> list[dict[str, object]]:
    try:
        resolved_limit = int(limit)
    except (TypeError, ValueError):
        resolved_limit = ACTIVITY_FEED_LIMIT
    resolved_limit = max(1, min(resolved_limit, 100))
    df = df_query(
        conn,
        dedent(
            """
            SELECT a.activity_id,
                   a.event_type,
                   a.entity_type,
                   a.entity_id,
                   a.description,
                   a.created_at,
                   u.username
            FROM activity_log a
            LEFT JOIN users u ON u.user_id = a.user_id
            ORDER BY datetime(a.created_at) DESC, a.activity_id DESC
            LIMIT ?
            """
        ),
        (resolved_limit,),
    )
    if df.empty:
        return []
    feed: list[dict[str, object]] = []
    for record in df.to_dict("records"):
        event_type = clean_text(record.get("event_type")) or "activity"
        label = NOTIFICATION_EVENT_LABELS.get(
            event_type, event_type.replace("_", " ").title()
        )
        feed.append(
            {
                "title": label,
                "message": clean_text(record.get("description")) or "",
                "timestamp": clean_text(record.get("created_at")) or "",
                "actor": clean_text(record.get("username")) or "Team member",
                "severity": "info",
                "event_type": event_type,
            }
        )
    return feed


def to_iso_date(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        value = stripped
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, datetime):
        return value.date().strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    try:
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    except Exception:
        return None
    if pd.isna(parsed):
        return None
    if isinstance(parsed, pd.DatetimeIndex):
        if len(parsed) == 0:
            return None
        parsed = parsed[0]
    return pd.Timestamp(parsed).normalize().strftime("%Y-%m-%d")


def format_money(value) -> Optional[str]:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    symbol = CURRENCY_SYMBOL.strip()
    if symbol:
        return f"{symbol} {amount:,.2f}"
    return f"{amount:,.2f}"


def format_amount_in_words(value: object) -> Optional[str]:
    amount = parse_amount(value)
    if amount is None:
        return None

    taka = int(amount)
    paisa = int(round((amount - taka) * 100))

    below_twenty = [
        "zero",
        "one",
        "two",
        "three",
        "four",
        "five",
        "six",
        "seven",
        "eight",
        "nine",
        "ten",
        "eleven",
        "twelve",
        "thirteen",
        "fourteen",
        "fifteen",
        "sixteen",
        "seventeen",
        "eighteen",
        "nineteen",
    ]
    tens_words = [
        "",
        "ten",
        "twenty",
        "thirty",
        "forty",
        "fifty",
        "sixty",
        "seventy",
        "eighty",
        "ninety",
    ]
    scales = ["", "thousand", "million", "billion", "trillion"]

    def _three_digit_words(number: int) -> list[str]:
        words: list[str] = []
        hundreds, remainder = divmod(number, 100)
        if hundreds:
            words.extend([below_twenty[hundreds], "hundred"])
            if remainder:
                words.append("and")
        if remainder:
            if remainder < 20:
                words.append(below_twenty[remainder])
            else:
                tens, ones = divmod(remainder, 10)
                words.append(tens_words[tens])
                if ones:
                    words.append(below_twenty[ones])
        return words

    def _number_to_words(number: int) -> str:
        if number == 0:
            return "zero"
        words: list[str] = []
        scale_index = 0
        while number > 0 and scale_index < len(scales):
            number, remainder = divmod(number, 1000)
            if remainder:
                chunk_words = _three_digit_words(remainder)
                if scales[scale_index]:
                    chunk_words.append(scales[scale_index])
                words = chunk_words + words
            scale_index += 1
        return " ".join(words)

    parts = [f"{_number_to_words(taka)} taka"]
    if paisa:
        parts.append(f"{_number_to_words(paisa)} paisa")

    return (" and ".join(parts)).capitalize()


def _coerce_float(value: object, default: float = 0.0) -> float:
    if isinstance(value, str):
        value = value.strip().replace(",", "")
        if value == "":
            return default
    try:
        number = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return default
    if not math.isfinite(number):
        return default
    return number


def parse_amount(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        amount = float(value)
        return round(amount, 2) if amount != 0 else amount
    text = clean_text(value)
    if not text:
        return None
    normalized = re.sub(r"[^0-9.\-]", "", text)
    if normalized in {"", ".", "-", "-."}:
        return None
    try:
        amount = float(normalized)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(amount):
        return None
    if math.isclose(amount, 0.0):
        return 0.0
    return round(amount, 2)


def parse_quantity(value, *, default: int = 1) -> int:
    if value is None:
        return default
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if not value:
            return default
    try:
        qty = float(value)
    except (TypeError, ValueError):
        return default
    if not math.isfinite(qty) or qty <= 0:
        return default
    return max(1, int(round(qty)))


def _format_editor_date(value: object) -> Optional[str]:
    iso_date = to_iso_date(value)
    if not iso_date:
        return None
    try:
        return datetime.fromisoformat(iso_date).strftime(DATE_FMT)
    except Exception:
        return iso_date


def _fetch_quotation_for_editor(conn, quotation_id: int) -> Optional[dict[str, object]]:
    cursor = conn.execute(
        dedent(
            """
            SELECT quotation_id, reference, quote_date, customer_name, customer_company,
                   customer_address, customer_district, customer_contact, attention_name,
                   attention_title, subject, salutation, introduction, closing, total_amount,
                   discount_pct, document_path, letter_template, salesperson_name,
                   salesperson_title, salesperson_contact, quote_type
              FROM quotations
             WHERE quotation_id=? AND deleted_at IS NULL
            """
        ),
        (quotation_id,),
    )
    row = cursor.fetchone()
    if not row:
        return None

    columns = [
        "quotation_id",
        "reference",
        "quote_date",
        "customer_name",
        "customer_company",
        "customer_address",
        "customer_district",
        "customer_contact",
        "attention_name",
        "attention_title",
        "subject",
        "salutation",
        "introduction",
        "closing",
        "total_amount",
        "discount_pct",
        "document_path",
        "letter_template",
        "salesperson_name",
        "salesperson_title",
        "salesperson_contact",
        "quote_type",
    ]
    return {col: row[idx] for idx, col in enumerate(columns)}


def _apply_editor_payload(conn, quotation_id: int, payload: Mapping[str, object]) -> bool:
    updates = {
        "reference": clean_text(payload.get("reference")),
        "quote_date": to_iso_date(payload.get("date")),
        "customer_company": clean_text(payload.get("customer_company")),
        "customer_name": clean_text(payload.get("customer_contact"))
        or clean_text(payload.get("attention")),
        "customer_address": clean_text(payload.get("address")),
        "attention_name": clean_text(payload.get("attention")),
        "subject": clean_text(payload.get("subject")),
        "salutation": clean_text(payload.get("salutation")),
        "introduction": clean_text(payload.get("introduction"))
        if payload.get("introduction")
        else None,
        "closing": clean_text(payload.get("closing")) if payload.get("closing") else None,
    }
    set_parts: list[str] = []
    values: list[object] = []
    for column, value in updates.items():
        set_parts.append(f"{column}=?")
        values.append(value)

    if not set_parts:
        return False

    values.append(quotation_id)
    cursor = conn.execute(
        f"UPDATE quotations SET {', '.join(set_parts)}, updated_at=datetime('now') WHERE quotation_id=? AND deleted_at IS NULL",
        tuple(values),
    )
    conn.commit()
    return cursor.rowcount > 0


def _build_editor_metadata(
    record: Mapping[str, object], payload: Mapping[str, object]
) -> OrderedDict:
    merged = dict(record)
    merged.update(
        {
            "quote_date": to_iso_date(payload.get("date")) or merged.get("quote_date"),
            "reference": clean_text(payload.get("reference")) or merged.get("reference"),
            "customer_company": clean_text(payload.get("customer_company"))
            or merged.get("customer_company"),
            "customer_address": payload.get("address") or merged.get("customer_address"),
            "customer_name": clean_text(payload.get("customer_contact"))
            or clean_text(payload.get("attention"))
            or merged.get("customer_name"),
            "attention_name": clean_text(payload.get("attention")) or merged.get("attention_name"),
            "subject": clean_text(payload.get("subject")) or merged.get("subject"),
            "salutation": clean_text(payload.get("salutation")) or merged.get("salutation"),
            "introduction": payload.get("introduction") or merged.get("introduction"),
            "closing": payload.get("closing") or merged.get("closing"),
        }
    )

    metadata = OrderedDict()
    metadata["Reference number"] = merged.get("reference")
    metadata["Date"] = _format_editor_date(merged.get("quote_date"))
    metadata["Customer contact name"] = merged.get("customer_name")
    metadata["Customer company"] = merged.get("customer_company")
    metadata["Customer address"] = merged.get("customer_address")
    metadata["Customer district"] = merged.get("customer_district")
    metadata["Customer contact"] = merged.get("customer_contact")
    metadata["Attention name"] = merged.get("attention_name")
    metadata["Attention title"] = merged.get("attention_title")
    metadata["Subject"] = merged.get("subject")
    metadata["Salutation"] = merged.get("salutation") or "Dear Sir,"
    metadata["Introduction"] = merged.get("introduction")
    metadata["Quote type"] = merged.get("quote_type") or "Quotation"
    metadata["Closing / thanks"] = merged.get("closing")
    metadata["Salesperson name"] = merged.get("salesperson_name")
    metadata["Salesperson title"] = merged.get("salesperson_title")
    metadata["Salesperson contact"] = merged.get("salesperson_contact")
    return metadata


def _generate_editor_pdf(
    quotation_id: int, payload: Mapping[str, object]
) -> Optional[bytes]:
    conn = get_conn()
    try:
        record = _fetch_quotation_for_editor(conn, quotation_id)
    finally:
        conn.close()

    if not record:
        return None

    metadata = _build_editor_metadata(record, payload)
    grand_total = _coerce_float(
        payload.get("total_amount") or record.get("total_amount"), 0.0
    )
    totals = {
        "grand_total": grand_total,
        "discount_total": 0.0,
        "gross_total": grand_total,
    }
    grand_total_label = format_money(grand_total) or f"{grand_total:,.2f}"
    grand_total_words = format_amount_in_words(grand_total)
    template_choice = clean_text(record.get("letter_template")) or "PS letterhead"

    return _build_quotation_pdf(
        metadata=metadata,
        items=[],
        totals=totals,
        grand_total_label=grand_total_label,
        template_choice=template_choice,
        grand_total_words=grand_total_words,
    )


class _QuotationEditorRequestHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, directory: str | None = None, **kwargs):  # type: ignore[override]
        super().__init__(*args, directory=directory, **kwargs)

    def log_message(self, format: str, *args):  # pragma: no cover - reduce noise
        return

    def _parse_quotation_id(self, path_parts: list[str]) -> Optional[int]:
        if len(path_parts) < 3:
            return None
        try:
            quotation_id = int(path_parts[2])
        except (TypeError, ValueError):
            return None
        return quotation_id if quotation_id > 0 else None

    def _send_json(self, status: int, payload: Mapping[str, object]):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_json(self) -> Optional[dict[str, object]]:
        try:
            length = int(self.headers.get("Content-Length") or 0)
        except (TypeError, ValueError):
            length = 0
        data = self.rfile.read(length) if length > 0 else b""
        if not data:
            return None
        try:
            parsed = json.loads(data.decode("utf-8"))
        except json.JSONDecodeError:
            return None
        if isinstance(parsed, dict):
            return parsed
        return None

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):  # noqa: N802 - following BaseHTTPRequestHandler signature
        parsed = urllib.parse.urlparse(self.path)
        path_parts = parsed.path.strip("/").split("/")
        if len(path_parts) >= 2 and path_parts[0] == "api" and path_parts[1] == "quotation":
            quotation_id = self._parse_quotation_id(path_parts)
            if quotation_id is None:
                self._send_json(400, {"error": "Invalid quotation ID"})
                return

            conn = get_conn()
            try:
                record = _fetch_quotation_for_editor(conn, quotation_id)
            finally:
                conn.close()

            if not record:
                self._send_json(404, {"error": "Quotation not found"})
                return

            payload = {
                "quotation_id": quotation_id,
                "date": to_iso_date(record.get("quote_date")),
                "reference": clean_text(record.get("reference")) or "",
                "customer_company": clean_text(record.get("customer_company")) or "",
                "attention": clean_text(record.get("attention_name")) or "",
                "subject": clean_text(record.get("subject")) or "",
                "address": record.get("customer_address") or "",
                "salutation": clean_text(record.get("salutation")) or "Dear Sir,",
                "introduction": record.get("introduction") or "",
                "closing": record.get("closing") or "",
                "customer_contact": clean_text(record.get("customer_name")) or "",
            }
            self._send_json(200, payload)
            return

        return super().do_GET()

    def do_POST(self):  # noqa: N802 - following BaseHTTPRequestHandler signature
        parsed = urllib.parse.urlparse(self.path)
        path_parts = parsed.path.strip("/").split("/")
        if len(path_parts) >= 3 and path_parts[0] == "api" and path_parts[1] == "quotation":
            quotation_id = self._parse_quotation_id(path_parts)
            if quotation_id is None:
                self._send_json(400, {"error": "Invalid quotation ID"})
                return

            payload = self._read_json() or {}
            if len(path_parts) >= 4 and path_parts[3] == "save":
                conn = get_conn()
                try:
                    updated = _apply_editor_payload(conn, quotation_id, payload)
                finally:
                    conn.close()

                if not updated:
                    self._send_json(404, {"error": "Quotation not found or unchanged"})
                    return
                self._send_json(200, {"status": "ok", "quotation_id": quotation_id})
                return

            if len(path_parts) >= 4 and path_parts[3] == "pdf":
                conn = get_conn()
                try:
                    _apply_editor_payload(conn, quotation_id, payload)
                finally:
                    conn.close()

                pdf_bytes = _generate_editor_pdf(quotation_id, payload)
                if not pdf_bytes:
                    self._send_json(404, {"error": "Quotation not found"})
                    return

                filename = f"quotation_{quotation_id}.pdf"
                self.send_response(200)
                self.send_header("Content-Type", "application/pdf")
                self.send_header(
                    "Content-Disposition", f"attachment; filename={filename}"
                )
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Content-Length", str(len(pdf_bytes)))
                self.end_headers()
                self.wfile.write(pdf_bytes)
                return

        self._send_json(404, {"error": "Unsupported endpoint"})


def _ensure_quotation_editor_server():
    global _quotation_editor_server, _quotation_editor_thread
    if _quotation_editor_server is not None:
        return

    try:
        handler_cls = partial(
            _QuotationEditorRequestHandler, directory=str(PROJECT_ROOT.resolve())
        )
        server = http.server.ThreadingHTTPServer(
            ("0.0.0.0", QUOTATION_EDITOR_PORT), handler_cls
        )
    except OSError:
        return

    _quotation_editor_server = server
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    _quotation_editor_thread = thread


def format_period_label(period_type: str) -> str:
    if not period_type:
        return "Unknown"
    key = str(period_type).strip().lower()
    return REPORT_PERIOD_OPTIONS.get(key, key.title())


def format_period_range(start: Optional[str], end: Optional[str]) -> str:
    def _label(value: Optional[str]) -> Optional[str]:
        if not value:
            return None
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        if isinstance(parsed, pd.DatetimeIndex):
            if len(parsed) == 0:
                return None
            parsed = parsed[0]
        return pd.Timestamp(parsed).strftime(DATE_FMT)

    start_label = _label(start)
    end_label = _label(end)
    if start_label and end_label:
        if start_label == end_label:
            return start_label
        return f"{start_label} → {end_label}"
    return start_label or end_label or "—"


def _clamp_percentage(value: float) -> float:
    if value < 0.0:
        return 0.0
    if value > 100.0:
        return 100.0
    return value


def _value_or_default(value: object, default: object) -> object:
    if value is None:
        return default
    try:
        if pd.isna(value):  # type: ignore[arg-type]
            return default
    except Exception:
        pass
    if isinstance(value, str) and value.strip() == "":
        return default
    return value


def normalize_product_entries(
    entries: Iterable[dict[str, object]]
) -> tuple[list[dict[str, object]], list[str]]:
    cleaned: list[dict[str, object]] = []
    labels: list[str] = []
    for entry in entries:
        name_clean = clean_text(entry.get("name")) if isinstance(entry, dict) else None
        model_clean = clean_text(entry.get("model")) if isinstance(entry, dict) else None
        serial_clean = clean_text(entry.get("serial")) if isinstance(entry, dict) else None
        quantity_raw = entry.get("quantity") if isinstance(entry, dict) else None
        price_raw = None
        if isinstance(entry, dict):
            price_raw = entry.get("unit_price") if "unit_price" in entry else entry.get("price")
        qty_val = _coerce_float(quantity_raw, 1.0)
        try:
            qty_val_int = int(round(qty_val))
        except Exception:
            qty_val_int = 1
        qty_val = max(qty_val_int, 1)
        unit_price = max(_coerce_float(price_raw, 0.0), 0.0)
        line_total = unit_price * qty_val if unit_price else None
        if not any([name_clean, model_clean, serial_clean]):
            continue
        cleaned.append(
            {
                "name": name_clean,
                "model": model_clean,
                "serial": serial_clean,
                "quantity": qty_val,
                "unit_price": unit_price if unit_price else None,
                "total": line_total if line_total else None,
            }
        )
        label_parts = [val for val in [name_clean, model_clean] if val]
        label = " - ".join(label_parts)
        if qty_val > 1:
            label = f"{label} ×{qty_val}" if label else f"×{qty_val}"
        if unit_price:
            price_block = f"Tk {unit_price:,.2f}"
            if line_total:
                price_block = f"{price_block} (Total Tk {line_total:,.2f})"
            label = f"{label} @ {price_block}" if label else price_block
        if serial_clean:
            label = f"{label} (Serial: {serial_clean})" if label else f"Serial: {serial_clean}"
        if label:
            labels.append(label)
    return cleaned, labels


def normalize_quotation_items(
    entries: Iterable[dict[str, object]]
) -> tuple[list[dict[str, object]], dict[str, float]]:
    cleaned: list[dict[str, object]] = []
    totals = {
        "gross_total": 0.0,
        "discount_total": 0.0,
        "grand_total": 0.0,
    }

    for entry in entries:
        if not isinstance(entry, dict):
            continue

        description = clean_text(entry.get("description") or entry.get("Description"))
        if not description:
            continue

        kva = clean_text(entry.get("kva"))
        model = clean_text(entry.get("model"))
        specs = clean_text(entry.get("specs")) or model
        note = clean_text(entry.get("note"))
        hsn = clean_text(entry.get("hsn"))
        unit = clean_text(entry.get("unit"))

        quantity = max(_coerce_float(entry.get("quantity"), 0.0), 0.0)
        rate = max(_coerce_float(entry.get("rate"), 0.0), 0.0)
        discount_pct = _clamp_percentage(_coerce_float(entry.get("discount"), 0.0))

        gross_amount = quantity * rate
        discount_amount = gross_amount * (discount_pct / 100.0)
        override_total = _coerce_float(entry.get("total_price"), None)
        line_total = (
            override_total
            if override_total is not None and override_total >= 0
            else max(gross_amount - discount_amount, 0.0)
        )

        description_label = dedupe_join([description, model], " – ") or description

        item = {
            "Sl No.": len(cleaned) + 1,
            "Description": description,
            "Description of Generator": description_label,
            "Quantity": quantity,
            "Qty.": quantity,
            "Unit": unit,
            "HSN/SAC": hsn,
            "KVA": kva,
            "Specs": specs,
            "Notes": note,
            "Rate": rate,
            "Unit Price, Tk.": rate,
            "Gross amount": gross_amount,
            "Discount (%)": discount_pct,
            "Discount amount": discount_amount,
            "Line total": line_total,
            "Total Price, Tk.": line_total,
        }

        # Remove optional empty fields to keep downstream tables tidy
        for optional_key in ["Unit", "HSN/SAC", "KVA", "Specs", "Notes"]:
            if not item.get(optional_key):
                item.pop(optional_key, None)

        cleaned.append(item)

        totals["gross_total"] += gross_amount
        totals["discount_total"] += discount_amount
        totals["grand_total"] += line_total

    return cleaned, totals


def format_period_span(
    start: Optional[str], end: Optional[str], *, joiner: str = " → "
) -> Optional[str]:
    start_clean = clean_text(start)
    end_clean = clean_text(end)
    if not start_clean and not end_clean:
        return None
    if start_clean and end_clean:
        if start_clean == end_clean:
            return start_clean
        return f"{start_clean}{joiner}{end_clean}"
    return start_clean or end_clean


def get_status_choice(prefix: str, fallback: str = DEFAULT_SERVICE_STATUS) -> str:
    choice = st.session_state.get(f"{prefix}_status_choice", fallback)
    if isinstance(choice, str) and choice in SERVICE_STATUS_OPTIONS:
        return choice
    return fallback


def ensure_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime().date()
    except Exception:
        pass
    try:
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    except Exception:
        parsed = None
    if parsed is None or pd.isna(parsed):
        return None
    if isinstance(parsed, pd.DatetimeIndex) and len(parsed) > 0:
        parsed = parsed[0]
    if isinstance(parsed, datetime):
        return parsed.date()
    try:
        return parsed.to_pydatetime().date()
    except Exception:
        return None


def determine_period_dates(
    status_choice: str, raw_value
) -> tuple[Optional[date], Optional[date], Optional[date]]:
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    if status_choice == "Completed":
        raw_items: list[Optional[date]]
        if isinstance(raw_value, (list, tuple)):
            raw_items = [ensure_date(v) for v in raw_value]
        else:
            raw_items = [ensure_date(raw_value)]
        clean_items = [item for item in raw_items if item is not None]
        if clean_items:
            start_date = clean_items[0]
            end_date = clean_items[-1]
            if end_date is None:
                end_date = start_date
            if start_date and end_date and end_date < start_date:
                start_date, end_date = end_date, start_date
    else:
        if isinstance(raw_value, (list, tuple)):
            raw_value = raw_value[0] if raw_value else None
        start_date = ensure_date(raw_value)
        end_date = None
    primary_date = start_date or end_date
    return primary_date, start_date, end_date


def determine_period_strings(
    status_choice: str, raw_value
) -> tuple[Optional[str], Optional[str], Optional[str]]:
    primary_date, start_date, end_date = determine_period_dates(status_choice, raw_value)

    def _to_str(value: Optional[date]) -> Optional[str]:
        return value.strftime("%Y-%m-%d") if value else None

    return _to_str(primary_date), _to_str(start_date), _to_str(end_date)


def is_pending_status(status: Optional[str]) -> bool:
    text = clean_text(status)
    if not text:
        return True
    normalized = text.lower()
    return normalized not in {"completed", "in progress"}


def status_input_widget(prefix: str, default_status: Optional[str] = None) -> str:
    lookup = {opt.lower(): opt for opt in SERVICE_STATUS_OPTIONS}
    default_choice = DEFAULT_SERVICE_STATUS
    custom_default = "Haven't started"
    default_clean = clean_text(default_status)
    if default_clean:
        normalized = default_clean.lower()
        if normalized in lookup and lookup[normalized] != "Haven't started":
            default_choice = lookup[normalized]
        elif normalized == "haven't started":
            default_choice = lookup[normalized]
            custom_default = lookup[normalized]
        else:
            default_choice = "Haven't started"
            custom_default = default_clean

    choice = st.selectbox(
        "Status",
        SERVICE_STATUS_OPTIONS,
        index=SERVICE_STATUS_OPTIONS.index(default_choice),
        key=f"{prefix}_status_choice",
    )
    if choice == "Haven't started":
        custom_value = st.text_input(
            "Custom status label",
            value=custom_default or "Haven't started",
            key=f"{prefix}_status_custom",
            help="Customize the saved status when a record hasn't started yet.",
        )
        return clean_text(custom_value) or "Haven't started"
    return choice


def link_delivery_order_to_customer(
    conn: sqlite3.Connection, do_number: Optional[str], customer_id: Optional[int]
) -> None:
    do_serial = clean_text(do_number)
    if not do_serial:
        return
    cur = conn.cursor()
    row = cur.execute(
        "SELECT customer_id FROM delivery_orders WHERE do_number = ? AND deleted_at IS NULL",
        (do_serial,),
    ).fetchone()
    if row is None:
        if customer_id is not None:
            cur.execute(
                "UPDATE customers SET delivery_order_code = ? WHERE customer_id = ?",
                (do_serial, int(customer_id)),
            )
        return
    previous_customer = int(row[0]) if row and row[0] is not None else None
    if customer_id is not None:
        cur.execute(
            "UPDATE delivery_orders SET customer_id = ? WHERE do_number = ?",
            (int(customer_id), do_serial),
        )
        cur.execute(
            "UPDATE customers SET delivery_order_code = ? WHERE customer_id = ?",
            (do_serial, int(customer_id)),
        )
    else:
        cur.execute(
            "UPDATE delivery_orders SET customer_id = NULL WHERE do_number = ?",
            (do_serial,),
        )
    if previous_customer and previous_customer != (int(customer_id) if customer_id is not None else None):
        cur.execute(
            "UPDATE customers SET delivery_order_code = NULL WHERE customer_id = ? AND delivery_order_code = ?",
            (previous_customer, do_serial),
        )


def _safe_rerun():
    try:
        st.rerun()
    except Exception:
        try:
            st.experimental_rerun()
        except Exception:
            pass


def _default_new_customer_products() -> list[dict[str, object]]:
    return [
        {
            "name": "",
            "model": "",
            "serial": "",
            "quantity": 1,
            "unit_price": 0.0,
        }
    ]


def _reset_new_customer_form_state() -> None:
    default_products = _default_new_customer_products()
    st.session_state["new_customer_products_rows"] = default_products
    for key in [
        "new_customer_name",
        "new_customer_company",
        "new_customer_phone",
        "new_customer_address",
        "new_customer_delivery_address",
        "new_customer_purchase_date",
        "new_customer_purchase_date_enabled",
        "new_customer_do_code",
        "new_customer_sales_person",
        "new_customer_remarks",
        "new_customer_amount_spent",
        "new_customer_pdf",
        "new_customer_create_delivery_order",
        "new_customer_create_work_done",
        "new_customer_work_done_number",
        "new_customer_work_done_notes",
        "new_customer_work_done_pdf",
        "new_customer_create_service",
        "new_customer_service_date",
        "new_customer_service_description",
        "new_customer_create_maintenance",
        "new_customer_maintenance_date",
        "new_customer_maintenance_description",
        "new_customer_products_table",
    ]:
        st.session_state.pop(key, None)


def _default_quotation_items() -> list[dict[str, object]]:
    return [
        {
            "description": "",
            "quantity": 1.0,
            "rate": 0.0,
            "total_price": 0.0,
        }
    ]


def _default_delivery_items() -> list[dict[str, object]]:
    return [
        {
            "description": "",
            "quantity": 1.0,
            "unit_price": 0.0,
            "discount": 0.0,
        }
    ]


def _default_simple_items() -> list[dict[str, object]]:
    return [
        {
            "description": "",
            "quantity": 1.0,
            "unit_price": 0.0,
        }
    ]


def _products_to_delivery_items(products: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    for entry in products or []:
        if not isinstance(entry, dict):
            continue
        name = clean_text(entry.get("name"))
        model = clean_text(entry.get("model"))
        description = " ".join(part for part in [name, model] if part).strip()
        if not description:
            continue
        items.append(
            {
                "description": description,
                "quantity": _coerce_float(entry.get("quantity"), 1.0),
                "unit_price": _coerce_float(entry.get("unit_price"), 0.0),
                "discount": 0.0,
            }
        )
    return items


def normalize_delivery_items(rows: Iterable[dict[str, object]]) -> tuple[list[dict[str, object]], float]:
    normalized: list[dict[str, object]] = []
    total_amount = 0.0
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        description = clean_text(row.get("description"))
        if not description:
            continue
        quantity = _coerce_float(row.get("quantity"), 1.0)
        if quantity < 0:
            quantity = 0.0
        unit_price = _coerce_float(row.get("unit_price"), 0.0)
        if unit_price < 0:
            unit_price = 0.0
        discount = _coerce_float(row.get("discount"), 0.0)
        if discount < 0:
            discount = 0.0
        if discount > 100:
            discount = 100.0
        line_total = quantity * unit_price * (1 - (discount / 100))
        total_amount += line_total
        normalized.append(
            {
                "description": description,
                "quantity": quantity,
                "unit_price": unit_price,
                "discount": discount,
                "line_total": line_total,
            }
        )
    return normalized, total_amount


def normalize_simple_items(rows: Iterable[dict[str, object]]) -> tuple[list[dict[str, object]], float]:
    normalized: list[dict[str, object]] = []
    total_amount = 0.0
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        description = clean_text(row.get("description"))
        if not description:
            continue
        quantity = max(_coerce_float(row.get("quantity"), 1.0), 0.0)
        unit_price = max(_coerce_float(row.get("unit_price"), 0.0), 0.0)
        line_total = quantity * unit_price
        total_amount += line_total
        normalized.append(
            {
                "description": description,
                "quantity": quantity,
                "unit_price": unit_price,
                "line_total": line_total,
            }
        )
    return normalized, total_amount


def format_simple_item_labels(items: Iterable[dict[str, object]]) -> list[str]:
    labels: list[str] = []
    for entry in items or []:
        if not isinstance(entry, dict):
            continue
        description = clean_text(entry.get("description"))
        if not description:
            continue
        quantity = _coerce_float(entry.get("quantity"), 1.0)
        unit_price = _coerce_float(entry.get("unit_price"), 0.0)
        label = description
        if quantity:
            qty_label = int(quantity) if math.isclose(quantity, round(quantity)) else quantity
            label = f"{label} ×{qty_label}"
        if unit_price:
            label = f"{label} @ Tk {unit_price:,.2f}"
        labels.append(label)
    return labels


def parse_delivery_items_payload(value: Optional[str]) -> list[dict[str, object]]:
    text = clean_text(value)
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except (TypeError, ValueError):
        return []
    rows: list[dict[str, object]] = []
    if isinstance(parsed, list):
        for entry in parsed:
            if not isinstance(entry, dict):
                continue
            rows.append(
                {
                    "description": clean_text(entry.get("description")) or "",
                    "quantity": _coerce_float(entry.get("quantity"), 1.0),
                    "unit_price": _coerce_float(entry.get("unit_price"), 0.0),
                    "discount": _coerce_float(entry.get("discount"), 0.0),
                    "line_total": _coerce_float(entry.get("line_total"), 0.0),
                }
            )
    return rows


def _reset_quotation_form_state() -> None:
    default_items = _default_quotation_items()
    st.session_state["quotation_item_rows"] = default_items
    st.session_state["quotation_preview_item_rows"] = []
    st.session_state["quotation_preview_items_dirty"] = True
    st.session_state.pop("quotation_form_initialized", None)
    for key in [
        "quotation_reference",
        "quotation_date",
        "quotation_prepared_by",
        "quotation_valid_days",
        "quotation_company_name",
        "quotation_company_details",
        "quotation_customer_name",
        "quotation_customer_contact",
        "quotation_customer_address",
        "quotation_project_name",
        "quotation_subject",
        "quotation_scope_notes",
        "quotation_terms",
        "quotation_status",
        "quotation_follow_up_status",
        "quotation_follow_up_notes",
        "quotation_follow_up_date",
        "quotation_follow_up_date_toggle",
        "quotation_follow_up_choice",
        "quotation_salesperson_title",
        "quotation_salesperson_contact",
        "quotation_attention_name",
        "quotation_attention_title",
        "quotation_salutation",
        "quotation_introduction",
        "quotation_closing",
        "quotation_quote_type",
        "quotation_customer_district",
        "quotation_customer_district_select",
        "quotation_letter_template",
        "quotation_admin_notes",
        "quotation_reminder_label",
        "quotation_customer_contact_name",
        "quotation_receipt_upload",
        "quotation_document_path",
        "quotation_payment_receipt_path",
        "quotation_manual_total",
    ]:
        st.session_state.pop(key, None)
    st.session_state.pop("quotation_result", None)


def _streamlit_runtime_active() -> bool:
    """Return True when running inside a Streamlit runtime."""

    runtime = None
    try:
        from streamlit import runtime as st_runtime

        runtime = st_runtime
    except Exception:
        runtime = None

    if runtime is not None:
        try:
            if runtime.exists():
                return True
        except Exception:
            pass

    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
    except Exception:
        return False

    try:
        return get_script_run_ctx() is not None
    except Exception:
        return False


def ensure_upload_dirs():
    for path in (
        UPLOADS_DIR,
        DELIVERY_ORDER_DIR,
        SERVICE_DOCS_DIR,
        MAINTENANCE_DOCS_DIR,
        CUSTOMER_DOCS_DIR,
        OPERATIONS_OTHER_DIR,
        SERVICE_BILL_DIR,
        REPORT_DOCS_DIR,
        QUOTATION_RECEIPT_DIR,
        QUOTATION_DOCS_DIR,
        DELIVERY_RECEIPT_DIR,
    ):
        path.mkdir(parents=True, exist_ok=True)


def _read_uploaded_bytes(uploaded_file) -> bytes:
    if uploaded_file is None:
        return b""
    data = b""
    if hasattr(uploaded_file, "getvalue"):
        try:
            data = uploaded_file.getvalue()
        except Exception:
            data = b""
    if data:
        return data
    if hasattr(uploaded_file, "seek"):
        try:
            uploaded_file.seek(0)
        except Exception:
            pass
    try:
        return uploaded_file.read()
    except Exception:
        return b""


def save_uploaded_file(
    uploaded_file,
    target_dir: Path,
    filename: Optional[str] = None,
    *,
    allowed_extensions: Optional[set[str]] = None,
    default_extension: str = ".pdf",
) -> Optional[Path]:
    if uploaded_file is None:
        return None
    ensure_upload_dirs()
    raw_name = filename or uploaded_file.name or "upload"
    raw_name = "".join(ch for ch in raw_name if ch.isalnum() or ch in (".", "_", "-"))
    stem = Path(raw_name).stem or "upload"
    ext = Path(raw_name).suffix.lower()

    if allowed_extensions is not None:
        normalized_allowed = {val.lower() for val in allowed_extensions}
        if ext not in normalized_allowed:
            ext = default_extension if default_extension.startswith(".") else f".{default_extension}"
        safe_name = f"{stem}{ext}"
    else:
        if not ext:
            ext = default_extension if default_extension.startswith(".") else f".{default_extension}"
        safe_name = f"{stem}{ext}"
    dest = target_dir / safe_name
    counter = 1
    while dest.exists():
        stem = dest.stem
        suffix = dest.suffix
        dest = target_dir / f"{stem}_{counter}{suffix}"
        counter += 1
    data = _read_uploaded_bytes(uploaded_file)
    if not data:
        return None
    with open(dest, "wb") as fh:
        fh.write(data)
    return dest


def store_uploaded_pdf(uploaded_file, target_dir: Path, filename: Optional[str] = None) -> Optional[str]:
    """Persist an uploaded PDF and return its path relative to ``BASE_DIR``.

    Streamlit's ``UploadedFile`` objects expose a ``read`` method and ``name``
    attribute. This helper mirrors ``save_uploaded_file`` but normalises the
    resulting path so callers can safely stash it in the database without
    worrying about absolute paths or platform differences.
    """

    saved_path = save_uploaded_file(uploaded_file, target_dir, filename=filename)
    if not saved_path:
        return None
    try:
        return str(saved_path.relative_to(BASE_DIR))
    except ValueError:
        return str(saved_path)


def store_uploaded_document(
    uploaded_file,
    target_dir: Path,
    *,
    filename_stem: str,
    allowed_extensions: Optional[set[str]] = None,
    default_extension: str = ".pdf",
) -> Optional[str]:
    """Persist an uploaded document (PDF or image) and return its path relative to ``BASE_DIR``."""

    if uploaded_file is None:
        return None
    suffix = Path(getattr(uploaded_file, "name", "") or "").suffix.lower()
    filename = f"{filename_stem}{suffix}" if suffix else filename_stem
    saved_path = save_uploaded_file(
        uploaded_file,
        target_dir,
        filename=filename,
        allowed_extensions=allowed_extensions or DOCUMENT_UPLOAD_EXTENSIONS,
        default_extension=default_extension,
    )
    if not saved_path:
        return None
    try:
        return str(saved_path.relative_to(BASE_DIR))
    except ValueError:
        return str(saved_path)


def store_payment_receipt(
    uploaded_file,
    *,
    identifier: Optional[str] = None,
    target_dir: Path = QUOTATION_RECEIPT_DIR,
) -> Optional[str]:
    """Persist an uploaded receipt (PDF or image) for paid records."""

    if uploaded_file is None:
        return None
    filename = identifier or uploaded_file.name or "receipt"
    saved_path = save_uploaded_file(
        uploaded_file,
        target_dir,
        filename=filename,
        allowed_extensions={".pdf", ".png", ".jpg", ".jpeg", ".webp"},
        default_extension=".pdf",
    )
    if not saved_path:
        return None
    try:
        return str(saved_path.relative_to(BASE_DIR))
    except ValueError:
        return str(saved_path)


def store_report_attachment(uploaded_file, *, identifier: Optional[str] = None) -> Optional[str]:
    """Persist a supporting document for a work report."""

    if uploaded_file is None:
        return None

    ensure_upload_dirs()
    raw_name = uploaded_file.name or "attachment"
    allowed_exts = {".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif", ".xlsx", ".xls"}
    suffix = Path(raw_name).suffix.lower()
    if suffix not in allowed_exts:
        suffix = ".pdf"
    stem = Path(raw_name).stem
    safe_stem = "".join(ch for ch in stem if ch.isalnum() or ch in ("_", "-")) or "attachment"
    safe_stem = safe_stem.strip("_") or "attachment"
    if identifier:
        ident = "".join(ch for ch in identifier if ch.isalnum() or ch in ("_", "-"))
        if ident:
            safe_stem = f"{ident}_{safe_stem}"
    dest = REPORT_DOCS_DIR / f"{safe_stem}{suffix}"
    counter = 1
    while dest.exists():
        dest = REPORT_DOCS_DIR / f"{safe_stem}_{counter}{suffix}"
        counter += 1
    data = _read_uploaded_bytes(uploaded_file)
    if not data:
        return None
    with open(dest, "wb") as fh:
        fh.write(data)
    try:
        return str(dest.relative_to(BASE_DIR))
    except ValueError:
        return str(dest)


def store_report_import_file(
    filename: str,
    payload: bytes,
    *,
    identifier: Optional[str] = None,
) -> Optional[str]:
    """Persist a report import spreadsheet for later download."""

    if not payload:
        return None

    ensure_upload_dirs()
    raw_name = filename or "report_import"
    allowed_exts = {".xlsx", ".xls", ".csv"}
    suffix = Path(raw_name).suffix.lower()
    if suffix not in allowed_exts:
        suffix = ".xlsx"
    stem = Path(raw_name).stem
    safe_stem = "".join(ch for ch in stem if ch.isalnum() or ch in ("_", "-")) or "report_import"
    safe_stem = safe_stem.strip("_") or "report_import"
    if identifier:
        ident = "".join(ch for ch in identifier if ch.isalnum() or ch in ("_", "-"))
        if ident:
            safe_stem = f"{ident}_{safe_stem}"
    dest = REPORT_DOCS_DIR / f"{safe_stem}{suffix}"
    counter = 1
    while dest.exists():
        dest = REPORT_DOCS_DIR / f"{safe_stem}_{counter}{suffix}"
        counter += 1
    with open(dest, "wb") as fh:
        fh.write(payload)
    try:
        return str(dest.relative_to(BASE_DIR))
    except ValueError:
        return str(dest)


def resolve_upload_path(path_str: Optional[str]) -> Optional[Path]:
    if not path_str:
        return None
    path = Path(path_str)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path


_ATTACHMENT_UNCHANGED = object()


def normalize_report_window(period_type: str, start_value, end_value) -> tuple[str, date, date]:
    """Return a canonical report period and date window."""

    key = (period_type or "").strip().lower()
    if key not in REPORT_PERIOD_OPTIONS:
        key = "daily"

    def _coerce(value) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        iso = to_iso_date(value)
        if iso:
            try:
                return datetime.strptime(iso, "%Y-%m-%d").date()
            except ValueError:
                pass
        try:
            parsed = pd.to_datetime(value, errors="coerce")
        except Exception:
            parsed = None
        if parsed is None or pd.isna(parsed):
            return None
        if isinstance(parsed, pd.DatetimeIndex):
            if len(parsed) == 0:
                return None
            parsed = parsed[0]
        if hasattr(parsed, "to_pydatetime"):
            parsed = parsed.to_pydatetime()
        if isinstance(parsed, datetime):
            return parsed.date()
        if isinstance(parsed, date):
            return parsed
        return None

    start_date = _coerce(start_value)
    end_date = _coerce(end_value)

    if key == "daily":
        anchor = start_date or end_date
        if anchor is None:
            raise ValueError("Select a date for the daily report.")
        start_date = end_date = anchor
    elif key == "weekly":
        anchor = start_date or end_date
        if anchor is None:
            raise ValueError("Select a week for the report.")
        start_date = anchor - timedelta(days=anchor.weekday())
        end_date = start_date + timedelta(days=6)
    else:
        anchor = start_date or end_date
        if anchor is None:
            raise ValueError("Select a month for the report.")
        start_date = anchor.replace(day=1)
        last_day = monthrange(start_date.year, start_date.month)[1]
        end_date = date(start_date.year, start_date.month, last_day)

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    return key, start_date, end_date


def _sanitize_path_component(value: Optional[str]) -> str:
    if not value:
        return "item"
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_.() ")
    cleaned = "".join(ch if ch in allowed else "_" for ch in str(value))
    cleaned = cleaned.strip()
    return cleaned or "item"


def _upload_extension(uploaded_file, *, default: str = ".pdf") -> str:
    ext = Path(getattr(uploaded_file, "name", "") or "").suffix.lower()
    if not ext:
        ext = default if default.startswith(".") else f".{default}"
    return ext


def build_customer_groups(conn, only_complete: bool = True):
    criteria = []
    params: list[object] = []
    if only_complete:
        criteria.append(customer_complete_clause())
    scope_clause, scope_params = customer_scope_filter()
    if scope_clause:
        criteria.append(scope_clause)
        params.extend(scope_params)
    where_clause = f"WHERE {' AND '.join(criteria)}" if criteria else ""
    df = df_query(
        conn,
        f"SELECT customer_id, TRIM(name) AS name FROM customers {where_clause}",
        tuple(params),
    )
    if df.empty:
        return [], {}
    df["name"] = df["name"].fillna("")
    df["norm_name"] = df["name"].astype(str).str.strip()
    df.sort_values(by=["norm_name", "customer_id"], inplace=True)
    groups = []
    label_by_id = {}
    for norm_name, group in df.groupby("norm_name", sort=False):
        ids = group["customer_id"].astype(int).tolist()
        primary = ids[0]
        raw_name = clean_text(group.iloc[0].get("name"))
        count = len(ids)
        base_label = raw_name or f"Customer #{primary}"
        if raw_name and count > 1:
            display_label = f"{base_label} ({count} records)"
        else:
            display_label = base_label
        groups.append(
            {
                "norm_name": norm_name,
                "primary_id": primary,
                "ids": ids,
                "raw_name": raw_name,
                "label": display_label,
                "count": count,
            }
        )
        for cid in ids:
            label_by_id[int(cid)] = display_label
    groups.sort(key=lambda g: (g["norm_name"] or "").lower())
    return groups, label_by_id


def fetch_customer_choices(conn, *, only_complete: bool = True):
    groups, label_by_id = build_customer_groups(conn, only_complete=only_complete)
    options = [None]
    labels = {None: "-- Select customer --"}
    group_map = {}
    for group in groups:
        primary = group["primary_id"]
        options.append(primary)
        labels[primary] = group["label"]
        group_map[primary] = group["ids"]
    return options, labels, group_map, label_by_id


def attach_documents(
    conn,
    table: str,
    fk_column: str,
    record_id: int,
    files,
    target_dir: Path,
    prefix: str,
    *,
    allowed_extensions: Optional[set[str]] = None,
    default_extension: str = ".pdf",
):
    if not files:
        return 0
    saved = 0
    try:
        cols = {row[1] for row in conn.execute(f"PRAGMA table_info({table})")}
    except Exception:
        cols = set()
    include_uploader = "uploaded_by" in cols
    uploader_id = current_user_id()
    for idx, uploaded in enumerate(files, start=1):
        if uploaded is None:
            continue
        original_name = uploaded.name or f"{prefix}_{idx}.pdf"
        safe_original = Path(original_name).name
        filename = f"{prefix}_{idx}_{safe_original}"
        stored_path = None
        if allowed_extensions:
            saved_path = save_uploaded_file(
                uploaded,
                target_dir,
                filename=filename,
                allowed_extensions=allowed_extensions,
                default_extension=default_extension,
            )
            if saved_path:
                try:
                    stored_path = str(saved_path.relative_to(BASE_DIR))
                except ValueError:
                    stored_path = str(saved_path)
        else:
            stored_path = store_uploaded_pdf(uploaded, target_dir, filename=filename)
        if not stored_path:
            continue
        if include_uploader:
            conn.execute(
                f"INSERT INTO {table} ({fk_column}, file_path, original_name, uploaded_by) VALUES (?, ?, ?, ?)",
                (int(record_id), stored_path, safe_original, uploader_id),
            )
        else:
            conn.execute(
                f"INSERT INTO {table} ({fk_column}, file_path, original_name) VALUES (?, ?, ?)",
                (int(record_id), stored_path, safe_original),
            )
        saved += 1
    return saved


def bundle_documents_zip(documents):
    if not documents:
        return None
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for doc in documents:
            path = doc.get("path")
            archive_name = doc.get("archive_name")
            if not path or not archive_name:
                continue
            if not path.exists():
                continue
            zf.write(path, archive_name)
    buffer.seek(0)
    return buffer


def bundle_customer_package(
    documents: list[dict[str, object]],
    summary_pdf: bytes,
    summary_name: str,
) -> io.BytesIO:
    buffer = io.BytesIO()
    safe_name = _sanitize_path_component(summary_name) or "customer"
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        if summary_pdf:
            zf.writestr(f"{safe_name}_summary.pdf", summary_pdf)
        for doc in documents:
            path = doc.get("path")
            archive_name = doc.get("archive_name")
            if not path or not archive_name:
                continue
            if not path.exists():
                continue
            zf.write(path, archive_name)
    buffer.seek(0)
    return buffer


def dedupe_join(values: Iterable[Optional[str]], joiner: str = ", ") -> str:
    seen = []
    for value in values:
        if value is None:
            continue
        val = str(value).strip()
        if not val:
            continue
        if val not in seen:
            seen.append(val)
    return joiner.join(seen)


def join_with_counts(values: Iterable[Optional[str]]) -> str:
    """Join values with commas, annotating duplicates with counts."""

    counts: dict[str, int] = {}
    order: list[str] = []
    for value in values:
        if value is None:
            continue
        cleaned = str(value).strip()
        if not cleaned:
            continue
        counts[cleaned] = counts.get(cleaned, 0) + 1
        if cleaned not in order:
            order.append(cleaned)

    parts: list[str] = []
    for item in order:
        count = counts.get(item, 0)
        if count > 1:
            parts.append(f"{item} (x{count})")
        else:
            parts.append(item)
    return ", ".join(parts)


def merge_customer_records(conn, customer_ids) -> bool:
    ids = []
    for cid in customer_ids:
        cid_int = int_or_none(cid)
        if cid_int is not None and cid_int not in ids:
            ids.append(cid_int)
    if len(ids) < 2:
        return False

    placeholders = ",".join(["?"] * len(ids))
    query = dedent(
        f"""
        SELECT customer_id, name, company_name, phone, address, delivery_address, remarks, purchase_date, product_info, delivery_order_code, sales_person, created_at
        FROM customers
        WHERE customer_id IN ({placeholders})
        """
    )
    df = df_query(conn, query, params=tuple(ids))
    if df.empty:
        return False

    df["created_at_dt"] = pd.to_datetime(df.get("created_at"), errors="coerce")
    df.sort_values(by=["created_at_dt", "customer_id"], inplace=True, na_position="last")
    base_row = df.iloc[0]
    base_id = int(base_row.get("customer_id"))
    other_ids = []
    for row in df.get("customer_id", pd.Series(dtype=object)).tolist():
        rid = int_or_none(row)
        if rid is not None and rid != base_id and rid not in other_ids:
            other_ids.append(rid)
    if not other_ids:
        return False

    name_values = [clean_text(v) for v in df.get("name", pd.Series(dtype=object)).tolist()]
    name_values = [v for v in name_values if v]
    company_values = [clean_text(v) for v in df.get("company_name", pd.Series(dtype=object)).tolist()]
    company_values = [v for v in company_values if v]
    address_values = [clean_text(v) for v in df.get("address", pd.Series(dtype=object)).tolist()]
    address_values = [v for v in address_values if v]
    delivery_values = [clean_text(v) for v in df.get("delivery_address", pd.Series(dtype=object)).tolist()]
    delivery_values = [v for v in delivery_values if v]
    remarks_values = [clean_text(v) for v in df.get("remarks", pd.Series(dtype=object)).tolist()]
    remarks_values = [v for v in remarks_values if v]
    phone_values = [clean_text(v) for v in df.get("phone", pd.Series(dtype=object)).tolist()]
    phone_values = [v for v in phone_values if v]
    phones_to_recalc: set[str] = set(phone_values)

    base_name = clean_text(base_row.get("name")) or (name_values[0] if name_values else None)
    base_company = clean_text(base_row.get("company_name")) or (company_values[0] if company_values else None)
    base_address = clean_text(base_row.get("address")) or (address_values[0] if address_values else None)
    base_delivery_address = clean_text(base_row.get("delivery_address")) or (delivery_values[0] if delivery_values else None)
    combined_remarks = dedupe_join(remarks_values)
    base_phone = clean_text(base_row.get("phone")) or (phone_values[0] if phone_values else None)

    do_codes = []
    product_lines = []
    fallback_products = []
    purchase_dates = []
    purchase_labels = []
    sales_people = []

    for record in df.to_dict("records"):
        date_raw = clean_text(record.get("purchase_date"))
        product_raw = clean_text(record.get("product_info"))
        do_raw = clean_text(record.get("delivery_order_code"))
        sales_raw = clean_text(record.get("sales_person"))
        if do_raw:
            do_codes.append(do_raw)
        if product_raw:
            fallback_products.append(product_raw)
        dt = parse_date_value(record.get("purchase_date"))
        if dt is not None:
            purchase_dates.append(dt)
            date_label = dt.strftime(DATE_FMT)
        else:
            date_label = date_raw
        if date_label:
            purchase_labels.append(date_label)
        if date_label and product_raw:
            product_lines.append(f"{date_label} – {product_raw}")
        elif product_raw:
            product_lines.append(product_raw)
        elif date_label:
            product_lines.append(date_label)
        if sales_raw:
            sales_people.append(sales_raw)

    earliest_purchase = min(purchase_dates).strftime("%Y-%m-%d") if purchase_dates else None
    combined_products = join_with_counts(product_lines or fallback_products)
    combined_do_codes = join_with_counts(do_codes)
    combined_sales = join_with_counts(sales_people)
    combined_purchase_labels = join_with_counts(purchase_labels)

    conn.execute(
        """
        UPDATE customers
        SET name=?, company_name=?, phone=?, address=?, delivery_address=?, remarks=?, purchase_date=?, product_info=?, delivery_order_code=?, sales_person=?, dup_flag=0
        WHERE customer_id=?
        """,
        (
            base_name,
            base_company,
            base_phone,
            base_address,
            base_delivery_address,
            clean_text(combined_remarks),
            earliest_purchase,
            clean_text(combined_products or combined_purchase_labels),
            clean_text(combined_do_codes),
            clean_text(combined_sales),
            base_id,
        ),
    )

    related_tables = (
        "orders",
        "warranties",
        "delivery_orders",
        "services",
        "maintenance_records",
        "needs",
    )
    for cid in other_ids:
        for table in related_tables:
            conn.execute(f"UPDATE {table} SET customer_id=? WHERE customer_id=?", (base_id, cid))
        conn.execute("UPDATE import_history SET customer_id=? WHERE customer_id=?", (base_id, cid))
        conn.execute("DELETE FROM customers WHERE customer_id=?", (cid,))

    if base_phone:
        phones_to_recalc.add(base_phone)
    if phones_to_recalc:
        for phone in phones_to_recalc:
            recalc_customer_duplicate_flag(conn, phone)
    conn.commit()
    return True


def auto_merge_matching_customers(conn) -> bool:
    """Automatically merge customers sharing the same name and address."""

    df = df_query(
        conn,
        dedent(
            """
            SELECT customer_id, name, company_name, phone, address, delivery_address,
                   purchase_date, product_info, delivery_order_code, sales_person, remarks, created_at
            FROM customers
            WHERE TRIM(COALESCE(name, '')) <> '' AND TRIM(COALESCE(address, '')) <> ''
            """
        ),
    )
    if df.empty:
        return False

    def _normalize(value: object) -> str:
        cleaned = clean_text(value) or ""
        return " ".join(cleaned.lower().split())

    df["_name_norm"] = df.get("name", pd.Series(dtype=object)).apply(_normalize)
    df["_address_norm"] = df.get("address", pd.Series(dtype=object)).apply(_normalize)

    merged_any = False
    grouped = df.groupby(["_name_norm", "_address_norm"], dropna=False)
    for _, group in grouped:
        ids = [int(cid) for cid in group.get("customer_id", []) if int_or_none(cid) is not None]
        if len(ids) < 2:
            continue
        if merge_customer_records(conn, ids):
            merged_any = True

    return merged_any


def delete_customer_record(conn, customer_id: int) -> None:
    """Delete a customer and related records, recalculating duplicate flags."""

    try:
        cid = int(customer_id)
    except (TypeError, ValueError):
        return

    cur = conn.execute(
        "SELECT name, phone, delivery_order_code, attachment_path FROM customers WHERE customer_id=?",
        (cid,),
    )
    row = cur.fetchone()
    if not row:
        return

    name_val = clean_text(row[0])
    phone_val = clean_text(row[1])
    do_code = clean_text(row[2])
    attachment_path = row[3]

    conn.execute("DELETE FROM customers WHERE customer_id=?", (cid,))
    if do_code:
        conn.execute(
            "DELETE FROM delivery_orders WHERE do_number=? AND (customer_id IS NULL OR customer_id=?)",
            (do_code, cid),
        )
    conn.execute(
        "UPDATE import_history SET deleted_at = datetime('now') WHERE customer_id=? AND deleted_at IS NULL",
        (cid,),
    )
    conn.commit()

    if phone_val:
        recalc_customer_duplicate_flag(conn, phone_val)
        conn.commit()

    if attachment_path:
        path = resolve_upload_path(attachment_path)
        if path and path.exists():
            try:
                path.unlink()
            except Exception:
                pass

    summary_bits: list[str] = []
    if name_val:
        summary_bits.append(name_val)
    if phone_val:
        summary_bits.append(f"phone {phone_val}")
    description = "; ".join(summary_bits) or f"ID #{cid}"
    log_activity(
        conn,
        event_type="customer_deleted",
        description=f"Deleted customer {description}",
        entity_type="customer",
        entity_id=cid,
    )


def collapse_warranty_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    work = df.copy()
    work["description"] = work.apply(
        lambda row: dedupe_join(
            [
                clean_text(row.get("product")),
                clean_text(row.get("model")),
                clean_text(row.get("serial")),
            ]
        ),
        axis=1,
    )
    if "remarks" in work.columns:
        work["remarks_clean"] = work["remarks"].apply(clean_text)
    else:
        work["remarks_clean"] = None
    if "staff" in work.columns:
        work["staff_clean"] = work["staff"].apply(clean_text)
    elif "sales_person" in work.columns:
        work["staff_clean"] = work["sales_person"].apply(clean_text)
    else:
        work["staff_clean"] = None
    issue_dt = pd.to_datetime(work.get("issue_date"), errors="coerce")
    expiry_dt = pd.to_datetime(work.get("expiry_date"), errors="coerce")
    work["issue_fmt"] = issue_dt.dt.strftime(DATE_FMT)
    work.loc[issue_dt.isna(), "issue_fmt"] = None
    work["expiry_fmt"] = expiry_dt.dt.strftime(DATE_FMT)
    work.loc[expiry_dt.isna(), "expiry_fmt"] = None
    work["expiry_dt"] = expiry_dt

    grouped = (
        work.groupby("customer", dropna=False)
        .apply(
            lambda g: pd.Series(
                {
                    "description": dedupe_join(g["description"].tolist()),
                    "issue_date": dedupe_join(g["issue_fmt"].tolist()),
                    "expiry_date": dedupe_join(g["expiry_fmt"].tolist()),
                    "remarks": dedupe_join(g["remarks_clean"].tolist()),
                    "staff": dedupe_join(g.get("staff_clean", pd.Series(dtype=object)).tolist()),
                    "_sort": g["expiry_dt"].min(),
                }
            )
        )
        .reset_index()
    )
    grouped = grouped.sort_values("_sort", na_position="last").drop(columns=["_sort"])
    grouped.rename(
        columns={
            "customer": "Customer",
            "description": "Description",
            "issue_date": "Issue date",
            "expiry_date": "Expiry date",
            "remarks": "Remarks",
            "staff": "Staff",
        },
        inplace=True,
    )
    if "Customer" in grouped.columns:
        grouped["Customer"] = grouped["Customer"].fillna("(unknown)")
    return grouped


def _build_customers_export(conn) -> pd.DataFrame:
    scope_clause, scope_params = customer_scope_filter("c")
    where_sql = f"WHERE {scope_clause}" if scope_clause else ""
    query = dedent(
        f"""
        SELECT c.customer_id,
               c.name,
               c.phone,
               c.address,
               c.amount_spent,
               c.purchase_date,
               c.product_info,
               c.delivery_order_code,
               c.sales_person,
               c.created_at,
               COALESCE(u.username, '(unknown)') AS uploaded_by
        FROM customers c
        LEFT JOIN users u ON u.user_id = c.created_by
        {where_sql}
        ORDER BY datetime(c.created_at) DESC, c.customer_id DESC
        """
    )
    df = df_query(conn, query, scope_params if scope_clause else ())
    df = fmt_dates(df, ["purchase_date", "created_at"])
    return df.rename(
        columns={
            "customer_id": "Customer ID",
            "name": "Customer",
            "phone": "Phone",
            "address": "Address",
            "amount_spent": "Amount spent",
            "purchase_date": "Purchase date",
            "product_info": "Product info",
            "delivery_order_code": "Delivery order",
            "sales_person": "Sales person",
            "created_at": "Created at",
            "uploaded_by": "Uploaded by",
        }
    )


def _build_delivery_orders_export(conn) -> pd.DataFrame:
    query = dedent(
        """
        SELECT d.do_number,
               COALESCE(c.name, '(unknown)') AS customer,
               d.description,
               d.sales_person,
                d.remarks,
                d.created_at
        FROM delivery_orders d
        LEFT JOIN customers c ON c.customer_id = d.customer_id
        WHERE COALESCE(d.record_type, 'delivery_order') = 'delivery_order'
          AND d.deleted_at IS NULL
        ORDER BY datetime(d.created_at) DESC, d.do_number DESC
        """
    )
    df = df_query(conn, query)
    df = fmt_dates(df, ["created_at"])
    return df.rename(
        columns={
            "do_number": "DO number",
            "customer": "Customer",
            "description": "Description",
            "sales_person": "Sales person",
            "remarks": "Remarks",
            "created_at": "Created at",
        }
    )


def _build_warranties_export(conn) -> pd.DataFrame:
    query = dedent(
        """
        SELECT w.warranty_id,
               COALESCE(c.name, '(unknown)') AS customer,
               COALESCE(p.name, '') AS product,
               p.model,
               w.serial,
               w.issue_date,
               w.expiry_date,
               w.status,
               w.remarks
        FROM warranties w
        LEFT JOIN customers c ON c.customer_id = w.customer_id
        LEFT JOIN products p ON p.product_id = w.product_id
        ORDER BY date(w.expiry_date) ASC, w.warranty_id ASC
        """
    )
    df = df_query(conn, query)
    df = fmt_dates(df, ["issue_date", "expiry_date"])
    if "status" in df.columns:
        df["status"] = df["status"].fillna("Active").apply(lambda x: str(x).title())
    return df.rename(
        columns={
            "warranty_id": "Warranty ID",
            "customer": "Customer",
            "product": "Product",
            "model": "Model",
            "serial": "Serial",
            "issue_date": "Issue date",
            "expiry_date": "Expiry date",
            "status": "Status",
            "remarks": "Remarks",
        }
    )


def _build_services_export(conn) -> pd.DataFrame:
    query = dedent(
        """
        SELECT s.service_id,
               s.do_number,
               COALESCE(c.name, cdo.name, '(unknown)') AS customer,
               s.service_date,
               s.service_start_date,
               s.service_end_date,
               s.service_product_info,
               s.description,
               s.status,
               s.remarks,
               s.condition_status,
               s.condition_remarks,
               s.bill_amount,
               s.bill_document_path,
               s.updated_at
        FROM services s
        LEFT JOIN customers c ON c.customer_id = s.customer_id
        LEFT JOIN delivery_orders d ON d.do_number = s.do_number
        LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
        WHERE s.deleted_at IS NULL
        ORDER BY datetime(s.service_date) DESC, s.service_id DESC
        """
    )
    df = df_query(conn, query)
    df = fmt_dates(df, ["service_date", "service_start_date", "service_end_date", "updated_at"])
    if "status" in df.columns:
        df["status"] = df["status"].apply(lambda x: clean_text(x) or DEFAULT_SERVICE_STATUS)
    return df.rename(
        columns={
            "service_id": "Service ID",
            "do_number": "DO number",
            "customer": "Customer",
            "service_date": "Service date",
            "service_start_date": "Service start date",
            "service_end_date": "Service end date",
            "service_product_info": "Products sold",
            "description": "Description",
            "status": "Status",
            "remarks": "Remarks",
            "condition_status": "Condition",
            "condition_remarks": "Condition notes",
            "bill_amount": "Bill amount",
            "bill_document_path": "Bill document",
            "updated_at": "Updated at",
        }
    )


def _build_maintenance_export(conn) -> pd.DataFrame:
    query = dedent(
        """
        SELECT m.maintenance_id,
               m.do_number,
               COALESCE(c.name, cdo.name, '(unknown)') AS customer,
               m.maintenance_date,
               m.maintenance_start_date,
               m.maintenance_end_date,
               m.maintenance_product_info,
               m.description,
               m.status,
               m.remarks,
               m.updated_at
        FROM maintenance_records m
        LEFT JOIN customers c ON c.customer_id = m.customer_id
        LEFT JOIN delivery_orders d ON d.do_number = m.do_number
        LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
        WHERE m.deleted_at IS NULL
        ORDER BY datetime(m.maintenance_date) DESC, m.maintenance_id DESC
        """
    )
    df = df_query(conn, query)
    df = fmt_dates(df, ["maintenance_date", "maintenance_start_date", "maintenance_end_date", "updated_at"])
    if "status" in df.columns:
        df["status"] = df["status"].apply(lambda x: clean_text(x) or DEFAULT_SERVICE_STATUS)
    return df.rename(
        columns={
            "maintenance_id": "Maintenance ID",
            "do_number": "DO number",
            "customer": "Customer",
            "maintenance_date": "Maintenance date",
            "maintenance_start_date": "Maintenance start date",
            "maintenance_end_date": "Maintenance end date",
            "maintenance_product_info": "Products sold",
            "description": "Description",
            "status": "Status",
            "remarks": "Remarks",
            "updated_at": "Updated at",
        }
    )


def _build_quotations_export(conn) -> pd.DataFrame:
    scope_clause, scope_params = _quotation_scope_filter()
    query = dedent(
        f"""
        SELECT quotation_id,
               reference,
               quote_date,
               customer_company,
               customer_name,
               customer_contact,
               customer_address,
               customer_district,
               attention_name,
               subject,
               total_amount,
               discount_pct,
               status,
               follow_up_status,
               follow_up_notes,
               follow_up_date,
               reminder_label,
               payment_receipt_path,
               salesperson_name,
               salesperson_title,
               salesperson_contact,
               remarks_internal,
               created_at,
               updated_at
        FROM quotations
        {scope_clause}
        ORDER BY datetime(quote_date) DESC, quotation_id DESC
        """
    )
    df = df_query(conn, query, scope_params)
    df = fmt_dates(df, ["quote_date", "follow_up_date", "created_at", "updated_at"])
    return df.rename(
        columns={
            "quotation_id": "Quotation ID",
            "reference": "Reference",
            "quote_date": "Quote date",
            "customer_company": "Customer",
            "customer_name": "Contact name",
            "customer_contact": "Contact details",
            "customer_address": "Address",
            "customer_district": "District",
            "attention_name": "Attention",
            "subject": "Subject",
            "total_amount": "Total amount",
            "discount_pct": "Discount (%)",
            "status": "Status",
            "follow_up_status": "Follow-up status",
            "follow_up_notes": "Follow-up notes",
            "follow_up_date": "Follow-up date",
            "reminder_label": "Reminder",
            "payment_receipt_path": "Receipt path",
            "salesperson_name": "Salesperson",
            "salesperson_title": "Salesperson title",
            "salesperson_contact": "Salesperson contact",
            "remarks_internal": "Internal remarks",
            "created_at": "Created at",
            "updated_at": "Updated at",
        }
    )


def _build_master_sheet(sheets: list[tuple[str, pd.DataFrame]]) -> pd.DataFrame:
    rows = [
        {
            "Sheet": "Export generated at",
            "Details": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    ]
    for sheet_name, df in sheets:
        count = len(df.index) if df is not None else 0
        label = "record" if count == 1 else "records"
        rows.append({"Sheet": sheet_name, "Details": f"{count} {label}"})
    return pd.DataFrame(rows, columns=["Sheet", "Details"])


def get_theme() -> str:
    theme = clean_text(st.session_state.get("theme")) or "light"
    return "dark" if theme.lower().startswith("dark") else "light"


def set_theme(value: object) -> None:
    if isinstance(value, str):
        normalized = value.strip().lower()
        theme = "dark" if normalized.startswith("dark") else "light"
    else:
        theme = "dark" if bool(value) else "light"
    st.session_state["theme"] = theme


def apply_theme_css() -> None:
    theme = get_theme()
    if theme == "dark":
        colors = {
            "bg": "#0e1117",
            "sidebar_bg": "#0b1220",
            "panel_bg": "#0f172a",
            "panel_border": "#1f2937",
            "text": "#f8fafc",
            "muted": "#94a3b8",
            "input_bg": "#111827",
            "input_border": "#334155",
            "accent": "#38bdf8",
            "button_bg": "#111827",
            "button_border": "#334155",
            "button_text": "#f8fafc",
            "button_hover": "#1f2937",
            "button_primary_bg": "#38bdf8",
            "button_primary_text": "#0b1220",
            "button_primary_hover": "#0ea5e9",
            "metric_bg": "rgba(255, 255, 255, 0.04)",
            "metric_border": "rgba(250, 250, 250, 0.12)",
            "table_header_bg": "#111827",
            "table_row_alt_bg": "rgba(148, 163, 184, 0.08)",
        }
    else:
        colors = {
            "bg": "#ffffff",
            "sidebar_bg": "#f8f9fb",
            "panel_bg": "#ffffff",
            "panel_border": "#e5e7eb",
            "text": "#111827",
            "muted": "#6b7280",
            "input_bg": "#ffffff",
            "input_border": "#d1d5db",
            "accent": "#1d3b64",
            "button_bg": "#ffffff",
            "button_border": "#d1d5db",
            "button_text": "#111827",
            "button_hover": "#f3f4f6",
            "button_primary_bg": "#1d3b64",
            "button_primary_text": "#ffffff",
            "button_primary_hover": "#1e4b82",
            "metric_bg": "#f7f9fc",
            "metric_border": "rgba(49, 51, 63, 0.08)",
            "table_header_bg": "#f3f4f6",
            "table_row_alt_bg": "rgba(15, 23, 42, 0.04)",
        }
    st.markdown(
        f"""
        <style>
        :root {{
            --ps-bg: {colors['bg']};
            --ps-sidebar-bg: {colors['sidebar_bg']};
            --ps-panel-bg: {colors['panel_bg']};
            --ps-panel-border: {colors['panel_border']};
            --ps-text: {colors['text']};
            --ps-muted: {colors['muted']};
            --ps-input-bg: {colors['input_bg']};
            --ps-input-border: {colors['input_border']};
            --ps-accent: {colors['accent']};
            --ps-metric-bg: {colors['metric_bg']};
            --ps-metric-border: {colors['metric_border']};
            --ps-button-bg: {colors['button_bg']};
            --ps-button-border: {colors['button_border']};
            --ps-button-text: {colors['button_text']};
            --ps-button-hover: {colors['button_hover']};
            --ps-button-primary-bg: {colors['button_primary_bg']};
            --ps-button-primary-text: {colors['button_primary_text']};
            --ps-button-primary-hover: {colors['button_primary_hover']};
            --ps-table-header-bg: {colors['table_header_bg']};
            --ps-table-row-alt-bg: {colors['table_row_alt_bg']};
            color-scheme: {theme};
        }}
        body,
        .stApp,
        section.main {{
            background-color: var(--ps-bg);
            color: var(--ps-text);
        }}
        [data-testid="stAppViewContainer"],
        [data-testid="stSidebar"] {{
            background-color: var(--ps-bg);
            color: var(--ps-text);
        }}
        [data-testid="stSidebar"] {{
            background-color: var(--ps-sidebar-bg);
        }}
        div[data-testid="stMarkdownContainer"] p,
        div[data-testid="stMarkdownContainer"] li,
        div[data-testid="stMarkdownContainer"] span,
        div[data-testid="stText"] {{
            overflow-wrap: anywhere;
            word-break: break-word;
            white-space: pre-wrap;
        }}
        .ps-ribbon-nav {{
            position: sticky;
            top: 1rem;
            background: var(--ps-sidebar-bg);
            border: 1px solid var(--ps-panel-border);
            border-radius: 18px;
            padding: 1rem 0.85rem;
            box-shadow: 0 18px 40px rgba(15, 23, 42, 0.12);
            display: none !important;
        }}
        .ps-ribbon-nav h3 {{
            margin-top: 0;
        }}
        .ps-ribbon-nav [role="radiogroup"] {{
            gap: 0.35rem;
        }}
        .ps-ribbon-nav [data-testid="stRadio"] label {{
            border-radius: 999px;
            padding: 0.35rem 0.75rem;
            border: 1px solid transparent;
            transition: all 0.2s ease;
        }}
        .ps-ribbon-nav [data-testid="stRadio"] label:hover {{
            background: var(--ps-button-hover);
        }}
        .ps-ribbon-nav [data-testid="stRadio"] label[data-selected="true"] {{
            border-color: var(--ps-button-border);
            background: var(--ps-panel-bg);
            font-weight: 600;
        }}
        .ps-ribbon-nav .stButton > button {{
            border-radius: 999px;
        }}
        .ps-mobile-nav {{
            display: none;
            position: sticky;
            top: 0.75rem;
            z-index: 1000;
        }}
        .ps-mobile-nav button {{
            padding: 0.2rem 0.6rem;
            border-radius: 999px;
            font-size: 0.85rem;
            min-height: unset;
        }}
        @media (max-width: 1200px) {{
            .ps-ribbon-nav {{
                display: block !important;
            }}
            [data-testid="stSidebar"] {{
                display: none !important;
            }}
        }}
        @media (max-width: 768px) {{
            .ps-mobile-nav {{
                display: block;
            }}
        }}
        [data-testid="stTextInput"] input,
        [data-testid="stTextArea"] textarea,
        [data-testid="stDateInput"] input,
        [data-testid="stNumberInput"] input,
        [data-testid="stTimeInput"] input {{
            background-color: var(--ps-input-bg);
            border-color: var(--ps-input-border);
            color: var(--ps-text);
        }}
        [data-testid="stSelectbox"] [data-baseweb="select"] > div,
        [data-testid="stMultiSelect"] [data-baseweb="select"] > div {{
            background-color: var(--ps-input-bg);
            border-color: var(--ps-input-border);
            color: var(--ps-text);
        }}
        [data-testid="stSelectbox"] [data-baseweb="select"] input,
        [data-testid="stMultiSelect"] [data-baseweb="select"] input {{
            color: var(--ps-text);
        }}
        [data-testid="stSelectbox"] [data-baseweb="select"] svg,
        [data-testid="stMultiSelect"] [data-baseweb="select"] svg {{
            color: var(--ps-muted);
            fill: var(--ps-muted);
        }}
        [data-testid="stSelectbox"] [role="listbox"],
        [data-testid="stMultiSelect"] [role="listbox"] {{
            background-color: var(--ps-panel-bg);
            color: var(--ps-text);
            border-color: var(--ps-panel-border);
        }}
        [data-testid="stSelectbox"] [role="option"],
        [data-testid="stMultiSelect"] [role="option"] {{
            color: var(--ps-text);
        }}
        [data-testid="stSelectbox"] [role="option"][aria-selected="true"],
        [data-testid="stMultiSelect"] [role="option"][aria-selected="true"] {{
            background-color: rgba(56, 189, 248, 0.18);
        }}
        [data-testid="stMarkdownContainer"] p,
        [data-testid="stMarkdownContainer"] span,
        [data-testid="stMarkdownContainer"] li,
        [data-testid="stMarkdownContainer"] label,
        [data-testid="stMarkdownContainer"] h1,
        [data-testid="stMarkdownContainer"] h2,
        [data-testid="stMarkdownContainer"] h3,
        [data-testid="stMarkdownContainer"] h4,
        [data-testid="stMarkdownContainer"] h5,
        [data-testid="stMarkdownContainer"] h6 {{
            color: var(--ps-text);
        }}
        [data-testid="stCaptionContainer"] p,
        [data-testid="stCaptionContainer"] span,
        .stCaption,
        .stCaption p {{
            color: var(--ps-muted) !important;
        }}
        [data-testid="stMetric"] {{
            background: var(--ps-metric-bg);
            border: 1px solid var(--ps-metric-border);
        }}
        [data-testid="stExpander"] details {{
            background-color: var(--ps-panel-bg);
            border: 1px solid var(--ps-panel-border);
            border-radius: 0.65rem;
        }}
        [data-testid="stExpander"] summary {{
            color: var(--ps-text);
            background-color: var(--ps-panel-bg);
            border-radius: 0.65rem;
        }}
        [data-testid="stExpander"] summary:hover {{
            background-color: var(--ps-button-hover);
            color: var(--ps-text);
        }}
        [data-testid="stExpander"] div[role="region"] {{
            background-color: var(--ps-panel-bg);
            color: var(--ps-text);
        }}
        div[data-testid="stButton"] > button,
        div[data-testid="stDownloadButton"] > button,
        div[data-testid="stFormSubmitButton"] > button,
        div[data-testid="stPopover"] button,
        div[data-testid="stForm"] button {{
            background-color: var(--ps-button-bg);
            border-color: var(--ps-button-border);
            color: var(--ps-button-text);
        }}
        div[data-testid="stButton"] > button:hover,
        div[data-testid="stDownloadButton"] > button:hover,
        div[data-testid="stFormSubmitButton"] > button:hover,
        div[data-testid="stForm"] button:hover {{
            background-color: var(--ps-button-hover);
            border-color: var(--ps-button-border);
            color: var(--ps-button-text);
        }}
        button[data-testid="baseButton-primary"] {{
            background-color: var(--ps-button-primary-bg) !important;
            border-color: var(--ps-button-primary-bg) !important;
            color: var(--ps-button-primary-text) !important;
        }}
        button[data-testid="baseButton-primary"]:hover {{
            background-color: var(--ps-button-primary-hover) !important;
            border-color: var(--ps-button-primary-hover) !important;
            color: var(--ps-button-primary-text) !important;
        }}
        [data-testid="stPopoverContent"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border: 1px solid var(--ps-panel-border) !important;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.18);
        }}
        [data-baseweb="popover"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-baseweb="popover"] > div {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-testid="stPopover"] [role="dialog"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border: 1px solid var(--ps-panel-border) !important;
        }}
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] p,
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] span,
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] li,
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] h1,
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] h2,
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] h3,
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] h4,
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] h5,
        [data-testid="stPopoverContent"] [data-testid="stMarkdownContainer"] h6 {{
            color: var(--ps-text) !important;
        }}
        [data-testid="stPopover"] [role="dialog"] p,
        [data-testid="stPopover"] [role="dialog"] span,
        [data-testid="stPopover"] [role="dialog"] li,
        [data-testid="stPopover"] [role="dialog"] label {{
            color: var(--ps-text) !important;
        }}
        [data-testid="stPopoverContent"] .stCaption,
        [data-testid="stPopoverContent"] .stCaption span,
        [data-testid="stPopover"] [role="dialog"] .stCaption,
        [data-testid="stPopover"] [role="dialog"] .stCaption span {{
            color: var(--ps-muted) !important;
        }}
        [data-testid="stPopoverContent"] hr,
        [data-testid="stPopover"] [role="dialog"] hr {{
            border-color: var(--ps-panel-border) !important;
        }}
        [data-testid="stAlert"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        [data-testid="stAlert"] p,
        [data-testid="stAlert"] span,
        [data-testid="stAlert"] li {{
            color: var(--ps-text) !important;
        }}
        [data-testid="stToast"],
        .stToast {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border: 1px solid var(--ps-panel-border) !important;
        }}
        [data-testid="stToast"] > div,
        .stToast > div {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-testid="stToast"] p,
        [data-testid="stToast"] span,
        [data-testid="stToast"] button,
        [data-testid="stToast"] svg,
        .stToast p,
        .stToast span,
        .stToast button,
        .stToast svg {{
            color: var(--ps-text) !important;
            fill: var(--ps-text) !important;
        }}
        [data-testid="stToast"] button,
        .stToast button {{
            background-color: var(--ps-button-bg) !important;
            border-color: var(--ps-button-border) !important;
        }}
        [data-baseweb="tooltip"] > div,
        [data-baseweb="tooltip"] [role="tooltip"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border: 1px solid var(--ps-panel-border) !important;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.18);
        }}
        [data-baseweb="tooltip"] p,
        [data-baseweb="tooltip"] span,
        [data-baseweb="tooltip"] li {{
            color: var(--ps-text) !important;
        }}
        [data-baseweb="tooltip"] svg {{
            fill: var(--ps-text) !important;
        }}
        [data-testid="stVegaLiteChart"] svg,
        [data-testid="stVegaLiteChart"] canvas {{
            background-color: var(--ps-panel-bg) !important;
        }}
        [data-testid="stVegaLiteChart"] text {{
            fill: var(--ps-muted) !important;
        }}
        [data-testid="stDataFrame"],
        [data-testid="stDataEditor"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border: 1px solid var(--ps-panel-border) !important;
            border-radius: 0.65rem;
        }}
        [data-testid="stDataFrame"] > div,
        [data-testid="stDataFrame"] [data-testid="stDataFrameResizable"],
        [data-testid="stDataFrame"] [data-testid="stDataFrameScrollable"],
        [data-testid="stDataEditor"] > div {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-testid="stDataFrame"] [data-baseweb="table"] > div,
        [data-testid="stDataFrame"] [data-baseweb="table"] [role="rowgroup"],
        [data-testid="stDataFrame"] [data-baseweb="table"] [role="grid"],
        [data-testid="stDataEditor"] [data-baseweb="table"] > div,
        [data-testid="stDataEditor"] [data-baseweb="table"] [role="rowgroup"],
        [data-testid="stDataEditor"] [data-baseweb="table"] [role="grid"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-testid="stDataFrame"] table,
        [data-testid="stDataEditor"] table,
        [data-testid="stDataFrame"] [role="grid"],
        [data-testid="stDataEditor"] [role="grid"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-testid="stDataFrame"] th,
        [data-testid="stDataFrame"] td,
        [data-testid="stDataEditor"] th,
        [data-testid="stDataEditor"] td {{
            border-color: var(--ps-panel-border) !important;
            color: var(--ps-text) !important;
            background-color: var(--ps-panel-bg) !important;
        }}
        [data-testid="stDataFrame"] [role="columnheader"],
        [data-testid="stDataFrame"] [role="gridcell"],
        [data-testid="stDataEditor"] [role="columnheader"],
        [data-testid="stDataEditor"] [role="gridcell"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        [data-testid="stDataFrame"] [role="row"],
        [data-testid="stDataFrame"] [role="row"] > div,
        [data-testid="stDataFrame"] [role="cell"],
        [data-testid="stDataEditor"] [role="row"],
        [data-testid="stDataEditor"] [role="row"] > div,
        [data-testid="stDataEditor"] [role="cell"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        .stDataFrame [role="columnheader"],
        .stDataFrame [role="gridcell"],
        .stDataFrame [role="row"],
        .stDataFrame [data-baseweb="table"],
        .stDataFrame [data-baseweb="table"] th,
        .stDataFrame [data-baseweb="table"] td {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        [data-testid="stDataFrame"] [data-baseweb="table"],
        [data-testid="stDataEditor"] [data-baseweb="table"],
        [data-testid="stDataFrame"] [data-baseweb="table"] tbody,
        [data-testid="stDataEditor"] [data-baseweb="table"] tbody,
        [data-testid="stDataFrame"] [data-baseweb="table"] thead,
        [data-testid="stDataEditor"] [data-baseweb="table"] thead {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-testid="stDataEditor"] input,
        [data-testid="stDataEditor"] textarea {{
            background-color: var(--ps-input-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-input-border) !important;
        }}
        [data-baseweb="table"] {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-baseweb="table"] th,
        [data-baseweb="table"] td {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        [data-baseweb="table"] th {{
            background-color: var(--ps-table-header-bg) !important;
        }}
        [data-baseweb="table"] [role="row"] {{
            background-color: var(--ps-panel-bg) !important;
        }}
        [data-baseweb="table"] [role="row"]:nth-child(even) {{
            background-color: var(--ps-table-row-alt-bg) !important;
        }}
        .stDataFrame [data-baseweb="table"] [role="row"]:nth-child(even),
        .stTable tbody tr:nth-child(even) {{
            background-color: var(--ps-table-row-alt-bg) !important;
        }}
        .stDataFrame [data-baseweb="table"] [role="columnheader"] {{
            background-color: var(--ps-table-header-bg) !important;
        }}
        [data-testid="stTable"] {{
            background-color: var(--ps-panel-bg) !important;
            border: 1px solid var(--ps-panel-border) !important;
            border-radius: 0.65rem;
            overflow: hidden;
        }}
        [data-testid="stTable"] table {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
        }}
        [data-testid="stTable"] th,
        [data-testid="stTable"] td {{
            border-color: var(--ps-panel-border) !important;
            color: var(--ps-text) !important;
            background-color: var(--ps-panel-bg) !important;
        }}
        [data-testid="stTable"] th {{
            background-color: var(--ps-table-header-bg) !important;
        }}
        [data-testid="stTable"] tbody tr:nth-child(even) {{
            background-color: var(--ps-table-row-alt-bg) !important;
        }}
        [data-testid="stMarkdownContainer"] table {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        [data-testid="stMarkdownContainer"] th,
        [data-testid="stMarkdownContainer"] td {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        [data-testid="stMarkdownContainer"] th {{
            background-color: var(--ps-table-header-bg) !important;
        }}
        table {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        table th,
        table td {{
            background-color: var(--ps-panel-bg) !important;
            color: var(--ps-text) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        table tbody tr {{
            background-color: var(--ps-panel-bg) !important;
        }}
        table thead th {{
            background-color: var(--ps-table-header-bg) !important;
        }}
        table tbody tr:nth-child(even) {{
            background-color: var(--ps-table-row-alt-bg) !important;
        }}
        [data-testid="stFileUploader"] section {{
            background-color: var(--ps-panel-bg) !important;
            border-color: var(--ps-panel-border) !important;
            color: var(--ps-text) !important;
        }}
        [data-testid="stFileUploader"] {{
            background-color: var(--ps-panel-bg) !important;
        }}
        [data-testid="stFileUploader"] section p,
        [data-testid="stFileUploader"] section span,
        [data-testid="stFileUploader"] section svg {{
            color: var(--ps-text) !important;
            fill: var(--ps-text) !important;
        }}
        [data-testid="stFileUploader"] ul,
        [data-testid="stFileUploader"] li,
        [data-testid="stFileUploader"] li span,
        [data-testid="stFileUploader"] li small,
        [data-testid="stFileUploader"] [data-testid="stFileUploaderFileName"] {{
            color: var(--ps-text) !important;
        }}
        [data-testid="stFileUploader"] li {{
            background-color: var(--ps-panel-bg) !important;
            border-color: var(--ps-panel-border) !important;
        }}
        [data-testid="stFileUploader"] button {{
            background-color: var(--ps-button-bg) !important;
            border-color: var(--ps-button-border) !important;
            color: var(--ps-button-text) !important;
        }}
        [data-testid="stFileUploader"] button:hover {{
            background-color: var(--ps-button-hover) !important;
            border-color: var(--ps-button-border) !important;
            color: var(--ps-button-text) !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _guard_double_submit(
    action_key: str,
    submitted: bool,
    *,
    cooldown_seconds: float = 3.0,
) -> bool:
    if not submitted:
        return False
    state_key = f"submit_guard_{action_key}"
    last_submit = st.session_state.get(state_key)
    now = time.time()
    if isinstance(last_submit, (int, float)) and now - last_submit < cooldown_seconds:
        st.warning("That was just submitted. Please wait a moment before trying again.")
        return False
    st.session_state[state_key] = now
    return True


def _find_login_cover_image() -> Optional[Path]:
    preferred = PROJECT_ROOT / "assets" / "login cover.png"
    if preferred.exists():
        return preferred
    candidates = [
        "cover photo (1)",
        "cover photo(1)",
        "cover_photo(1)",
        "cover_photo",
        "cover-photo",
        "cover",
    ]
    extensions = [".png", ".jpg", ".jpeg", ".webp", ".gif"]
    search_roots = [PROJECT_ROOT, BASE_DIR]
    for root in search_roots:
        for base in candidates:
            for ext in extensions:
                candidate = root / f"{base}{ext}"
                if candidate.exists():
                    return candidate
    for root in search_roots:
        for path in root.rglob("cover photo (1).png"):
            if path.is_file():
                return path
        for path in root.rglob("cover photo(1).*"):
            if path.is_file():
                return path
    return None


def _list_database_tables(conn: sqlite3.Connection) -> list[str]:
    tables_df = df_query(
        conn,
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table'
          AND name NOT LIKE 'sqlite_%'
        ORDER BY name
        """,
    )
    if tables_df.empty:
        return []
    return [clean_text(name) for name in tables_df["name"].tolist() if clean_text(name)]


def _format_generic_table_dates(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    formatted = df.copy()
    for col in formatted.columns:
        label = str(col).lower()
        if "date" in label:
            formatted[col] = pd.to_datetime(formatted[col], errors="coerce").dt.strftime(DATE_FMT)
        elif label.endswith("_at") or label in {"created_at", "updated_at"}:
            formatted[col] = pd.to_datetime(formatted[col], errors="coerce").dt.strftime(
                "%Y-%m-%d %H:%M:%S"
            )
    return formatted


def _build_generic_table_export(conn: sqlite3.Connection, table_name: str) -> pd.DataFrame:
    df = df_query(conn, f'SELECT * FROM "{table_name}"')
    return _format_generic_table_dates(df)


def _build_report_cadence_summary(conn: sqlite3.Connection) -> pd.DataFrame:
    df = df_query(
        conn,
        """
        SELECT period_type, COUNT(*) AS report_count
        FROM work_reports
        GROUP BY period_type
        ORDER BY period_type
        """,
    )
    if df.empty:
        return df
    df["Cadence"] = df["period_type"].apply(format_period_label)
    return df.rename(columns={"report_count": "Report count"})[
        ["Cadence", "Report count"]
    ]


def _build_report_coverage_summary(conn: sqlite3.Connection) -> pd.DataFrame:
    df = df_query(
        conn,
        """
        SELECT COALESCE(u.username, 'User #' || wr.user_id) AS username,
               wr.period_type,
               COUNT(*) AS report_count
        FROM work_reports wr
        LEFT JOIN users u ON u.user_id = wr.user_id
        GROUP BY u.username, wr.user_id, wr.period_type
        ORDER BY LOWER(COALESCE(u.username, 'user')), wr.period_type
        """,
    )
    if df.empty:
        return df
    df["Cadence"] = df["period_type"].apply(format_period_label)
    pivot = (
        df.pivot_table(
            index="username",
            columns="Cadence",
            values="report_count",
            aggfunc="sum",
            fill_value=0,
        )
        .astype(int)
        .reset_index()
        .rename(columns={"username": "Team member"})
    )
    pivot.columns.name = None
    return pivot


def _build_admin_kpi_snapshot(conn: sqlite3.Connection) -> pd.DataFrame:
    staff_df = df_query(
        conn,
        dedent(
            """
            SELECT user_id, COALESCE(username, 'User #' || user_id) AS username
            FROM users
            WHERE LOWER(COALESCE(role, 'staff')) <> 'admin'
            ORDER BY LOWER(username)
            """
        ),
    )

    if staff_df.empty:
        return staff_df

    today = date.today()
    start_month = today.replace(day=1)
    days_elapsed = max((today - start_month).days + 1, 1)

    monthly_df = df_query(
        conn,
        dedent(
            """
            SELECT user_id,
                   COUNT(*) AS monthly_reports,
                   MAX(date(period_start)) AS last_report_date
            FROM work_reports
            WHERE period_type='daily'
              AND strftime('%Y-%m', period_start) = strftime('%Y-%m', 'now')
            GROUP BY user_id
            """
        ),
    ).rename(columns={"last_report_date": "last_report_month"})

    lifetime_df = df_query(
        conn,
        dedent(
            """
            SELECT user_id,
                   COUNT(*) AS total_reports,
                   MIN(date(period_start)) AS first_report_date,
                   MAX(date(period_start)) AS last_report_date
            FROM work_reports
            WHERE period_type='daily'
            GROUP BY user_id
            """
        ),
    ).rename(columns={"last_report_date": "last_report_all"})

    def _parse_date(value: object) -> Optional[date]:
        try:
            parsed = pd.to_datetime(value, errors="coerce")
        except Exception:
            return None
        if pd.isna(parsed):
            return None
        try:
            return parsed.date()
        except AttributeError:
            return None

    staff_df["user_id"] = staff_df["user_id"].apply(lambda val: int(float(val)))
    staff_df = staff_df.rename(columns={"username": "Team member"})

    merged = (
        staff_df.merge(monthly_df, on="user_id", how="left")
        .merge(lifetime_df, on="user_id", how="left")
    )

    merged[["monthly_reports", "total_reports"]] = merged[
        ["monthly_reports", "total_reports"]
    ].fillna(0)

    kpi_rows: list[dict[str, object]] = []
    for _, row in merged.iterrows():
        monthly_reports = int(_coerce_float(row.get("monthly_reports"), 0.0))
        total_reports = int(_coerce_float(row.get("total_reports"), 0.0))
        last_seen = _parse_date(row.get("last_report_month")) or _parse_date(
            row.get("last_report_all")
        )
        first_seen = _parse_date(row.get("first_report_date"))

        days_since_last = (today - last_seen).days if last_seen else days_elapsed + 7
        months_active = (
            (today.year - first_seen.year) * 12 + today.month - first_seen.month + 1
            if first_seen
            else 1
        )
        months_active = max(months_active, 1)

        monthly_completion = (monthly_reports / days_elapsed) * 100
        recency_penalty = min(days_since_last * 3.0, 60.0)
        momentum_boost = min((total_reports / months_active) * 2.5, 25.0)
        monthly_score = max(
            0.0,
            min(100.0, monthly_completion - recency_penalty + 20.0 + momentum_boost),
        )

        lifetime_velocity = (total_reports / months_active) / 20.0
        lifetime_score = max(
            0.0,
            min(
                100.0,
                (lifetime_velocity * 100.0) - min(days_since_last * 1.5, 40.0) + 20.0,
            ),
        )

        kpi_rows.append(
            {
                "Team member": clean_text(row.get("Team member"))
                or f"User #{int(row.get('user_id'))}",
                "Monthly KPI": f"{monthly_score:,.0f}/100",
                "Lifetime KPI": f"{lifetime_score:,.0f}/100",
            }
        )

    return pd.DataFrame(kpi_rows)


def _safe_sheet_name(name: str, used: set[str]) -> str:
    safe_name = (name or "Sheet").strip()
    if not safe_name:
        safe_name = "Sheet"
    safe_name = re.sub(r"[\\/*?:\[\]]", " ", safe_name)
    safe_name = " ".join(safe_name.split()).strip() or "Sheet"
    safe_name = safe_name[:31]
    candidate = safe_name
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{safe_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def export_database_to_excel(conn, include_all: bool = False) -> bytes:
    sheet_builders = [
        ("Customers", _build_customers_export),
        ("Delivery orders", _build_delivery_orders_export),
        ("Warranties", _build_warranties_export),
        ("Services", _build_services_export),
        ("Maintenance", _build_maintenance_export),
    ]
    if include_all:
        sheet_builders.extend(
            [
                ("Quotations", _build_quotations_export),
                ("Report cadence", _build_report_cadence_summary),
                ("Report coverage", _build_report_coverage_summary),
                ("Admin KPI snapshot", _build_admin_kpi_snapshot),
            ]
        )

    sheet_data: list[tuple[str, pd.DataFrame]] = []
    for name, builder in sheet_builders:
        try:
            df = builder(conn)
        except Exception:
            df = pd.DataFrame()
        sheet_data.append((name, df))

    if include_all:
        for table_name in _list_database_tables(conn):
            df = _build_generic_table_export(conn, table_name)
            sheet_data.append((f"Table: {table_name}", df))

    master_df = _build_master_sheet(sheet_data)
    ordered_sheets = [("Master", master_df)] + sheet_data

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        used_names: set[str] = set()
        for sheet_name, df in ordered_sheets:
            safe_name = _safe_sheet_name(sheet_name, used_names)
            if df is None or df.empty:
                df_to_write = pd.DataFrame()
            else:
                df_to_write = df
            df_to_write.to_excel(writer, sheet_name=safe_name, index=False)
    buffer.seek(0)
    return buffer.getvalue()


def export_full_archive(
    conn: Optional[sqlite3.Connection] = None, excel_bytes: Optional[bytes] = None
) -> bytes:
    """Bundle all user data, uploads, and database exports into one archive.

    The archive now includes:
    - The live SQLite database file.
    - A full SQL dump for recovery.
    - The Excel export of every table.
    - Every file under the application storage directory (uploads, receipts, etc.).
    """

    def _hash_bytes(payload: bytes) -> str:
        return hashlib.sha256(payload).hexdigest()

    def _hash_file(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    buffer = io.BytesIO()
    close_conn = False
    active_conn = conn
    if active_conn is None:
        active_conn = sqlite3.connect(DB_PATH)
        close_conn = True

    try:
        dump_buffer = io.StringIO()
        for line in active_conn.iterdump():
            dump_buffer.write(f"{line}\n")
        dump_bytes = dump_buffer.getvalue().encode("utf-8")

        if excel_bytes is None:
            excel_bytes = export_database_to_excel(active_conn, include_all=True)

        db_path = Path(DB_PATH)
        resource_paths = [
            PROJECT_ROOT / "import_template.xlsx",
            PROJECT_ROOT / "ps_letterhead.png",
            PROJECT_ROOT / "letterhead",
            PROJECT_ROOT / "assets",
        ]

        def _iter_files(path: Path) -> Iterable[Path]:
            if path.is_file():
                yield path
            elif path.is_dir():
                yield from (p for p in path.rglob("*") if p.is_file())

        raw_storage_files = [
            path
            for path in (BASE_DIR.rglob("*") if BASE_DIR.exists() else [])
            if path.is_file()
        ]
        storage_files = [
            path
            for path in raw_storage_files
            if not (db_path.exists() and path.resolve() == db_path.resolve())
        ]
        resource_counts: dict[Path, int] = {}
        resource_files: list[Path] = []
        for res in resource_paths:
            files = list(_iter_files(res))
            resource_counts[res] = len(files)
            resource_files.extend(files)

        checksum_lines: list[str] = []
        with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            if db_path.exists():
                arcname = f"database/{db_path.name}"
                zf.write(db_path, arcname=arcname)
                checksum_lines.append(f"{_hash_file(db_path)}  {arcname}")

            if dump_bytes:
                zf.writestr("exports/ps_crm.sql", dump_bytes)
                checksum_lines.append(
                    f"{_hash_bytes(dump_bytes)}  exports/ps_crm.sql"
                )

            if excel_bytes:
                zf.writestr(
                    "exports/ps_crm.xlsx",
                    excel_bytes,
                )
                checksum_lines.append(
                    f"{_hash_bytes(excel_bytes)}  exports/ps_crm.xlsx"
                )

            if BASE_DIR.exists():
                for path in storage_files:
                    arcname = Path("storage") / path.relative_to(BASE_DIR)
                    zf.write(path, arcname=str(arcname))
                    checksum_lines.append(f"{_hash_file(path)}  {arcname}")

            for res in resource_paths:
                for file_path in _iter_files(res):
                    arcname = Path("resources") / file_path.relative_to(res.parent)
                    zf.write(file_path, arcname=str(arcname))
                    checksum_lines.append(f"{_hash_file(file_path)}  {arcname}")

            total_exports = 0
            if dump_bytes:
                total_exports += 1
            if excel_bytes:
                total_exports += 1
            manifest_lines = [
                f"Export generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                (
                    "Database: includes users, staff accounts, and application records. "
                    "Protect this archive to preserve privacy."
                ),
                f"Database path: {db_path} (included: {'yes' if db_path.exists() else 'no'})",
                "SQL dump: exports/ps_crm.sql",
                "Excel export: exports/ps_crm.xlsx",
                f"Storage directory: {BASE_DIR} (files: {len(storage_files)})",
                f"Checksum file: checksums.txt (entries: {len(checksum_lines)})",
            ]
            for res in resource_paths:
                manifest_lines.append(
                    f"Resource: {res} (files: {resource_counts.get(res, 0)})"
                )
            manifest_lines.append(
                "Total archive files: "
                f"{len(storage_files) + len(resource_files) + (1 if db_path.exists() else 0) + total_exports + 1}"
            )
            zf.writestr("manifest.txt", "\n".join(manifest_lines))
            if checksum_lines:
                zf.writestr("checksums.txt", "\n".join(checksum_lines))
    finally:
        if close_conn:
            active_conn.close()

    buffer.seek(0)
    return buffer.getvalue()




def fetch_warranty_window(conn, start_days: int, end_days: int) -> pd.DataFrame:
    scope_clause, scope_params = customer_scope_filter("c")
    filters = [
        "w.status='active'",
        "date(w.expiry_date) BETWEEN date('now', ?) AND date('now', ?)",
        "p.name IS NOT NULL",
        "TRIM(p.name) != ''",
    ]
    params: list[object] = []
    if scope_clause:
        filters.append(scope_clause)
        params.extend(scope_params)
    where_clause = " AND ".join(filters)
    query = dedent(
        f"""
        SELECT c.name AS customer,
               p.name AS product,
               p.model,
               w.serial,
               COALESCE(w.issue_date, c.purchase_date) AS issue_date,
               w.expiry_date,
               w.remarks,
               COALESCE(c.sales_person, u.username) AS staff
        FROM warranties w
        LEFT JOIN customers c ON c.customer_id = w.customer_id
        LEFT JOIN products p ON p.product_id = w.product_id
        LEFT JOIN users u ON u.user_id = c.created_by
        WHERE {where_clause}
        ORDER BY date(w.expiry_date) ASC
        """
    )
    start = f"+{start_days} day"
    end = f"+{end_days} day"
    return df_query(conn, query, (start, end, *params))


def format_warranty_table(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    work = df.copy()
    expiry_raw = pd.to_datetime(work.get("expiry_date"), errors="coerce")
    today = pd.Timestamp.now().normalize()
    status_labels = []
    work["Description"] = work.apply(
        lambda row: dedupe_join(
            [
                clean_text(row.get("product")),
                clean_text(row.get("model")),
                clean_text(row.get("serial")),
            ]
        ),
        axis=1,
    )
    for idx in work.index:
        exp = expiry_raw.loc[idx] if expiry_raw is not None and idx in expiry_raw.index else pd.NaT
        if pd.notna(exp) and exp.normalize() < today:
            status_labels.append("Expired")
        else:
            base_status = clean_text(work.loc[idx, "status"]) if "status" in work.columns else None
            status_labels.append((base_status or "Active").title())
    work["Status"] = status_labels
    for col in ("product", "model", "serial"):
        if col in work.columns:
            work.drop(columns=[col], inplace=True)
    if "status" in work.columns:
        work.drop(columns=["status"], inplace=True)
    rename_map = {
        "customer": "Customer",
        "issue_date": "Issue date",
        "expiry_date": "Expiry date",
        "remarks": "Remarks",
        "staff": "Staff",
    }
    work.rename(columns={k: v for k, v in rename_map.items() if k in work.columns}, inplace=True)
    for col in ("dup_flag", "id", "duplicate"):
        if col in work.columns:
            work.drop(columns=[col], inplace=True)
    return work


def _pdf_escape_text(value: str) -> str:
    replacements = [("\\", "\\\\"), ("(", "\\("), (")", "\\)")]
    escaped = value
    for old, new in replacements:
        escaped = escaped.replace(old, new)
    return escaped


def _build_simple_pdf_document(lines: list[str]) -> bytes:
    if not lines:
        lines = [""]
    commands = ["BT", "/F1 12 Tf", "72 770 Td"]
    for idx, line in enumerate(lines):
        escaped = _pdf_escape_text(line)
        if idx == 0:
            commands.append(f"({escaped}) Tj")
        else:
            commands.append("0 -14 Td")
            commands.append(f"({escaped}) Tj")
    commands.append("ET")
    stream_bytes = "\n".join(commands).encode("latin-1", "replace")

    buffer = io.BytesIO()
    buffer.write(b"%PDF-1.4\n")
    offsets = []

    def write_obj(obj_id: int, body: bytes) -> None:
        offsets.append(buffer.tell())
        buffer.write(f"{obj_id} 0 obj\n".encode("latin-1"))
        buffer.write(body)
        if not body.endswith(b"\n"):
            buffer.write(b"\n")
        buffer.write(b"endobj\n")

    write_obj(1, b"<< /Type /Catalog /Pages 2 0 R >>\n")
    write_obj(2, b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>\n")
    write_obj(
        3,
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>\n",
    )
    stream_obj = b"<< /Length %d >>\nstream\n" % len(stream_bytes) + stream_bytes + b"\nendstream\n"
    write_obj(4, stream_obj)
    write_obj(5, b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\n")

    xref_offset = buffer.tell()
    buffer.write(f"xref\n0 {len(offsets) + 1}\n".encode("latin-1"))
    buffer.write(b"0000000000 65535 f \n")
    for off in offsets:
        buffer.write(f"{off:010d} 00000 n \n".encode("latin-1"))
    buffer.write(b"trailer\n")
    buffer.write(f"<< /Size {len(offsets) + 1} /Root 1 0 R >>\n".encode("latin-1"))
    buffer.write(b"startxref\n")
    buffer.write(f"{xref_offset}\n".encode("latin-1"))
    buffer.write(b"%%EOF\n")
    return buffer.getvalue()


def generate_customer_summary_pdf(customer_name: str, info: dict, warranties: Optional[pd.DataFrame], services: pd.DataFrame, maintenance: pd.DataFrame) -> bytes:
    lines: list[str] = [f"Customer Summary – {customer_name}", ""]
    lines.extend(
        [
            f"Phone: {clean_text(info.get('phone')) or '-'}",
            f"Address: {clean_text(info.get('address')) or '-'}",
            f"Purchase: {clean_text(info.get('purchase_dates')) or '-'}",
            f"Product: {clean_text(info.get('products')) or '-'}",
            f"Delivery order: {clean_text(info.get('do_codes')) or '-'}",
            "",
        ]
    )

    def extend_section(title: str, rows: list[str]) -> None:
        lines.append(title)
        if not rows:
            lines.append("  (no records)")
        else:
            for row in rows:
                lines.append(f"  • {row}")
        lines.append("")

    warranty_rows: list[str] = []
    if warranties is not None and isinstance(warranties, pd.DataFrame) and not warranties.empty:
        for _, row in warranties.iterrows():
            warranty_rows.append(
                " | ".join(
                    [
                        f"Description: {clean_text(row.get('Description')) or '-'}",
                        f"Issue: {clean_text(row.get('Issue date')) or '-'}",
                        f"Expiry: {clean_text(row.get('Expiry date')) or '-'}",
                        f"Status: {clean_text(row.get('Status')) or '-'}",
                    ]
                )
            )

    service_rows: list[str] = []
    if isinstance(services, pd.DataFrame) and not services.empty:
        for _, row in services.iterrows():
            service_rows.append(
                " | ".join(
                    [
                        f"DO: {clean_text(row.get('do_number')) or '-'}",
                        f"Date: {clean_text(row.get('service_date')) or '-'}",
                        f"Desc: {clean_text(row.get('description')) or '-'}",
                        f"Remarks: {clean_text(row.get('remarks')) or '-'}",
                    ]
                )
            )

    maintenance_rows: list[str] = []
    if isinstance(maintenance, pd.DataFrame) and not maintenance.empty:
        for _, row in maintenance.iterrows():
            maintenance_rows.append(
                " | ".join(
                    [
                        f"DO: {clean_text(row.get('do_number')) or '-'}",
                        f"Date: {clean_text(row.get('maintenance_date')) or '-'}",
                        f"Desc: {clean_text(row.get('description')) or '-'}",
                        f"Remarks: {clean_text(row.get('remarks')) or '-'}",
                    ]
                )
            )

    extend_section("Warranties", warranty_rows)
    extend_section("Service history", service_rows)
    extend_section("Maintenance history", maintenance_rows)

    return _build_simple_pdf_document(lines)


def _streamlit_flag_options_from_env() -> dict[str, object]:
    """Derive Streamlit bootstrap flag options from environment variables."""

    flag_options: dict[str, object] = {}

    port_env = os.getenv("PORT")
    if port_env:
        try:
            port = int(port_env)
        except (TypeError, ValueError):
            port = None
        if port and port > 0:
            flag_options["server.port"] = port

    address_env = os.getenv("HOST") or os.getenv("BIND_ADDRESS")
    flag_options["server.address"] = address_env or "0.0.0.0"

    headless_env = os.getenv("STREAMLIT_SERVER_HEADLESS")
    if headless_env is None:
        flag_options["server.headless"] = True
    else:
        flag_options["server.headless"] = headless_env.strip().lower() in (
            "1",
            "true",
            "yes",
            "on",
        )

    return flag_options


def _bootstrap_streamlit_app() -> None:
    """Launch the Streamlit app when executed via ``python app.py``."""

    try:
        from streamlit.web import bootstrap
    except Exception:
        return

    try:
        bootstrap.run(
            os.path.abspath(__file__),
            False,
            [],
            _streamlit_flag_options_from_env(),
        )
    except Exception:
        pass


def recalc_customer_duplicate_flag(conn, phone):
    if not phone or str(phone).strip() == "":
        return
    cur = conn.execute(
        "SELECT customer_id, purchase_date FROM customers WHERE phone = ?",
        (str(phone).strip(),),
    )
    rows = cur.fetchall()
    if not rows:
        return

    grouped: dict[Optional[str], list[int]] = {}
    for cid, purchase_date in rows:
        try:
            cid_int = int(cid)
        except (TypeError, ValueError):
            continue
        key = clean_text(purchase_date) or None
        grouped.setdefault(key, []).append(cid_int)

    updates: list[tuple[int, int]] = []
    for cid_list in grouped.values():
        dup_flag = 1 if len(cid_list) > 1 else 0
        updates.extend((dup_flag, cid) for cid in cid_list)

    if updates:
        conn.executemany(
            "UPDATE customers SET dup_flag=? WHERE customer_id=?",
            updates,
        )


def init_ui():
    st.set_page_config(
        page_title="PS Business Suites",
        page_icon="🧰",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.session_state["ocr_uploads_enabled"] = True
    if "user" not in st.session_state:
        st.session_state.user = None
    if st.session_state.user:
        st.set_option("client.toolbarMode", "minimal")
        st.title("PS Engineering – Business Suites")
        st.caption("Customers • Warranties • Needs • Summaries")
        st.markdown(
            """
            <style>
            #MainMenu,
            header,
            div[data-testid="stStatusWidget"],
            div[data-testid="stDecoration"],
            div[data-testid="stToolbar"] {
                display: none !important;
            }
            [data-testid="stSidebar"] {
                display: block !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
    st.markdown(
        """
        <style>
        [data-testid="stMetric"] {
            background: var(--ps-metric-bg);
            border-radius: 0.8rem;
            padding: 0.85rem;
            border: 1px solid var(--ps-metric-border);
        }
        [data-testid="stMetricValue"] {
            color: var(--ps-text);
        }
        [data-testid="stMetricLabel"] {
            color: var(--ps-muted);
        }
        div[data-testid="stPopover"] > button {
            border: none !important;
            background: transparent !important;
            font-size: 1.25rem;
            padding: 0.15rem 0.35rem !important;
            color: var(--ps-accent) !important;
        }
        .ps-notification-popover {
            display: flex;
            justify-content: flex-end;
        }
        .ps-notification-popover button:hover {
            background: var(--ps-button-hover) !important;
        }
        .ps-notification-section-title {
            font-size: 0.9rem;
            font-weight: 600;
            color: var(--ps-accent);
            margin-bottom: 0.25rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
# ---------- Auth ----------
SESSION_TOKEN_PARAM = "session"


def _session_duration_days() -> float:
    raw_value = os.getenv("PS_CRM_SESSION_DAYS", "7")
    try:
        parsed = float(raw_value)
    except (TypeError, ValueError):
        return 7.0
    return parsed if parsed > 0 else 7.0


SESSION_DURATION_DAYS = _session_duration_days()


def _get_session_token_from_url() -> Optional[str]:
    params = st.query_params
    token_values = params.get(SESSION_TOKEN_PARAM, [])
    if isinstance(token_values, list):
        return token_values[0] if token_values else None
    return token_values or None


def _set_session_token_in_url(token: Optional[str]) -> None:
    if token:
        st.query_params[SESSION_TOKEN_PARAM] = token
    else:
        st.query_params.clear()


def _purge_expired_sessions(conn) -> None:
    conn.execute(
        """
        DELETE FROM user_sessions
        WHERE expires_at IS NOT NULL
          AND datetime(expires_at) <= datetime('now')
        """
    )
    conn.commit()


def _load_user_from_session(conn, token: str) -> Optional[dict[str, object]]:
    if not token:
        return None
    row = df_query(
        conn,
        """
        SELECT u.user_id, u.username, u.role, u.phone, u.title, u.email, s.expires_at
        FROM user_sessions s
        JOIN users u ON u.user_id = s.user_id
        WHERE s.token = ?
          AND (s.expires_at IS NULL OR datetime(s.expires_at) > datetime('now'))
        """,
        (token,),
    )
    if row.empty:
        return None
    return {
        "user_id": int(row.iloc[0]["user_id"]),
        "username": row.iloc[0]["username"],
        "role": row.iloc[0]["role"],
        "phone": clean_text(row.iloc[0].get("phone")),
        "title": clean_text(row.iloc[0].get("title")),
        "email": clean_text(row.iloc[0].get("email")),
    }


def _touch_session(conn, token: str) -> None:
    if not token:
        return
    conn.execute(
        """
        UPDATE user_sessions
        SET last_seen=datetime('now'),
            expires_at=datetime('now', ?)
        WHERE token=?
        """,
        (f"+{SESSION_DURATION_DAYS} days", token),
    )
    conn.commit()


def _create_session(conn, user_id: int) -> str:
    token = secrets.token_urlsafe(32)
    conn.execute(
        """
        INSERT INTO user_sessions (token, user_id, expires_at)
        VALUES (?, ?, datetime('now', ?))
        """,
        (token, int(user_id), f"+{SESSION_DURATION_DAYS} days"),
    )
    conn.commit()
    return token


def _restore_user_session(conn) -> None:
    if st.session_state.get("user"):
        return
    token = _get_session_token_from_url()
    if not token:
        return
    user = _load_user_from_session(conn, token)
    if user:
        st.session_state.user = user
        st.session_state["session_token"] = token
        _touch_session(conn, token)
    else:
        _set_session_token_in_url(None)


def _ensure_session_token(conn, user: dict[str, object]) -> None:
    token = st.session_state.get("session_token") or _get_session_token_from_url()
    if token:
        st.session_state["session_token"] = token
        _touch_session(conn, token)
        return
    user_id = int(user.get("user_id") or 0)
    if user_id:
        new_token = _create_session(conn, user_id)
        st.session_state["session_token"] = new_token
        _set_session_token_in_url(new_token)


def _clear_session_for_logout(conn) -> None:
    token = st.session_state.get("session_token") or _get_session_token_from_url()
    if token:
        conn.execute("DELETE FROM user_sessions WHERE token = ?", (token,))
        conn.commit()
    _set_session_token_in_url(None)
    preserved = {}
    if "theme" in st.session_state:
        preserved["theme"] = st.session_state.get("theme")
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    for k, val in preserved.items():
        st.session_state[k] = val


def _request_logout() -> None:
    st.session_state["logout_requested"] = True


def login_box(conn, *, render_id=None):
    apply_theme_css()
    if st.session_state.user:
        _ensure_session_token(conn, st.session_state.user)
        st.sidebar.markdown("### Login")
        st.sidebar.success(f"Logged in as {st.session_state.user['username']} ({st.session_state.user['role']})")
        return True
    # Hide Streamlit chrome on the login screen (auth gate UI only).
    st.set_option("client.toolbarMode", "viewer")
    cover_image = _find_login_cover_image()
    cover_css = ""
    if cover_image:
        try:
            with open(cover_image, "rb") as handle:
                encoded = base64.b64encode(handle.read()).decode("utf-8")
            cover_css = (
                "background-image: url('data:image/png;base64,"
                f"{encoded}');"
            )
        except OSError:
            cover_css = ""
    app_bg = "#ffffff"
    panel_bg = "#ffffff"
    panel_text = "#111827"
    input_bg = "#ffffff"
    input_border = "#d1d5db"
    placeholder_color = "rgba(75, 85, 99, 0.9)"
    button_bg = "#ffffff"
    button_text = "#111827"
    button_hover = "#f3f4f6"
    st.markdown(
        f"""
        <style>
        #MainMenu,
        header,
        footer,
        div[data-testid="stToolbar"],
        div[data-testid="stStatusWidget"],
        div[data-testid="stDecoration"] {{
            display: none !important;
        }}
        [data-testid="stSidebar"] {{
            display: none !important;
        }}
        [data-testid="stAppViewContainer"] {{
            {cover_css}
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-color: {app_bg};
        }}
        section.main > div {{
            padding-top: 6rem;
        }}
        div[data-testid="stForm"] {{
            background: {panel_bg};
            color: {panel_text};
            padding: 1.75rem 2rem;
            border-radius: 18px;
            max-width: 420px;
            margin: 6.25rem auto 0 auto;
            box-shadow: 0 18px 60px rgba(0, 0, 0, 0.25);
        }}
        div[data-testid="stForm"] h3 {{
            color: {panel_text} !important;
        }}
        div[data-testid="stForm"] label {{
            color: {panel_text} !important;
        }}
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] p,
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] span,
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] li,
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] h1,
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] h2,
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] h3,
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] h4,
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] h5,
        div[data-testid="stForm"] [data-testid="stMarkdownContainer"] h6 {{
            color: {panel_text} !important;
        }}
        div[data-testid="stForm"] input {{
            background-color: {input_bg};
            color: {panel_text} !important;
            border: 1px solid {input_border};
        }}
        div[data-testid="stForm"] input::placeholder {{
            color: {placeholder_color} !important;
        }}
        div[data-testid="stForm"] button {{
            background: {button_bg};
            color: {button_text} !important;
            border: 1px solid {input_border};
        }}
        div[data-testid="stForm"] button span {{
            color: {button_text} !important;
        }}
        div[data-testid="stForm"] button:hover {{
            background: {button_hover};
            color: {button_text} !important;
        }}
        div[data-testid="stForm"] button:hover span {{
            color: {button_text} !important;
        }}
        div[data-testid="stForm"] [data-baseweb="input"] button {{
            background-color: transparent !important;
            color: {panel_text} !important;
            border-left: 1px solid {input_border};
        }}
        div[data-testid="stForm"] [data-baseweb="input"] button:hover {{
            background-color: rgba(148, 163, 184, 0.15) !important;
        }}
        div[data-testid="stForm"] [data-baseweb="input"] button span {{
            color: {panel_text} !important;
        }}
        div[data-testid="stForm"] [data-baseweb="input"] button svg {{
            fill: {panel_text} !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )
    with st.form("login_form"):
        st.markdown("### Welcome back")
        u = st.text_input("Username")
        p = st.text_input("Password", type="password")
        ok = st.form_submit_button("Login", use_container_width=True)
    if ok:
        row = df_query(
            conn,
            "SELECT user_id, username, pass_hash, role, phone, title, email FROM users WHERE username = ?",
            (u,),
        )
        if not row.empty and hashlib.sha256(p.encode("utf-8")).hexdigest() == row.iloc[0]["pass_hash"]:
            user_payload = {
                "user_id": int(row.iloc[0]["user_id"]),
                "username": row.iloc[0]["username"],
                "role": row.iloc[0]["role"],
                "phone": clean_text(row.iloc[0].get("phone")),
                "title": clean_text(row.iloc[0].get("title")),
                "email": clean_text(row.iloc[0].get("email")),
            }
            st.session_state.user = user_payload
            _ensure_session_token(conn, user_payload)
            st.session_state.page = "Dashboard"
            st.session_state.just_logged_in = True
            _safe_rerun()
        else:
            st.sidebar.error("Invalid credentials")
    st.stop()

def ensure_auth(role=None):
    if role and st.session_state.user and st.session_state.user["role"] != role:
        st.warning("You do not have permission to access this page.")
        st.stop()


def _render_admin_record_history(conn):
    user = get_current_user()
    if clean_text(user.get("role")) != "admin":
        return

    st.markdown("#### Team record history")
    st.caption(
        "Review recent submissions across reports, delivery orders, work done, and maintenance/service."
    )

    users_df = df_query(
        conn,
        "SELECT user_id, username FROM users ORDER BY LOWER(COALESCE(username, 'user'))",
    )
    member_options: list[Optional[int]] = [None]
    member_labels: dict[Optional[int], str] = {None: "All team members"}
    if not users_df.empty:
        for _, row in users_df.iterrows():
            try:
                uid = int(row.get("user_id"))
            except Exception:
                continue
            label = clean_text(row.get("username")) or f"User #{uid}"
            member_options.append(uid)
            member_labels[uid] = label

    selected_user = st.selectbox(
        "Filter by team member",
        member_options,
        format_func=lambda uid: member_labels.get(uid, "All team members"),
        key="dashboard_record_history_user",
    )

    tabs = st.tabs(
        ["Reports", "Delivery orders", "Work done", "Maintenance & Service"]
    )

    tab_map: list[tuple[Any, tuple[str, ...], str]] = [
        (tabs[0], ("report",), "No reports recorded yet."),
        (tabs[1], ("delivery_order",), "No delivery orders recorded yet."),
        (tabs[2], ("work_done",), "No work done entries recorded yet."),
        (
            tabs[3],
            ("service", "maintenance"),
            "No maintenance or service updates recorded yet.",
        ),
    ]

    for tab, entities, empty_message in tab_map:
        with tab:
            history_df = _fetch_entity_activity(
                conn, entities, user_filter=selected_user, limit=50
            )
            if history_df.empty:
                st.info(empty_message)
                continue

            history_df["When"] = history_df["created_at"].apply(
                lambda value: format_time_ago(value)
                or format_period_range(value, value)
            )
            history_df["Record type"] = history_df["entity_type"].apply(
                lambda val: (clean_text(val) or "Record").replace("_", " ").title()
            )
            display_df = history_df.rename(
                columns={"actor": "Team member", "description": "Details"}
            )
            st.dataframe(
                display_df[
                    ["Team member", "Record type", "Details", "When"]
                ],
                use_container_width=True,
                hide_index=True,
            )


def _render_admin_kpi_panel(conn) -> None:
    user = get_current_user()
    if clean_text(user.get("role")) != "admin":
        return

    staff_df = df_query(
        conn,
        dedent(
            """
            SELECT user_id, COALESCE(username, 'User #' || user_id) AS username
            FROM users
            WHERE LOWER(COALESCE(role, 'staff')) <> 'admin'
            ORDER BY LOWER(username)
            """
        ),
    )

    if staff_df.empty:
        st.caption("No staff accounts available for KPI tracking yet.")
        return

    today = date.today()
    start_month = today.replace(day=1)
    days_elapsed = max((today - start_month).days + 1, 1)

    monthly_df = df_query(
        conn,
        dedent(
            """
            SELECT user_id,
                   COUNT(*) AS monthly_reports,
                   MAX(date(period_start)) AS last_report_date
            FROM work_reports
            WHERE period_type='daily'
              AND strftime('%Y-%m', period_start) = strftime('%Y-%m', 'now')
            GROUP BY user_id
            """
        ),
    ).rename(columns={"last_report_date": "last_report_month"})

    lifetime_df = df_query(
        conn,
        dedent(
            """
            SELECT user_id,
                   COUNT(*) AS total_reports,
                   MIN(date(period_start)) AS first_report_date,
                   MAX(date(period_start)) AS last_report_date
            FROM work_reports
            WHERE period_type='daily'
            GROUP BY user_id
            """
        ),
    ).rename(columns={"last_report_date": "last_report_all"})

    def _parse_date(value: object) -> Optional[date]:
        try:
            parsed = pd.to_datetime(value, errors="coerce")
        except Exception:
            return None

        if pd.isna(parsed):
            return None

        try:
            return parsed.date()
        except AttributeError:
            return None

    staff_df["user_id"] = staff_df["user_id"].apply(lambda val: int(float(val)))
    staff_df = staff_df.rename(columns={"username": "Team member"})

    merged = (
        staff_df.merge(monthly_df, on="user_id", how="left")
        .merge(lifetime_df, on="user_id", how="left")
    )

    merged[["monthly_reports", "total_reports"]] = merged[
        ["monthly_reports", "total_reports"]
    ].fillna(0)

    kpi_rows: list[dict[str, object]] = []
    for _, row in merged.iterrows():
        monthly_reports = int(_coerce_float(row.get("monthly_reports"), 0.0))
        total_reports = int(_coerce_float(row.get("total_reports"), 0.0))
        last_seen = _parse_date(row.get("last_report_month")) or _parse_date(
            row.get("last_report_all")
        )
        first_seen = _parse_date(row.get("first_report_date"))

        days_since_last = (today - last_seen).days if last_seen else days_elapsed + 7
        months_active = (
            (today.year - first_seen.year) * 12 + today.month - first_seen.month + 1
            if first_seen
            else 1
        )
        months_active = max(months_active, 1)

        monthly_completion = (monthly_reports / days_elapsed) * 100
        recency_penalty = min(days_since_last * 3.0, 60.0)
        momentum_boost = min((total_reports / months_active) * 2.5, 25.0)
        monthly_score = max(
            0.0,
            min(100.0, monthly_completion - recency_penalty + 20.0 + momentum_boost),
        )

        lifetime_velocity = (total_reports / months_active) / 20.0
        lifetime_score = max(
            0.0,
            min(
                100.0,
                (lifetime_velocity * 100.0) - min(days_since_last * 1.5, 40.0) + 20.0,
            ),
        )

        kpi_rows.append(
            {
                "Team member": clean_text(row.get("Team member"))
                or f"User #{int(row.get('user_id'))}",
                "Monthly KPI": f"{monthly_score:,.0f}/100",
                "Lifetime KPI": f"{lifetime_score:,.0f}/100",
            }
        )

    kpi_table = pd.DataFrame(kpi_rows)
    if kpi_table.empty:
        st.caption("KPI scores will appear after the team logs their first daily reports.")
        return

    kpi_table = kpi_table[["Team member", "Monthly KPI", "Lifetime KPI"]]

    st.markdown("##### 🛰️ Admin KPI snapshot")
    st.caption("Simplified lifetime and monthly scores for quick admin review.")
    st.dataframe(
        kpi_table,
        use_container_width=True,
        hide_index=True,
    )


# ---------- Pages ----------
def dashboard(conn):
    st.subheader("📊 Dashboard")
    st.markdown(
        "<div style='text-align: right; font-size: 0.6rem; color: #888;'>by ZAD</div>",
        unsafe_allow_html=True,
    )
    header_cols = st.columns((0.85, 0.15))
    with header_cols[1]:
        render_notification_bell(conn)
    user = st.session_state.user or {}
    is_admin = user.get("role") == "admin"
    current_actor_id = current_user_id()
    allowed_customers = accessible_customer_ids(conn)
    scope_clause, scope_params = customer_scope_filter("c")
    sales_scope_clause, sales_scope_params = sales_scope_filter("d")
    sales_metrics = fetch_sales_metrics(conn, sales_scope_clause, sales_scope_params)

    announcement = df_query(
        conn,
        dedent(
            """
            SELECT dr.remark_id, dr.note, dr.created_at, COALESCE(u.username, 'User') AS author
            FROM dashboard_remarks dr
            LEFT JOIN users u ON u.user_id = dr.user_id
            ORDER BY datetime(dr.created_at) DESC
            LIMIT 1
            """
        ),
    )
    current_announcement = announcement.iloc[0] if not announcement.empty else None

    st.markdown("#### Admin announcement for all staff")
    if current_announcement is None:
        st.caption("No announcement is currently set.")
    else:
        st.info(
            f"**{current_announcement['note']}**\n\n"
            f"Posted by {current_announcement['author']} on "
            f"{fmt_dates(announcement, ['created_at']).iloc[0]['created_at']}",
            icon="📢",
        )

    if is_admin:
        if st.session_state.pop("dashboard_remark_reset", False):
            st.session_state["dashboard_remark_text"] = ""
        st.session_state.setdefault(
            "dashboard_remark_text",
            clean_text(current_announcement["note"]) if current_announcement is not None else "",
        )

        st.markdown("##### Update or clear the announcement")
        with st.form("dashboard_remark_form"):
            remark_text = st.text_area(
                "Message visible to all staff",
                help="Admins can share reminders or updates. The latest message stays pinned until replaced or deleted.",
                key="dashboard_remark_text",
            )
            submit_remark = st.form_submit_button("Save announcement", type="primary")

        if submit_remark:
            cleaned_note = clean_text(remark_text)
            if not cleaned_note:
                st.warning("Please enter a message before saving.")
            else:
                conn.execute(
                    "INSERT INTO dashboard_remarks (user_id, note) VALUES (?, ?)",
                    (current_user_id(), cleaned_note),
                )
                conn.commit()
                st.success("Announcement saved for all staff.")
                st.session_state["dashboard_remark_reset"] = True
                _safe_rerun()

        if current_announcement is not None:
            if st.button("Delete current announcement", type="secondary"):
                conn.execute(
                    "DELETE FROM dashboard_remarks WHERE remark_id = ?",
                    (int(current_announcement["remark_id"]),),
                )
                conn.commit()
                st.success("Announcement removed.")
                st.session_state["dashboard_remark_reset"] = True
                _safe_rerun()

    if "show_today_expired" not in st.session_state:
        st.session_state.show_today_expired = False

    if is_admin:
        col1, col2, col3, col4 = st.columns(4)
        complete_count = int(
            df_query(conn, f"SELECT COUNT(*) c FROM customers WHERE {customer_complete_clause()}").iloc[0]["c"]
        )
        scrap_count = int(
            df_query(conn, f"SELECT COUNT(*) c FROM customers WHERE {customer_incomplete_clause()}").iloc[0]["c"]
        )
        with col1:
            st.metric("Customers", complete_count)
        with col2:
            st.metric("Scraps", scrap_count)
        with col3:
            st.metric(
                "Active Warranties",
                int(
                    df_query(
                        conn,
                        "SELECT COUNT(*) c FROM warranties WHERE status='active' AND date(expiry_date) >= date('now')",
                    ).iloc[0]["c"]
                ),
            )
        with col4:
            expired_count = int(
                df_query(
                    conn,
                    "SELECT COUNT(*) c FROM warranties WHERE status='active' AND date(expiry_date) < date('now')",
                ).iloc[0]["c"]
            )
            st.metric("Expired", expired_count)

        st.markdown("#### Sales performance")
        sales_cols = st.columns(3)
        sales_cols[0].metric("Daily sales", format_sales_amount(sales_metrics["daily"]))
        sales_cols[1].metric("Weekly sales", format_sales_amount(sales_metrics["weekly"]))
        sales_cols[2].metric("Monthly sales", format_sales_amount(sales_metrics["monthly"]))

        st.markdown("#### Daily report coverage")
        report_date_value = st.date_input(
            "Review date",
            value=date.today(),
            key="dashboard_daily_report_date",
            help="Identify who submitted a daily report on the selected date.",
        )
        report_iso = to_iso_date(report_date_value) or date.today().isoformat()
        staff_df = df_query(
            conn,
            dedent(
                """
                SELECT user_id, username
                FROM users
                WHERE LOWER(COALESCE(role, 'staff')) <> 'admin'
                ORDER BY LOWER(username)
                """
            ),
        )

        if staff_df.empty:
            st.info("No staff accounts available for coverage tracking yet.")
        else:
            staff_df["user_id"] = staff_df["user_id"].apply(lambda val: int(float(val)))
            staff_df["username"] = staff_df.apply(
                lambda row: clean_text(row.get("username")) or f"User #{int(row['user_id'])}",
                axis=1,
            )
            submitted_df = df_query(
                conn,
                dedent(
                    """
                    SELECT DISTINCT user_id
                    FROM work_reports
                    WHERE period_type='daily' AND date(period_start)=date(?)
                    """
                ),
                (report_iso,),
            )
            submitted_ids: set[int] = set()
            if not submitted_df.empty:
                submitted_ids = {
                    int(float(uid))
                    for uid in submitted_df["user_id"].dropna().tolist()
                }

            staff_df["Submitted"] = staff_df["user_id"].apply(lambda uid: uid in submitted_ids)
            total_staff = int(staff_df.shape[0])
            submitted_total = int(staff_df["Submitted"].sum())
            missing_total = total_staff - submitted_total

            coverage_cols = st.columns(3)
            coverage_cols[0].metric("Team members", total_staff)
            coverage_cols[1].metric("Reports filed", submitted_total)
            coverage_cols[2].metric("Missing reports", missing_total)

            missing_df = staff_df[~staff_df["Submitted"]]
            if missing_total:
                st.warning("Daily reports pending for the following team members:")
                st.markdown("\n".join(f"- {name}" for name in missing_df["username"]))
            else:
                st.success("All tracked team members have filed their daily report.")

            status_table = staff_df.rename(
                columns={"username": "Team member", "Submitted": "Daily report"}
            )
            status_table["Daily report"] = status_table["Daily report"].map(
                {True: "Submitted", False: "Missing"}
            )
            st.dataframe(
                status_table[["Team member", "Daily report"]],
                use_container_width=True,
            )
            st.caption(
                f"Coverage for {format_period_range(report_iso, report_iso)} • Admins are excluded from this list."
            )
            staff_options = [int(uid) for uid in staff_df["user_id"].tolist()]
            staff_labels = {
                int(row["user_id"]): row["username"] for _, row in staff_df.iterrows()
            }
            if staff_options:
                st.markdown("##### Review daily submissions")
                submitted_options = [
                    int(uid)
                    for uid in staff_df.loc[staff_df["Submitted"], "user_id"].tolist()
                ]
                default_staff_id = (
                    submitted_options[0] if submitted_options else staff_options[0]
                )
                try:
                    default_index = staff_options.index(default_staff_id)
                except ValueError:
                    default_index = 0
                selected_staff_id = int(
                    st.selectbox(
                        "Team member report",
                        staff_options,
                        index=default_index,
                        format_func=lambda uid: staff_labels.get(
                            int(uid), f"User #{int(uid)}"
                        ),
                        key="dashboard_daily_report_user",
                    )
                )
                selected_staff_name = staff_labels.get(
                    selected_staff_id, f"User #{selected_staff_id}"
                )
                report_detail = df_query(
                    conn,
                    dedent(
                        """
                        SELECT report_id,
                               tasks,
                               remarks,
                               research,
                               grid_payload,
                               attachment_path,
                               import_file_path,
                               report_template,
                               period_start,
                               period_end,
                               created_at,
                               updated_at
                        FROM work_reports
                        WHERE user_id=?
                          AND period_type='daily'
                          AND date(period_start) = date(?)
                        ORDER BY datetime(updated_at) DESC, report_id DESC
                        LIMIT 1
                        """
                    ),
                    (selected_staff_id, report_iso),
                )
                has_marked_submitted = bool(
                    staff_df.loc[staff_df["user_id"] == selected_staff_id, "Submitted"].any()
                )
                if report_detail.empty:
                    if has_marked_submitted:
                        st.warning(
                            f"No daily report could be located for {selected_staff_name}.",
                            icon="⚠️",
                        )
                    else:
                        st.info(
                            f"{selected_staff_name} has not submitted a daily report for "
                            f"{format_period_range(report_iso, report_iso)}.",
                        )
                else:
                    record = report_detail.iloc[0].to_dict()
                    template_key = _normalize_report_template(
                        record.get("report_template")
                    )
                    template_label = REPORT_TEMPLATE_LABELS.get(
                        template_key, "Service report"
                    )
                    st.markdown(f"**Template:** {template_label}")
                    st.markdown(
                        f"**Period:** {format_period_range(record.get('period_start'), record.get('period_end'))}"
                    )
                    submitted_label = (
                        format_time_ago(record.get("created_at"))
                        or format_period_range(record.get("created_at"), record.get("created_at"))
                    )
                    updated_label = (
                        format_time_ago(record.get("updated_at"))
                        or format_period_range(record.get("updated_at"), record.get("updated_at"))
                    )
                    grid_rows = parse_report_grid_payload(
                        record.get("grid_payload"), template_key=template_key
                    )
                    grid_df = format_report_grid_rows_for_display(
                        grid_rows, empty_ok=True, template_key=template_key
                    )
                    if not grid_df.empty:
                        st.markdown("**Submitted entries**")
                        st.dataframe(grid_df, use_container_width=True, hide_index=True)
                    summary_labels = {
                        "service": {
                            "tasks": "Tasks completed",
                            "remarks": "Remarks / blockers",
                            "research": "Research / learnings",
                        },
                        "sales": {
                            "tasks": "Tasks completed",
                            "remarks": "Remarks / blockers",
                            "research": "Research / learnings",
                        },
                        "follow_up": {
                            "tasks": "Notes",
                            "remarks": "Status",
                            "research": "Product detail",
                        },
                    }
                    label_map = summary_labels.get(
                        template_key, summary_labels["service"]
                    )
                    summary_cols = st.columns(3)
                    summary_fields = [
                        (label_map["tasks"], "tasks"),
                        (label_map["remarks"], "remarks"),
                        (label_map["research"], "research"),
                    ]
                    for col, (label, field) in zip(summary_cols, summary_fields):
                        with col:
                            st.markdown(f"**{label}**")
                            st.write(clean_text(record.get(field)) or "—")
                    summary_lines = [
                        f"Daily submission for {selected_staff_name}",
                        f"Template: {template_label}",
                        f"Date: {format_period_range(record.get('period_start'), record.get('period_end'))}",
                        "",
                        f"{label_map['tasks']}:",
                        clean_text(record.get("tasks")) or "—",
                        "",
                        f"{label_map['remarks']}:",
                        clean_text(record.get("remarks")) or "—",
                        "",
                        f"{label_map['research']}:",
                        clean_text(record.get("research")) or "—",
                        "",
                    ]
                    if not grid_df.empty:
                        summary_lines.extend(
                            [
                                "Entries:",
                                grid_df.to_csv(index=False).strip(),
                                "",
                            ]
                        )
                    summary_lines.extend(
                        [
                        f"Submitted: {submitted_label}",
                        f"Last updated: {updated_label}",
                        ]
                    )
                    summary_payload = "\n".join(summary_lines)
                    download_name = _sanitize_path_component(
                        f"daily_submission_{selected_staff_name}_{report_iso}"
                    )
                    st.download_button(
                        "Download daily submission",
                        data=summary_payload.encode("utf-8"),
                        file_name=f"{download_name or 'daily_submission'}.txt",
                        mime="text/plain",
                        key=f"dashboard_daily_submission_{record.get('report_id')}",
                    )
                    st.caption(
                        f"Submitted {submitted_label} • Last updated {updated_label}"
                    )
                    import_file_value = clean_text(record.get("import_file_path"))
                    if import_file_value:
                        import_path = resolve_upload_path(import_file_value)
                        import_bytes = None
                        if import_path and import_path.exists():
                            try:
                                import_bytes = import_path.read_bytes()
                            except OSError:
                                import_bytes = None
                        if import_bytes:
                            st.download_button(
                                "Download imported file",
                                data=import_bytes,
                                file_name=(
                                    import_path.name if import_path else "imported_report"
                                ),
                                key=f"dashboard_report_import_{record.get('report_id')}",
                            )
                        else:
                            st.warning(
                                "The imported file could not be found on disk.",
                                icon="⚠️",
                            )
                    else:
                        st.caption("No import file saved for this report.")
                    attachment_value = clean_text(record.get("attachment_path"))
                    if attachment_value:
                        attachment_path = resolve_upload_path(attachment_value)
                        attachment_bytes = None
                        if attachment_path and attachment_path.exists():
                            try:
                                attachment_bytes = attachment_path.read_bytes()
                            except OSError:
                                attachment_bytes = None
                        if attachment_bytes:
                            st.download_button(
                                "Download attachment",
                                data=attachment_bytes,
                                file_name=(
                                    attachment_path.name if attachment_path else "attachment"
                                ),
                                key=f"dashboard_report_attachment_{record.get('report_id')}",
                            )
                        else:
                            st.warning(
                                "The attached file could not be found on disk.",
                                icon="⚠️",
                            )
                    else:
                        st.caption("No attachment uploaded for this report.")
    else:
        st.info("Staff view: focus on upcoming activities below.")
        st.markdown("#### Your sales snapshot")
        staff_sales_cols = st.columns(3)
        staff_sales_cols[0].metric("Today", format_sales_amount(sales_metrics["daily"]))
        staff_sales_cols[1].metric("Last 7 days", format_sales_amount(sales_metrics["weekly"]))
        staff_sales_cols[2].metric("This month", format_sales_amount(sales_metrics["monthly"]))

        staff_sales_df = df_query(
            conn,
            """
            SELECT d.do_number,
                   d.customer_id,
                   d.created_by,
                   d.record_type,
                   d.items_payload,
                   d.description,
                   d.total_amount,
                   d.status,
                   d.created_at,
                   COALESCE(c.company_name, c.name, '(customer)') AS customer
            FROM delivery_orders d
            LEFT JOIN customers c ON c.customer_id = d.customer_id
            WHERE d.deleted_at IS NULL
              AND COALESCE(d.record_type, 'delivery_order') IN ('delivery_order', 'work_done')
            ORDER BY datetime(d.created_at) DESC, d.do_number DESC
            LIMIT 30
            """,
        )
        staff_sales_df = filter_delivery_orders_for_view(
            staff_sales_df,
            allowed_customers,
            record_types={"delivery_order", "work_done"},
        )
        staff_service_df = df_query(
            conn,
            """
            SELECT s.service_id,
                   s.do_number,
                   s.customer_id,
                   s.created_by,
                   s.service_product_info,
                   s.description,
                   s.bill_amount,
                   s.payment_status,
                   s.updated_at,
                   COALESCE(c.company_name, c.name, '(customer)') AS customer
            FROM services s
            LEFT JOIN customers c ON c.customer_id = s.customer_id
            WHERE s.deleted_at IS NULL
            ORDER BY datetime(s.updated_at) DESC, s.service_id DESC
            LIMIT 30
            """,
        )
        staff_maintenance_df = df_query(
            conn,
            """
            SELECT m.maintenance_id,
                   m.do_number,
                   m.customer_id,
                   m.created_by,
                   m.maintenance_product_info,
                   m.description,
                   m.total_amount,
                   m.payment_status,
                   m.updated_at,
                   COALESCE(c.company_name, c.name, '(customer)') AS customer
            FROM maintenance_records m
            LEFT JOIN customers c ON c.customer_id = m.customer_id
            WHERE m.deleted_at IS NULL
            ORDER BY datetime(m.updated_at) DESC, m.maintenance_id DESC
            LIMIT 30
            """,
        )
        if allowed_customers is not None:
            staff_service_df = staff_service_df[
                staff_service_df["customer_id"].apply(
                    lambda value: pd.notna(value) and int(value) in allowed_customers
                )
            ]
            staff_maintenance_df = staff_maintenance_df[
                staff_maintenance_df["customer_id"].apply(
                    lambda value: pd.notna(value) and int(value) in allowed_customers
                )
            ]

        def _format_payment_label(value: Optional[str]) -> str:
            cleaned = clean_text(value)
            if not cleaned:
                return "Pending"
            return cleaned.replace("_", " ").title()

        sales_frames: list[pd.DataFrame] = []
        if not staff_sales_df.empty:
            staff_sales_df["Reference"] = staff_sales_df["do_number"]
            staff_sales_df["Products"] = staff_sales_df.apply(
                lambda row: dedupe_join(
                    format_simple_item_labels(parse_delivery_items_payload(row.get("items_payload")))
                )
                or clean_text(row.get("description"))
                or "—",
                axis=1,
            )
            staff_sales_df["Sales date"] = pd.to_datetime(
                staff_sales_df["created_at"], errors="coerce"
            ).dt.date
            staff_sales_df["When"] = staff_sales_df["created_at"].apply(
                lambda value: format_time_ago(value) or format_period_range(value, value)
            )
            staff_sales_df["Total (BDT)"] = staff_sales_df["total_amount"].apply(format_sales_amount)
            staff_sales_df["Status"] = staff_sales_df["status"].apply(_format_payment_label)
            staff_sales_df["sort_date"] = pd.to_datetime(
                staff_sales_df["created_at"], errors="coerce"
            )
            sales_frames.append(
                staff_sales_df[
                    [
                        "Reference",
                        "customer",
                        "Products",
                        "Sales date",
                        "Total (BDT)",
                        "Status",
                        "When",
                        "sort_date",
                    ]
                ].rename(columns={"customer": "Customer"})
            )

        if not staff_service_df.empty:
            staff_service_df["Reference"] = staff_service_df.apply(
                lambda row: clean_text(row.get("do_number"))
                or f"Service #{int(row.get('service_id'))}"
                if pd.notna(row.get("service_id"))
                else "Service",
                axis=1,
            )
            staff_service_df["Products"] = staff_service_df.apply(
                lambda row: clean_text(row.get("service_product_info"))
                or clean_text(row.get("description"))
                or "—",
                axis=1,
            )
            staff_service_df["Sales date"] = pd.to_datetime(
                staff_service_df["updated_at"], errors="coerce"
            ).dt.date
            staff_service_df["When"] = staff_service_df["updated_at"].apply(
                lambda value: format_time_ago(value) or format_period_range(value, value)
            )
            staff_service_df["Total (BDT)"] = staff_service_df["bill_amount"].apply(format_sales_amount)
            staff_service_df["Status"] = staff_service_df["payment_status"].apply(_format_payment_label)
            staff_service_df["sort_date"] = pd.to_datetime(
                staff_service_df["updated_at"], errors="coerce"
            )
            sales_frames.append(
                staff_service_df[
                    [
                        "Reference",
                        "customer",
                        "Products",
                        "Sales date",
                        "Total (BDT)",
                        "Status",
                        "When",
                        "sort_date",
                    ]
                ].rename(columns={"customer": "Customer"})
            )

        if not staff_maintenance_df.empty:
            staff_maintenance_df["Reference"] = staff_maintenance_df.apply(
                lambda row: clean_text(row.get("do_number"))
                or f"Maintenance #{int(row.get('maintenance_id'))}"
                if pd.notna(row.get("maintenance_id"))
                else "Maintenance",
                axis=1,
            )
            staff_maintenance_df["Products"] = staff_maintenance_df.apply(
                lambda row: clean_text(row.get("maintenance_product_info"))
                or clean_text(row.get("description"))
                or "—",
                axis=1,
            )
            staff_maintenance_df["Sales date"] = pd.to_datetime(
                staff_maintenance_df["updated_at"], errors="coerce"
            ).dt.date
            staff_maintenance_df["When"] = staff_maintenance_df["updated_at"].apply(
                lambda value: format_time_ago(value) or format_period_range(value, value)
            )
            staff_maintenance_df["Total (BDT)"] = staff_maintenance_df["total_amount"].apply(
                format_sales_amount
            )
            staff_maintenance_df["Status"] = staff_maintenance_df["payment_status"].apply(
                _format_payment_label
            )
            staff_maintenance_df["sort_date"] = pd.to_datetime(
                staff_maintenance_df["updated_at"], errors="coerce"
            )
            sales_frames.append(
                staff_maintenance_df[
                    [
                        "Reference",
                        "customer",
                        "Products",
                        "Sales date",
                        "Total (BDT)",
                        "Status",
                        "When",
                        "sort_date",
                    ]
                ].rename(columns={"customer": "Customer"})
            )

        if not sales_frames:
            st.caption("No personal sales have been logged yet.")
        else:
            sales_snapshot = (
                pd.concat(sales_frames, ignore_index=True)
                .sort_values(by="sort_date", ascending=False)
                .head(30)
            )
            st.dataframe(
                sales_snapshot[
                    [
                        "Reference",
                        "Customer",
                        "Products",
                        "Sales date",
                        "Total (BDT)",
                        "Status",
                        "When",
                    ]
                ].rename(
                    columns={"Reference": "DO/Work No./Service/Maintenance"}
                ),
                use_container_width=True,
                hide_index=True,
            )

    report_scope = ""
    report_params: tuple[object, ...] = ()
    viewer_id = current_user_id()
    if not is_admin:
        if viewer_id is not None:
            report_scope = "WHERE wr.user_id = ?"
            report_params = (viewer_id,)
        else:
            report_scope = "WHERE 1=0"

    report_metrics = df_query(
        conn,
        dedent(
            f"""
            SELECT COUNT(*) AS total_reports,
                   SUM(CASE WHEN date(wr.created_at) >= date('now', '-6 days') THEN 1 ELSE 0 END) AS weekly_reports,
                   SUM(CASE WHEN strftime('%Y-%m', wr.created_at) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) AS monthly_reports
            FROM work_reports wr
            {report_scope}
            """
        ),
        report_params,
    )
    recent_reports = df_query(
        conn,
        dedent(
            f"""
            SELECT wr.report_id,
                   wr.period_type,
                   wr.period_start,
                   wr.period_end,
                   wr.created_at,
                   wr.report_template,
                   COALESCE(u.username, 'User #' || wr.user_id) AS owner
            FROM work_reports wr
            LEFT JOIN users u ON u.user_id = wr.user_id
            {report_scope}
            ORDER BY datetime(wr.created_at) DESC, wr.report_id DESC
            LIMIT 6
            """
        ),
        report_params,
    )

    if not report_metrics.empty:
        total_reports = int(report_metrics.iloc[0].get("total_reports") or 0)
        weekly_reports = int(report_metrics.iloc[0].get("weekly_reports") or 0)
        monthly_reports = int(report_metrics.iloc[0].get("monthly_reports") or 0)
    else:
        total_reports = 0
        weekly_reports = 0
        monthly_reports = 0

    st.markdown("#### Report submissions")
    report_metric_cols = st.columns(3)
    report_metric_cols[0].metric("Total reports", total_reports)
    report_metric_cols[1].metric("Weekly reports", weekly_reports)
    report_metric_cols[2].metric("Monthly reports", monthly_reports)

    if not recent_reports.empty:
        st.markdown("#### Recent report submissions")
        recent_reports["Template"] = recent_reports["report_template"].apply(
            lambda value: REPORT_TEMPLATE_LABELS.get(
                _normalize_report_template(value), "Service report"
            )
        )
        recent_reports["Submitted"] = recent_reports["created_at"].apply(
            lambda value: format_period_range(value, value)
        )
        recent_reports["Cadence"] = recent_reports["period_type"].apply(
            lambda val: REPORT_PERIOD_OPTIONS.get(clean_text(val) or "", str(val).title())
        )
        recent_reports["When"] = recent_reports["created_at"].apply(
            lambda value: format_time_ago(value) or format_period_range(value, value)
        )
        display_cols = ["Team member", "Template", "Cadence", "Submitted", "When"]
        recent_display = recent_reports.rename(columns={"owner": "Team member"})
        st.dataframe(
            recent_display[display_cols],
            use_container_width=True,
            hide_index=True,
        )

    uploads_scope = report_scope.replace("WHERE", "AND", 1) if report_scope else ""
    uploads_params = tuple(report_params)
    uploads_df = df_query(
        conn,
        dedent(
            f"""
            SELECT wr.report_id,
                   wr.attachment_path,
                   wr.import_file_path,
                   wr.period_start,
                   wr.period_end,
                   wr.report_template,
                   COALESCE(u.username, 'User #' || wr.user_id) AS owner,
                   wr.created_at
            FROM work_reports wr
            LEFT JOIN users u ON u.user_id = wr.user_id
            WHERE wr.attachment_path IS NOT NULL AND wr.attachment_path != ''
            {uploads_scope}
            ORDER BY datetime(wr.created_at) DESC
            LIMIT 5
            """
        ),
        uploads_params,
    )

    if not uploads_df.empty:
        st.markdown("#### Latest report uploads")
        uploads_df["Template"] = uploads_df["report_template"].apply(
            lambda value: REPORT_TEMPLATE_LABELS.get(
                _normalize_report_template(value), "Service report"
            )
        )
        uploads_df["Period"] = uploads_df.apply(
            lambda row: format_period_range(row.get("period_start"), row.get("period_end")),
            axis=1,
        )
        uploads_df["When"] = uploads_df["created_at"].apply(
            lambda value: format_time_ago(value) or format_period_range(value, value)
        )
        for _, row in uploads_df.iterrows():
            file_entries = []
            attachment_value = clean_text(row.get("attachment_path"))
            if attachment_value:
                file_entries.append(("Attachment", attachment_value))
            import_value = clean_text(row.get("import_file_path"))
            if import_value:
                file_entries.append(("Import", import_value))
            if not file_entries:
                continue
            label_prefix = (
                f"{row.get('owner')} • {row.get('Template')} • {row.get('Period')}"
            )
            for idx, (kind, file_value) in enumerate(file_entries, start=1):
                path = resolve_upload_path(file_value)
                if not path or not path.exists():
                    continue
                try:
                    payload = path.read_bytes()
                except OSError:
                    continue
                label = f"{label_prefix} • {kind}"
                st.download_button(
                    label,
                    data=payload,
                    file_name=path.name,
                    key=f"recent_attachment_{row.get('report_id')}_{idx}",
                )

    if is_admin:
        recent_template_reports = df_query(
            conn,
            dedent(
                """
                SELECT wr.report_id,
                       wr.period_start,
                       wr.period_end,
                       wr.report_template,
                       wr.grid_payload,
                       wr.attachment_path,
                       wr.import_file_path,
                       wr.created_at,
                       COALESCE(u.username, 'User #' || wr.user_id) AS owner
                FROM work_reports wr
                LEFT JOIN users u ON u.user_id = wr.user_id
                WHERE LOWER(COALESCE(wr.report_template, '')) IN ('service', 'sales', 'follow_up')
                ORDER BY datetime(wr.created_at) DESC, wr.report_id DESC
                LIMIT 6
                """
            ),
        )
        if not recent_template_reports.empty:
            st.markdown("#### Service & follow-up report snapshots")
            recent_template_reports["Template"] = recent_template_reports[
                "report_template"
            ].apply(
                lambda value: REPORT_TEMPLATE_LABELS.get(
                    _normalize_report_template(value), "Service report"
                )
            )
            recent_template_reports["Period"] = recent_template_reports.apply(
                lambda row: format_period_range(
                    row.get("period_start"), row.get("period_end")
                ),
                axis=1,
            )
            for _, row in recent_template_reports.iterrows():
                header = (
                    f"{row.get('owner')} • {row.get('Template')} • {row.get('Period')}"
                )
                with st.expander(header, expanded=False):
                    template_key = _normalize_report_template(row.get("report_template"))
                    grid_rows = parse_report_grid_payload(
                        row.get("grid_payload"), template_key=template_key
                    )
                    grid_df = format_report_grid_rows_for_display(
                        grid_rows, empty_ok=True, template_key=template_key
                    )
                    if grid_df.empty:
                        st.caption("No structured entries were captured for this report.")
                    else:
                        st.dataframe(grid_df, use_container_width=True, hide_index=True)
                        csv_payload = grid_df.to_csv(index=False).encode("utf-8")
                        st.download_button(
                            "Download report entries (CSV)",
                            data=csv_payload,
                            file_name="report_entries.csv",
                            key=f"dashboard_report_csv_{row.get('report_id')}",
                        )
                    file_options = [
                        ("Import file", row.get("import_file_path")),
                        ("Attachment", row.get("attachment_path")),
                    ]
                    for label, path_value in file_options:
                        clean_path = clean_text(path_value)
                        if not clean_path:
                            continue
                        resolved = resolve_upload_path(clean_path)
                        if not resolved or not resolved.exists():
                            continue
                        try:
                            file_bytes = resolved.read_bytes()
                        except OSError:
                            continue
                        st.download_button(
                            f"Download {label.lower()}",
                            data=file_bytes,
                            file_name=resolved.name,
                            key=f"dashboard_report_file_{row.get('report_id')}_{label}",
                        )

        uploaded_quotes = df_query(
            conn,
            dedent(
                """
                SELECT q.quotation_id,
                       q.reference,
                       q.customer_company,
                       q.customer_name,
                       q.customer_contact,
                       q.total_amount,
                       q.status,
                       q.quote_date,
                       q.document_path,
                       q.follow_up_status,
                       q.follow_up_date,
                       q.follow_up_notes,
                       q.salesperson_name,
                       q.updated_at,
                       COALESCE(u.username, 'User #' || q.created_by) AS owner
                FROM quotations q
                LEFT JOIN users u ON u.user_id = q.created_by
                WHERE q.document_path IS NOT NULL AND q.document_path != ''
                  AND q.deleted_at IS NULL
                ORDER BY datetime(q.updated_at) DESC, datetime(q.quote_date) DESC, q.quotation_id DESC
                LIMIT 12
                """
            ),
        )

        if not uploaded_quotes.empty:
            uploaded_quotes = fmt_dates(
                uploaded_quotes, ["quote_date", "follow_up_date", "updated_at"]
            )
            uploaded_quotes["quotation_id"] = uploaded_quotes["quotation_id"].apply(
                lambda value: int(_coerce_float(value, 0))
            )
            with st.expander("Uploaded quotation details", expanded=False):
                option_labels = {
                    int(row["quotation_id"]): (
                        f"{clean_text(row.get('reference')) or 'Quotation'} • "
                        f"{clean_text(row.get('customer_company')) or clean_text(row.get('customer_name')) or 'Customer'}"
                    )
                    for _, row in uploaded_quotes.iterrows()
                }
                quote_options = list(option_labels.keys())
                selected_quote_id = st.selectbox(
                    "Select a quotation to inspect",
                    quote_options,
                    format_func=lambda qid: option_labels.get(qid, f"Quotation #{qid}"),
                    key="dashboard_uploaded_quote_selector",
                )

                quote_match = uploaded_quotes[
                    uploaded_quotes["quotation_id"] == selected_quote_id
                ].iloc[0]
                left, right = st.columns((2, 1))
                customer_lines = [
                    f"**Customer:** {clean_text(quote_match.get('customer_company')) or clean_text(quote_match.get('customer_name')) or '—'}",
                    f"**Contact:** {clean_text(quote_match.get('customer_contact')) or '—'}",
                ]
                if quote_match.get("follow_up_status"):
                    follow_label = clean_text(quote_match.get("follow_up_status"))
                    follow_date = clean_text(quote_match.get("follow_up_date")) or "—"
                    customer_lines.append(
                        f"**Follow-up:** {follow_label.title()} • {follow_date}"
                    )
                if quote_match.get("follow_up_notes"):
                    customer_lines.append(
                        f"**Admin remarks:** {clean_text(quote_match.get('follow_up_notes'))}"
                    )
                left.markdown("\n".join(customer_lines))

                status_label = clean_text(quote_match.get("status")) or "Pending"
                amount_label = format_money(quote_match.get("total_amount")) or f"{_coerce_float(quote_match.get('total_amount'), 0.0):,.2f}"
                right.metric(
                    "Total (BDT)",
                    amount_label,
                    help=f"Status: {status_label.title()}",
                )
                right.caption(
                    f"Sales: {clean_text(quote_match.get('salesperson_name')) or clean_text(quote_match.get('owner')) or '—'}"
                )
                right.caption(
                    f"Last updated {format_time_ago(quote_match.get('updated_at')) or quote_match.get('updated_at') or '—'}"
                )

                doc_path = clean_text(quote_match.get("document_path"))
                resolved_path = resolve_upload_path(doc_path)
                if resolved_path and resolved_path.exists():
                    try:
                        payload = resolved_path.read_bytes()
                    except OSError:
                        payload = None
                    if payload:
                        st.download_button(
                            "Download uploaded quotation",
                            payload,
                            file_name=resolved_path.name,
                            key=f"dashboard_quote_download_{selected_quote_id}",
                        )
                    else:
                        st.warning(
                            "The uploaded quotation file could not be read.", icon="⚠️"
                        )
                else:
                    st.caption("Uploaded quotation file not found on disk.")

    if is_admin:
        _render_admin_record_history(conn)

        st.markdown(
            """
            <style>
            @media (max-width: 768px) {
              [data-testid="stHorizontalBlock"] {
                flex-wrap: wrap;
              }
              [data-testid="stHorizontalBlock"] > div {
                flex: 1 1 100% !important;
                min-width: 100% !important;
              }
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("#### Team quotations (downloadable)")
        quote_search = st.text_input(
            "Search quotations",
            key="admin_quote_search",
            help="Search by reference, customer, or salesperson.",
        )
        quote_params: list[object] = []
        quote_filters = ["q.deleted_at IS NULL"]
        if quote_search:
            query_value = f"%{quote_search.strip()}%"
            quote_filters.append(
                "("
                "q.reference LIKE ? OR q.customer_company LIKE ? OR q.salesperson_name LIKE ? OR u.username LIKE ?"
                ")"
            )
            quote_params.extend([query_value, query_value, query_value, query_value])
        quote_where = " AND ".join(quote_filters)
        quote_limit = 50 if quote_search else 20
        staff_quotes = df_query(
            conn,
            dedent(
                f"""
                SELECT q.quotation_id,
                       q.reference,
                       q.customer_company,
                       q.total_amount,
                       q.status,
                       q.quote_date,
                       q.document_path,
                       q.salesperson_name,
                       u.username
                FROM quotations q
                LEFT JOIN users u ON u.user_id = q.created_by
                WHERE {quote_where}
                ORDER BY datetime(q.quote_date) DESC, q.quotation_id DESC
                LIMIT {quote_limit}
                """
            ),
            tuple(quote_params),
        )
        if not staff_quotes.empty:
            staff_quotes = fmt_dates(staff_quotes, ["quote_date"])
            if not quote_search:
                st.caption("Showing the latest 20 quotations. Use search to find older records.")
            with st.container():
                st.markdown(
                    "<div style='max-height: 320px; overflow-y: auto;'>",
                    unsafe_allow_html=True,
                )
                for _, row in staff_quotes.iterrows():
                    cols = st.columns((1.5, 1, 1, 1))
                    cols[0].markdown(
                        f"**{clean_text(row.get('reference')) or 'Quotation'}**\n"
                        f"{clean_text(row.get('customer_company')) or '(customer)'}"
                    )
                    cols[1].write(
                        clean_text(row.get("salesperson_name"))
                        or clean_text(row.get("username"))
                        or "—"
                    )
                    total_value = format_money(row.get("total_amount")) or f"{_coerce_float(row.get('total_amount'), 0.0):,.2f}"
                    cols[2].write(
                        f"{clean_text(row.get('status')).title() if row.get('status') else 'Pending'}\n"
                        f"{row.get('quote_date')}\n{total_value}"
                    )

                    download_key = f"dash_quote_{int(row.get('quotation_id'))}"
                    doc_path = clean_text(row.get("document_path"))
                    file_path = BASE_DIR / doc_path if doc_path else None

                    # Prefer a PDF version for dashboard downloads; fall back to any
                    # matching PDF file for the quotation ID if the stored path is
                    # missing or points to a non-PDF asset. If only an Excel workbook
                    # exists, regenerate the PDF from the workbook contents and persist
                    # it for future downloads.
                    pdf_candidate = None
                    regenerated_bytes: Optional[bytes] = None
                    if file_path and file_path.suffix.lower() != ".pdf":
                        alt_pdf = file_path.with_suffix(".pdf")
                        if alt_pdf.exists():
                            pdf_candidate = alt_pdf
                        elif file_path.exists():
                            regenerated_bytes = _regenerate_quotation_pdf_from_workbook(
                                file_path
                            )
                    if pdf_candidate is None and regenerated_bytes is None:
                        try:
                            quote_id = int(row.get("quotation_id"))
                        except Exception:
                            quote_id = None
                        if quote_id:
                            for alt_path in QUOTATION_RECEIPT_DIR.glob(
                                f"quotation_{quote_id}_*.pdf"
                            ):
                                pdf_candidate = alt_path
                                break
                    if pdf_candidate:
                        file_path = pdf_candidate

                    download_payload: Optional[bytes] = None
                    download_name: Optional[str] = None

                    if regenerated_bytes:
                        try:
                            quote_id = int(row.get("quotation_id"))
                        except Exception:
                            quote_id = None
                        persisted_path = None
                        if quote_id:
                            persisted_path = _persist_quotation_pdf(
                                quote_id,
                                regenerated_bytes,
                                clean_text(row.get("reference")),
                            )
                            if persisted_path:
                                try:
                                    conn.execute(
                                        "UPDATE quotations SET document_path=? WHERE quotation_id=? AND deleted_at IS NULL",
                                        (persisted_path, quote_id),
                                    )
                                    conn.commit()
                                    file_path = BASE_DIR / persisted_path
                                except sqlite3.Error:
                                    pass
                        download_payload = regenerated_bytes
                        download_name = (
                            file_path.name
                            if file_path and file_path.exists()
                            else f"{Path(doc_path).stem}.pdf" if doc_path else "quotation.pdf"
                        )

                    if download_payload is None and file_path and file_path.exists():
                        try:
                            download_payload = file_path.read_bytes()
                        except OSError:
                            download_payload = None
                        download_name = file_path.name

                    if download_payload:
                        cols[3].download_button(
                            "Download PDF" if download_name.endswith(".pdf") else "Download",
                            download_payload,
                            file_name=download_name,
                            key=download_key,
                            mime="application/pdf"
                            if download_name.endswith(".pdf")
                            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    else:
                        cols[3].caption("File missing")
                st.markdown("</div>", unsafe_allow_html=True)
        else:
            st.info("No quotations found for the current filters.")

    quote_scope, quote_params = _quotation_scope_filter()
    quote_clause = quote_scope.replace("WHERE", "WHERE", 1)
    quote_metrics = df_query(
        conn,
        dedent(
            f"""
            SELECT COUNT(*) AS total_quotes,
                   SUM(CASE WHEN date(quote_date) >= date('now', '-6 days') THEN 1 ELSE 0 END) AS weekly_quotes,
                   SUM(CASE WHEN strftime('%Y-%m', quote_date) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) AS monthly_quotes,
                   SUM(CASE WHEN LOWER(status) = 'paid' THEN 1 ELSE 0 END) AS paid_quotes
            FROM quotations
            {quote_clause}
            """
        ),
        quote_params,
    )
    quotes_df = df_query(
        conn,
        dedent(
            f"""
            SELECT quotation_id,
                   reference,
                   quote_date,
                   total_amount,
                   status,
                   COALESCE(
                       NULLIF(TRIM(customer_company), ''),
                       NULLIF(TRIM(customer_name), ''),
                       '—'
                   ) AS customer,
                   COALESCE(
                       NULLIF(TRIM(subject), ''),
                       NULLIF(TRIM(remarks_internal), '')
                   ) AS product_details
            FROM quotations
            {quote_clause}
            ORDER BY datetime(quote_date) DESC, quotation_id DESC
            LIMIT 20
            """
        ),
        quote_params,
    )
    st.markdown("#### Quotation insights")
    if quote_metrics.empty:
        total_quotes = 0
        weekly_quotes = 0
        monthly_quotes = 0
        paid_quotes = 0
    else:
        total_quotes = int(quote_metrics.iloc[0].get("total_quotes") or 0)
        weekly_quotes = int(quote_metrics.iloc[0].get("weekly_quotes") or 0)
        monthly_quotes = int(quote_metrics.iloc[0].get("monthly_quotes") or 0)
        paid_quotes = int(quote_metrics.iloc[0].get("paid_quotes") or 0)
    conversion = (paid_quotes / total_quotes) * 100 if total_quotes else 0.0
    metrics_cols = st.columns(5)
    metrics_cols[0].metric("Quotations created", total_quotes)
    metrics_cols[1].metric("Weekly quotations", weekly_quotes)
    metrics_cols[2].metric("Monthly quotations", monthly_quotes)
    metrics_cols[3].metric("Paid / converted", paid_quotes)
    metrics_cols[4].metric("Conversion rate", f"{conversion:.1f}%")

    if quotes_df.empty:
        st.info("No quotations available for the selected scope yet.")
    else:
        quotes_df = fmt_dates(quotes_df, ["quote_date"])
        quotes_df["customer"] = quotes_df["customer"].apply(
            lambda value: clean_text(value) or "—"
        )
        quotes_df["product_details"] = quotes_df["product_details"].apply(
            lambda value: clean_text(value) or "—"
        )
        quotes_df["total_amount"] = quotes_df["total_amount"].apply(
            lambda value: format_money(value) or f"{_coerce_float(value, 0.0):,.2f}"
        )
        st.dataframe(
            quotes_df.rename(
                columns={
                    "reference": "Reference",
                    "quote_date": "Date",
                    "total_amount": "Total (BDT)",
                    "status": "Status",
                    "customer": "Customer",
                    "product_details": "Product details",
                }
            ),
            use_container_width=True,
            hide_index=True,
        )

    staff_scope_clause = ""
    staff_scope_params: tuple[object, ...] = ()
    if not is_admin:
        user_id = current_user_id()
        if user_id is None:
            staff_scope_clause = "1=0"
        else:
            staff_scope_clause = "customer_id IN (SELECT customer_id FROM customers WHERE created_by=?)"
            staff_scope_params = (user_id,)

    warranty_where = "status='active' AND date(expiry_date) < date('now')"
    if staff_scope_clause:
        warranty_where = f"{warranty_where} AND {staff_scope_clause}"
    month_expired_current, month_expired_previous = month_bucket_counts(
        conn,
        "warranties",
        "expiry_date",
        where=warranty_where,
        params=staff_scope_params,
    )
    month_expired = month_expired_current
    expired_delta = format_metric_delta(month_expired_current, month_expired_previous)

    service_where = staff_scope_clause if staff_scope_clause else None
    service_month_current, service_month_previous = month_bucket_counts(
        conn,
        "services",
        "service_date",
        where=service_where,
        params=staff_scope_params if staff_scope_clause else None,
    )
    service_month = service_month_current
    service_delta = format_metric_delta(service_month_current, service_month_previous)

    maintenance_where = staff_scope_clause if staff_scope_clause else None
    maintenance_month_current, maintenance_month_previous = month_bucket_counts(
        conn,
        "maintenance_records",
        "maintenance_date",
        where=maintenance_where,
        params=staff_scope_params if staff_scope_clause else None,
    )
    maintenance_month = maintenance_month_current
    maintenance_delta = format_metric_delta(
        maintenance_month_current, maintenance_month_previous
    )
    today_expired_df = df_query(
        conn,
        dedent(
            f"""
            SELECT c.name AS customer,
                   p.name AS product,
                   p.model,
                   w.serial,
                   COALESCE(w.issue_date, c.purchase_date) AS issue_date,
                   w.expiry_date,
                   COALESCE(c.sales_person, u.username) AS staff
            FROM warranties w
            LEFT JOIN customers c ON c.customer_id = w.customer_id
            LEFT JOIN products p ON p.product_id = w.product_id
            LEFT JOIN users u ON u.user_id = c.created_by
            WHERE w.status='active' AND date(w.expiry_date) = date('now')
            {f"AND {scope_clause}" if scope_clause else ""}
            ORDER BY date(w.expiry_date) ASC
            """
        ),
        scope_params,
    )
    today_expired_count = len(today_expired_df.index)
    col5, col6, col7, col8 = st.columns(4)
    with col5:
        st.metric("Expired this month", month_expired, delta=expired_delta)
    with col6:
        st.metric("Services this month", service_month, delta=service_delta)
    with col7:
        st.metric(
            "Maintenance this month",
            maintenance_month,
            delta=maintenance_delta,
        )
    with col8:
        st.metric("Expired today", today_expired_count)
        toggle_label = "Show list" if not st.session_state.get("show_today_expired") else "Hide list"
        if st.button(toggle_label, key="toggle_expired_today"):
            st.session_state.show_today_expired = not st.session_state.get("show_today_expired")
            show_today_expired = st.session_state.show_today_expired
        else:
            show_today_expired = st.session_state.get("show_today_expired")

    if not today_expired_df.empty:
        notice = collapse_warranty_rows(today_expired_df)
        lines = []
        for _, row in notice.iterrows():
            customer = row.get("Customer") or "(unknown)"
            description = row.get("Description") or ""
            staff_label = row.get("Staff") or ""
            if description:
                staff_note = f" (by {staff_label})" if staff_label else ""
                lines.append(f"- {customer}: {description}{staff_note}")
            else:
                staff_note = f" (by {staff_label})" if staff_label else ""
                lines.append(f"- {customer}{staff_note}")
        st.warning("⚠️ Warranties expiring today:\n" + "\n".join(lines))

    show_today_expired = st.session_state.get("show_today_expired")
    if show_today_expired:
        if today_expired_df.empty:
            st.info("No warranties expire today.")
        else:
            today_detail = fmt_dates(today_expired_df, ["issue_date", "expiry_date"])
            today_table = format_warranty_table(today_detail)
            st.markdown("#### Warranties expiring today")
            st.dataframe(today_table, use_container_width=True)

    if is_admin:
        if "show_deleted_panel" not in st.session_state:
            st.session_state.show_deleted_panel = False

        export_state = st.session_state.setdefault("admin_export_state", {})
        downloads_enabled = st.checkbox(
            "Enable admin downloads",
            key="admin_downloads_enabled",
        )
        if downloads_enabled and st.button("Prepare downloads", use_container_width=True):
            # Generate admin export bytes only on explicit request (IDM-safe).
            excel_bytes = export_database_to_excel(conn, include_all=True)
            export_state["excel_bytes"] = excel_bytes
            export_state["archive_bytes"] = export_full_archive(conn, excel_bytes)
        excel_bytes = export_state.get("excel_bytes")
        archive_bytes = export_state.get("archive_bytes")
        if downloads_enabled:
            if not excel_bytes and not archive_bytes:
                st.info("Click “Prepare downloads” to generate export files.")
            download_cols = st.columns([0.5, 0.5])
            with download_cols[0]:
                if excel_bytes:
                    st.download_button(
                        "⬇️ Download full database (Excel)",
                        excel_bytes,
                        file_name="ps_crm.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                else:
                    st.info("Excel export not ready yet.")
            with download_cols[1]:
                if archive_bytes:
                    st.download_button(
                        "⬇️ Download full archive (.zip)",
                        archive_bytes,
                        file_name="ps_crm_full.zip",
                        mime="application/zip",
                        help="Bundles the database, uploads, and receipts into one portable file.",
                    )
                else:
                    st.info("Archive export not ready yet.")
        backup_status = get_backup_status(BACKUP_DIR)
        if backup_status:
            backup_label = backup_status.get("last_backup_at") or "Unknown time"
            backup_file = backup_status.get("last_backup_file") or "unknown file"
            backup_lines = [
                f"Last automatic backup: {backup_label} • {backup_file} "
                f"(stored in {backup_status.get('backup_dir')})"
            ]
            mirror_dir = backup_status.get("mirror_dir")
            if mirror_dir:
                backup_lines.append(f"Backup mirror: {mirror_dir}")
            mirror_error = backup_status.get("mirror_error")
            if mirror_error:
                backup_lines.append(f"Backup mirror error: {mirror_error}")
            st.caption(" • ".join(backup_lines))
            if not mirror_dir:
                st.info(
                    "Tip: set PS_BACKUP_MIRROR_DIR to store automatic backups on a separate Linode volume "
                    "so older archives (including staff/users database data and uploads) stay available after redeploys.",
                    icon="💾",
                )
        if st.session_state.get("auto_backup_error"):
            st.warning(
                f"Automatic backup failed: {st.session_state['auto_backup_error']}"
            )

        toggle_label = (
            "🗑️ Deleted data"
            if not st.session_state.get("show_deleted_panel")
            else "Hide deleted data"
        )
        toggle_cols = st.columns([0.78, 0.22])
        with toggle_cols[1]:
            if st.button(
                toggle_label,
                key="toggle_deleted_panel",
                help="Admins can review deleted import records here.",
            ):
                st.session_state.show_deleted_panel = not st.session_state.get(
                    "show_deleted_panel", False
                )

        if st.session_state.get("show_deleted_panel"):
            deleted_df = df_query(
                conn,
                """
                SELECT ih.import_id,
                       ih.imported_at,
                       ih.customer_name,
                       ih.phone,
                       ih.product_label,
                       ih.original_date,
                       ih.do_number,
                       ih.deleted_at,
                       ih.deleted_by,
                       u.username AS deleted_by_name
                FROM import_history ih
                LEFT JOIN users u ON u.user_id = ih.deleted_by
                WHERE ih.deleted_at IS NOT NULL
                ORDER BY datetime(ih.deleted_at) DESC
                """,
            )

            if deleted_df.empty:
                st.info("No deleted import entries found.")
            else:
                formatted_deleted = fmt_dates(
                    deleted_df,
                    ["imported_at", "original_date", "deleted_at"],
                )
                deleted_bytes = io.BytesIO()
                with pd.ExcelWriter(deleted_bytes, engine="openpyxl") as writer:
                    formatted_deleted.to_excel(
                        writer, index=False, sheet_name="deleted_imports"
                    )
                deleted_bytes.seek(0)

                st.markdown("#### Deleted import history")
                st.caption(
                    "Only administrators can access this view. Download the Excel file for a full audit trail."
                )
                st.download_button(
                    "Download deleted imports",
                    deleted_bytes.getvalue(),
                    file_name="deleted_imports.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="deleted_imports_dl",
                )
                preview_cols = [
                    "import_id",
                    "imported_at",
                    "customer_name",
                    "phone",
                    "product_label",
                    "do_number",
                    "original_date",
                    "deleted_at",
                    "deleted_by_name",
                ]
                st.dataframe(
                    formatted_deleted[preview_cols],
                    use_container_width=True,
                )

            deleted_ops = df_query(
                conn,
                """
                SELECT d.do_number,
                       COALESCE(d.record_type, 'delivery_order') AS record_type,
                       COALESCE(c.name, '(unknown)') AS customer,
                       d.description,
                       d.file_path,
                       d.payment_receipt_path,
                       d.deleted_at,
                       d.deleted_by,
                       u.username AS deleted_by_name
                FROM delivery_orders d
                LEFT JOIN customers c ON c.customer_id = d.customer_id
                LEFT JOIN users u ON u.user_id = d.deleted_by
                WHERE d.deleted_at IS NOT NULL
                ORDER BY datetime(d.deleted_at) DESC
                """,
            )
            if deleted_ops.empty:
                st.caption("No deleted delivery/work done records found.")
            else:
                formatted_ops = fmt_dates(deleted_ops, ["deleted_at"])
                st.markdown("#### Deleted delivery/work done records")
                ops_bytes = io.BytesIO()
                with pd.ExcelWriter(ops_bytes, engine="openpyxl") as writer:
                    formatted_ops.to_excel(
                        writer, index=False, sheet_name="deleted_operations"
                    )
                ops_bytes.seek(0)
                st.download_button(
                    "Download deleted delivery/work done",
                    ops_bytes.getvalue(),
                    file_name="deleted_delivery_work_done.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="deleted_ops_dl",
                )
                st.dataframe(formatted_ops, use_container_width=True)

            deleted_documents = df_query(
                conn,
                """
                SELECT d.document_id,
                       d.customer_id,
                       COALESCE(c.name, c.company_name, '(customer)') AS customer,
                       d.doc_type,
                       d.original_name,
                       d.file_path,
                       d.uploaded_at,
                       d.deleted_at,
                       d.deleted_by,
                       u.username AS deleted_by_name
                FROM customer_documents d
                LEFT JOIN customers c ON c.customer_id = d.customer_id
                LEFT JOIN users u ON u.user_id = d.deleted_by
                WHERE d.deleted_at IS NOT NULL
                ORDER BY datetime(d.deleted_at) DESC
                """,
            )
            if deleted_documents.empty:
                st.caption("No deleted operations documents found.")
            else:
                formatted_docs = fmt_dates(deleted_documents, ["uploaded_at", "deleted_at"])
                st.markdown("#### Deleted operations documents")
                docs_bytes = io.BytesIO()
                with pd.ExcelWriter(docs_bytes, engine="openpyxl") as writer:
                    formatted_docs.to_excel(
                        writer, index=False, sheet_name="deleted_documents"
                    )
                docs_bytes.seek(0)
                st.download_button(
                    "Download deleted operations documents",
                    docs_bytes.getvalue(),
                    file_name="deleted_operations_documents.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="deleted_ops_docs_dl",
                )
                st.dataframe(formatted_docs, use_container_width=True)

            deleted_other_docs = df_query(
                conn,
                """
                SELECT o.document_id,
                       o.customer_id,
                       COALESCE(c.name, c.company_name, '(customer)') AS customer,
                       o.description,
                       o.original_name,
                       o.file_path,
                       o.uploaded_at,
                       o.updated_at,
                       o.deleted_at,
                       o.deleted_by,
                       u.username AS deleted_by_name
                FROM operations_other_documents o
                LEFT JOIN customers c ON c.customer_id = o.customer_id
                LEFT JOIN users u ON u.user_id = o.deleted_by
                WHERE o.deleted_at IS NOT NULL
                ORDER BY datetime(o.deleted_at) DESC
                """,
            )
            if deleted_other_docs.empty:
                st.caption("No deleted other operations uploads found.")
            else:
                formatted_other_docs = fmt_dates(
                    deleted_other_docs, ["uploaded_at", "updated_at", "deleted_at"]
                )
                st.markdown("#### Deleted other operations uploads")
                other_docs_bytes = io.BytesIO()
                with pd.ExcelWriter(other_docs_bytes, engine="openpyxl") as writer:
                    formatted_other_docs.to_excel(
                        writer, index=False, sheet_name="deleted_other_uploads"
                    )
                other_docs_bytes.seek(0)
                st.download_button(
                    "Download deleted other operations uploads",
                    other_docs_bytes.getvalue(),
                    file_name="deleted_other_operations_uploads.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="deleted_other_ops_docs_dl",
                )
                st.dataframe(formatted_other_docs, use_container_width=True)

    def _resolve_recent_pdf_bytes(path_value: Optional[str]) -> tuple[Optional[bytes], Optional[str]]:
        clean_path = clean_text(path_value)
        if not clean_path:
            return None, None
        resolved_path = resolve_upload_path(clean_path)
        if not resolved_path:
            return None, None
        pdf_path = resolved_path
        if pdf_path.suffix.lower() != ".pdf":
            alt_pdf = pdf_path.with_suffix(".pdf")
            if alt_pdf.exists():
                pdf_path = alt_pdf
            else:
                return None, None
        if not pdf_path.exists():
            return None, None
        try:
            return pdf_path.read_bytes(), pdf_path.name
        except OSError:
            return None, None

    def _render_recent_pdf_downloads(
        label_prefix: str,
        snapshot_df: pd.DataFrame,
        key_prefix: str,
    ) -> None:
        if snapshot_df.empty:
            return
        download_rows = []
        for _, row in snapshot_df.iterrows():
            attachments = []
            for label, path_value in (
                ("Document", row.get("file_path")),
                ("Receipt", row.get("payment_receipt_path")),
            ):
                payload, filename = _resolve_recent_pdf_bytes(path_value)
                if payload and filename:
                    attachments.append((label, payload, filename))
            if attachments:
                download_rows.append((row, attachments))
        if not download_rows:
            return
        st.markdown("##### Download PDFs")
        for idx, (row, attachments) in enumerate(download_rows):
            reference = clean_text(row.get("do_number")) or "Record"
            customer_label = clean_text(row.get("customer")) or "Customer"
            expander_title = f"{label_prefix} {reference} • {customer_label}"
            with st.expander(expander_title, expanded=False):
                for label, payload, filename in attachments:
                    safe_key = (
                        _sanitize_path_component(f"{reference}_{label}_{idx}") or "file"
                    )
                    st.download_button(
                        f"Download {label} PDF",
                        payload,
                        file_name=filename,
                        key=f"{key_prefix}_{safe_key}",
                        mime="application/pdf",
                    )

    st.markdown("---")
    st.subheader("🔎 Quick snapshots")
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Upcoming expiries (next 60 days)",
        "Recent services",
        "Recent maintenance",
        "Recent delivery orders",
        "Recent work orders",
    ])

    with tab1:
        range_col, projection_col = st.columns((2, 1))
        with range_col:
            days_window = st.slider(
                "Upcoming window (days)",
                min_value=7,
                max_value=180,
                value=60,
                step=1,
                help="Adjust how far ahead to look for upcoming warranty expiries.",
            )
        with projection_col:
            months_projection = st.slider(
                "Projection window (months)",
                min_value=1,
                max_value=12,
                value=6,
                help="Preview the workload trend for active warranties.",
            )

        upcoming = fetch_warranty_window(conn, 0, int(days_window))
        upcoming = format_warranty_table(upcoming)
        upcoming_count = int(len(upcoming.index)) if upcoming is not None else 0

        metric_col1, metric_col2 = st.columns((1, 1))
        with metric_col1:
            st.metric("Upcoming expiries", upcoming_count)
        with metric_col2:
            st.metric("Days in view", int(days_window))

        st.caption(
            f"Active warranties scheduled to expire in the next {int(days_window)} days."
        )

        if upcoming is None or upcoming.empty:
            st.info("No active warranties are due within the selected window.")
        else:
            show_all = False
            if len(upcoming.index) > 10:
                show_all = st.checkbox(
                    "Show all upcoming expiries", key="show_all_upcoming"
                )
            upcoming_display = upcoming if show_all else upcoming.head(10)
            st.dataframe(upcoming_display, use_container_width=True)

            csv_bytes = upcoming.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Download upcoming expiries (CSV)",
                csv_bytes,
                file_name="upcoming_warranties.csv",
                mime="text/csv",
                key="download_upcoming_csv",
            )

        projection_df = upcoming_warranty_projection(conn, int(months_projection))
        st.caption("Projected monthly warranty expiries")
        if projection_df.empty:
            st.info("No active warranties are scheduled to expire in the selected projection window.")
        else:
            st.bar_chart(projection_df.set_index("Month"))
            peak_row = projection_df.loc[projection_df["Expiring warranties"].idxmax()]
            peak_value = int(peak_row["Expiring warranties"])
            if peak_value > 0:
                st.success(
                    f"Peak month: {peak_row['Month']} with {peak_value} scheduled expiries."
                )
            else:
                st.info("All selected months currently show zero scheduled expiries.")

    with tab2:
        show_all_services = st.checkbox(
            "Show all services", key="dashboard_services_show_all"
        )
        services_limit = "" if show_all_services else "LIMIT 200"
        recent_services = df_query(
            conn,
            f"""
            SELECT s.do_number,
                   s.customer_id,
                   d.customer_id AS do_customer_id,
                   s.service_date,
                   COALESCE(c.name, cdo.name, '(unknown)') AS customer,
                   s.description
            FROM services s
            LEFT JOIN customers c ON c.customer_id = s.customer_id
            LEFT JOIN delivery_orders d ON d.do_number = s.do_number
            LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
            WHERE s.deleted_at IS NULL
            ORDER BY datetime(s.service_date) DESC, s.service_id DESC
            {services_limit}
            """,
        )
        if allowed_customers is not None:
            recent_services = recent_services[
                recent_services.apply(
                    lambda row: any(
                        cid in allowed_customers
                        for cid in [
                            int(row.get("customer_id"))
                            if pd.notna(row.get("customer_id"))
                            else None,
                            int(row.get("do_customer_id"))
                            if pd.notna(row.get("do_customer_id"))
                            else None,
                        ]
                        if cid is not None
                    ),
                    axis=1,
                )
            ]
        recent_services = fmt_dates(recent_services, ["service_date"])
        st.dataframe(
            recent_services.rename(
                columns={
                    "do_number": "DO Serial",
                    "service_date": "Service date",
                    "customer": "Customer",
                    "description": "Description",
                }
            ).drop(columns=["customer_id", "do_customer_id"], errors="ignore"),
            use_container_width=True,
            height=320,
        )

    with tab3:
        show_all_maintenance = st.checkbox(
            "Show all maintenance records", key="dashboard_maintenance_show_all"
        )
        maintenance_limit = "" if show_all_maintenance else "LIMIT 200"
        recent_maintenance = df_query(
            conn,
            f"""
            SELECT m.do_number,
                   m.customer_id,
                   d.customer_id AS do_customer_id,
                   m.maintenance_date,
                   COALESCE(c.name, cdo.name, '(unknown)') AS customer,
                   m.description
            FROM maintenance_records m
            LEFT JOIN customers c ON c.customer_id = m.customer_id
            LEFT JOIN delivery_orders d ON d.do_number = m.do_number
            LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
            WHERE m.deleted_at IS NULL
            ORDER BY datetime(m.maintenance_date) DESC, m.maintenance_id DESC
            {maintenance_limit}
            """,
        )
        if allowed_customers is not None:
            recent_maintenance = recent_maintenance[
                recent_maintenance.apply(
                    lambda row: any(
                        cid in allowed_customers
                        for cid in [
                            int(row.get("customer_id"))
                            if pd.notna(row.get("customer_id"))
                            else None,
                            int(row.get("do_customer_id"))
                            if pd.notna(row.get("do_customer_id"))
                            else None,
                        ]
                        if cid is not None
                    ),
                    axis=1,
                )
            ]
        recent_maintenance = fmt_dates(recent_maintenance, ["maintenance_date"])
        st.dataframe(
            recent_maintenance.rename(
                columns={
                    "do_number": "DO Serial",
                    "maintenance_date": "Maintenance date",
                    "customer": "Customer",
                    "description": "Description",
                }
            ).drop(columns=["customer_id", "do_customer_id"], errors="ignore"),
            use_container_width=True,
            height=320,
        )

    with tab4:
        show_all_delivery = st.checkbox(
            "Show all delivery orders", key="dashboard_delivery_show_all"
        )
        delivery_limit = "" if show_all_delivery else "LIMIT 200"
        recent_delivery_orders = df_query(
            conn,
            f"""
            SELECT d.do_number,
                   d.customer_id,
                   d.created_at,
                   d.status,
                   d.total_amount,
                   d.file_path,
                   d.payment_receipt_path,
                   COALESCE(c.name, '(unknown)') AS customer,
                   d.description
            FROM delivery_orders d
            LEFT JOIN customers c ON c.customer_id = d.customer_id
            WHERE d.deleted_at IS NULL
              AND COALESCE(d.record_type, 'delivery_order') = 'delivery_order'
            ORDER BY datetime(d.created_at) DESC
            {delivery_limit}
            """,
        )
        if allowed_customers is not None:
            recent_delivery_orders = recent_delivery_orders[
                recent_delivery_orders["customer_id"].apply(
                    lambda value: pd.notna(value) and int(value) in allowed_customers
                )
            ]
        if recent_delivery_orders.empty:
            st.info("No recent delivery orders found.")
        else:
            recent_delivery_orders = fmt_dates(recent_delivery_orders, ["created_at"])
            recent_delivery_orders["total_amount"] = recent_delivery_orders[
                "total_amount"
            ].apply(
                lambda value: format_money(value)
                or (f"{_coerce_float(value, 0.0):,.2f}" if pd.notna(value) else "—")
            )
            st.dataframe(
                recent_delivery_orders.rename(
                    columns={
                        "do_number": "DO Serial",
                        "created_at": "Created",
                        "status": "Status",
                        "total_amount": "Total",
                        "customer": "Customer",
                        "description": "Description",
                    }
                ).drop(
                    columns=["customer_id", "file_path", "payment_receipt_path"],
                    errors="ignore",
                ),
                use_container_width=True,
                height=320,
            )
            _render_recent_pdf_downloads(
                "Delivery order",
                recent_delivery_orders,
                "recent_delivery_pdf",
            )

    with tab5:
        show_all_work = st.checkbox(
            "Show all work orders", key="dashboard_work_show_all"
        )
        work_limit = "" if show_all_work else "LIMIT 200"
        recent_work_orders = df_query(
            conn,
            f"""
            SELECT d.do_number,
                   d.customer_id,
                   d.created_at,
                   d.status,
                   d.total_amount,
                   d.file_path,
                   d.payment_receipt_path,
                   COALESCE(c.name, '(unknown)') AS customer,
                   d.description
            FROM delivery_orders d
            LEFT JOIN customers c ON c.customer_id = d.customer_id
            WHERE d.deleted_at IS NULL
              AND COALESCE(d.record_type, 'delivery_order') = 'work_done'
            ORDER BY datetime(d.created_at) DESC
            {work_limit}
            """,
        )
        if allowed_customers is not None:
            recent_work_orders = recent_work_orders[
                recent_work_orders["customer_id"].apply(
                    lambda value: pd.notna(value) and int(value) in allowed_customers
                )
            ]
        if recent_work_orders.empty:
            st.info("No recent work orders found.")
        else:
            recent_work_orders = fmt_dates(recent_work_orders, ["created_at"])
            recent_work_orders["total_amount"] = recent_work_orders["total_amount"].apply(
                lambda value: format_money(value)
                or (f"{_coerce_float(value, 0.0):,.2f}" if pd.notna(value) else "—")
            )
            st.dataframe(
                recent_work_orders.rename(
                    columns={
                        "do_number": "Work order",
                        "created_at": "Created",
                        "status": "Status",
                        "total_amount": "Total",
                        "customer": "Customer",
                        "description": "Description",
                    }
                ).drop(
                    columns=["customer_id", "file_path", "payment_receipt_path"],
                    errors="ignore",
                ),
                use_container_width=True,
                height=320,
            )
            _render_recent_pdf_downloads(
                "Work order",
                recent_work_orders,
                "recent_work_pdf",
            )

    if is_admin:
        st.markdown("---")
        _render_admin_kpi_panel(conn)


def show_expiry_notifications(conn):
    is_admin = current_user_is_admin()

    if not st.session_state.get("just_logged_in"):
        return

    scope_clause, scope_params = customer_scope_filter("c")
    allowed_customers = accessible_customer_ids(conn)
    scheduled_services = df_query(
        conn,
        """
        SELECT s.service_id,
               s.customer_id,
               d.customer_id AS do_customer_id,
               s.do_number,
               COALESCE(s.service_start_date, s.service_date) AS start_date,
               s.status,
               s.description,
               COALESCE(c.name, cdo.name, '(unknown)') AS customer
        FROM services s
        LEFT JOIN customers c ON c.customer_id = s.customer_id
        LEFT JOIN delivery_orders d ON d.do_number = s.do_number
        LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
        WHERE s.deleted_at IS NULL
          AND COALESCE(s.service_start_date, s.service_date) IS NOT NULL
          AND date(COALESCE(s.service_start_date, s.service_date)) = date('now')
        ORDER BY datetime(COALESCE(s.service_start_date, s.service_date)) ASC, s.service_id ASC
        """,
    )
    scheduled_maintenance = df_query(
        conn,
        """
        SELECT m.maintenance_id,
               m.customer_id,
               d.customer_id AS do_customer_id,
               m.do_number,
               COALESCE(m.maintenance_start_date, m.maintenance_date) AS start_date,
               m.status,
               m.description,
               COALESCE(c.name, cdo.name, '(unknown)') AS customer
        FROM maintenance_records m
        LEFT JOIN customers c ON c.customer_id = m.customer_id
        LEFT JOIN delivery_orders d ON d.do_number = m.do_number
        LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
        WHERE m.deleted_at IS NULL
          AND COALESCE(m.maintenance_start_date, m.maintenance_date) IS NOT NULL
          AND date(COALESCE(m.maintenance_start_date, m.maintenance_date)) = date('now')
        ORDER BY datetime(COALESCE(m.maintenance_start_date, m.maintenance_date)) ASC, m.maintenance_id ASC
        """,
    )

    if not scheduled_services.empty:
        scheduled_services = scheduled_services[
            scheduled_services["status"].apply(is_pending_status)
        ]
    if not scheduled_maintenance.empty:
        scheduled_maintenance = scheduled_maintenance[
            scheduled_maintenance["status"].apply(is_pending_status)
        ]

    scheduled_services = fmt_dates(scheduled_services, ["start_date"])
    scheduled_maintenance = fmt_dates(scheduled_maintenance, ["start_date"])
    if allowed_customers is not None:
        def _belongs(row):
            candidates = []
            for key in ("customer_id", "do_customer_id"):
                value = row.get(key)
                if pd.notna(value):
                    try:
                        candidates.append(int(value))
                    except (TypeError, ValueError):
                        continue
            return any(cid in allowed_customers for cid in candidates)

        scheduled_services = scheduled_services[scheduled_services.apply(_belongs, axis=1)]
        scheduled_maintenance = scheduled_maintenance[scheduled_maintenance.apply(_belongs, axis=1)]
    scheduled_services.drop(columns=["customer_id", "do_customer_id"], inplace=True, errors="ignore")
    scheduled_maintenance.drop(columns=["customer_id", "do_customer_id"], inplace=True, errors="ignore")

    upcoming_sections: list[pd.DataFrame] = []
    upcoming_messages: list[str] = []

    due_notes = df_query(
        conn,
        """
        SELECT n.note_id,
               n.customer_id,
               n.note,
               n.remind_on,
               c.name AS customer
        FROM customer_notes n
        JOIN customers c ON c.customer_id = n.customer_id
        WHERE n.is_done = 0
          AND n.remind_on IS NOT NULL
          AND date(n.remind_on) <= date('now')
        ORDER BY date(n.remind_on) ASC, datetime(n.created_at) ASC
        """,
    )
    due_notes = fmt_dates(due_notes, ["remind_on"])
    if allowed_customers is not None:
        due_notes = due_notes[due_notes["customer_id"].apply(lambda value: pd.notna(value) and int(value) in allowed_customers)]
    due_notes.drop(columns=["customer_id"], inplace=True, errors="ignore")
    notes_display = pd.DataFrame()
    if not due_notes.empty:
        notes_display = due_notes.rename(
            columns={
                "customer": "Customer",
                "note": "Remark",
                "remind_on": "Due date",
            }
        )[
            ["Customer", "Remark", "Due date"]
        ]
        for record in due_notes.to_dict("records"):
            customer_ref = clean_text(record.get("customer")) or "(unknown)"
            note_text = clean_text(record.get("note")) or "Follow-up due"
            due_label = clean_text(record.get("remind_on")) or datetime.now().strftime(DATE_FMT)
            upcoming_messages.append(
                f"Reminder: follow up with {customer_ref} – {note_text} (due {due_label})."
            )

    if not scheduled_services.empty:
        service_records = scheduled_services.to_dict("records")
        for record in service_records:
            do_ref = clean_text(record.get("do_number"))
            if not do_ref:
                try:
                    service_identifier = int(record.get("service_id"))
                    do_ref = f"Service #{service_identifier}"
                except Exception:
                    do_ref = "Service"
            customer_ref = clean_text(record.get("customer")) or "(unknown)"
            start_label = clean_text(record.get("start_date")) or datetime.now().strftime(DATE_FMT)
            upcoming_messages.append(
                f"Service {do_ref} for {customer_ref} starts today ({start_label})."
            )
        service_display = scheduled_services.copy()
        service_display = service_display.rename(
            columns={
                "do_number": "DO Serial",
                "start_date": "Start date",
                "status": "Status",
                "description": "Description",
                "customer": "Customer",
            }
        )
        service_display.insert(0, "Type", "Service")
        service_display = service_display.drop(columns=["service_id"], errors="ignore")
        upcoming_sections.append(
            service_display[[
                "Type",
                "DO Serial",
                "Customer",
                "Start date",
                "Status",
                "Description",
            ]]
        )

    if not scheduled_maintenance.empty:
        maintenance_records = scheduled_maintenance.to_dict("records")
        for record in maintenance_records:
            do_ref = clean_text(record.get("do_number"))
            if not do_ref:
                try:
                    maintenance_identifier = int(record.get("maintenance_id"))
                    do_ref = f"Maintenance #{maintenance_identifier}"
                except Exception:
                    do_ref = "Maintenance"
            customer_ref = clean_text(record.get("customer")) or "(unknown)"
            start_label = clean_text(record.get("start_date")) or datetime.now().strftime(DATE_FMT)
            upcoming_messages.append(
                f"Maintenance {do_ref} for {customer_ref} starts today ({start_label})."
            )
        maintenance_display = scheduled_maintenance.copy()
        maintenance_display = maintenance_display.rename(
            columns={
                "do_number": "DO Serial",
                "start_date": "Start date",
                "status": "Status",
                "description": "Description",
                "customer": "Customer",
            }
        )
        maintenance_display.insert(0, "Type", "Maintenance")
        maintenance_display = maintenance_display.drop(columns=["maintenance_id"], errors="ignore")
        upcoming_sections.append(
            maintenance_display[[
                "Type",
                "DO Serial",
                "Customer",
                "Start date",
                "Status",
                "Description",
            ]]
        )

    upcoming_df = (
        pd.concat(upcoming_sections, ignore_index=True)
        if upcoming_sections
        else pd.DataFrame()
    )

    scope_filter_clause = f" AND {scope_clause}" if scope_clause else ""
    total_expired_query = dedent(
        f"""
        SELECT COUNT(*) c
        FROM warranties w
        LEFT JOIN customers c ON c.customer_id = w.customer_id
        WHERE date(w.expiry_date) < date('now'){scope_filter_clause}
        """
    )
    total_expired = int(df_query(conn, total_expired_query, scope_params).iloc[0]["c"])
    month_expired = 0
    formatted = pd.DataFrame()
    if total_expired > 0:
        month_expired_query = dedent(
            f"""
            SELECT COUNT(*) c
            FROM warranties w
            LEFT JOIN customers c ON c.customer_id = w.customer_id
            WHERE date(w.expiry_date) < date('now')
              AND strftime('%Y-%m', w.expiry_date) = strftime('%Y-%m', 'now'){scope_filter_clause}
            """
        )
        month_expired = int(df_query(conn, month_expired_query, scope_params).iloc[0]["c"])
        expired_recent_query = dedent(
            f"""
            SELECT c.name AS customer, p.name AS product, p.model, w.serial, w.issue_date, w.expiry_date
            FROM warranties w
            LEFT JOIN customers c ON c.customer_id = w.customer_id
            LEFT JOIN products p ON p.product_id = w.product_id
            WHERE date(w.expiry_date) < date('now'){scope_filter_clause}
            ORDER BY date(w.expiry_date) DESC
            LIMIT 12
            """
        )
        expired_recent = df_query(conn, expired_recent_query, scope_params)
        formatted = format_warranty_table(expired_recent)

    show_upcoming = not upcoming_df.empty
    show_expired = total_expired > 0
    show_notes = not notes_display.empty

    if not show_upcoming and not show_expired and not show_notes:
        st.session_state.just_logged_in = False
        return

    if show_upcoming:
        upcoming_preview = upcoming_df.head(5)
        upcoming_details = []
        for record in upcoming_preview.to_dict("records"):
            type_label = clean_text(record.get("Type")) or ""
            customer_label = clean_text(record.get("Customer")) or "(unknown)"
            start_label = clean_text(record.get("Start date")) or ""
            description_label = clean_text(record.get("Description")) or clean_text(
                record.get("Status")
            ) or ""
            detail_parts = [part for part in [type_label, customer_label, start_label] if part]
            detail_line = " • ".join(detail_parts)
            if description_label:
                detail_line = (
                    f"{detail_line} – {description_label}" if detail_line else description_label
                )
            upcoming_details.append(detail_line)
        push_runtime_notification(
            "Today's schedule",
            f"{len(upcoming_df.index)} task(s) scheduled for today.",
            severity="info",
            details=upcoming_details,
        )

    if show_notes:
        notes_preview = notes_display.head(5)
        notes_details = []
        for record in notes_preview.to_dict("records"):
            customer_label = clean_text(record.get("Customer")) or "(unknown)"
            due_label = clean_text(record.get("Due date")) or ""
            remark_label = clean_text(record.get("Remark")) or ""
            detail_line = " • ".join(
                part for part in [customer_label, due_label, remark_label] if part
            )
            notes_details.append(detail_line)
        push_runtime_notification(
            "Follow-up reminders",
            f"{len(notes_display.index)} customer reminder(s) due.",
            severity="warning",
            details=notes_details,
        )

    if show_expired:
        expiry_preview = formatted.head(5) if isinstance(formatted, pd.DataFrame) else pd.DataFrame()
        expiry_details = []
        if isinstance(expiry_preview, pd.DataFrame) and not expiry_preview.empty:
            for record in expiry_preview.to_dict("records"):
                customer_label = clean_text(record.get("Customer")) or "(unknown)"
                product_label = clean_text(record.get("Product")) or clean_text(
                    record.get("Model")
                ) or ""
                expiry_label = clean_text(record.get("Expiry date")) or ""
                detail_line = " • ".join(
                    part for part in [customer_label, product_label, expiry_label] if part
                )
                expiry_details.append(detail_line)
        push_runtime_notification(
            "Expired warranties",
            f"{total_expired} warranty record(s) need attention ({month_expired} this month).",
            severity="warning",
            details=expiry_details,
        )

    for message in upcoming_messages:
        try:
            st.toast(message)
        except Exception:
            break
    try:
        if show_expired:
            st.toast(f"{total_expired} warranties require attention.")
    except Exception:
        pass

    st.session_state.just_logged_in = False


def _render_notification_entry(entry: dict[str, object], *, include_actor: bool = False) -> None:
    severity = str(entry.get("severity") or "info").lower()
    icon = {
        "warning": "🟠",
        "error": "🔴",
        "success": "🟢",
    }.get(severity, "🔵")
    title = clean_text(entry.get("title")) or "Notification"
    message = clean_text(entry.get("message")) or ""
    st.markdown(f"{icon} **{title}**")
    if message:
        st.write(message)
    details = entry.get("details") or []
    for detail in list(details)[:5]:
        st.caption(f"• {detail}")
    footer_bits: list[str] = []
    if include_actor:
        actor = clean_text(entry.get("actor"))
        if actor:
            footer_bits.append(actor)
    time_label = format_time_ago(entry.get("timestamp"))
    if time_label:
        footer_bits.append(time_label)
    if footer_bits:
        st.caption(" · ".join(footer_bits))


def _render_notification_section(
    entries: list[dict[str, object]],
    *,
    include_actor: bool = False,
    heading: Optional[str] = None,
) -> None:
    if not entries:
        return
    if heading:
        st.markdown(
            f"<div class='ps-notification-section-title'>{heading}</div>",
            unsafe_allow_html=True,
        )
    first = True
    for entry in entries:
        if not first:
            st.divider()
        _render_notification_entry(entry, include_actor=include_actor)
        first = False


def _render_notification_body(
    alerts: list[dict[str, object]],
    activity: list[dict[str, object]],
) -> None:
    if not alerts and not activity:
        st.caption("No notifications yet. Updates will appear here as your team works.")
        return
    _render_notification_section(alerts, heading="Alerts")
    if alerts and activity:
        st.divider()
    _render_notification_section(activity, include_actor=True, heading="Recent activity")


def render_notification_bell(conn) -> None:
    user = get_current_user()
    if not user:
        return

    is_admin = user.get("role") == "admin"
    alerts = list(reversed(get_runtime_notifications()))
    user_id = current_user_id()
    alerts.extend(_build_staff_alerts(conn, user_id=user_id))
    activity = fetch_activity_feed(conn, limit=ACTIVITY_FEED_LIMIT) if is_admin else []

    total = len(alerts) + len(activity)
    label = "🔔" if total == 0 else f"🔔 {total}"
    container = st.container()
    with container:
        for alert in alerts:
            if alert.get("severity") in {"warning", "error", "danger"}:
                toast_msg = alert.get("message") or alert.get("title")
                if toast_msg:
                    st.toast(toast_msg, icon="⚠️")
        st.markdown("<div class='ps-notification-popover'>", unsafe_allow_html=True)
        popover = getattr(st, "popover", None)
        if callable(popover):
            with popover(label, help="View alerts and staff activity", use_container_width=True):
                _render_notification_body(alerts, activity)
        else:
            with st.expander(f"{label} Notifications", expanded=False):
                _render_notification_body(alerts, activity)
        st.markdown("</div>", unsafe_allow_html=True)


def render_customer_quick_edit_section(
    conn,
    *,
    key_prefix: str,
    include_leads: bool = True,
    include_leads_in_main: bool = False,
    show_heading: bool = True,
    show_editor: bool = True,
    show_filters: bool = True,
    show_id: bool = True,
    enable_pagination: bool = False,
    limit_rows: Optional[int] = None,
    show_do_code: bool = True,
    show_duplicate: bool = True,
    action_icon_only: bool = False,
    use_popover: bool = True,
    include_quotation_upload: bool = True,
    enable_uploads: bool = True,
) -> pd.DataFrame:
    pagination_enabled = enable_pagination and show_editor and not include_leads
    if show_filters:
        sort_dir = st.radio(
            "Sort by created date",
            ["Newest first", "Oldest first"],
            horizontal=True,
            key=f"{key_prefix}_sort_dir",
        )
        order = "DESC" if sort_dir == "Newest first" else "ASC"
        q = st.text_input(
            "Search (name/phone/address/product/DO)",
            key=f"{key_prefix}_search",
        )
        if pagination_enabled:
            page_size = st.selectbox(
                "Rows per page",
                options=[25, 50, 100],
                index=1,
                key=f"{key_prefix}_page_size",
            )
    else:
        order = "DESC"
        q = ""
        if pagination_enabled:
            page_size = 50
    scope_clause, scope_params = customer_scope_filter("c")
    search_clause = dedent(
        """
        (? = ''
         OR c.name LIKE '%'||?||'%'
         OR c.company_name LIKE '%'||?||'%'
         OR c.phone LIKE '%'||?||'%'
         OR c.address LIKE '%'||?||'%'
         OR c.delivery_address LIKE '%'||?||'%'
         OR c.remarks LIKE '%'||?||'%'
         OR c.product_info LIKE '%'||?||'%'
         OR c.delivery_order_code LIKE '%'||?||'%'
         OR c.sales_person LIKE '%'||?||'%')
        """
    ).strip()
    where_parts = [search_clause]
    params: list[object] = [q, q, q, q, q, q, q, q, q, q]
    if scope_clause:
        where_parts.append(scope_clause)
        params.extend(scope_params)
    where_sql = " AND ".join(where_parts)
    limit_clause = ""
    total_count = None
    if pagination_enabled:
        total_count = int(
            df_query(
                conn,
                f"SELECT COUNT(*) AS cnt FROM customers c WHERE {where_sql}",
                tuple(params),
            ).iloc[0]["cnt"]
        )
        total_pages = max(1, math.ceil(total_count / page_size)) if page_size else 1
        page_number = st.number_input(
            "Page",
            min_value=1,
            max_value=total_pages,
            value=1,
            step=1,
            key=f"{key_prefix}_page",
        )
        offset = (page_number - 1) * page_size
        limit_clause = f"LIMIT {page_size} OFFSET {offset}"
        st.caption(f"Showing page {page_number} of {total_pages} ({total_count} customers).")
    elif limit_rows:
        limit_clause = f"LIMIT {int(limit_rows)}"
    df_raw = df_query(
        conn,
        f"""
        SELECT
            c.customer_id AS id,
            c.name,
            c.company_name,
            c.phone,
            c.address,
            c.delivery_address,
            c.remarks,
            c.purchase_date,
            c.product_info,
            c.delivery_order_code,
            c.sales_person,
            c.attachment_path,
            c.created_at,
            c.dup_flag,
            c.created_by,
            COALESCE(u.username, '(unknown)') AS uploaded_by
        FROM customers c
        LEFT JOIN users u ON u.user_id = c.created_by
        WHERE {where_sql}
        ORDER BY datetime(c.created_at) {order}
        {limit_clause}
    """,
        tuple(params),
    )
    lead_mask = df_raw.get("remarks", pd.Series(dtype=object)).apply(_is_lead_customer)
    lead_df = df_raw[lead_mask].copy()
    if not include_leads_in_main:
        df_raw = df_raw[~lead_mask].copy()
    user = st.session_state.user or {}
    is_admin = user.get("role") == "admin"
    current_actor_id = current_user_id()
    if include_leads and not lead_df.empty:
        st.markdown("### Leads (Chasing)")
        lead_view = lead_df.copy()
        lead_view["created_at"] = pd.to_datetime(
            lead_view["created_at"], errors="coerce"
        )
        lead_columns = [
            col
            for col in [
                "id",
                "name",
                "company_name",
                "phone",
                "address",
                "remarks",
                "created_at",
                "uploaded_by",
            ]
            if col in lead_view.columns
        ]
        lead_view = lead_view[lead_columns]
        st.dataframe(
            lead_view,
            use_container_width=True,
            column_config={
                "id": st.column_config.Column("ID"),
                "name": st.column_config.TextColumn("Name"),
                "company_name": st.column_config.TextColumn("Company"),
                "phone": st.column_config.TextColumn("Phone"),
                "address": st.column_config.TextColumn("Address"),
                "remarks": st.column_config.TextColumn("Lead status"),
                "created_at": st.column_config.DatetimeColumn(
                    "Created", format="DD-MM-YYYY HH:mm"
                ),
                "uploaded_by": st.column_config.Column("Uploaded by"),
            },
        )
    if show_heading and show_editor:
        heading_label = "### Quick edit or delete"
        if action_icon_only:
            heading_label = "### Quick edit"
        st.markdown(heading_label)
    if df_raw.empty:
        if show_editor:
            st.info("No customers found for the current filters.")
        return df_raw
    if not show_editor:
        return df_raw

    original_map: dict[int, dict] = {}
    for record in df_raw.to_dict("records"):
        cid = int_or_none(record.get("id"))
        if cid is not None:
            original_map[cid] = record
    editor_df = df_raw.copy()
    editor_df["purchase_date"] = pd.to_datetime(editor_df["purchase_date"], errors="coerce")
    editor_df["created_at"] = pd.to_datetime(editor_df["created_at"], errors="coerce")
    if "dup_flag" in editor_df.columns:
        editor_df["duplicate"] = editor_df["dup_flag"].apply(
            lambda x: "🔁 duplicate phone" if int_or_none(x) == 1 else ""
        )
    else:
        editor_df["duplicate"] = ""
    editor_df["Action"] = "Keep"
    column_order = [
        col
        for col in [
            "id",
            "name",
            "company_name",
            "phone",
            "address",
            "delivery_address",
            "remarks",
            "purchase_date",
            "product_info",
            "delivery_order_code" if show_do_code else None,
            "duplicate" if show_duplicate else None,
            "Action",
        ]
        if col and col in editor_df.columns
    ]
    editor_df = editor_df[column_order]
    customer_ids = [int(cid) for cid in editor_df.get("id", pd.Series(dtype=int)).tolist()]

    def _build_quick_view_documents(customer_ids: list[int]) -> dict[int, list[dict[str, object]]]:
        docs_map: dict[int, list[dict[str, object]]] = {cid: [] for cid in customer_ids}
        seen: dict[int, set[tuple[str, str]]] = {cid: set() for cid in customer_ids}
        if not customer_ids:
            return docs_map
        placeholders = ",".join(["?"] * len(customer_ids))

        def _add_doc(cid: int, label: str, path: Optional[str], uploaded_at: Optional[object] = None) -> None:
            if not path:
                return
            label_clean = clean_text(label) or "Document"
            path_clean = clean_text(path) or ""
            if not path_clean:
                return
            doc_key = (label_clean, path_clean)
            seen.setdefault(cid, set())
            if doc_key in seen[cid]:
                return
            seen[cid].add(doc_key)
            docs_map.setdefault(cid, []).append(
                {"label": label_clean, "path": path_clean, "uploaded_at": uploaded_at}
            )

        customer_docs = df_query(
            conn,
            f"""
            SELECT customer_id, doc_type, file_path, original_name, uploaded_at
            FROM customer_documents
            WHERE customer_id IN ({placeholders}) AND deleted_at IS NULL
            ORDER BY datetime(uploaded_at) DESC, document_id DESC
            """,
            tuple(customer_ids),
        )
        for _, row in customer_docs.iterrows():
            cid = int_or_none(row.get("customer_id"))
            if cid is None:
                continue
            doc_type = clean_text(row.get("doc_type")) or "Document"
            original_name = clean_text(row.get("original_name")) or "(document)"
            label = f"{doc_type}: {original_name}"
            _add_doc(cid, label, clean_text(row.get("file_path")), row.get("uploaded_at"))

        delivery_docs = df_query(
            conn,
            f"""
            SELECT customer_id, do_number, record_type, file_path, payment_receipt_path, updated_at
            FROM delivery_orders
            WHERE customer_id IN ({placeholders}) AND deleted_at IS NULL
            ORDER BY datetime(updated_at) DESC
            """,
            tuple(customer_ids),
        )
        for _, row in delivery_docs.iterrows():
            cid = int_or_none(row.get("customer_id"))
            if cid is None:
                continue
            do_number = clean_text(row.get("do_number")) or "-"
            record_type = clean_text(row.get("record_type")) or "delivery_order"
            label_base = "Work done" if record_type == "work_done" else "Delivery order"
            _add_doc(
                cid,
                f"{label_base} {do_number} document",
                clean_text(row.get("file_path")),
                row.get("updated_at"),
            )
            _add_doc(
                cid,
                f"{label_base} {do_number} receipt",
                clean_text(row.get("payment_receipt_path")),
                row.get("updated_at"),
            )

        service_docs = df_query(
            conn,
            f"""
            SELECT s.customer_id,
                   s.service_id,
                   s.do_number,
                   s.bill_document_path,
                   s.payment_receipt_path,
                   s.updated_at,
                   sd.file_path AS attachment_path,
                   sd.original_name AS attachment_name,
                   sd.uploaded_at AS attachment_uploaded_at
            FROM services s
            LEFT JOIN service_documents sd ON sd.service_id = s.service_id
            WHERE s.customer_id IN ({placeholders})
              AND s.deleted_at IS NULL
            ORDER BY datetime(s.updated_at) DESC
            """,
            tuple(customer_ids),
        )
        for _, row in service_docs.iterrows():
            cid = int_or_none(row.get("customer_id"))
            if cid is None:
                continue
            do_number = clean_text(row.get("do_number")) or "-"
            _add_doc(
                cid,
                f"Service {do_number} bill",
                clean_text(row.get("bill_document_path")),
                row.get("updated_at"),
            )
            _add_doc(
                cid,
                f"Service {do_number} receipt",
                clean_text(row.get("payment_receipt_path")),
                row.get("updated_at"),
            )
            attachment_name = clean_text(row.get("attachment_name")) or "(attachment)"
            _add_doc(
                cid,
                f"Service {do_number} attachment: {attachment_name}",
                clean_text(row.get("attachment_path")),
                row.get("attachment_uploaded_at"),
            )

        maintenance_docs = df_query(
            conn,
            f"""
            SELECT m.customer_id,
                   m.maintenance_id,
                   m.do_number,
                   m.payment_receipt_path,
                   m.updated_at,
                   md.file_path AS attachment_path,
                   md.original_name AS attachment_name,
                   md.uploaded_at AS attachment_uploaded_at
            FROM maintenance_records m
            LEFT JOIN maintenance_documents md ON md.maintenance_id = m.maintenance_id
            WHERE m.customer_id IN ({placeholders})
              AND m.deleted_at IS NULL
            ORDER BY datetime(m.updated_at) DESC
            """,
            tuple(customer_ids),
        )
        for _, row in maintenance_docs.iterrows():
            cid = int_or_none(row.get("customer_id"))
            if cid is None:
                continue
            do_number = clean_text(row.get("do_number")) or "-"
            _add_doc(
                cid,
                f"Maintenance {do_number} receipt",
                clean_text(row.get("payment_receipt_path")),
                row.get("updated_at"),
            )
            attachment_name = clean_text(row.get("attachment_name")) or "(attachment)"
            _add_doc(
                cid,
                f"Maintenance {do_number} attachment: {attachment_name}",
                clean_text(row.get("attachment_path")),
                row.get("attachment_uploaded_at"),
            )

        return docs_map

    docs_map = _build_quick_view_documents(customer_ids)
    doc_type_options = []
    if enable_uploads:
        doc_type_options = [
            "Delivery order",
            "Work done",
            "Service",
            "Maintenance",
            "Other",
        ]
        if include_quotation_upload:
            doc_type_options.insert(2, "Quotation")
    base_widths = [
        1.2,
        1.2,
        1.0,
        1.4,
        1.4,
        1.2,
        1.0,
        1.6,
    ]
    if show_do_code:
        base_widths.append(1.0)
    if show_duplicate:
        base_widths.append(0.9)
    if enable_uploads:
        base_widths.extend([1.0, 2.2, 1.0])
    else:
        base_widths.extend([1.0, 1.0])
    widths = [0.5, *base_widths] if show_id else base_widths
    header_cols = st.columns(tuple(widths))
    header_idx = 0
    if show_id:
        header_cols[header_idx].write("**ID**")
        header_idx += 1
    header_cols[header_idx + 0].write("**Name**")
    header_cols[header_idx + 1].write("**Company**")
    header_cols[header_idx + 2].write("**Phone**")
    header_cols[header_idx + 3].write("**Billing address**")
    header_cols[header_idx + 4].write("**Delivery address**")
    header_cols[header_idx + 5].write("**Remarks**")
    header_cols[header_idx + 6].write("**Purchase date (not follow-up)**")
    header_cols[header_idx + 7].write("**Product**")
    col_offset = 8
    if show_do_code:
        header_cols[header_idx + col_offset].write("**DO code**")
        col_offset += 1
    if show_duplicate:
        header_cols[header_idx + col_offset].write("**Duplicate**")
        col_offset += 1
    header_cols[header_idx + col_offset].write("**View**")
    action_offset = col_offset + 1
    if enable_uploads:
        header_cols[header_idx + col_offset + 1].write("**Upload**")
        action_offset = col_offset + 2
    action_label = "**Action**" if not action_icon_only else "**Delete**"
    header_cols[header_idx + action_offset].write(action_label)
    editor_rows: list[dict[str, object]] = []
    for row in editor_df.to_dict("records"):
        cid = int_or_none(row.get("id"))
        if cid is None:
            continue
        row_cols = st.columns(tuple(widths))
        row_idx = 0
        if show_id:
            row_cols[row_idx].write(cid)
            row_idx += 1
        name_key = f"{key_prefix}_quick_name_{cid}"
        company_key = f"{key_prefix}_quick_company_{cid}"
        phone_key = f"{key_prefix}_quick_phone_{cid}"
        address_key = f"{key_prefix}_quick_address_{cid}"
        delivery_key = f"{key_prefix}_quick_delivery_{cid}"
        remarks_key = f"{key_prefix}_quick_remarks_{cid}"
        purchase_key = f"{key_prefix}_quick_purchase_{cid}"
        product_key = f"{key_prefix}_quick_product_{cid}"
        do_key = f"{key_prefix}_quick_do_{cid}"
        file_type_key = f"{key_prefix}_quick_file_type_{cid}"
        upload_key = f"{key_prefix}_quick_upload_{cid}"
        upload_btn_key = f"{key_prefix}_quick_upload_btn_{cid}"
        action_key = f"{key_prefix}_quick_action_{cid}"
        purchase_date_value = row.get("purchase_date")
        purchase_date_label = ""
        if isinstance(purchase_date_value, pd.Timestamp) and not pd.isna(purchase_date_value):
            purchase_date_label = purchase_date_value.strftime(DATE_FMT)
        row_cols[row_idx + 0].text_input(
            "Name",
            value=clean_text(row.get("name")) or "",
            key=name_key,
            label_visibility="collapsed",
        )
        row_cols[row_idx + 1].text_input(
            "Company",
            value=clean_text(row.get("company_name")) or "",
            key=company_key,
            label_visibility="collapsed",
        )
        row_cols[row_idx + 2].text_input(
            "Phone",
            value=clean_text(row.get("phone")) or "",
            key=phone_key,
            label_visibility="collapsed",
        )
        row_cols[row_idx + 3].text_input(
            "Billing address",
            value=clean_text(row.get("address")) or "",
            key=address_key,
            label_visibility="collapsed",
        )
        row_cols[row_idx + 4].text_input(
            "Delivery address",
            value=clean_text(row.get("delivery_address")) or "",
            key=delivery_key,
            label_visibility="collapsed",
        )
        row_cols[row_idx + 5].text_input(
            "Remarks",
            value=clean_text(row.get("remarks")) or "",
            key=remarks_key,
            label_visibility="collapsed",
        )
        row_cols[row_idx + 6].text_input(
            "Purchase date",
            value=purchase_date_label,
            key=purchase_key,
            label_visibility="collapsed",
            disabled=not is_admin,
            help="Only admins can edit the purchase date.",
        )
        row_cols[row_idx + 7].text_input(
            "Product",
            value=clean_text(row.get("product_info")) or "",
            key=product_key,
            label_visibility="collapsed",
        )
        col_offset = 8
        if show_do_code:
            row_cols[row_idx + col_offset].text_input(
                "DO code",
                value=clean_text(row.get("delivery_order_code")) or "",
                key=do_key,
                label_visibility="collapsed",
            )
            col_offset += 1
        if show_duplicate:
            row_cols[row_idx + col_offset].write(clean_text(row.get("duplicate")) or "")
            col_offset += 1
        view_docs = docs_map.get(cid, [])
        view_target = row_cols[row_idx + col_offset]
        view_container = getattr(st, "popover", None)
        view_label = "📎 View" if view_docs else "View"
        if use_popover and callable(view_container):
            view_panel = view_target.popover(view_label, use_container_width=True)
        else:
            view_panel = view_target.expander(view_label, expanded=False)
        with view_panel:
            if not view_docs:
                st.caption("No documents available.")
            else:
                for idx, doc in enumerate(view_docs):
                    label = clean_text(doc.get("label")) or "Document"
                    uploaded_at = pd.to_datetime(
                        doc.get("uploaded_at"), errors="coerce"
                    )
                    suffix = (
                        f" ({uploaded_at.strftime('%d-%m-%Y')})"
                        if pd.notna(uploaded_at)
                        else ""
                    )
                    path = resolve_upload_path(doc.get("path"))
                    if path and path.exists():
                        st.download_button(
                            f"{label}{suffix}",
                            data=path.read_bytes(),
                            file_name=path.name,
                            key=f"{key_prefix}_quick_view_{cid}_{idx}",
                        )
                    else:
                        st.caption(f"{label}{suffix} (file missing)")
        col_offset += 1
        if enable_uploads:
            upload_container = getattr(st, "popover", None)
            upload_target = row_cols[row_idx + col_offset]
            if use_popover and callable(upload_container):
                upload_panel = upload_target.popover("Upload", use_container_width=True)
            else:
                upload_panel = upload_target.expander("Upload", expanded=False)
            with upload_panel:
                with st.form(f"{upload_key}_form"):
                    doc_type = st.selectbox(
                        "Document type",
                        options=doc_type_options,
                        key=file_type_key,
                    )
                    upload_file = st.file_uploader(
                        "Upload document",
                        type=None,
                        accept_multiple_files=False,
                        key=upload_key,
                    )
                    details = _render_doc_detail_inputs(
                        doc_type,
                        key_prefix=f"{upload_key}_details",
                        defaults=row,
                    )
                    doc_type_emoji = {
                        "Delivery order": "🚚",
                        "Work done": "✅",
                        "Quotation": "🧾",
                        "Service": "🛠️",
                        "Maintenance": "🧰",
                        "Other": "📎",
                    }
                    upload_label = f"{doc_type_emoji.get(doc_type, '📎')} Upload {doc_type}"
                    upload_clicked = st.form_submit_button(upload_label)
                if _guard_double_submit(upload_btn_key, upload_clicked):
                    if upload_file is None:
                        st.warning("Select a file to upload.")
                    else:
                        saved = _save_customer_document_upload(
                            conn,
                            customer_id=cid,
                            customer_record=row,
                            doc_type=doc_type,
                            upload_file=upload_file,
                            details=details,
                        )
                        if saved:
                            st.success("Document uploaded.")
                            _safe_rerun()
            col_offset += 1
        st.session_state.setdefault(action_key, "Keep")
        action_value = st.session_state.get(action_key, "Keep")
        action_col = row_cols[row_idx + col_offset]
        if action_icon_only:
            row_created_by = int_or_none(original_map.get(cid, {}).get("created_by"))
            can_delete_row = is_admin or (
                current_actor_id is not None and row_created_by == current_actor_id
            )
            if not can_delete_row:
                st.session_state[action_key] = "Keep"
                action_value = "Keep"
                action_col.button(
                    "🔒",
                    key=f"{action_key}_trash",
                    help="Only admins or record owners can delete",
                    disabled=True,
                    use_container_width=True,
                )
            else:
                delete_help = (
                    "Toggle delete selection"
                    if is_admin
                    else "Toggle delete selection (your own customers)"
                )
                if action_col.button(
                    "🗑️",
                    key=f"{action_key}_trash",
                    help=delete_help,
                    use_container_width=True,
                ):
                    st.session_state[action_key] = (
                        "Keep" if action_value == "Delete" else "Delete"
                    )
                    action_value = st.session_state[action_key]
                if action_value == "Delete":
                    action_col.caption("Marked")
        else:
            action_col.selectbox(
                "Action",
                options=["Keep", "Delete"],
                key=action_key,
                label_visibility="collapsed",
            )
            action_value = st.session_state.get(action_key)
        editor_rows.append(
            {
                "id": cid,
                "name": st.session_state.get(name_key),
                "company_name": st.session_state.get(company_key),
                "phone": st.session_state.get(phone_key),
                "address": st.session_state.get(address_key),
                "delivery_address": st.session_state.get(delivery_key),
                "remarks": st.session_state.get(remarks_key),
                "purchase_date": st.session_state.get(purchase_key),
                "product_info": st.session_state.get(product_key),
                "delivery_order_code": st.session_state.get(do_key)
                if show_do_code
                else clean_text(row.get("delivery_order_code")),
                "sales_person": row.get("sales_person"),
                "Action": action_value,
            }
        )
    if not is_admin:
        st.caption(
            "Delete actions are limited to admins or the staff member who created the customer."
        )
    if is_admin and not editor_df.empty:
        delete_labels: dict[int, str] = {}
        for record in editor_df.to_dict("records"):
            cid = int_or_none(record.get("id"))
            if cid is None:
                continue
            name_val = clean_text(record.get("name")) or "(no name)"
            phone_val = clean_text(record.get("phone")) or "-"
            delete_labels[cid] = f"#{cid} – {name_val} | {phone_val}"
        delete_choices = sorted(
            delete_labels.keys(), key=lambda cid: delete_labels[cid].lower()
        )
        delete_state_key = f"{key_prefix}_bulk_delete_ids"
        st.session_state.setdefault(delete_state_key, [])
        with st.form(f"{key_prefix}_bulk_customer_delete"):
            selected_delete_ids = st.multiselect(
                "Select customers to delete",
                delete_choices,
                key=delete_state_key,
                format_func=lambda cid: delete_labels.get(
                    int(cid), f"Customer #{cid}"
                ),
                help="Removes the selected customers and their related records.",
            )
            bulk_delete_submit = st.form_submit_button(
                "Delete selected customers",
                disabled=(not is_admin) or (not selected_delete_ids),
                type="secondary",
            )
        if bulk_delete_submit and selected_delete_ids:
            if not is_admin:
                st.error("Only admins can delete customer records.")
                return df_raw
            deleted_count = 0
            for cid in selected_delete_ids:
                try:
                    delete_customer_record(conn, int(cid))
                    deleted_count += 1
                except Exception as err:
                    st.error(f"Unable to delete customer #{cid}: {err}")
            if deleted_count:
                st.session_state[delete_state_key] = []
                st.warning(f"Deleted {deleted_count} customer(s).")
                _safe_rerun()
    if st.button(
        "Apply table updates",
        type="primary",
        key=f"{key_prefix}_apply_updates",
    ):
        editor_result = pd.DataFrame(editor_rows)
        if editor_result.empty:
            st.info("No rows to update.")
        else:
            phones_to_recalc: set[str] = set()
            updates = deletes = 0
            errors: list[str] = []
            made_updates = False
            activity_events: list[tuple[str, int, str]] = []
            for row in editor_result.to_dict("records"):
                cid = int_or_none(row.get("id"))
                if cid is None or cid not in original_map:
                    continue
                action = str(row.get("Action") or "Keep").strip().lower()
                if action == "delete":
                    row_created_by = int_or_none(original_map.get(cid, {}).get("created_by"))
                    can_delete_row = is_admin or (
                        current_actor_id is not None and row_created_by == current_actor_id
                    )
                    if can_delete_row:
                        delete_customer_record(conn, cid)
                        deletes += 1
                    else:
                        errors.append(
                            f"Only admins or record owners can delete customers (ID #{cid})."
                        )
                    continue
                new_name = clean_text(row.get("name"))
                new_company = clean_text(row.get("company_name"))
                new_phone = clean_text(row.get("phone"))
                new_address = clean_text(row.get("address"))
                new_delivery_address = clean_text(row.get("delivery_address"))
                new_remarks = clean_text(row.get("remarks"))
                purchase_str, _ = date_strings_from_input(row.get("purchase_date"))
                product_label = clean_text(row.get("product_info"))
                new_do = clean_text(row.get("delivery_order_code"))
                new_sales_person = clean_text(row.get("sales_person"))
                original_row = original_map[cid]
                old_name = clean_text(original_row.get("name"))
                old_company = clean_text(original_row.get("company_name"))
                old_phone = clean_text(original_row.get("phone"))
                old_address = clean_text(original_row.get("address"))
                old_delivery_address = clean_text(original_row.get("delivery_address"))
                old_remarks = clean_text(original_row.get("remarks"))
                old_purchase = clean_text(original_row.get("purchase_date"))
                old_product = clean_text(original_row.get("product_info"))
                old_do = clean_text(original_row.get("delivery_order_code"))
                old_sales_person = clean_text(original_row.get("sales_person"))
                changes: list[str] = []
                if (
                    new_name == old_name
                    and new_company == old_company
                    and new_phone == old_phone
                    and new_address == old_address
                    and new_delivery_address == old_delivery_address
                    and new_remarks == old_remarks
                    and purchase_str == old_purchase
                    and product_label == old_product
                    and new_do == old_do
                    and new_sales_person == old_sales_person
                ):
                    continue
                conn.execute(
                    "UPDATE customers SET name=?, company_name=?, phone=?, address=?, delivery_address=?, remarks=?, purchase_date=?, product_info=?, delivery_order_code=?, sales_person=?, dup_flag=0 WHERE customer_id=?",
                    (
                        new_name,
                        new_company,
                        new_phone,
                        new_address,
                        new_delivery_address,
                        new_remarks,
                        purchase_str,
                        product_label,
                        new_do,
                        new_sales_person,
                        cid,
                    ),
                )
                if new_name != old_name:
                    changes.append("name")
                if new_company != old_company:
                    changes.append("company")
                if new_phone != old_phone:
                    changes.append("phone")
                if new_address != old_address:
                    changes.append("billing address")
                if new_delivery_address != old_delivery_address:
                    changes.append("delivery address")
                if new_remarks != old_remarks:
                    changes.append("remarks")
                if purchase_str != old_purchase:
                    changes.append("purchase date")
                if product_label != old_product:
                    changes.append("products")
                if new_do != old_do:
                    changes.append("DO code")
                if new_sales_person != old_sales_person:
                    changes.append("sales person")
                if new_do:
                    conn.execute(
                        """
                        INSERT INTO delivery_orders (do_number, customer_id, order_id, description, sales_person, remarks, file_path)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(do_number) DO UPDATE SET
                            customer_id=excluded.customer_id,
                            description=excluded.description,
                            sales_person=excluded.sales_person,
                            remarks=excluded.remarks
                        """,
                        (
                            new_do,
                            cid,
                            None,
                            product_label,
                            new_sales_person,
                            new_remarks,
                            None,
                        ),
                    )
                if old_do and old_do != new_do:
                    conn.execute(
                        "DELETE FROM delivery_orders WHERE do_number=? AND (customer_id IS NULL OR customer_id=?)",
                        (old_do, cid),
                    )
                conn.execute(
                    "UPDATE import_history SET customer_name=?, phone=?, address=?, delivery_address=?, product_label=?, do_number=?, original_date=? WHERE customer_id=? AND deleted_at IS NULL",
                    (
                        new_name,
                        new_phone,
                        new_address,
                        new_delivery_address,
                        product_label,
                        new_do,
                        purchase_str,
                        cid,
                    ),
                )
                if old_phone and old_phone != new_phone:
                    phones_to_recalc.add(old_phone)
                if new_phone:
                    phones_to_recalc.add(new_phone)
                updates += 1
                made_updates = True
                if changes:
                    display_name = new_name or old_name or f"Customer #{cid}"
                    summary = ", ".join(changes)
                    activity_events.append(
                        (
                            "customer_updated",
                            cid,
                            f"Updated {display_name} ({summary})",
                        )
                    )
            if made_updates:
                conn.commit()
            if phones_to_recalc:
                for phone_value in phones_to_recalc:
                    recalc_customer_duplicate_flag(conn, phone_value)
                conn.commit()
            for event_type, entity_id, description in activity_events:
                log_activity(
                    conn,
                    event_type=event_type,
                    description=description,
                    entity_type="customer",
                    entity_id=int(entity_id),
                )
            if errors:
                for err in errors:
                    st.error(err)
            if updates or deletes:
                st.success(f"Updated {updates} row(s) and deleted {deletes} row(s).")
                if not errors:
                    _safe_rerun()
            elif not errors:
                st.info("No changes detected.")
    return df_raw


def _render_doc_detail_inputs(
    doc_type: str,
    *,
    key_prefix: str,
    defaults: Optional[dict[str, object]] = None,
) -> dict[str, object]:
    defaults = defaults or {}
    details: dict[str, object] = {}
    current_user = get_current_user() or {}
    user_label = clean_text(current_user.get("username")) or ""
    if doc_type == "Quotation":
        details["reference"] = st.text_input(
            "Quotation reference",
            key=f"{key_prefix}_quotation_reference",
        )
        items_key = f"{key_prefix}_quotation_items"
        st.session_state.setdefault(items_key, _default_quotation_items())
        items_df = pd.DataFrame(st.session_state.get(items_key, []))
        for col in ["description", "quantity", "rate", "total_price"]:
            if col not in items_df.columns:
                items_df[col] = 0.0 if col != "description" else ""
        items_df["total_price"] = items_df.apply(
            lambda row: max(
                _coerce_float(row.get("quantity"), 0.0)
                * _coerce_float(row.get("rate"), 0.0),
                0.0,
            ),
            axis=1,
        )
        st.markdown("**Quotation items**")
        edited_items = st.data_editor(
            items_df[["description", "quantity", "rate"]],
            key=f"{items_key}_editor",
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "description": st.column_config.TextColumn("Product"),
                "quantity": st.column_config.NumberColumn("Qty", min_value=0.0, step=1.0, format="%d"),
                "rate": st.column_config.NumberColumn(
                    "Unit price", min_value=0.0, step=100.0, format="%.2f"
                ),
            },
        )
        items_records = (
            edited_items.to_dict("records")
            if isinstance(edited_items, pd.DataFrame)
            else edited_items
        )
        details["items"] = items_records
        st.session_state[items_key] = items_records
        default_purchase_date = parse_date_value(defaults.get("purchase_date"))
        details["quote_date"] = st.date_input(
            "Quotation date",
            value=default_purchase_date or date.today(),
            key=f"{key_prefix}_quotation_date",
            format="DD-MM-YYYY",
        )
        details["payment_status"] = st.selectbox(
            "Payment status",
            options=["pending", "paid", "rejected"],
            key=f"{key_prefix}_quotation_payment_status",
            format_func=lambda status: status.title(),
        )
        details["follow_up_notes"] = st.text_area(
            "Follow-up note",
            key=f"{key_prefix}_quotation_follow_up_notes",
            help="Internal note to track the next action or update.",
        )
        details["person_in_charge"] = st.text_input(
            "Person in charge (optional)",
            value=user_label,
            key=f"{key_prefix}_quotation_person_in_charge",
        )
        details["receipt_upload"] = st.file_uploader(
            "Payment receipt (required for paid)",
            type=["pdf", "png", "jpg", "jpeg", "webp"],
            key=f"{key_prefix}_quotation_receipt",
            help="Attach the receipt when marking this quotation as paid.",
        )
    elif doc_type in ("Delivery order", "Work done"):
        items_key = f"{key_prefix}_delivery_items"
        st.session_state.setdefault(items_key, _default_simple_items())
        items_df = pd.DataFrame(st.session_state.get(items_key, []))
        for col in ["description", "quantity", "unit_price", "total"]:
            if col not in items_df.columns:
                items_df[col] = 0.0 if col != "description" else ""
        items_df["total"] = items_df.apply(
            lambda row: max(
                _coerce_float(row.get("quantity"), 0.0)
                * _coerce_float(row.get("unit_price"), 0.0),
                0.0,
            ),
            axis=1,
        )
        st.markdown("**Products**")
        edited_items = st.data_editor(
            items_df[["description", "quantity", "unit_price"]],
            key=f"{items_key}_editor",
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "description": st.column_config.TextColumn("Product"),
                "quantity": st.column_config.NumberColumn("Qty", min_value=0.0, step=1.0, format="%d"),
                "unit_price": st.column_config.NumberColumn(
                    "Unit price", min_value=0.0, step=100.0, format="%.2f"
                ),
            },
        )
        items_records = (
            edited_items.to_dict("records")
            if isinstance(edited_items, pd.DataFrame)
            else edited_items
        )
        details["items"] = items_records
        st.session_state[items_key] = items_records
        label = "Delivery order number" if doc_type == "Delivery order" else "Work done number"
        details["do_number"] = st.text_input(
            label,
            key=f"{key_prefix}_do_number",
        )
        details["status"] = st.selectbox(
            "Status",
            options=DELIVERY_STATUS_OPTIONS,
            key=f"{key_prefix}_do_status",
            format_func=lambda option: DELIVERY_STATUS_LABELS.get(option, option.title()),
        )
        details["person_in_charge"] = st.text_input(
            "Person in charge (optional)",
            value=clean_text(defaults.get("sales_person")) or user_label,
            key=f"{key_prefix}_do_person_in_charge",
        )
        details["description"] = st.text_area(
            "Description",
            key=f"{key_prefix}_do_description",
        )
        details["remarks"] = st.text_area(
            "Remarks",
            key=f"{key_prefix}_do_remarks",
        )
        details["advance_receipt_upload"] = None
        details["receipt_upload"] = None
        status_value = normalize_delivery_status(details.get("status"))
        details["advance_taken"] = False
        if status_value == "advanced":
            details["advance_taken"] = st.checkbox(
                "Advance payment was received",
                key=f"{key_prefix}_do_advance_taken",
                help="Enable if an advance receipt should be attached.",
            )
            if details.get("advance_taken"):
                details["advance_receipt_upload"] = st.file_uploader(
                    "Advance receipt (highly recommended)",
                    type=["pdf", "png", "jpg", "jpeg", "webp"],
                    key=f"{key_prefix}_do_advance_receipt",
                    help="Add the advance receipt when a deposit was collected.",
                )
        if status_value == "paid":
            details["receipt_upload"] = st.file_uploader(
                "Full payment receipt (highly recommended)",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"{key_prefix}_do_receipt",
                help="Attach the final receipt for paid delivery orders/work done.",
            )
    elif doc_type == "Service":
        items_key = f"{key_prefix}_service_items"
        st.session_state.setdefault(items_key, _default_simple_items())
        items_df = pd.DataFrame(st.session_state.get(items_key, []))
        for col in ["description", "quantity", "unit_price", "total"]:
            if col not in items_df.columns:
                items_df[col] = 0.0 if col != "description" else ""
        items_df["total"] = items_df.apply(
            lambda row: max(
                _coerce_float(row.get("quantity"), 0.0)
                * _coerce_float(row.get("unit_price"), 0.0),
                0.0,
            ),
            axis=1,
        )
        st.markdown("**Products**")
        edited_items = st.data_editor(
            items_df[["description", "quantity", "unit_price"]],
            key=f"{items_key}_editor",
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "description": st.column_config.TextColumn("Product"),
                "quantity": st.column_config.NumberColumn("Qty", min_value=0.0, step=1.0, format="%d"),
                "unit_price": st.column_config.NumberColumn(
                    "Unit price", min_value=0.0, step=100.0, format="%.2f"
                ),
            },
        )
        items_records = (
            edited_items.to_dict("records")
            if isinstance(edited_items, pd.DataFrame)
            else edited_items
        )
        details["items"] = items_records
        st.session_state[items_key] = items_records
        details["service_date"] = st.date_input(
            "Service date",
            value=date.today(),
            key=f"{key_prefix}_service_date",
            format="DD-MM-YYYY",
        )
        details["description"] = st.text_area(
            "Service description",
            key=f"{key_prefix}_service_description",
        )
        details["status"] = st.selectbox(
            "Progress status",
            SERVICE_STATUS_OPTIONS,
            key=f"{key_prefix}_service_status",
        )
        details["remarks"] = st.text_area(
            "Remarks",
            key=f"{key_prefix}_service_remarks",
        )
        details["payment_status"] = st.selectbox(
            "Payment status",
            ["pending", "advanced", "paid"],
            key=f"{key_prefix}_service_payment_status",
        )
        details["person_in_charge"] = st.text_input(
            "Person in charge (optional)",
            value=user_label,
            key=f"{key_prefix}_service_person_in_charge",
        )
        if details.get("payment_status") in {"advanced", "paid"}:
            details["receipt_upload"] = st.file_uploader(
                "Payment receipt (highly recommended)",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"{key_prefix}_service_receipt",
            )
    elif doc_type == "Maintenance":
        items_key = f"{key_prefix}_maintenance_items"
        st.session_state.setdefault(items_key, _default_simple_items())
        items_df = pd.DataFrame(st.session_state.get(items_key, []))
        for col in ["description", "quantity", "unit_price", "total"]:
            if col not in items_df.columns:
                items_df[col] = 0.0 if col != "description" else ""
        items_df["total"] = items_df.apply(
            lambda row: max(
                _coerce_float(row.get("quantity"), 0.0)
                * _coerce_float(row.get("unit_price"), 0.0),
                0.0,
            ),
            axis=1,
        )
        st.markdown("**Products**")
        edited_items = st.data_editor(
            items_df[["description", "quantity", "unit_price"]],
            key=f"{items_key}_editor",
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "description": st.column_config.TextColumn("Product"),
                "quantity": st.column_config.NumberColumn("Qty", min_value=0.0, step=1.0, format="%d"),
                "unit_price": st.column_config.NumberColumn(
                    "Unit price", min_value=0.0, step=100.0, format="%.2f"
                ),
            },
        )
        items_records = (
            edited_items.to_dict("records")
            if isinstance(edited_items, pd.DataFrame)
            else edited_items
        )
        details["items"] = items_records
        st.session_state[items_key] = items_records
        details["maintenance_date"] = st.date_input(
            "Maintenance date",
            value=date.today(),
            key=f"{key_prefix}_maintenance_date",
            format="DD-MM-YYYY",
        )
        details["description"] = st.text_area(
            "Maintenance description",
            key=f"{key_prefix}_maintenance_description",
        )
        details["status"] = st.selectbox(
            "Progress status",
            SERVICE_STATUS_OPTIONS,
            key=f"{key_prefix}_maintenance_status",
        )
        details["remarks"] = st.text_area(
            "Remarks",
            key=f"{key_prefix}_maintenance_remarks",
        )
        details["payment_status"] = st.selectbox(
            "Payment status",
            ["pending", "advanced", "paid"],
            key=f"{key_prefix}_maintenance_payment_status",
        )
        details["person_in_charge"] = st.text_input(
            "Person in charge (optional)",
            value=user_label,
            key=f"{key_prefix}_maintenance_person_in_charge",
        )
        if details.get("payment_status") in {"advanced", "paid"}:
            details["receipt_upload"] = st.file_uploader(
                "Payment receipt (highly recommended)",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"{key_prefix}_maintenance_receipt",
            )
    elif doc_type == "Other":
        items_key = f"{key_prefix}_other_items"
        st.session_state.setdefault(items_key, _default_simple_items())
        items_df = pd.DataFrame(st.session_state.get(items_key, []))
        for col in ["description", "quantity", "unit_price", "total"]:
            if col not in items_df.columns:
                items_df[col] = 0.0 if col != "description" else ""
        items_df["total"] = items_df.apply(
            lambda row: max(
                _coerce_float(row.get("quantity"), 0.0)
                * _coerce_float(row.get("unit_price"), 0.0),
                0.0,
            ),
            axis=1,
        )
        st.markdown("**Items purchased (optional)**")
        edited_items = st.data_editor(
            items_df[["description", "quantity", "unit_price"]],
            key=f"{items_key}_editor",
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "description": st.column_config.TextColumn("Item"),
                "quantity": st.column_config.NumberColumn("Qty", min_value=0.0, step=1.0, format="%d"),
                "unit_price": st.column_config.NumberColumn(
                    "Unit price", min_value=0.0, step=100.0, format="%.2f"
                ),
            },
        )
        items_records = (
            edited_items.to_dict("records")
            if isinstance(edited_items, pd.DataFrame)
            else edited_items
        )
        details["items"] = items_records
        st.session_state[items_key] = items_records
        details["description"] = st.text_area(
            "Description",
            key=f"{key_prefix}_other_description",
        )
    return details


def _clear_operations_upload_state(
    *,
    file_key: str,
    details_key_prefix: str,
    doc_type: str,
) -> None:
    keys_to_clear = [
        file_key,
        f"{file_key}_ocr_token",
        f"{file_key}_ocr_text",
        f"{file_key}_ocr_warnings",
    ]
    if doc_type in ("Delivery order", "Work done"):
        keys_to_clear.extend(
            [
                f"{details_key_prefix}_delivery_items",
                f"{details_key_prefix}_delivery_items_editor",
                f"{details_key_prefix}_do_number",
                f"{details_key_prefix}_do_status",
                f"{details_key_prefix}_do_person_in_charge",
                f"{details_key_prefix}_do_description",
                f"{details_key_prefix}_do_remarks",
                f"{details_key_prefix}_do_advance_taken",
                f"{details_key_prefix}_do_advance_receipt",
                f"{details_key_prefix}_do_receipt",
            ]
        )
    elif doc_type == "Service":
        keys_to_clear.extend(
            [
                f"{details_key_prefix}_service_items",
                f"{details_key_prefix}_service_items_editor",
                f"{details_key_prefix}_service_date",
                f"{details_key_prefix}_service_description",
                f"{details_key_prefix}_service_status",
                f"{details_key_prefix}_service_remarks",
                f"{details_key_prefix}_service_payment_status",
                f"{details_key_prefix}_service_person_in_charge",
                f"{details_key_prefix}_service_receipt",
            ]
        )
    elif doc_type == "Maintenance":
        keys_to_clear.extend(
            [
                f"{details_key_prefix}_maintenance_items",
                f"{details_key_prefix}_maintenance_items_editor",
                f"{details_key_prefix}_maintenance_date",
                f"{details_key_prefix}_maintenance_description",
                f"{details_key_prefix}_maintenance_status",
                f"{details_key_prefix}_maintenance_remarks",
                f"{details_key_prefix}_maintenance_payment_status",
                f"{details_key_prefix}_maintenance_person_in_charge",
                f"{details_key_prefix}_maintenance_receipt",
            ]
        )
    elif doc_type == "Other":
        keys_to_clear.extend(
            [
                f"{details_key_prefix}_other_items",
                f"{details_key_prefix}_other_items_editor",
                f"{details_key_prefix}_other_description",
            ]
        )
    for key in keys_to_clear:
        st.session_state.pop(key, None)


def _save_customer_document_upload(
    conn,
    *,
    customer_id: int,
    customer_record: dict[str, object],
    doc_type: str,
    upload_file,
    details: dict[str, object],
) -> bool:
    current_user = get_current_user() or {}
    user_label = clean_text(current_user.get("username")) or ""
    if doc_type in ("Delivery order", "Work done"):
        do_number = clean_text(details.get("do_number"))
        if not do_number:
            st.error("Provide a delivery/work done number before saving.")
            return False
        status_value = normalize_delivery_status(details.get("status"))
        if status_value == "advanced" and details.get("advance_receipt_upload") is None:
            st.warning("Advance receipt is highly recommended for advanced records.")
        if status_value == "paid":
            if details.get("receipt_upload") is None:
                st.warning("Full payment receipt is highly recommended for paid records.")
            if details.get("advance_taken") and details.get("advance_receipt_upload") is None:
                st.warning("Advance receipt is highly recommended when an advance was taken.")
    if doc_type == "Quotation":
        status_value = clean_text(details.get("payment_status")) or "pending"
        if not clean_text(details.get("reference")):
            st.error("Quotation reference is required.")
            return False
        if not clean_text(details.get("follow_up_notes")):
            st.error("Follow-up note is required for quotations.")
            return False
        items_input = details.get("items") or []
        items_clean, totals_data = normalize_quotation_items(items_input)
        if not items_clean:
            st.error("Add at least one quotation item before saving.")
            return False
        if status_value == "paid" and details.get("receipt_upload") is None:
            st.error("Upload a receipt before marking this quotation as paid.")
            return False
        details["items"] = items_clean
        details["total_amount"] = totals_data["grand_total"]
    if doc_type == "Service":
        if not clean_text(details.get("description")):
            st.error("Service description is required.")
            return False
        items_input = details.get("items") or []
        items_clean, _ = normalize_simple_items(items_input)
        if not items_clean:
            st.error("Add at least one service product item before saving.")
            return False
        if details.get("payment_status") in {"advanced", "paid"} and details.get("receipt_upload") is None:
            st.warning("Payment receipt is highly recommended for advanced or paid service records.")
        details["items"] = items_clean
    if doc_type == "Maintenance":
        if not clean_text(details.get("description")):
            st.error("Maintenance description is required.")
            return False
        items_input = details.get("items") or []
        items_clean, _ = normalize_simple_items(items_input)
        if not items_clean:
            st.error("Add at least one maintenance product item before saving.")
            return False
        if details.get("payment_status") in {"advanced", "paid"} and details.get("receipt_upload") is None:
            st.warning("Payment receipt is highly recommended for advanced or paid maintenance records.")
        details["items"] = items_clean
    if doc_type == "Other":
        if not clean_text(details.get("description")):
            st.error("Description is required for other uploads.")
            return False
        items_input = details.get("items") or []
        items_clean, _ = normalize_simple_items(items_input)
        details["items"] = items_clean

    doc_dir_map = {
        "Delivery order": DELIVERY_ORDER_DIR,
        "Work done": DELIVERY_ORDER_DIR,
        "Quotation": QUOTATION_DOCS_DIR,
        "Service": SERVICE_DOCS_DIR,
        "Maintenance": MAINTENANCE_DOCS_DIR,
        "Other": OPERATIONS_OTHER_DIR,
    }
    target_dir = doc_dir_map.get(doc_type, CUSTOMER_DOCS_DIR)
    target_dir.mkdir(parents=True, exist_ok=True)
    doc_type_slug = _sanitize_path_component(doc_type.lower().replace(" ", "_")) or "document"
    original_name = upload_file.name or f"{doc_type_slug}_{customer_id}.pdf"
    safe_original = Path(original_name).name
    filename = f"{doc_type_slug}_{customer_id}_{safe_original}"
    saved_path = save_uploaded_file(
        upload_file,
        target_dir,
        filename=filename,
        allowed_extensions={".pdf", ".png", ".jpg", ".jpeg", ".webp"},
        default_extension=".pdf",
    )
    if not saved_path:
        st.error("Unable to save the uploaded file.")
        return False
    try:
        stored_path = str(saved_path.relative_to(BASE_DIR))
    except ValueError:
        stored_path = str(saved_path)

    uploader_id = current_user_id()
    conn.execute(
        """
        INSERT INTO customer_documents (
            customer_id, doc_type, file_path, original_name, uploaded_by
        ) VALUES (?, ?, ?, ?, ?)
        """,
        (
            int(customer_id),
            doc_type,
            stored_path,
            safe_original,
            uploader_id,
        ),
    )

    new_product_labels: list[str] = []
    if doc_type == "Quotation":
        receipt_path = None
        receipt_upload = details.get("receipt_upload")
        if receipt_upload is not None:
            safe_ref = _sanitize_path_component(
                clean_text(details.get("reference")) or f"quotation_{customer_id}"
            )
            receipt_path = store_payment_receipt(
                receipt_upload,
                identifier=f"{safe_ref}_receipt",
                target_dir=QUOTATION_RECEIPT_DIR,
            )
        items_payload = None
        items_clean = details.get("items") or []
        if items_clean:
            items_payload = json.dumps(items_clean, ensure_ascii=False)
        payload = {
            "reference": clean_text(details.get("reference")),
            "quote_date": to_iso_date(details.get("quote_date")),
            "customer_name": clean_text(customer_record.get("name")),
            "customer_company": clean_text(customer_record.get("company_name")),
            "customer_address": clean_text(customer_record.get("address")),
            "customer_contact": clean_text(customer_record.get("phone")),
            "total_amount": _coerce_float(details.get("total_amount"), 0.0),
            "status": clean_text(details.get("payment_status")) or "pending",
            "payment_receipt_path": receipt_path,
            "follow_up_status": "Pending",
            "follow_up_notes": clean_text(details.get("follow_up_notes")),
            "salesperson_name": clean_text(details.get("person_in_charge")),
            "document_path": stored_path,
            "items_payload": items_payload,
            "created_by": uploader_id,
        }
        _save_quotation_record(conn, payload)
    elif doc_type in ("Delivery order", "Work done"):
        do_number = clean_text(details.get("do_number"))
        status_value = normalize_delivery_status(details.get("status"))
        receipt_path = None
        receipt_upload = details.get("receipt_upload")
        advance_receipt_upload = details.get("advance_receipt_upload")
        if status_value == "advanced" and advance_receipt_upload is not None:
            receipt_upload = advance_receipt_upload
        if receipt_upload is not None:
            receipt_path = store_payment_receipt(
                receipt_upload,
                identifier=f"{_sanitize_path_component(do_number)}_receipt",
                target_dir=DELIVERY_RECEIPT_DIR,
            )
        if advance_receipt_upload is not None:
            advance_path = store_payment_receipt(
                advance_receipt_upload,
                identifier=f"{_sanitize_path_component(do_number)}_advance_receipt",
                target_dir=DELIVERY_RECEIPT_DIR,
            )
            if advance_path:
                conn.execute(
                    """
                    INSERT INTO customer_documents (
                        customer_id, doc_type, file_path, original_name, uploaded_by
                    ) VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        int(customer_id),
                        f"{doc_type} advance receipt",
                        advance_path,
                        Path(advance_receipt_upload.name).name,
                        uploader_id,
                    ),
                )
        items_clean, total_amount = normalize_simple_items(details.get("items") or [])
        if items_clean:
            new_product_labels = format_simple_item_labels(items_clean)
        conn.execute(
            """
            INSERT INTO delivery_orders (
                do_number, customer_id, description, sales_person, remarks, file_path,
                items_payload, total_amount, record_type, status, payment_receipt_path, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            ON CONFLICT(do_number) DO UPDATE SET
                customer_id=excluded.customer_id,
                description=excluded.description,
                sales_person=excluded.sales_person,
                remarks=excluded.remarks,
                file_path=COALESCE(excluded.file_path, delivery_orders.file_path),
                items_payload=COALESCE(excluded.items_payload, delivery_orders.items_payload),
                total_amount=COALESCE(excluded.total_amount, delivery_orders.total_amount),
                record_type=excluded.record_type,
                status=excluded.status,
                payment_receipt_path=COALESCE(excluded.payment_receipt_path, delivery_orders.payment_receipt_path),
                updated_at=datetime('now')
            """,
            (
                do_number,
                int(customer_id),
                clean_text(details.get("description")),
                clean_text(details.get("person_in_charge")) or clean_text(user_label),
                clean_text(details.get("remarks")),
                stored_path,
                json.dumps(items_clean, ensure_ascii=False) if items_clean else None,
                total_amount if items_clean else None,
                "work_done" if doc_type == "Work done" else "delivery_order",
                status_value,
                receipt_path,
            ),
        )
    elif doc_type == "Service":
        items_clean, total_amount = normalize_simple_items(details.get("items") or [])
        if items_clean:
            new_product_labels = format_simple_item_labels(items_clean)
        product_info = ", ".join(new_product_labels)
        receipt_path = None
        receipt_upload = details.get("receipt_upload")
        if receipt_upload is not None:
            receipt_path = store_payment_receipt(
                receipt_upload,
                identifier=f"{_sanitize_path_component(str(customer_id))}_service_receipt",
            )
        remarks = clean_text(details.get("remarks"))
        person_in_charge = clean_text(details.get("person_in_charge"))
        if person_in_charge:
            remarks = dedupe_join([remarks, f"Person in charge: {person_in_charge}"], " | ")
        cur = conn.execute(
            """
            INSERT INTO services (
                customer_id, service_date, description, remarks, service_product_info,
                status, payment_status, payment_receipt_path, bill_amount, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(customer_id),
                to_iso_date(details.get("service_date")),
                clean_text(details.get("description")),
                remarks,
                product_info,
                clean_text(details.get("status")) or DEFAULT_SERVICE_STATUS,
                details.get("payment_status") or "pending",
                receipt_path,
                total_amount if total_amount else None,
                uploader_id,
            ),
        )
        service_id = cur.lastrowid
        conn.execute(
            """
            INSERT INTO service_documents (service_id, file_path, original_name)
            VALUES (?, ?, ?)
            """,
            (service_id, stored_path, safe_original),
        )
    elif doc_type == "Maintenance":
        items_clean, total_amount = normalize_simple_items(details.get("items") or [])
        if items_clean:
            new_product_labels = format_simple_item_labels(items_clean)
        product_info = ", ".join(new_product_labels)
        receipt_path = None
        receipt_upload = details.get("receipt_upload")
        if receipt_upload is not None:
            receipt_path = store_payment_receipt(
                receipt_upload,
                identifier=f"{_sanitize_path_component(str(customer_id))}_maintenance_receipt",
            )
        remarks = clean_text(details.get("remarks"))
        person_in_charge = clean_text(details.get("person_in_charge"))
        if person_in_charge:
            remarks = dedupe_join([remarks, f"Person in charge: {person_in_charge}"], " | ")
        cur = conn.execute(
            """
            INSERT INTO maintenance_records (
                customer_id, maintenance_date, description, remarks, maintenance_product_info,
                status, payment_status, payment_receipt_path, total_amount, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(customer_id),
                to_iso_date(details.get("maintenance_date")),
                clean_text(details.get("description")),
                remarks,
                product_info,
                clean_text(details.get("status")) or DEFAULT_SERVICE_STATUS,
                details.get("payment_status") or "pending",
                receipt_path,
                total_amount if total_amount else None,
                uploader_id,
            ),
        )
        maintenance_id = cur.lastrowid
        conn.execute(
            """
            INSERT INTO maintenance_documents (maintenance_id, file_path, original_name)
            VALUES (?, ?, ?)
            """,
            (maintenance_id, stored_path, safe_original),
        )
    elif doc_type == "Other":
        items_clean = details.get("items") or []
        items_payload = json.dumps(items_clean, ensure_ascii=False) if items_clean else None
        conn.execute(
            """
            INSERT INTO operations_other_documents (
                customer_id, description, items_payload, file_path, original_name, uploaded_by
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                int(customer_id),
                clean_text(details.get("description")),
                items_payload,
                stored_path,
                safe_original,
                uploader_id,
            ),
        )

    if new_product_labels:
        existing_products = clean_text(customer_record.get("product_info"))
        merged_products = existing_products or ""
        for label in new_product_labels:
            if label and label not in merged_products:
                merged_products = dedupe_join([merged_products, label], " ; ")
        if merged_products != existing_products:
            conn.execute(
                "UPDATE customers SET product_info=? WHERE customer_id=?",
                (merged_products, int(customer_id)),
            )

    conn.commit()
    return True


def render_customer_document_uploader(
    conn,
    *,
    key_prefix: str,
) -> None:
    st.markdown("### Customer document uploads")
    customer_options, customer_labels, _, _ = fetch_customer_choices(conn, only_complete=False)
    customer_choices = [cid for cid in customer_options if cid is not None]
    if not customer_choices:
        st.info("No customers available for document uploads yet.")
        return

    upload_container = getattr(st, "popover", None)
    if callable(upload_container):
        container = upload_container("Upload documents", use_container_width=True)
    else:
        container = st.expander("Upload documents", expanded=True)

    with container:
        selected_customer = st.selectbox(
            "Customer",
            customer_choices,
            format_func=lambda cid: customer_labels.get(cid, f"Customer #{cid}"),
            key=f"{key_prefix}_customer",
        )
        customer_seed = df_query(
            conn,
            """
            SELECT name, company_name, phone, address, delivery_address, sales_person
            FROM customers
            WHERE customer_id=?
            """,
            (int(selected_customer),),
        )
        customer_record = (
            customer_seed.iloc[0].to_dict() if not customer_seed.empty else {}
        )
        upload_cols = st.columns(2)
        with upload_cols[0]:
            st.markdown("**Quotation**")
            quote_file = st.file_uploader(
                "Quotation upload",
                type=None,
                accept_multiple_files=False,
                key=f"{key_prefix}_quote_file",
                help="Upload quotation PDFs or images.",
            )
            _apply_ocr_autofill(
                upload=quote_file,
                ocr_key_prefix=f"{key_prefix}_quote_file",
                doc_type="Quotation",
                details_key_prefix=f"{key_prefix}_quote_details",
            )
            quote_details = _render_doc_detail_inputs(
                "Quotation",
                key_prefix=f"{key_prefix}_quote_details",
                defaults=customer_record,
            )
            submit_quote = st.button(
                "Save quotation",
                key=f"{key_prefix}_quote_save",
            )
            if _guard_double_submit(f"{key_prefix}_quote_save", submit_quote):
                if quote_file is None:
                    st.error("Select a quotation document to upload.")
                else:
                    saved = _save_customer_document_upload(
                        conn,
                        customer_id=int(selected_customer),
                        customer_record=customer_record,
                        doc_type="Quotation",
                        upload_file=quote_file,
                        details=quote_details,
                    )
                    if saved:
                        st.success("Quotation uploaded.")
                        _safe_rerun()

        with upload_cols[1]:
            st.markdown("**Work done**")
            work_done_file = st.file_uploader(
                "Work done upload",
                type=None,
                accept_multiple_files=False,
                key=f"{key_prefix}_work_done_file",
                help="Upload completed work slips or PDFs.",
            )
            _apply_ocr_autofill(
                upload=work_done_file,
                ocr_key_prefix=f"{key_prefix}_work_done_file",
                doc_type="Work done",
                details_key_prefix=f"{key_prefix}_work_done_details",
            )
            work_done_details = _render_doc_detail_inputs(
                "Work done",
                key_prefix=f"{key_prefix}_work_done_details",
                defaults=customer_record,
            )
            submit_work_done = st.button(
                "Save work done",
                key=f"{key_prefix}_work_done_save",
            )
            if _guard_double_submit(f"{key_prefix}_work_done_save", submit_work_done):
                if work_done_file is None:
                    st.error("Select a work done document to upload.")
                else:
                    saved = _save_customer_document_upload(
                        conn,
                        customer_id=int(selected_customer),
                        customer_record=customer_record,
                        doc_type="Work done",
                        upload_file=work_done_file,
                        details=work_done_details,
                    )
                    if saved:
                        st.success("Work done uploaded.")
                        _clear_operations_upload_state(
                            file_key=f"{key_prefix}_work_done_file",
                            details_key_prefix=f"{key_prefix}_work_done_details",
                            doc_type="Work done",
                        )

        upload_cols = st.columns(2)
        with upload_cols[0]:
            st.markdown("**Delivery order (DO)**")
            with st.form(key=f"{key_prefix}_do_form", clear_on_submit=False):
                do_file = st.file_uploader(
                    "Delivery order upload",
                    type=None,
                    accept_multiple_files=False,
                    key=f"{key_prefix}_do_file",
                    help="Upload the delivery order PDF or image.",
                )
                _apply_ocr_autofill(
                    upload=do_file,
                    ocr_key_prefix=f"{key_prefix}_do_file",
                    doc_type="Delivery order",
                    details_key_prefix=f"{key_prefix}_do_details",
                )
                do_details = _render_doc_detail_inputs(
                    "Delivery order",
                    key_prefix=f"{key_prefix}_do_details",
                    defaults=customer_record,
                )
                submit_do = st.form_submit_button(
                    "Save delivery order",
                )
            if _guard_double_submit(f"{key_prefix}_do_save", submit_do):
                if do_file is None:
                    st.error("Select a delivery order document to upload.")
                else:
                    saved = _save_customer_document_upload(
                        conn,
                        customer_id=int(selected_customer),
                        customer_record=customer_record,
                        doc_type="Delivery order",
                        upload_file=do_file,
                        details=do_details,
                    )
                    if saved:
                        st.success("Delivery order uploaded.")
                        _clear_operations_upload_state(
                            file_key=f"{key_prefix}_do_file",
                            details_key_prefix=f"{key_prefix}_do_details",
                            doc_type="Delivery order",
                        )
                        _clear_operations_upload_state(
                            file_key=f"{key_prefix}_do_file",
                            details_key_prefix=f"{key_prefix}_do_details",
                            doc_type="Delivery order",
                        )

        with upload_cols[1]:
            st.markdown("**Service / Maintenance**")
            service_choice = st.selectbox(
                "Type",
                ["Service", "Maintenance"],
                key=f"{key_prefix}_service_choice",
            )
            service_file = st.file_uploader(
                "Service/Maintenance upload",
                type=None,
                accept_multiple_files=False,
                key=f"{key_prefix}_service_file",
                help="Upload service or maintenance documents.",
            )
            _apply_ocr_autofill(
                upload=service_file,
                ocr_key_prefix=f"{key_prefix}_service_file",
                doc_type=service_choice,
                details_key_prefix=f"{key_prefix}_service_details",
            )
            service_details = _render_doc_detail_inputs(
                service_choice,
                key_prefix=f"{key_prefix}_service_details",
                defaults=customer_record,
            )
            submit_service = st.button(
                "Save service/maintenance",
                key=f"{key_prefix}_service_save",
            )
            if _guard_double_submit(f"{key_prefix}_service_save", submit_service):
                if service_file is None:
                    st.error("Select a service/maintenance document to upload.")
                else:
                    saved = _save_customer_document_upload(
                        conn,
                        customer_id=int(selected_customer),
                        customer_record=customer_record,
                        doc_type=service_choice,
                        upload_file=service_file,
                        details=service_details,
                    )
                    if saved:
                        st.success(f"{service_choice} uploaded.")
                        _clear_operations_upload_state(
                            file_key=f"{key_prefix}_service_file",
                            details_key_prefix=f"{key_prefix}_service_details",
                            doc_type=service_choice,
                        )

    docs_df = df_query(
        conn,
        """
        SELECT document_id, doc_type, file_path, original_name, uploaded_at
        FROM customer_documents
        WHERE customer_id=?
          AND deleted_at IS NULL
        ORDER BY datetime(uploaded_at) DESC, document_id DESC
        """,
        (int(selected_customer),),
    )
    st.markdown("#### Existing documents")
    if docs_df.empty:
        st.caption("No documents uploaded for this customer yet.")
        return
    for _, row in docs_df.iterrows():
        path = resolve_upload_path(row.get("file_path"))
        label = clean_text(row.get("original_name")) or "(document)"
        doc_type = clean_text(row.get("doc_type")) or "Document"
        uploaded_at = pd.to_datetime(row.get("uploaded_at"), errors="coerce")
        suffix = f" ({uploaded_at.strftime('%d-%m-%Y')})" if pd.notna(uploaded_at) else ""
        if path and path.exists():
            st.download_button(
                f"{doc_type}: {label}{suffix}",
                data=path.read_bytes(),
                file_name=path.name,
                key=f"{key_prefix}_download_{int(row['document_id'])}",
            )
        else:
            st.caption(f"{doc_type}: {label}{suffix} (file missing)")


def render_operations_document_uploader(
    conn,
    *,
    key_prefix: str,
) -> None:
    st.markdown("### Operations document uploads")
    st.markdown(
        """
        <style>
        @media (max-width: 768px) {
          [data-testid="stHorizontalBlock"] {
            flex-wrap: wrap;
          }
          [data-testid="stHorizontalBlock"] > div {
            flex: 1 1 100% !important;
            min-width: 100% !important;
          }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    customer_options, customer_labels, _, _ = fetch_customer_choices(conn, only_complete=False)
    customer_choices = [cid for cid in customer_options if cid is not None]
    if not customer_choices:
        st.info("No customers available for document uploads yet.")
        return

    selected_customer = st.selectbox(
        "Customer",
        customer_choices,
        format_func=lambda cid: customer_labels.get(cid, f"Customer #{cid}"),
        key=f"{key_prefix}_customer",
    )
    customer_seed = df_query(
        conn,
        """
        SELECT name, company_name, phone, address, delivery_address, sales_person
        FROM customers
        WHERE customer_id=?
        """,
        (int(selected_customer),),
    )
    customer_record = customer_seed.iloc[0].to_dict() if not customer_seed.empty else {}

    with st.expander("Operations uploads", expanded=False):
        upload_cols = st.columns(2)
        with upload_cols[0]:
            st.markdown("**Delivery order (DO)**")
            with st.form(key=f"{key_prefix}_do_form", clear_on_submit=False):
                do_file = st.file_uploader(
                    "Delivery order upload",
                    type=None,
                    accept_multiple_files=False,
                    key=f"{key_prefix}_do_file",
                    help="Upload the delivery order PDF or image.",
                )
                _apply_ocr_autofill(
                    upload=do_file,
                    ocr_key_prefix=f"{key_prefix}_do_file",
                    doc_type="Delivery order",
                    details_key_prefix=f"{key_prefix}_do_details",
                )
                do_details = _render_doc_detail_inputs(
                    "Delivery order",
                    key_prefix=f"{key_prefix}_do_details",
                    defaults=customer_record,
                )
                submit_do = st.form_submit_button(
                    "Save delivery order",
                )
            if _guard_double_submit(f"{key_prefix}_do_save", submit_do):
                if do_file is None:
                    st.error("Select a delivery order document to upload.")
                else:
                    saved = _save_customer_document_upload(
                        conn,
                        customer_id=int(selected_customer),
                        customer_record=customer_record,
                        doc_type="Delivery order",
                        upload_file=do_file,
                        details=do_details,
                    )
                    if saved:
                        st.success("Delivery order uploaded.")

        with upload_cols[1]:
            st.markdown("**Work done**")
            with st.form(key=f"{key_prefix}_work_done_form", clear_on_submit=False):
                work_done_file = st.file_uploader(
                    "Work done upload",
                    type=None,
                    accept_multiple_files=False,
                    key=f"{key_prefix}_work_done_file",
                    help="Upload completed work slips or PDFs.",
                )
                _apply_ocr_autofill(
                    upload=work_done_file,
                    ocr_key_prefix=f"{key_prefix}_work_done_file",
                    doc_type="Work done",
                    details_key_prefix=f"{key_prefix}_work_done_details",
                )
                work_done_details = _render_doc_detail_inputs(
                    "Work done",
                    key_prefix=f"{key_prefix}_work_done_details",
                    defaults=customer_record,
                )
                submit_work_done = st.form_submit_button(
                    "Save work done",
                )
            if _guard_double_submit(f"{key_prefix}_work_done_save", submit_work_done):
                if work_done_file is None:
                    st.error("Select a work done document to upload.")
                else:
                    saved = _save_customer_document_upload(
                        conn,
                        customer_id=int(selected_customer),
                        customer_record=customer_record,
                        doc_type="Work done",
                        upload_file=work_done_file,
                        details=work_done_details,
                    )
                    if saved:
                        st.success("Work done uploaded.")
                        _clear_operations_upload_state(
                            file_key=f"{key_prefix}_work_done_file",
                            details_key_prefix=f"{key_prefix}_work_done_details",
                            doc_type="Work done",
                        )

        upload_cols = st.columns(2)
        with upload_cols[0]:
            st.markdown("**Service**")
            with st.form(key=f"{key_prefix}_service_form", clear_on_submit=False):
                service_file = st.file_uploader(
                    "Service upload",
                    type=None,
                    accept_multiple_files=False,
                    key=f"{key_prefix}_service_file",
                    help="Upload service documents.",
                )
                _apply_ocr_autofill(
                    upload=service_file,
                    ocr_key_prefix=f"{key_prefix}_service_file",
                    doc_type="Service",
                    details_key_prefix=f"{key_prefix}_service_details",
                )
                service_details = _render_doc_detail_inputs(
                    "Service",
                    key_prefix=f"{key_prefix}_service_details",
                    defaults=customer_record,
                )
                submit_service = st.form_submit_button(
                    "Save service",
                )
            if _guard_double_submit(f"{key_prefix}_service_save", submit_service):
                if service_file is None:
                    st.error("Select a service document to upload.")
                else:
                    saved = _save_customer_document_upload(
                        conn,
                        customer_id=int(selected_customer),
                        customer_record=customer_record,
                        doc_type="Service",
                        upload_file=service_file,
                        details=service_details,
                    )
                    if saved:
                        st.success("Service uploaded.")
                        _clear_operations_upload_state(
                            file_key=f"{key_prefix}_service_file",
                            details_key_prefix=f"{key_prefix}_service_details",
                            doc_type="Service",
                        )

        with upload_cols[1]:
            st.markdown("**Maintenance**")
            with st.form(key=f"{key_prefix}_maintenance_form", clear_on_submit=False):
                maintenance_file = st.file_uploader(
                    "Maintenance upload",
                    type=None,
                    accept_multiple_files=False,
                    key=f"{key_prefix}_maintenance_file",
                    help="Upload maintenance documents.",
                )
                _apply_ocr_autofill(
                    upload=maintenance_file,
                    ocr_key_prefix=f"{key_prefix}_maintenance_file",
                    doc_type="Maintenance",
                    details_key_prefix=f"{key_prefix}_maintenance_details",
                )
                maintenance_details = _render_doc_detail_inputs(
                    "Maintenance",
                    key_prefix=f"{key_prefix}_maintenance_details",
                    defaults=customer_record,
                )
                submit_maintenance = st.form_submit_button(
                    "Save maintenance",
                )
            if _guard_double_submit(f"{key_prefix}_maintenance_save", submit_maintenance):
                if maintenance_file is None:
                    st.error("Select a maintenance document to upload.")
                else:
                    saved = _save_customer_document_upload(
                        conn,
                        customer_id=int(selected_customer),
                        customer_record=customer_record,
                        doc_type="Maintenance",
                        upload_file=maintenance_file,
                        details=maintenance_details,
                    )
                    if saved:
                        st.success("Maintenance uploaded.")
                        _clear_operations_upload_state(
                            file_key=f"{key_prefix}_maintenance_file",
                            details_key_prefix=f"{key_prefix}_maintenance_details",
                            doc_type="Maintenance",
                        )

        st.markdown("**Others**")
        with st.form(key=f"{key_prefix}_other_form", clear_on_submit=False):
            other_file = st.file_uploader(
                "Other document upload",
                type=None,
                accept_multiple_files=False,
                key=f"{key_prefix}_other_file",
                help="Upload any other supporting document.",
            )
            other_details = _render_doc_detail_inputs(
                "Other",
                key_prefix=f"{key_prefix}_other_details",
                defaults=customer_record,
            )
            submit_other = st.form_submit_button(
                "Save other upload",
            )
        if _guard_double_submit(f"{key_prefix}_other_save", submit_other):
            if other_file is None:
                st.error("Select a document to upload.")
            else:
                saved = _save_customer_document_upload(
                    conn,
                    customer_id=int(selected_customer),
                    customer_record=customer_record,
                    doc_type="Other",
                    upload_file=other_file,
                    details=other_details,
                )
                if saved:
                    st.success("Other document uploaded.")
                    _clear_operations_upload_state(
                        file_key=f"{key_prefix}_other_file",
                        details_key_prefix=f"{key_prefix}_other_details",
                        doc_type="Other",
                    )

    docs_df = df_query(
        conn,
        """
        SELECT document_id, doc_type, file_path, original_name, uploaded_at, uploaded_by
        FROM customer_documents
        WHERE customer_id=?
          AND deleted_at IS NULL
        ORDER BY datetime(uploaded_at) DESC, document_id DESC
        """,
        (int(selected_customer),),
    )
    if not docs_df.empty:
        docs_df = docs_df.drop_duplicates(
            subset=["file_path", "doc_type", "original_name"], keep="first"
        )
    docs_df_all = docs_df.copy()
    st.markdown("#### Existing documents")
    if docs_df.empty:
        st.caption("No documents uploaded for this customer yet.")
    else:
        doc_type_filters = [
            ("All", "All documents"),
            ("Delivery order", "Delivery order (DO)"),
            ("Work done", "Work done (WO)"),
            ("Service", "Service"),
            ("Maintenance", "Maintenance"),
            ("Other", "Other"),
        ]
        type_label_map = {key: label for key, label in doc_type_filters}
        with st.sidebar:
            selected_doc_filter = st.radio(
                "Operations documents",
                [key for key, _ in doc_type_filters],
                format_func=lambda key: type_label_map.get(key, key),
                key=f"{key_prefix}_doc_filter",
            )
        scoped_docs = docs_df.copy()
        if selected_doc_filter != "All":
            scoped_docs = scoped_docs[
                scoped_docs["doc_type"].fillna("") == selected_doc_filter
            ]
        scoped_docs = scoped_docs.sort_values(
            by=["uploaded_at", "document_id"], ascending=[False, False]
        )
        st.caption("Showing the latest 20 uploads for the selected category.")
        header_cols = st.columns([0.55, 0.25, 0.2])
        header_cols[0].markdown("**Document**")
        header_cols[1].markdown("**Uploaded**")
        header_cols[2].markdown("**Download**")
        latest_docs = scoped_docs.head(20)
        for _, row in latest_docs.iterrows():
            path = resolve_upload_path(row.get("file_path"))
            label = clean_text(row.get("original_name")) or "(document)"
            doc_type = clean_text(row.get("doc_type")) or "Document"
            uploaded_at = pd.to_datetime(row.get("uploaded_at"), errors="coerce")
            uploaded_label = uploaded_at.strftime("%d-%m-%Y") if pd.notna(uploaded_at) else "—"
            row_cols = st.columns([0.55, 0.25, 0.2])
            row_cols[0].write(f"{doc_type}: {label}")
            row_cols[1].write(uploaded_label)
            if path and path.exists():
                row_cols[2].download_button(
                    "Download",
                    data=path.read_bytes(),
                    file_name=path.name,
                    key=f"{key_prefix}_download_{int(row['document_id'])}",
                )
            else:
                row_cols[2].caption("Missing")

        st.markdown("#### Edit or delete an uploaded document")
        search_query = st.text_input(
            "Search documents (name or type)",
            key=f"{key_prefix}_doc_search",
        )
        edit_scope = docs_df_all.copy()
        if selected_doc_filter != "All":
            edit_scope = edit_scope[
                edit_scope["doc_type"].fillna("") == selected_doc_filter
            ]
        if search_query:
            query_value = search_query.strip().lower()
            edit_scope = edit_scope[
                edit_scope.apply(
                    lambda row: query_value
                    in " ".join(
                        [
                            clean_text(row.get("doc_type")),
                            clean_text(row.get("original_name")),
                            clean_text(row.get("file_path")),
                        ]
                    ).lower(),
                    axis=1,
                )
            ]
        else:
            edit_scope = edit_scope.sort_values(
                by=["uploaded_at", "document_id"], ascending=[False, False]
            ).head(20)
        doc_records = edit_scope.to_dict("records")
        if not doc_records:
            st.caption("No documents match that search.")
            return
        doc_labels = {
            int(row["document_id"]): " • ".join(
                part
                for part in [
                    clean_text(row.get("doc_type")) or "Document",
                    clean_text(row.get("original_name")) or "(file)",
                    pd.to_datetime(row.get("uploaded_at"), errors="coerce").strftime("%d-%m-%Y")
                    if pd.notna(pd.to_datetime(row.get("uploaded_at"), errors="coerce"))
                    else "",
                ]
                if part
            )
            for row in doc_records
        }
        doc_choices = list(doc_labels.keys())
        selected_doc_id = st.selectbox(
            "Select a document",
            doc_choices,
            format_func=lambda rid: doc_labels.get(rid, f"Document #{rid}"),
            key=f"{key_prefix}_doc_edit_select",
        )
        selected_doc = next(
            row for row in doc_records if int(row["document_id"]) == int(selected_doc_id)
        )
        existing_doc_path = resolve_upload_path(selected_doc.get("file_path"))
        if existing_doc_path and existing_doc_path.exists():
            st.download_button(
                "Download current document",
                data=existing_doc_path.read_bytes(),
                file_name=existing_doc_path.name,
                key=f"{key_prefix}_doc_download_{int(selected_doc_id)}",
            )
        actor_id = current_user_id()
        can_edit = actor_id is not None
        delivery_record = None
        delivery_items_df = None
        delivery_doc_type = clean_text(selected_doc.get("doc_type"))
        if delivery_doc_type in {"Delivery order", "Work done"}:
            record_type = "work_done" if delivery_doc_type == "Work done" else "delivery_order"
            delivery_df = df_query(
                conn,
                """
                SELECT do_number, description, remarks, sales_person, status, items_payload, record_type
                FROM delivery_orders
                WHERE file_path=?
                  AND record_type=?
                  AND deleted_at IS NULL
                """,
                (clean_text(selected_doc.get("file_path")), record_type),
            )
            if not delivery_df.empty:
                delivery_record = delivery_df.iloc[0].to_dict()
                delivery_items = parse_delivery_items_payload(delivery_record.get("items_payload"))
                if not delivery_items:
                    delivery_items = _default_simple_items()
                delivery_items_df = pd.DataFrame(delivery_items)
                for col in ["description", "quantity", "unit_price", "total"]:
                    if col not in delivery_items_df.columns:
                        delivery_items_df[col] = 0.0 if col != "description" else ""
                delivery_items_df["total"] = delivery_items_df.apply(
                    lambda row: max(
                        _coerce_float(row.get("quantity"), 0.0)
                        * _coerce_float(row.get("unit_price"), 0.0),
                        0.0,
                    ),
                    axis=1,
                )
            else:
                st.info("No delivery/work done details found for this upload yet.")
        with st.form(f"{key_prefix}_doc_edit_form"):
            doc_type_options = ["Delivery order", "Work done", "Service", "Maintenance", "Other"]
            selected_doc_type = clean_text(selected_doc.get("doc_type")) or "Other"
            if selected_doc_type not in doc_type_options:
                selected_doc_type = "Other"
            doc_type_choice = selected_doc_type
            st.text_input(
                "Document type",
                value=selected_doc_type,
                key=f"{key_prefix}_doc_edit_type",
                disabled=True,
            )
            replace_doc_file = st.file_uploader(
                "Replace document (optional)",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"{key_prefix}_doc_edit_file",
            )
            delivery_description_input = None
            delivery_remarks_input = None
            delivery_status_choice = None
            delivery_sales_person = None
            edited_delivery_items = None
            if doc_type_choice in {"Delivery order", "Work done"} and delivery_record:
                st.markdown("**Edit delivery/work done details**")
                st.text_input(
                    "Document number",
                    value=clean_text(delivery_record.get("do_number")),
                    disabled=True,
                    key=f"{key_prefix}_doc_edit_do_number",
                )
                current_status = normalize_delivery_status(delivery_record.get("status"))
                if current_status not in DELIVERY_STATUS_OPTIONS:
                    current_status = DELIVERY_STATUS_OPTIONS[0]
                delivery_status_choice = st.selectbox(
                    "Status",
                    options=DELIVERY_STATUS_OPTIONS,
                    index=DELIVERY_STATUS_OPTIONS.index(current_status),
                    key=f"{key_prefix}_doc_edit_status",
                    format_func=lambda option: DELIVERY_STATUS_LABELS.get(option, option.title()),
                )
                delivery_sales_person = st.text_input(
                    "Person in charge",
                    value=clean_text(delivery_record.get("sales_person")),
                    key=f"{key_prefix}_doc_edit_sales_person",
                )
                delivery_description_input = st.text_area(
                    "Description",
                    value=clean_text(delivery_record.get("description")),
                    key=f"{key_prefix}_doc_edit_description",
                )
                delivery_remarks_input = st.text_area(
                    "Remarks",
                    value=clean_text(delivery_record.get("remarks")),
                    key=f"{key_prefix}_doc_edit_remarks",
                )
                edited_delivery_items = st.data_editor(
                    delivery_items_df[["description", "quantity", "unit_price"]],
                    num_rows="dynamic",
                    hide_index=True,
                    use_container_width=True,
                    key=f"{key_prefix}_doc_edit_items",
                    column_config={
                        "description": st.column_config.TextColumn("Product"),
                        "quantity": st.column_config.NumberColumn(
                            "Qty", min_value=0.0, step=1.0, format="%d"
                        ),
                        "unit_price": st.column_config.NumberColumn(
                            "Unit price", min_value=0.0, step=100.0, format="%.2f"
                        ),
                    },
                )
            save_doc_changes = st.form_submit_button(
                "Save document changes",
                type="primary",
                disabled=not can_edit,
            )

        if save_doc_changes:
            if not can_edit:
                st.error("Only staff members can edit this document.")
                return
            stored_path = clean_text(selected_doc.get("file_path"))
            original_name = clean_text(selected_doc.get("original_name"))
            if replace_doc_file is not None:
                doc_dir_map = {
                    "Delivery order": DELIVERY_ORDER_DIR,
                    "Work done": DELIVERY_ORDER_DIR,
                    "Service": SERVICE_DOCS_DIR,
                    "Maintenance": MAINTENANCE_DOCS_DIR,
                    "Other": OPERATIONS_OTHER_DIR,
                }
                target_dir = doc_dir_map.get(doc_type_choice, CUSTOMER_DOCS_DIR)
                target_dir.mkdir(parents=True, exist_ok=True)
                doc_type_slug = (
                    _sanitize_path_component(doc_type_choice.lower().replace(" ", "_"))
                    or "document"
                )
                safe_original = Path(replace_doc_file.name or "document.pdf").name
                filename = f"{doc_type_slug}_{int(selected_customer)}_{safe_original}"
                saved_path = save_uploaded_file(
                    replace_doc_file,
                    target_dir,
                    filename=filename,
                    allowed_extensions={".pdf", ".png", ".jpg", ".jpeg", ".webp"},
                    default_extension=".pdf",
                )
                if not saved_path:
                    st.error("Unable to save the replacement file.")
                    return
                try:
                    stored_path = str(saved_path.relative_to(BASE_DIR))
                except ValueError:
                    stored_path = str(saved_path)
                original_name = safe_original
            conn.execute(
                """
                UPDATE customer_documents
                SET doc_type=?,
                    file_path=?,
                    original_name=?,
                    updated_at=datetime('now'),
                    updated_by=?
                WHERE document_id=?
                """,
                (
                    doc_type_choice,
                    stored_path,
                    original_name,
                    actor_id,
                    int(selected_doc_id),
                ),
            )
            conn.commit()
            if doc_type_choice in {"Delivery order", "Work done"} and delivery_record:
                items_records = (
                    edited_delivery_items.to_dict("records")
                    if isinstance(edited_delivery_items, pd.DataFrame)
                    else []
                )
                cleaned_items, total_amount = normalize_simple_items(items_records)
                items_payload = json.dumps(cleaned_items, ensure_ascii=False) if cleaned_items else None
                record_type = clean_text(delivery_record.get("record_type")) or (
                    "work_done" if doc_type_choice == "Work done" else "delivery_order"
                )
                conn.execute(
                    """
                    UPDATE delivery_orders
                    SET description=?,
                        remarks=?,
                        sales_person=?,
                        status=?,
                        items_payload=?,
                        total_amount=?,
                        file_path=?,
                        updated_at=datetime('now')
                    WHERE do_number=?
                      AND deleted_at IS NULL
                    """,
                    (
                        clean_text(delivery_description_input),
                        clean_text(delivery_remarks_input),
                        clean_text(delivery_sales_person),
                        normalize_delivery_status(delivery_status_choice),
                        items_payload,
                        total_amount if cleaned_items else None,
                        stored_path,
                        clean_text(delivery_record.get("do_number")),
                    ),
                )
                conn.commit()
                log_activity(
                    conn,
                    event_type="work_done_updated"
                    if record_type == "work_done"
                    else "delivery_order_updated",
                    description=f"{'Work done' if record_type == 'work_done' else 'Delivery order'} "
                    f"{clean_text(delivery_record.get('do_number'))} updated",
                    entity_type="delivery_order",
                    entity_id=clean_text(delivery_record.get("do_number")),
                    user_id=actor_id,
                )
            log_activity(
                conn,
                event_type="customer_document_updated",
                description=f"Document #{int(selected_doc_id)} updated",
                entity_type="customer_document",
                entity_id=int(selected_doc_id),
                user_id=actor_id,
            )
            st.success("Document updated.")
            _safe_rerun()

        st.markdown("#### Delete document")
        confirm_delete = st.checkbox(
            "I understand this document will be deleted.",
            key=f"{key_prefix}_doc_delete_confirm",
        )
        if st.button(
            "Delete document",
            type="secondary",
            disabled=not (confirm_delete and can_edit),
            key=f"{key_prefix}_doc_delete_button",
        ):
            if not can_edit:
                st.error("Only staff members can delete this document.")
                return
            conn.execute(
                """
                UPDATE customer_documents
                SET deleted_at=datetime('now'),
                    deleted_by=?
                WHERE document_id=?
                  AND deleted_at IS NULL
                """,
                (actor_id, int(selected_doc_id)),
            )
            conn.commit()
            log_activity(
                conn,
                event_type="customer_document_deleted",
                description=f"Document #{int(selected_doc_id)} deleted",
                entity_type="customer_document",
                entity_id=int(selected_doc_id),
                user_id=actor_id,
            )
            st.warning("Document deleted.")
            _safe_rerun()

    other_df = df_query(
        conn,
        """
        SELECT document_id, description, items_payload, file_path, original_name,
               uploaded_at, uploaded_by, updated_at, updated_by
        FROM operations_other_documents
        WHERE customer_id=?
          AND deleted_at IS NULL
        ORDER BY datetime(uploaded_at) DESC, document_id DESC
        """,
        (int(selected_customer),),
    )
    if not other_df.empty:
        st.markdown("#### Other purchase history")
        for _, row in other_df.iterrows():
            label = clean_text(row.get("description")) or "Other purchase"
            uploaded_at = pd.to_datetime(row.get("uploaded_at"), errors="coerce")
            updated_at = pd.to_datetime(row.get("updated_at"), errors="coerce")
            date_label = uploaded_at.strftime("%d-%m-%Y") if pd.notna(uploaded_at) else ""
            header = f"{label} {f'({date_label})' if date_label else ''}".strip()
            with st.expander(header, expanded=False):
                if pd.notna(updated_at) and updated_at != uploaded_at:
                    st.caption(f"Last updated: {updated_at.strftime('%d-%m-%Y')}")
                items_payload = clean_text(row.get("items_payload"))
                if items_payload:
                    try:
                        items_rows = json.loads(items_payload)
                    except (TypeError, ValueError):
                        items_rows = []
                    if items_rows:
                        st.dataframe(pd.DataFrame(items_rows), use_container_width=True, hide_index=True)
                path = resolve_upload_path(row.get("file_path"))
                if path and path.exists():
                    st.download_button(
                        "Download document",
                        data=path.read_bytes(),
                        file_name=path.name,
                        key=f"{key_prefix}_other_download_{int(row['document_id'])}",
                    )
        st.markdown("#### Edit or delete other uploads")
        other_records = other_df.to_dict("records")
        other_labels = {
            int(row["document_id"]): " • ".join(
                part
                for part in [
                    clean_text(row.get("description")) or f"Other #{int(row['document_id'])}",
                    pd.to_datetime(row.get("uploaded_at"), errors="coerce").strftime("%d-%m-%Y")
                    if pd.notna(pd.to_datetime(row.get("uploaded_at"), errors="coerce"))
                    else "",
                ]
                if part
            )
            for row in other_records
        }
        other_choices = list(other_labels.keys())
        selected_other_id = st.selectbox(
            "Select an other upload",
            other_choices,
            format_func=lambda rid: other_labels.get(rid, f"Other #{rid}"),
            key=f"{key_prefix}_other_edit_select_inline",
        )
        selected_other = next(
            row for row in other_records if int(row["document_id"]) == int(selected_other_id)
        )
        existing_other_path = resolve_upload_path(selected_other.get("file_path"))
        if existing_other_path and existing_other_path.exists():
            st.download_button(
                "Download current other document",
                data=existing_other_path.read_bytes(),
                file_name=existing_other_path.name,
                key=f"{key_prefix}_other_download_edit_{int(selected_other_id)}",
            )
        existing_items_payload = clean_text(selected_other.get("items_payload"))
        try:
            existing_items = json.loads(existing_items_payload) if existing_items_payload else []
        except (TypeError, ValueError):
            existing_items = []
        if not isinstance(existing_items, list):
            existing_items = []
        if not existing_items:
            existing_items = _default_simple_items()
        items_df = pd.DataFrame(existing_items)
        for col in ["description", "quantity", "unit_price", "total"]:
            if col not in items_df.columns:
                items_df[col] = 0.0 if col != "description" else ""
        items_df["total"] = items_df.apply(
            lambda row: max(
                _coerce_float(row.get("quantity"), 0.0)
                * _coerce_float(row.get("unit_price"), 0.0),
                0.0,
            ),
            axis=1,
        )
        actor_id = current_user_id()
        can_edit = actor_id is not None
        with st.form(f"{key_prefix}_other_edit_form_inline"):
            description_input = st.text_area(
                "Description",
                value=clean_text(selected_other.get("description")) or "",
                key=f"{key_prefix}_other_edit_desc_inline",
            )
            edited_items = st.data_editor(
                items_df[["description", "quantity", "unit_price"]],
                num_rows="dynamic",
                hide_index=True,
                use_container_width=True,
                key=f"{key_prefix}_other_edit_items_inline",
                column_config={
                    "description": st.column_config.TextColumn("Item"),
                    "quantity": st.column_config.NumberColumn(
                        "Qty", min_value=0.0, step=1.0, format="%d"
                    ),
                    "unit_price": st.column_config.NumberColumn(
                        "Unit price", min_value=0.0, step=100.0, format="%.2f"
                    ),
                },
            )
            replace_file = st.file_uploader(
                "Replace other document (optional)",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"{key_prefix}_other_edit_file_inline",
            )
            save_changes = st.form_submit_button(
                "Save other upload",
                type="primary",
                disabled=not can_edit,
            )

        if save_changes:
            if not can_edit:
                st.error("Only staff members can edit this record.")
                return
            items_records = (
                edited_items.to_dict("records")
                if isinstance(edited_items, pd.DataFrame)
                else []
            )
            cleaned_items, _ = normalize_simple_items(items_records)
            items_payload = json.dumps(cleaned_items, ensure_ascii=False) if cleaned_items else None
            stored_path = clean_text(selected_other.get("file_path"))
            original_name = clean_text(selected_other.get("original_name"))
            if replace_file is not None:
                target_dir = OPERATIONS_OTHER_DIR
                target_dir.mkdir(parents=True, exist_ok=True)
                safe_original = Path(replace_file.name or "other_document.pdf").name
                filename = f"other_{int(selected_customer)}_{safe_original}"
                saved_path = save_uploaded_file(
                    replace_file,
                    target_dir,
                    filename=filename,
                    allowed_extensions={".pdf", ".png", ".jpg", ".jpeg", ".webp"},
                    default_extension=".pdf",
                )
                if saved_path:
                    try:
                        stored_path = str(saved_path.relative_to(BASE_DIR))
                    except ValueError:
                        stored_path = str(saved_path)
                    original_name = safe_original
            conn.execute(
                """
                UPDATE operations_other_documents
                SET description=?,
                    items_payload=?,
                    file_path=?,
                    original_name=?,
                    updated_at=datetime('now'),
                    updated_by=?
                WHERE document_id=?
                  AND deleted_at IS NULL
                """,
                (
                    clean_text(description_input),
                    items_payload,
                    stored_path,
                    original_name,
                    actor_id,
                    int(selected_other_id),
                ),
            )
            conn.commit()
            log_activity(
                conn,
                event_type="operations_other_updated",
                description=f"Other record #{int(selected_other_id)} updated",
                entity_type="operations_other",
                entity_id=int(selected_other_id),
                user_id=actor_id,
            )
            st.success("Other upload updated.")
            _safe_rerun()

        st.markdown("#### Delete other upload")
        confirm_delete = st.checkbox(
            "I understand this record will be removed from active views.",
            key=f"{key_prefix}_other_delete_confirm_inline",
        )
        if st.button(
            "Delete other upload",
            type="secondary",
            disabled=not (confirm_delete and can_edit),
            key=f"{key_prefix}_other_delete_button_inline",
        ):
            if not can_edit:
                st.error("Only staff members can delete this record.")
                return
            conn.execute(
                """
                UPDATE operations_other_documents
                SET deleted_at=datetime('now'),
                    deleted_by=?
                WHERE document_id=?
                  AND deleted_at IS NULL
                """,
                (actor_id, int(selected_other_id)),
            )
            conn.commit()
            log_activity(
                conn,
                event_type="operations_other_deleted",
                description=f"Other record #{int(selected_other_id)} deleted",
                entity_type="operations_other",
                entity_id=int(selected_other_id),
                user_id=actor_id,
            )
            st.warning("Other upload deleted.")
            _safe_rerun()


def _render_operations_other_manager(conn, *, key_prefix: str) -> None:
    st.markdown("### Other operations records")
    customer_options, customer_labels, _, _ = fetch_customer_choices(conn, only_complete=False)
    if not customer_options:
        st.info("No customers available for other operation records yet.")
        return

    filter_cols = st.columns((1.1, 1.4))
    with filter_cols[0]:
        customer_filter = st.selectbox(
            "Filter by customer",
            options=[None] + [cid for cid in customer_options if cid is not None],
            format_func=lambda cid: customer_labels.get(cid, "(any)"),
            key=f"{key_prefix}_other_customer_filter",
        )
    with filter_cols[1]:
        search_text = st.text_input(
            "Search description or file name",
            key=f"{key_prefix}_other_search",
        )

    other_df = df_query(
        conn,
        """
        SELECT o.document_id,
               o.customer_id,
               o.description,
               o.items_payload,
               o.file_path,
               o.original_name,
               o.uploaded_at,
               o.updated_at,
               o.uploaded_by,
               o.updated_by,
               COALESCE(c.name, c.company_name, '(customer)') AS customer,
               COALESCE(u.username, '(user)') AS uploaded_by_name,
               COALESCE(uu.username, '(user)') AS updated_by_name
        FROM operations_other_documents o
        LEFT JOIN customers c ON c.customer_id = o.customer_id
        LEFT JOIN users u ON u.user_id = o.uploaded_by
        LEFT JOIN users uu ON uu.user_id = o.updated_by
        WHERE o.deleted_at IS NULL
        ORDER BY datetime(o.uploaded_at) DESC, o.document_id DESC
        """,
    )
    if customer_filter:
        other_df = other_df[other_df["customer_id"] == int(customer_filter)]
    if search_text:
        needle = search_text.lower()
        other_df = other_df[
            other_df.apply(
                lambda row: any(
                    needle in str(row.get(col, "")).lower()
                    for col in ["description", "original_name", "customer"]
                ),
                axis=1,
            )
        ]

    if other_df.empty:
        st.caption("No other operations records found for the selected filters.")
        return

    def _summarize_items(payload: object) -> str:
        if not payload:
            return ""
        try:
            items = json.loads(payload) if isinstance(payload, str) else payload
        except (TypeError, ValueError):
            return ""
        if not isinstance(items, list):
            return ""
        summary_bits = []
        for item in items:
            if not isinstance(item, dict):
                continue
            desc = clean_text(item.get("description")) or clean_text(item.get("name"))
            qty = _coerce_float(item.get("quantity"), 0.0)
            if desc:
                qty_label = f" x{int(qty)}" if qty else ""
                summary_bits.append(f"{desc}{qty_label}")
        return ", ".join(summary_bits)

    display_df = other_df.copy()
    display_df["items_summary"] = display_df["items_payload"].apply(_summarize_items)
    display_df = fmt_dates(display_df, ["uploaded_at", "updated_at"])
    display_df["Document"] = display_df["file_path"].apply(
        lambda fp: "📎" if clean_text(fp) else ""
    )
    display_df = display_df.rename(
        columns={
            "description": "Description",
            "customer": "Customer",
            "items_summary": "Items",
            "uploaded_at": "Uploaded",
            "updated_at": "Updated",
            "uploaded_by_name": "Uploaded by",
            "updated_by_name": "Updated by",
        }
    )
    st.dataframe(
        display_df[
            [
                "Description",
                "Customer",
                "Items",
                "Uploaded",
                "Updated",
                "Document",
                "Uploaded by",
                "Updated by",
            ]
        ],
        use_container_width=True,
        hide_index=True,
    )

    other_records = other_df.to_dict("records")
    record_labels = {
        int(row["document_id"]): " • ".join(
            part
            for part in [
                clean_text(row.get("description")) or f"Other #{int(row['document_id'])}",
                clean_text(row.get("customer")),
            ]
            if part
        )
        for row in other_records
    }
    record_choices = list(record_labels.keys())

    st.markdown("#### Edit other record")
    selected_id = st.selectbox(
        "Select a record",
        record_choices,
        format_func=lambda rid: record_labels.get(rid, f"Other #{rid}"),
        key=f"{key_prefix}_other_edit_select",
    )
    selected_record = next(
        row for row in other_records if int(row["document_id"]) == int(selected_id)
    )
    existing_file_path = resolve_upload_path(selected_record.get("file_path"))
    if existing_file_path and existing_file_path.exists():
        st.download_button(
            "Download current document",
            data=existing_file_path.read_bytes(),
            file_name=existing_file_path.name,
            key=f"{key_prefix}_other_download_{int(selected_id)}",
        )
    existing_items_payload = clean_text(selected_record.get("items_payload"))
    try:
        existing_items = json.loads(existing_items_payload) if existing_items_payload else []
    except (TypeError, ValueError):
        existing_items = []
    if not isinstance(existing_items, list):
        existing_items = []
    if not existing_items:
        existing_items = _default_simple_items()

    items_df = pd.DataFrame(existing_items)
    for col in ["description", "quantity", "unit_price", "total"]:
        if col not in items_df.columns:
            items_df[col] = 0.0 if col != "description" else ""
    items_df["total"] = items_df.apply(
        lambda row: max(
            _coerce_float(row.get("quantity"), 0.0)
            * _coerce_float(row.get("unit_price"), 0.0),
            0.0,
        ),
        axis=1,
    )
    actor_id = current_user_id()
    can_edit = actor_id is not None
    with st.form(f"{key_prefix}_other_edit_form"):
        description_input = st.text_area(
            "Description",
            value=clean_text(selected_record.get("description")) or "",
            key=f"{key_prefix}_other_edit_desc",
        )
        edited_items = st.data_editor(
            items_df[["description", "quantity", "unit_price"]],
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            key=f"{key_prefix}_other_edit_items",
            column_config={
                "description": st.column_config.TextColumn("Item"),
                "quantity": st.column_config.NumberColumn("Qty", min_value=0.0, step=1.0, format="%d"),
                "unit_price": st.column_config.NumberColumn(
                    "Unit price", min_value=0.0, step=100.0, format="%.2f"
                ),
            },
        )
        replace_file = st.file_uploader(
            "Replace document (optional)",
            type=["pdf", "png", "jpg", "jpeg", "webp"],
            key=f"{key_prefix}_other_edit_file",
        )
        save_changes = st.form_submit_button(
            "Save other record",
            type="primary",
            disabled=not can_edit,
        )

    if save_changes:
        if not can_edit:
            st.error("Only staff members can edit this record.")
            return
        items_records = (
            edited_items.to_dict("records") if isinstance(edited_items, pd.DataFrame) else []
        )
        cleaned_items, _ = normalize_simple_items(items_records)
        items_payload = json.dumps(cleaned_items, ensure_ascii=False) if cleaned_items else None
        stored_path = clean_text(selected_record.get("file_path"))
        original_name = clean_text(selected_record.get("original_name"))
        if replace_file is not None:
            target_dir = OPERATIONS_OTHER_DIR
            target_dir.mkdir(parents=True, exist_ok=True)
            safe_original = Path(replace_file.name or "other_document.pdf").name
            filename = f"other_{int(selected_record['customer_id'])}_{safe_original}"
            saved_path = save_uploaded_file(
                replace_file,
                target_dir,
                filename=filename,
                allowed_extensions={".pdf", ".png", ".jpg", ".jpeg", ".webp"},
                default_extension=".pdf",
            )
            if saved_path:
                try:
                    stored_path = str(saved_path.relative_to(BASE_DIR))
                except ValueError:
                    stored_path = str(saved_path)
                original_name = safe_original
        conn.execute(
            """
            UPDATE operations_other_documents
            SET description=?,
                items_payload=?,
                file_path=?,
                original_name=?,
                updated_at=datetime('now'),
                updated_by=?
            WHERE document_id=?
              AND deleted_at IS NULL
            """,
            (
                clean_text(description_input),
                items_payload,
                stored_path,
                original_name,
                actor_id,
                int(selected_id),
            ),
        )
        conn.commit()
        log_activity(
            conn,
            event_type="operations_other_updated",
            description=f"Other record #{int(selected_id)} updated",
            entity_type="operations_other",
            entity_id=int(selected_id),
            user_id=actor_id,
        )
        st.success("Other record updated.")
        _safe_rerun()

    st.markdown("#### Delete other record")
    can_delete = can_edit
    confirm_delete = st.checkbox(
        "I understand this record will be removed from active views.",
        key=f"{key_prefix}_other_delete_confirm",
    )
    if st.button(
        "Delete other record",
        type="secondary",
        disabled=not (confirm_delete and can_delete),
        key=f"{key_prefix}_other_delete_button",
    ):
        if not can_delete:
            st.error("Only staff members can delete this record.")
            return
        conn.execute(
            """
            UPDATE operations_other_documents
            SET deleted_at=datetime('now'),
                deleted_by=?
            WHERE document_id=?
              AND deleted_at IS NULL
            """,
            (actor_id, int(selected_id)),
        )
        conn.commit()
        log_activity(
            conn,
            event_type="operations_other_deleted",
            description=f"Other record #{int(selected_id)} deleted",
            entity_type="operations_other",
            entity_id=int(selected_id),
            user_id=actor_id,
        )
        st.warning("Other record deleted.")
        _safe_rerun()


def operations_page(conn):
    st.subheader("🛠️ Operations")
    st.caption(
        "Review delivery orders, work done, service, maintenance, and other operational records in one place."
    )
    render_operations_document_uploader(conn, key_prefix="operations_page")

    st.markdown("---")
    tabs = st.tabs(
        [
            "Delivery orders",
            "Work done",
            "Service",
            "Maintenance",
            "Other uploads",
        ]
    )
    with tabs[0]:
        st.markdown("### Delivery orders")
        delivery_orders_page(
            conn,
            show_heading=False,
            record_type_label="Delivery order",
            record_type_key="delivery_order",
        )
    with tabs[1]:
        st.markdown("### Work done")
        delivery_orders_page(
            conn,
            show_heading=False,
            record_type_label="Work done",
            record_type_key="work_done",
        )
    with tabs[2]:
        st.markdown("### Service records")
        _render_service_section(conn, show_heading=False)
    with tabs[3]:
        st.markdown("### Maintenance records")
        _render_maintenance_section(conn, show_heading=False)
    with tabs[4]:
        _render_operations_other_manager(conn, key_prefix="operations_page")

    if current_user_is_admin():
        with st.expander("Admin activity log", expanded=False):
            activity = fetch_activity_feed(conn, limit=50)
            if not activity:
                st.caption("No recent activity yet.")
            else:
                activity_df = pd.DataFrame(activity)
                activity_df = fmt_dates(activity_df, ["timestamp"])
                activity_df = activity_df.rename(
                    columns={
                        "timestamp": "When",
                        "actor": "Staff",
                        "event_type": "Event",
                        "message": "Details",
                    }
                )
                st.dataframe(
                    activity_df[["When", "Staff", "Event", "Details"]],
                    use_container_width=True,
                    hide_index=True,
                )


def customers_page(conn):
    st.subheader("👥 Customers")
    feedback = st.session_state.pop("new_customer_feedback", None)
    if feedback:
        level, message = feedback
        if level == "success":
            st.success(message)
        elif level == "info":
            st.info(message)
        elif level == "warning":
            st.warning(message)
        else:
            st.write(message)

    with st.expander("Add new customer"):
        products_state = st.session_state.get(
            "new_customer_products_rows",
            _default_new_customer_products(),
        )
        st.session_state.setdefault(
            "new_customer_products_rows", products_state
        )
        user_seed = get_current_user()
        salesperson_seed = ""
        if user_seed:
            bits = [clean_text(user_seed.get("username")), clean_text(user_seed.get("phone"))]
            salesperson_seed = " ".join(bit for bit in bits if bit)
        st.session_state.setdefault("new_customer_sales_person", salesperson_seed)
        with st.form("new_customer"):
            name = st.text_input("Customer name *", key="new_customer_name")
            company = st.text_input(
                "Company name",
                key="new_customer_company",
                help="Optional organisation or business associated with this customer.",
            )
            phone = st.text_input("Phone", key="new_customer_phone")
            address = st.text_area(
                "Billing address",
                key="new_customer_address",
                help="Primary mailing or billing address for this customer.",
            )
            delivery_address = st.text_area(
                "Delivery address",
                key="new_customer_delivery_address",
                help="Where goods should be delivered. Leave blank if same as billing.",
            )
            purchase_default = st.session_state.get("new_customer_purchase_date")
            if isinstance(purchase_default, datetime):
                purchase_default = purchase_default.date()
            if not isinstance(purchase_default, date):
                purchase_default = None
            purchase_date_enabled = st.checkbox(
                "Set purchase date",
                value=bool(purchase_default),
                key="new_customer_purchase_date_enabled",
                help="Enable this if the customer has already purchased.",
            )
            purchase_date = None
            if purchase_date_enabled:
                purchase_date = st.date_input(
                    "Purchase date",
                    value=purchase_default or datetime.now().date(),
                    key="new_customer_purchase_date",
                    help="Used as the warranty issue date when creating new warranty records.",
                )
            else:
                st.session_state["new_customer_purchase_date"] = None
            remarks = st.text_area(
                "Remarks",
                key="new_customer_remarks",
                help="Internal notes or special instructions for this customer.",
            )
            amount_spent_input = st.number_input(
                "Amount spent",
                min_value=0.0,
                step=100.0,
                format="%.2f",
                key="new_customer_amount_spent",
                help="Record how much the customer has spent so far.",
            )
            st.markdown("#### Products / services purchased")
            st.caption(
                "Use the **Add row** option below to record each product or service purchased."
            )
            products_df = pd.DataFrame(products_state)
            required_columns = ["name", "model", "serial", "quantity", "unit_price", "total"]
            for column in required_columns:
                if column not in products_df.columns:
                    default_value = 0.0 if column in ["unit_price", "total"] else ""
                    if column == "quantity":
                        default_value = 1
                    products_df[column] = default_value
            products_df = products_df[required_columns]
            products_df["total"] = products_df.apply(
                lambda row: max(
                    _coerce_float(row.get("quantity"), 1.0)
                    * _coerce_float(row.get("unit_price"), 0.0),
                    0.0,
                ),
                axis=1,
            )
            edited_products = st.data_editor(
                products_df.drop(columns=["total"], errors="ignore"),
                key="new_customer_products_table",
                num_rows="dynamic",
                hide_index=True,
                use_container_width=True,
                column_config={
                    "name": st.column_config.TextColumn(
                        "Product / service",
                        help="Name or brief description of the item purchased.",
                    ),
                    "model": st.column_config.TextColumn(
                        "Model",
                        help="Add model or variant details to help identify the product.",
                    ),
                    "serial": st.column_config.TextColumn(
                        "Serial / ID",
                        help="Serial number or unique identifier (optional).",
                    ),
                    "quantity": st.column_config.NumberColumn(
                        "Quantity",
                        min_value=1,
                        step=1,
                        format="%d",
                    ),
                    "unit_price": st.column_config.NumberColumn(
                        "Unit price",
                        min_value=0.0,
                        step=100.0,
                        format="%.2f",
                    ),
                },
            )
            editor_df = edited_products if isinstance(edited_products, pd.DataFrame) else pd.DataFrame(edited_products)
            editor_df["total"] = editor_df.apply(
                lambda row: max(
                    _coerce_float(row.get("quantity"), 1.0)
                    * _coerce_float(row.get("unit_price"), 0.0),
                    0.0,
                ),
                axis=1,
            )
            product_entries = editor_df.to_dict("records")
            st.session_state["new_customer_products_rows"] = product_entries
            with st.expander("Attachments & advanced details", expanded=True):
                do_code = st.text_input(
                    "Delivery order (DO) code (optional)",
                    key="new_customer_do_code",
                    help="Link the customer to an existing delivery order if available.",
                )
                work_done_number = st.text_input(
                    "Work done number (optional)",
                    key="new_customer_work_done_number",
                    help="Reference code used when creating a work done record.",
                )
                service_reference = st.text_input(
                    "Service reference (optional)",
                    key="new_customer_service_reference",
                    help="Code or tag to use when creating the linked service record.",
                )
                maintenance_reference = st.text_input(
                    "Maintenance reference (optional)",
                    key="new_customer_maintenance_reference",
                    help="Code or tag for the maintenance record created with this customer.",
                )
                sales_person_input = st.text_input(
                    "Sales person",
                    value=st.session_state.get("new_customer_sales_person", salesperson_seed),
                    key="new_customer_sales_person",
                    help="Record who handled this sale for quick reference later.",
                )
                customer_pdf = st.file_uploader(
                    "Attach customer document (PDF or image)",
                    type=["pdf", "png", "jpg", "jpeg", "webp", "gif"],
                    key="new_customer_pdf",
                    help="Upload signed agreements, invoices, photos, or other supporting paperwork.",
                )
                st.markdown("---")
                create_cols = st.columns(4)
                if (
                    "new_customer_create_delivery_order" not in st.session_state
                    and do_code
                ):
                    st.session_state["new_customer_create_delivery_order"] = True
                create_delivery_order = create_cols[0].checkbox(
                    "Delivery order",
                    value=bool(st.session_state.get("new_customer_create_delivery_order")) or bool(do_code),
                    key="new_customer_create_delivery_order",
                )
                create_work_done = create_cols[1].checkbox(
                    "Work done",
                    value=bool(st.session_state.get("new_customer_create_work_done")),
                    key="new_customer_create_work_done",
                )
                create_service = create_cols[2].checkbox(
                    "Service",
                    value=bool(st.session_state.get("new_customer_create_service")),
                    key="new_customer_create_service",
                )
                create_maintenance = create_cols[3].checkbox(
                    "Maintenance",
                    value=bool(st.session_state.get("new_customer_create_maintenance")),
                    key="new_customer_create_maintenance",
                )

                if create_delivery_order:
                    st.subheader("Delivery order details")
                    st.caption(
                        "Uses the delivery order code and product list exactly as on the Delivery orders page."
                    )
                    do_status = st.selectbox(
                        "Delivery order status",
                        options=DELIVERY_STATUS_OPTIONS,
                        key="new_customer_do_status",
                        help="Mark as paid to require an accompanying receipt.",
                        format_func=lambda option: DELIVERY_STATUS_LABELS.get(option, option.title()),
                    )
                    st.file_uploader(
                        "Delivery order receipt (required if paid)",
                        type=["pdf", "png", "jpg", "jpeg", "webp"],
                        key="new_customer_do_receipt",
                        help="Upload proof of payment when marking the DO as paid.",
                    )

                if create_work_done:
                    st.subheader("Work done details")
                    st.caption(
                        "Matches the work done form so you can fill in the reference, remarks and PDF attachment."
                    )
                    work_done_cols = st.columns((1, 2, 2))
                    with work_done_cols[0]:
                        st.caption(
                            f"Work done number: {clean_text(work_done_number) or '— set above —'}"
                        )
                    work_done_status = work_done_cols[1].selectbox(
                        "Work done status",
                        options=DELIVERY_STATUS_OPTIONS,
                        key="new_customer_work_done_status",
                        help="Mark as paid to require an accompanying receipt.",
                        format_func=lambda option: DELIVERY_STATUS_LABELS.get(option, option.title()),
                    )
                    work_done_pdf = work_done_cols[2].file_uploader(
                        "Attach work done (PDF or image)",
                        type=["pdf", "png", "jpg", "jpeg", "webp"],
                        key="new_customer_work_done_pdf",
                    )
                    work_done_receipt = st.file_uploader(
                        "Payment receipt (required for paid work done)",
                        type=["pdf", "png", "jpg", "jpeg", "webp"],
                        key="new_customer_work_done_receipt",
                        help="Upload proof of payment if this work done is already paid.",
                    )
                    work_done_notes = st.text_area(
                        "Work done description / remarks",
                        key="new_customer_work_done_notes",
                    )
                else:
                    work_done_status = st.session_state.get("new_customer_work_done_status", "due")
                    work_done_pdf = st.session_state.get("new_customer_work_done_pdf")
                    work_done_receipt = st.session_state.get("new_customer_work_done_receipt")
                    work_done_notes = st.session_state.get("new_customer_work_done_notes")

                if create_service or create_maintenance:
                    st.subheader("After-sales records")

                if create_service:
                    service_cols = st.columns((1, 1, 1))
                    service_date_default = purchase_date or datetime.now().date()
                    service_date_input = service_cols[0].date_input(
                        "Service date",
                        value=service_date_default,
                        key="new_customer_service_date",
                    )
                    service_description = service_cols[1].text_area(
                        "Service description",
                        key="new_customer_service_description",
                        help="Mirror of the service page description field.",
                    )
                    service_status = service_cols[2].selectbox(
                        "Service payment status",
                        options=["pending", "paid"],
                        key="new_customer_service_payment_status",
                        help="Track whether the service has been paid.",
                    )
                    service_receipt = st.file_uploader(
                        "Service receipt (required if paid)",
                        type=["pdf", "png", "jpg", "jpeg", "webp"],
                        key="new_customer_service_receipt",
                        help="Upload payment receipt when marking the service as paid.",
                    )
                else:
                    service_date_input = st.session_state.get("new_customer_service_date")
                    service_description = st.session_state.get("new_customer_service_description")
                    service_status = st.session_state.get("new_customer_service_payment_status", "pending")
                    service_receipt = st.session_state.get("new_customer_service_receipt")

                if create_maintenance:
                    maintenance_cols = st.columns((1, 1, 1))
                    maintenance_date_default = purchase_date or datetime.now().date()
                    maintenance_date_input = maintenance_cols[0].date_input(
                        "Maintenance date",
                        value=maintenance_date_default,
                        key="new_customer_maintenance_date",
                    )
                    maintenance_description = maintenance_cols[1].text_area(
                        "Maintenance description",
                        key="new_customer_maintenance_description",
                        help="Same fields as the maintenance page for easy entry.",
                    )
                    maintenance_status = maintenance_cols[2].selectbox(
                        "Maintenance payment status",
                        options=["pending", "paid"],
                        key="new_customer_maintenance_payment_status",
                        help="Track whether maintenance has been paid.",
                    )
                    maintenance_receipt = st.file_uploader(
                        "Maintenance receipt (required if paid)",
                        type=["pdf", "png", "jpg", "jpeg", "webp"],
                        key="new_customer_maintenance_receipt",
                        help="Upload payment receipt when marking maintenance as paid.",
                    )
                else:
                    maintenance_date_input = st.session_state.get("new_customer_maintenance_date")
                    maintenance_description = st.session_state.get("new_customer_maintenance_description")
                    maintenance_status = st.session_state.get(
                        "new_customer_maintenance_payment_status", "pending"
                    )
                    maintenance_receipt = st.session_state.get("new_customer_maintenance_receipt")
            action_cols = st.columns((1, 1))
            submitted = action_cols[0].form_submit_button(
                "Save new customer", type="primary"
            )
            reset_form = action_cols[1].form_submit_button("Reset form")
            if reset_form:
                _reset_new_customer_form_state()
                st.session_state["new_customer_feedback"] = (
                    "info",
                    "Customer form cleared. You can start again with a blank form.",
                )
                _safe_rerun()
                return
            if _guard_double_submit("new_customer_save", submitted):
                errors: list[str] = []
                if not name.strip():
                    errors.append("Customer name is required before saving.")
                do_serial = clean_text(do_code)
                work_done_serial = clean_text(work_done_number)
                if create_work_done and not work_done_serial:
                    errors.append("Work done number is required when creating a record.")
                if create_service and not clean_text(service_description):
                    errors.append("Add a short service description to create the service record.")
                if create_maintenance and not clean_text(maintenance_description):
                    errors.append(
                        "Add a short maintenance description to create the maintenance record."
                    )
                if errors:
                    for msg in errors:
                        st.error(msg)
                    return
                cur = conn.cursor()
                name_val = clean_text(name)
                company_val = clean_text(company)
                phone_val = clean_text(phone)
                address_val = clean_text(address)
                delivery_address_val = clean_text(delivery_address)
                remarks_val = clean_text(remarks)
                sales_person_value = clean_text(sales_person_input) or salesperson_seed
                cleaned_products, product_labels = normalize_product_entries(product_entries)
                product_label = "\n".join(product_labels) if product_labels else None
                product_items = _products_to_delivery_items(cleaned_products)
                delivery_items_payload = None
                delivery_total = 0.0
                do_status_value = (
                    normalize_delivery_status(st.session_state.get("new_customer_do_status"))
                )
                if product_items:
                    normalized_items, delivery_total = normalize_delivery_items(product_items)
                    if normalized_items:
                        delivery_items_payload = json.dumps(
                            normalized_items, ensure_ascii=False
                        )
                    else:
                        delivery_total = 0.0
                purchase_str = purchase_date.strftime("%Y-%m-%d") if purchase_date else None
                amount_value = parse_amount(amount_spent_input)
                if amount_value == 0.0 and (amount_spent_input is None or amount_spent_input == 0.0):
                    amount_value = None
                created_by = current_user_id()
                cur.execute(
                    "INSERT INTO customers (name, company_name, phone, address, delivery_address, remarks, purchase_date, product_info, delivery_order_code, sales_person, amount_spent, created_by, dup_flag) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)",
                    (
                        name_val,
                        company_val,
                        phone_val,
                        address_val,
                        delivery_address_val,
                        remarks_val,
                        purchase_str,
                        product_label,
                        do_serial,
                        sales_person_value,
                        amount_value,
                        created_by,
                    ),
                )
                cid = cur.lastrowid
                conn.commit()
                if cleaned_products:
                    for prod in cleaned_products:
                        if not prod.get("name"):
                            continue
                        cur.execute(
                            "SELECT product_id FROM products WHERE name=? AND IFNULL(model,'')=IFNULL(?, '') LIMIT 1",
                            (prod.get("name"), prod.get("model")),
                        )
                        row = cur.fetchone()
                        if row:
                            pid = row[0]
                        else:
                            cur.execute(
                                "INSERT INTO products (name, model, serial) VALUES (?, ?, ?)",
                                (
                                    prod.get("name"),
                                    prod.get("model"),
                                    prod.get("serial"),
                                ),
                            )
                            pid = cur.lastrowid
                        issue = purchase_date.strftime("%Y-%m-%d") if purchase_date else None
                        expiry = (
                            (purchase_date + timedelta(days=365)).strftime("%Y-%m-%d")
                            if purchase_date
                            else None
                        )
                        cur.execute(
                            "INSERT INTO warranties (customer_id, product_id, serial, issue_date, expiry_date, status, remarks) VALUES (?, ?, ?, ?, ?, 'active', ?)",
                            (cid, pid, prod.get("serial"), issue, expiry, remarks_val),
                        )
                    conn.commit()
                if do_serial and create_delivery_order:
                    stored_path = None
                    do_receipt_path = None
                    cur = conn.cursor()
                    existing = cur.execute(
                        """
                        SELECT customer_id, file_path, items_payload, total_amount, payment_receipt_path
                        FROM delivery_orders
                        WHERE do_number = ? AND COALESCE(record_type, 'delivery_order') = 'delivery_order'
                          AND deleted_at IS NULL
                        """,
                        (do_serial,),
                    ).fetchone()
                    existing_receipt = clean_text(existing[4]) if existing else None
                    if do_status_value == "paid":
                        do_receipt = st.session_state.get("new_customer_do_receipt")
                        if do_receipt:
                            do_receipt_path = store_payment_receipt(
                                do_receipt,
                                identifier=f"{_sanitize_path_component(do_serial) or 'do'}_receipt",
                                target_dir=DELIVERY_RECEIPT_DIR,
                            )
                        if not do_receipt_path and existing_receipt:
                            do_receipt_path = existing_receipt
                        if not do_receipt_path:
                            st.error("Upload a payment receipt before marking the DO as paid.")
                            return
                    product_summary = (
                        cleaned_products[0].get("name") if cleaned_products else product_label
                    )
                    sales_clean = clean_text(sales_person_input)
                    if existing:
                        existing_customer, existing_path, existing_items, existing_total, existing_receipt = existing
                        if do_status_value == "paid" and not do_receipt_path:
                            do_receipt_path = clean_text(existing_receipt)
                        if existing_customer and int(existing_customer) != int(cid):
                            st.warning(
                                "Delivery order code already linked to another customer. Upload skipped."
                            )
                            if stored_path and stored_path != existing_path:
                                new_path = resolve_upload_path(stored_path)
                                if new_path and new_path.exists():
                                    try:
                                        new_path.unlink()
                                    except Exception:
                                        pass
                        else:
                            final_path = stored_path or existing_path
                            if stored_path and existing_path and stored_path != existing_path:
                                old_path = resolve_upload_path(existing_path)
                                if old_path and old_path.exists():
                                    try:
                                        old_path.unlink()
                                    except Exception:
                                        pass
                            conn.execute(
                                """
                                UPDATE delivery_orders
                                   SET customer_id=?,
                                       description=?,
                                       sales_person=?,
                                       remarks=?,
                                       file_path=?,
                                       items_payload=COALESCE(?, items_payload),
                                       total_amount=COALESCE(?, total_amount),
                                       status=?,
                                       payment_receipt_path=COALESCE(?, payment_receipt_path),
                                       record_type='delivery_order'
                                 WHERE do_number=? AND COALESCE(record_type, 'delivery_order') = 'delivery_order'
                                """,
                                (
                                    cid,
                                    product_summary,
                                    sales_clean,
                                    remarks_val,
                                    final_path,
                                    delivery_items_payload,
                                    delivery_total if delivery_items_payload else existing_total,
                                    do_status_value,
                                    do_receipt_path,
                                    do_serial,
                                ),
                            )
                            conn.commit()
                    else:
                        conn.execute(
                            """
                            INSERT INTO delivery_orders (
                                do_number,
                                customer_id,
                                order_id,
                                description,
                                sales_person,
                                remarks,
                                file_path,
                                items_payload,
                                total_amount,
                                status,
                                payment_receipt_path,
                                record_type
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'delivery_order')
                            """,
                            (
                                do_serial,
                                cid,
                                None,
                                product_summary,
                                sales_clean,
                                remarks_val,
                                stored_path,
                                delivery_items_payload,
                                delivery_total if delivery_items_payload else None,
                                do_status_value,
                                do_receipt_path,
                            ),
                        )
                        conn.commit()
                if create_work_done and work_done_serial:
                    if not product_items:
                        st.warning(
                            "Add at least one product row before creating a work done record."
                        )
                    else:
                        work_done_items, work_done_total = normalize_delivery_items(
                            product_items
                        )
                        if not work_done_items:
                            st.warning(
                                "Work done could not be saved because no valid products were provided."
                            )
                        else:
                            work_done_payload = json.dumps(
                                work_done_items, ensure_ascii=False
                            )
                            work_done_path = None
                            if work_done_pdf is not None:
                                safe_name = _sanitize_path_component(work_done_serial)
                                doc_ext = _upload_extension(work_done_pdf, default=".pdf")
                                work_done_path = save_uploaded_file(
                                    work_done_pdf,
                                    DELIVERY_ORDER_DIR,
                                    filename=f"work_done_{safe_name}{doc_ext}",
                                    allowed_extensions={".pdf", ".png", ".jpg", ".jpeg", ".webp"},
                                    default_extension=".pdf",
                                )
                                if work_done_path:
                                    try:
                                        work_done_path = str(work_done_path.relative_to(BASE_DIR))
                                    except ValueError:
                                        work_done_path = str(work_done_path)
                            work_done_saved = False
                            work_done_status_value = normalize_delivery_status(work_done_status)
                            work_done_receipt_path = None
                            if work_done_status_value == "paid":
                                work_done_receipt_path = store_payment_receipt(
                                    work_done_receipt,
                                    identifier=f"{_sanitize_path_component(work_done_serial) or 'work_done'}_receipt",
                                    target_dir=DELIVERY_RECEIPT_DIR,
                                )
                            existing_work_done = df_query(
                                conn,
                                "SELECT record_type, file_path, payment_receipt_path FROM delivery_orders WHERE do_number = ? AND deleted_at IS NULL",
                                (work_done_serial,),
                            )
                            existing_work_done_receipt = None
                            if not existing_work_done.empty:
                                existing_work_done_receipt = clean_text(
                                    existing_work_done.iloc[0].get("payment_receipt_path")
                                )
                            if work_done_status_value == "paid" and not (
                                work_done_receipt_path or existing_work_done_receipt
                            ):
                                st.error(
                                    "Upload a payment receipt before marking the work done as paid."
                                )
                                return
                            work_done_description = clean_text(work_done_notes) or product_label
                            if existing_work_done.empty:
                                conn.execute(
                                    """
                                    INSERT INTO delivery_orders (
                                        do_number,
                                        customer_id,
                                        description,
                                        sales_person,
                                        remarks,
                                        file_path,
                                        items_payload,
                                        total_amount,
                                        status,
                                        payment_receipt_path,
                                        record_type
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'work_done')
                                    """,
                                    (
                                        work_done_serial,
                                        cid,
                                        work_done_description,
                                        sales_person_value,
                                        clean_text(work_done_notes),
                                        work_done_path,
                                        work_done_payload,
                                        work_done_total,
                                        work_done_status_value,
                                        work_done_receipt_path,
                                    ),
                                )
                                work_done_saved = True
                            else:
                                existing_type = clean_text(
                                    existing_work_done.iloc[0].get("record_type")
                                ) or "delivery_order"
                                existing_work_done_receipt = clean_text(
                                    existing_work_done.iloc[0].get("payment_receipt_path")
                                )
                                if work_done_status_value == "paid" and not work_done_receipt_path:
                                    work_done_receipt_path = existing_work_done_receipt
                                if existing_type != "work_done":
                                    st.error(
                                        "A delivery order already uses this number. Choose a different work done number."
                                    )
                                    if work_done_path:
                                        new_path = resolve_upload_path(work_done_path)
                                        if new_path and new_path.exists():
                                            try:
                                                new_path.unlink()
                                            except Exception:
                                                pass
                                else:
                                    existing_path = clean_text(
                                        existing_work_done.iloc[0].get("file_path")
                                    )
                                    final_path = work_done_path or existing_path
                                    if work_done_path and existing_path and work_done_path != existing_path:
                                        old_path = resolve_upload_path(existing_path)
                                        if old_path and old_path.exists():
                                            try:
                                                old_path.unlink()
                                            except Exception:
                                                pass
                                    conn.execute(
                                        """
                                        UPDATE delivery_orders
                                           SET customer_id=?,
                                               description=?,
                                               sales_person=?,
                                               remarks=?,
                                               file_path=?,
                                               items_payload=?,
                                               total_amount=?,
                                               status=?,
                                               payment_receipt_path=COALESCE(?, payment_receipt_path),
                                               record_type='work_done'
                                         WHERE do_number=? AND COALESCE(record_type, 'delivery_order') = 'work_done'
                                        """,
                                        (
                                            cid,
                                            work_done_description,
                                            sales_person_value,
                                            clean_text(work_done_notes),
                                            final_path,
                                            work_done_payload,
                                            work_done_total,
                                            work_done_status_value,
                                            work_done_receipt_path,
                                            work_done_serial,
                                        ),
                                    )
                                    work_done_saved = True
                            if work_done_saved:
                                formatted_work_done = format_money(work_done_total)
                                if not formatted_work_done and work_done_total is not None:
                                    try:
                                        formatted_work_done = f"{float(work_done_total):,.2f}"
                                    except (TypeError, ValueError):
                                        formatted_work_done = ""
                                log_activity(
                                    conn,
                                    event_type="work_done_created",
                                    description=(
                                        f"Work done {work_done_serial} saved for {name_val or 'customer'}"
                                        f" ({formatted_work_done})"
                                    ),
                                    entity_type="work_done",
                                    entity_id=None,
                                )
                                conn.commit()

                if create_service:
                    service_date_str = to_iso_date(service_date_input) or purchase_str
                    if not service_date_str:
                        service_date_str = datetime.utcnow().strftime("%Y-%m-%d")
                    service_reference = clean_text(service_reference) or do_serial
                    service_status_value = clean_text(service_status) or "pending"
                    if service_status_value not in {"pending", "paid"}:
                        service_status_value = "pending"
                    service_receipt_path = None
                    if service_status_value == "paid":
                        service_receipt_path = store_payment_receipt(
                            service_receipt,
                            identifier=f"{_sanitize_path_component(service_reference) or 'service'}_receipt",
                            target_dir=SERVICE_BILL_DIR,
                        )
                        if not service_receipt_path:
                            st.error(
                                "Upload a payment receipt before marking the service as paid."
                            )
                            return
                    conn.execute(
                        """
                        INSERT INTO services (
                            do_number,
                            customer_id,
                            service_date,
                            service_start_date,
                            service_end_date,
                            description,
                            status,
                            remarks,
                            service_product_info,
                            payment_status,
                            payment_receipt_path,
                            updated_at,
                            created_by
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), ?)
                        """,
                        (
                            service_reference,
                            cid,
                            service_date_str,
                            service_date_str,
                            service_date_str,
                            clean_text(service_description),
                            DEFAULT_SERVICE_STATUS,
                            remarks_val,
                            product_label,
                            service_status_value,
                            service_receipt_path,
                            created_by,
                        ),
                    )
                    log_activity(
                        conn,
                        event_type="service_created",
                        description=f"Service logged for {name_val or 'customer'}",
                        entity_type="service",
                        entity_id=None,
                    )
                    conn.commit()

                if create_maintenance:
                    maintenance_date_str = to_iso_date(maintenance_date_input) or purchase_str
                    if not maintenance_date_str:
                        maintenance_date_str = datetime.utcnow().strftime("%Y-%m-%d")
                    maintenance_reference = clean_text(maintenance_reference) or do_serial
                    maintenance_status_value = clean_text(maintenance_status) or "pending"
                    if maintenance_status_value not in {"pending", "paid"}:
                        maintenance_status_value = "pending"
                    maintenance_receipt_path = None
                    if maintenance_status_value == "paid":
                        maintenance_receipt_path = store_payment_receipt(
                            maintenance_receipt,
                            identifier=f"{_sanitize_path_component(maintenance_reference) or 'maintenance'}_receipt",
                            target_dir=MAINTENANCE_DOCS_DIR,
                        )
                        if not maintenance_receipt_path:
                            st.error(
                                "Upload a payment receipt before marking the maintenance as paid."
                            )
                            return
                    conn.execute(
                        """
                        INSERT INTO maintenance_records (
                            do_number,
                            customer_id,
                            maintenance_date,
                            maintenance_start_date,
                            maintenance_end_date,
                            description,
                            status,
                            remarks,
                            maintenance_product_info,
                            total_amount,
                            payment_status,
                            payment_receipt_path,
                            updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                        """,
                        (
                            maintenance_reference,
                            cid,
                            maintenance_date_str,
                            maintenance_date_str,
                            maintenance_date_str,
                            clean_text(maintenance_description),
                            DEFAULT_SERVICE_STATUS,
                            remarks_val,
                            product_label,
                            None,
                            maintenance_status_value,
                            maintenance_receipt_path,
                        ),
                    )
                    log_activity(
                        conn,
                        event_type="maintenance_created",
                        description=f"Maintenance logged for {name_val or 'customer'}",
                        entity_type="maintenance",
                        entity_id=None,
                    )
                    conn.commit()
                if customer_pdf is not None:
                    stored_path = store_uploaded_document(
                        customer_pdf,
                        CUSTOMER_DOCS_DIR,
                        filename_stem=f"customer_{cid}",
                    )
                    if stored_path:
                        conn.execute(
                            "UPDATE customers SET attachment_path=? WHERE customer_id=?",
                            (stored_path, cid),
                        )
                        conn.commit()
                if phone_val:
                    recalc_customer_duplicate_flag(conn, phone_val)
                    conn.commit()
                display_name = name_val or f"Customer #{int(cid)}"
                product_count = len(
                    [prod for prod in cleaned_products if prod.get("name")]
                )
                details = (
                    f"{display_name} with {product_count} product(s)"
                    if product_count
                    else display_name
                )
                log_activity(
                    conn,
                    event_type="customer_created",
                    description=f"Added {details}",
                    entity_type="customer",
                    entity_id=int(cid),
                )
                _reset_new_customer_form_state()
                st.session_state["new_customer_feedback"] = (
                    "success",
                    f"Customer {name_val or 'record'} saved successfully.",
                )
                _safe_rerun()
                return
    df_raw = render_customer_quick_edit_section(
        conn,
        key_prefix="customers",
        include_leads=True,
        include_leads_in_main=True,
        show_heading=False,
        show_editor=False,
        show_filters=False,
        enable_uploads=False,
    )
    user = st.session_state.user or {}
    is_admin = user.get("role") == "admin"
    st.markdown("### Detailed editor & attachments")
    with st.expander("Open detailed editor", expanded=False):
        df_form = fmt_dates(df_raw.copy(), ["created_at", "purchase_date"])
        if df_form.empty:
            st.info("No customers to edit yet.")
        else:
            records_fmt = df_form.to_dict("records")
            raw_map = {int(row["id"]): row for row in df_raw.to_dict("records") if int_or_none(row.get("id")) is not None}
            option_ids = [int(row["id"]) for row in records_fmt]
            labels = {}
            for row in records_fmt:
                cid = int(row["id"])
                label_name = clean_text(row.get("name")) or "(no name)"
                label_phone = clean_text(row.get("phone")) or "-"
                labels[cid] = f"{label_name} – {label_phone}"
            selected_customer_id = st.selectbox(
                "Select customer",
                option_ids,
                format_func=lambda cid: labels.get(int(cid), str(cid)),
            )
            selected_raw = raw_map[int(selected_customer_id)]
            selected_fmt = next(r for r in records_fmt if int(r["id"]) == int(selected_customer_id))
            attachment_path = selected_raw.get("attachment_path")
            resolved_attachment = resolve_upload_path(attachment_path)
            if resolved_attachment and resolved_attachment.exists():
                st.download_button(
                    "Download current document",
                    data=resolved_attachment.read_bytes(),
                    file_name=resolved_attachment.name,
                    key=f"cust_pdf_dl_{selected_customer_id}",
                )
            else:
                st.caption("No customer document attached yet.")
            is_admin = user.get("role") == "admin"
            uploader_name = clean_text(selected_raw.get("uploaded_by"))
            if is_admin:
                st.caption(f"Uploaded by: {uploader_name or '(unknown)'}")
            with st.form(f"edit_customer_{selected_customer_id}"):
                name_edit = st.text_input("Name", value=clean_text(selected_raw.get("name")) or "")
                company_edit = st.text_input(
                    "Company",
                    value=clean_text(selected_raw.get("company_name")) or "",
                )
                phone_edit = st.text_input("Phone", value=clean_text(selected_raw.get("phone")) or "")
                address_edit = st.text_area(
                    "Billing address",
                    value=clean_text(selected_raw.get("address")) or "",
                )
                delivery_address_edit = st.text_area(
                    "Delivery address",
                    value=clean_text(selected_raw.get("delivery_address")) or "",
                )
                remarks_edit = st.text_area(
                    "Remarks",
                    value=clean_text(selected_raw.get("remarks")) or "",
                )
                purchase_edit = st.text_input(
                    "Purchase date (DD-MM-YYYY)", value=clean_text(selected_fmt.get("purchase_date")) or ""
                )
                product_edit = st.text_input("Product", value=clean_text(selected_raw.get("product_info")) or "")
                do_edit = st.text_input(
                    "Delivery order code", value=clean_text(selected_raw.get("delivery_order_code")) or ""
                )
                sales_person_edit = st.text_input(
                    "Sales person", value=clean_text(selected_raw.get("sales_person")) or ""
                )
                new_pdf = st.file_uploader(
                    "Attach/replace customer document (PDF or image)",
                    type=["pdf", "png", "jpg", "jpeg", "webp", "gif"],
                    key=f"edit_customer_pdf_{selected_customer_id}",
                )
                col1, col2 = st.columns(2)
                save_customer = col1.form_submit_button("Save changes", type="primary")
                delete_customer = col2.form_submit_button("Delete customer", disabled=not is_admin)
            if save_customer:
                old_phone = clean_text(selected_raw.get("phone"))
                new_name = clean_text(name_edit)
                new_company = clean_text(company_edit)
                new_phone = clean_text(phone_edit)
                new_address = clean_text(address_edit)
                new_delivery_address = clean_text(delivery_address_edit)
                new_remarks = clean_text(remarks_edit)
                purchase_str, _ = date_strings_from_input(purchase_edit)
                product_label = clean_text(product_edit)
                new_do = clean_text(do_edit)
                old_do = clean_text(selected_raw.get("delivery_order_code"))
                new_sales_person = clean_text(sales_person_edit)
                new_attachment_path = attachment_path
                if new_pdf is not None:
                    stored_path = store_uploaded_document(
                        new_pdf,
                        CUSTOMER_DOCS_DIR,
                        filename_stem=f"customer_{selected_customer_id}",
                    )
                    if stored_path:
                        new_attachment_path = stored_path
                        if attachment_path:
                            old_path = resolve_upload_path(attachment_path)
                            if old_path and old_path.exists():
                                new_path = resolve_upload_path(stored_path)
                                if not new_path or new_path != old_path:
                                    try:
                                        old_path.unlink()
                                    except Exception:
                                        pass
                conn.execute(
                    "UPDATE customers SET name=?, company_name=?, phone=?, address=?, delivery_address=?, remarks=?, purchase_date=?, product_info=?, delivery_order_code=?, sales_person=?, attachment_path=?, dup_flag=0 WHERE customer_id=?",
                    (
                        new_name,
                        new_company,
                        new_phone,
                        new_address,
                        new_delivery_address,
                        new_remarks,
                        purchase_str,
                        product_label,
                        new_do,
                        new_sales_person,
                        new_attachment_path,
                        int(selected_customer_id),
                    ),
                )
                if new_do:
                    conn.execute(
                        """
                        INSERT INTO delivery_orders (do_number, customer_id, order_id, description, sales_person, remarks, file_path)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(do_number) DO UPDATE SET
                            customer_id=excluded.customer_id,
                            description=excluded.description,
                            sales_person=excluded.sales_person,
                            remarks=excluded.remarks
                        """,
                        (
                            new_do,
                            int(selected_customer_id),
                            None,
                            product_label,
                            new_sales_person,
                            new_remarks,
                            None,
                        ),
                    )
                if old_do and old_do != new_do:
                    conn.execute(
                        "DELETE FROM delivery_orders WHERE do_number=? AND (customer_id IS NULL OR customer_id=?)",
                        (old_do, int(selected_customer_id)),
                    )
                conn.execute(
                    "UPDATE import_history SET customer_name=?, phone=?, address=?, delivery_address=?, product_label=?, do_number=?, original_date=? WHERE customer_id=? AND deleted_at IS NULL",
                    (
                        new_name,
                        new_phone,
                        new_address,
                        new_delivery_address,
                        product_label,
                        new_do,
                        purchase_str,
                        int(selected_customer_id),
                    ),
                )
                conn.commit()
                if old_phone and old_phone != new_phone:
                    recalc_customer_duplicate_flag(conn, old_phone)
                if new_phone:
                    recalc_customer_duplicate_flag(conn, new_phone)
                conn.commit()
                st.success("Customer updated.")
                _safe_rerun()
            if delete_customer:
                if is_admin:
                    delete_customer_record(conn, int(selected_customer_id))
                    st.warning("Customer deleted.")
                    _safe_rerun()
                else:
                    st.error("Only admins can delete customers.")
            st.markdown("#### Follow-ups & reminders")
            with st.form(f"customer_note_add_{selected_customer_id}"):
                new_note_text = st.text_area(
                    "Add a remark", placeholder="e.g. Call back next week with pricing update"
                )
                enable_follow_up = st.checkbox(
                    "Schedule follow-up reminder",
                    value=False,
                    key=f"customer_note_followup_{selected_customer_id}",
                )
                default_date = datetime.now().date()
                reminder_date = st.date_input(
                    "Reminder date",
                    value=default_date,
                    key=f"customer_note_reminder_{selected_customer_id}",
                    disabled=not enable_follow_up,
                )
                add_note = st.form_submit_button("Add remark", type="primary")
            if add_note:
                note_value = clean_text(new_note_text)
                if not note_value:
                    st.error("Remark text is required.")
                else:
                    reminder_value = to_iso_date(reminder_date) if enable_follow_up else None
                    conn.execute(
                        "INSERT INTO customer_notes (customer_id, note, remind_on) VALUES (?, ?, ?)",
                        (int(selected_customer_id), note_value, reminder_value),
                    )
                    conn.commit()
                    st.success("Remark added.")
                    _safe_rerun()

            notes_df = df_query(
                conn,
                """
                SELECT note_id, note, remind_on, is_done, created_at, updated_at
                FROM customer_notes
                WHERE customer_id = ?
                ORDER BY datetime(COALESCE(remind_on, created_at)) DESC, note_id DESC
                """,
                (int(selected_customer_id),),
            )
            if notes_df.empty:
                st.caption("No saved remarks yet.")
            else:
                notes_original = {
                    int(row["note_id"]): row for row in notes_df.to_dict("records") if int_or_none(row.get("note_id")) is not None
                }
                editor_df = notes_df.copy()
                editor_df["remind_on"] = pd.to_datetime(editor_df["remind_on"], errors="coerce")
                editor_df["created_at"] = pd.to_datetime(editor_df["created_at"], errors="coerce")
                editor_df["updated_at"] = pd.to_datetime(editor_df["updated_at"], errors="coerce")
                editor_df["Done"] = editor_df.get("is_done", 0).fillna(0).astype(int).apply(lambda v: bool(v))
                editor_df["Action"] = "Keep"
                column_order = [
                    col
                    for col in [
                        "note_id",
                        "note",
                        "remind_on",
                        "Done",
                        "created_at",
                        "updated_at",
                        "Action",
                    ]
                    if col in editor_df.columns
                ]
                editor_view = editor_df[column_order]
                note_editor_state = st.data_editor(
                    editor_view,
                    hide_index=True,
                    num_rows="fixed",
                    use_container_width=True,
                    column_config={
                        "note_id": st.column_config.Column("ID", disabled=True),
                        "note": st.column_config.TextColumn("Remark"),
                        "remind_on": st.column_config.DateColumn(
                            "Reminder date", format="DD-MM-YYYY", required=False
                        ),
                        "Done": st.column_config.CheckboxColumn("Completed"),
                        "created_at": st.column_config.DatetimeColumn(
                            "Created", format="DD-MM-YYYY HH:mm", disabled=True
                        ),
                        "updated_at": st.column_config.DatetimeColumn(
                            "Updated", format="DD-MM-YYYY HH:mm", disabled=True
                        ),
                        "Action": st.column_config.SelectboxColumn(
                            "Action", options=["Keep", "Delete"], required=True
                        ),
                    },
                    key=f"customer_notes_editor_{selected_customer_id}",
                )
                if st.button(
                    "Apply note updates",
                    key=f"apply_customer_notes_{selected_customer_id}",
                ):
                    note_result = (
                        note_editor_state
                        if isinstance(note_editor_state, pd.DataFrame)
                        else pd.DataFrame(note_editor_state)
                    )
                    if note_result.empty:
                        st.info("No notes to update.")
                    else:
                        changes = False
                        errors: list[str] = []
                        for row in note_result.to_dict("records"):
                            note_id = int_or_none(row.get("note_id"))
                            if note_id is None or note_id not in notes_original:
                                continue
                            action = str(row.get("Action") or "Keep").strip().lower()
                            if action == "delete":
                                conn.execute(
                                    "DELETE FROM customer_notes WHERE note_id = ? AND customer_id = ?",
                                    (note_id, int(selected_customer_id)),
                                )
                                changes = True
                                continue
                            new_note_text = clean_text(row.get("note"))
                            if not new_note_text:
                                errors.append(f"Remark #{note_id} cannot be empty.")
                                continue
                            reminder_iso = to_iso_date(row.get("remind_on"))
                            completed_flag = bool(row.get("Done"))
                            original = notes_original[note_id]
                            original_note = clean_text(original.get("note"))
                            original_reminder = to_iso_date(original.get("remind_on"))
                            original_done = bool(int_or_none(original.get("is_done")) or 0)
                            if (
                                new_note_text == original_note
                                and reminder_iso == original_reminder
                                and completed_flag == original_done
                            ):
                                continue
                            conn.execute(
                                """
                                UPDATE customer_notes
                                SET note = ?, remind_on = ?, is_done = ?, updated_at = datetime('now')
                                WHERE note_id = ? AND customer_id = ?
                                """,
                                (
                                    new_note_text,
                                    reminder_iso,
                                    1 if completed_flag else 0,
                                    note_id,
                                    int(selected_customer_id),
                                ),
                            )
                            changes = True
                        if errors:
                            for err in errors:
                                st.error(err)
                        if changes and not errors:
                            conn.commit()
                            st.success("Notes updated.")
                            _safe_rerun()
                        elif not changes and not errors:
                            st.info("No changes detected.")
                        elif changes:
                            conn.commit()
                            st.warning("Some changes were saved, but please review the errors above.")
    st.markdown("---")
    render_customer_document_uploader(conn, key_prefix="customers_docs")
    scope_clause, scope_params = customer_scope_filter("c")
    st.markdown("**Recently Added Customers**")
    recent_where = f"WHERE {scope_clause}" if scope_clause else ""
    recent_params = scope_params if scope_clause else ()
    recent_df = df_query(
        conn,
        f"""
        SELECT
            c.customer_id AS id,
            c.name,
            c.company_name,
            c.phone,
            c.address,
            c.delivery_address,
            c.remarks,
            c.purchase_date,
            c.product_info,
            c.delivery_order_code,
            c.sales_person,
            c.amount_spent,
            c.created_at,
            COALESCE(u.username, '(unknown)') AS uploaded_by
        FROM customers c
        LEFT JOIN users u ON u.user_id = c.created_by
        {recent_where}
        ORDER BY datetime(c.created_at) DESC LIMIT 200
    """,
        recent_params,
    )
    recent_df = fmt_dates(recent_df, ["created_at", "purchase_date"])
    recent_df = recent_df.rename(
        columns={
            "sales_person": "Sales person",
            "amount_spent": "Amount spent",
            "uploaded_by": "Uploaded by",
        }
    )
    st.dataframe(recent_df.drop(columns=["id"], errors="ignore"))
def warranties_page(conn):
    st.subheader("🛡️ Warranties")
    is_admin = current_user_is_admin()
    sort_dir = st.radio("Sort by expiry date", ["Soonest first", "Latest first"], horizontal=True)
    order = "ASC" if sort_dir == "Soonest first" else "DESC"
    q = st.text_input("Search (customer/product/model/serial)")

    base = dedent(
        """
        SELECT w.warranty_id as id, c.name as customer, p.name as product, p.model, w.serial,
               w.issue_date, w.expiry_date, w.status, w.remarks, w.dup_flag,
               COALESCE(c.sales_person, u.username) AS staff
        FROM warranties w
        LEFT JOIN customers c ON c.customer_id = w.customer_id
        LEFT JOIN products p ON p.product_id = w.product_id
        LEFT JOIN users u ON u.user_id = c.created_by
        WHERE {filters}
        ORDER BY date(w.expiry_date) {order}
        """
    )

    search_filter = "(? = '' OR c.name LIKE '%'||?||'%' OR p.name LIKE '%'||?||'%' OR p.model LIKE '%'||?||'%' OR w.serial LIKE '%'||?||'%')"
    status_filter = "(w.status IS NULL OR w.status <> 'deleted')"
    product_filter = "(p.name IS NOT NULL AND TRIM(p.name) != '')"
    scope_clause, scope_params = customer_scope_filter("c")

    def build_filters(date_condition: str) -> tuple[str, tuple[object, ...]]:
        clauses = [search_filter, status_filter, product_filter, date_condition]
        params = [q, q, q, q, q]
        if scope_clause:
            clauses.append(scope_clause)
            params.extend(scope_params)
        return " AND ".join(clauses), tuple(params)

    active_filters, active_params = build_filters("date(w.expiry_date) >= date('now')")
    active_query = base.format(filters=active_filters, order=order)
    active = df_query(conn, active_query, active_params)
    active = fmt_dates(active, ["issue_date","expiry_date"])
    if "dup_flag" in active.columns:
        active = active.assign(Duplicate=active["dup_flag"].apply(lambda x: "🔁 duplicate serial" if int(x)==1 else ""))
        active.drop(columns=["dup_flag"], inplace=True)
    active = format_warranty_table(active)
    if not is_admin and "Staff" in active.columns:
        active = active.drop(columns=["Staff"])
    st.markdown("**Active Warranties**")
    st.dataframe(active, use_container_width=True)

    expired_filters, expired_params = build_filters("date(w.expiry_date) < date('now')")
    expired_query = base.format(filters=expired_filters, order="DESC")
    expired = df_query(conn, expired_query, expired_params)
    expired = fmt_dates(expired, ["issue_date","expiry_date"])
    if "dup_flag" in expired.columns:
        expired = expired.assign(Duplicate=expired["dup_flag"].apply(lambda x: "🔁 duplicate serial" if int(x)==1 else ""))
        expired.drop(columns=["dup_flag"], inplace=True)
    expired = format_warranty_table(expired)
    if not is_admin and "Staff" in expired.columns:
        expired = expired.drop(columns=["Staff"])
    st.markdown("**Expired Warranties**")
    st.dataframe(expired, use_container_width=True)

    st.markdown("---")
    st.subheader("🔔 Upcoming Expiries")
    col1, col2 = st.columns(2)
    soon3 = collapse_warranty_rows(fetch_warranty_window(conn, 0, 3))
    soon60 = collapse_warranty_rows(fetch_warranty_window(conn, 0, 60))
    with col1:
        st.caption("Next **3** days")
        st.dataframe(soon3, use_container_width=True)
    with col2:
        st.caption("Next **60** days")
        st.dataframe(soon60, use_container_width=True)


def _render_service_section(conn, *, show_heading: bool = True):
    if show_heading:
        st.subheader("🛠️ Service Records")
    _, customer_label_map = build_customer_groups(conn, only_complete=False)
    customer_options, customer_labels, _, label_by_id = fetch_customer_choices(conn)
    viewer_id = current_user_id()
    do_df = df_query(
        conn,
        """
        SELECT d.do_number, d.customer_id, d.created_by, COALESCE(c.name, '(unknown)') AS customer_name, d.description, d.remarks, d.record_type
        FROM delivery_orders d
        LEFT JOIN customers c ON c.customer_id = d.customer_id
        WHERE COALESCE(d.record_type, 'delivery_order') = 'delivery_order'
          AND d.deleted_at IS NULL
        ORDER BY datetime(d.created_at) DESC
        """,
    )
    allowed_customers = accessible_customer_ids(conn)
    do_df = filter_delivery_orders_for_view(
        do_df, allowed_customers, record_types={"delivery_order"}
    )
    do_options = [None]
    do_labels = {None: "No delivery order (manual entry)"}
    do_customer_map = {}
    do_customer_name_map = {}
    for _, row in do_df.iterrows():
        do_num = clean_text(row.get("do_number"))
        if not do_num:
            continue
        cust_id = int(row["customer_id"]) if not pd.isna(row.get("customer_id")) else None
        summary = clean_text(row.get("description"))
        cust_name = customer_label_map.get(cust_id) if cust_id else clean_text(row.get("customer_name"))
        label_parts = [do_num]
        if cust_name:
            label_parts.append(f"({cust_name})")
        if summary:
            snippet = summary[:40]
            if len(summary) > 40:
                snippet += "…"
            label_parts.append(f"– {snippet}")
        label = " ".join(part for part in label_parts if part)
        do_options.append(do_num)
        do_labels[do_num] = label
        do_customer_map[do_num] = cust_id
        do_customer_name_map[do_num] = cust_name or "(not linked)"

    with st.form("service_form"):
        selected_do = st.selectbox(
            "Delivery order",
            options=do_options,
            format_func=lambda do: do_labels.get(do, str(do)),
        )
        default_customer = do_customer_map.get(selected_do)
        state_key = "service_customer_link"
        last_do_key = "service_customer_last_do"
        linked_customer = default_customer
        if default_customer is not None:
            st.session_state[last_do_key] = selected_do
            st.session_state[state_key] = default_customer
            customer_label = (
                customer_labels.get(default_customer)
                or customer_label_map.get(default_customer)
                or label_by_id.get(default_customer)
                or do_customer_name_map.get(selected_do)
                or f"Customer #{default_customer}"
            )
            st.text_input("Customer", value=customer_label, disabled=True)
        else:
            choices = list(customer_options)
            if st.session_state.get(last_do_key) != selected_do:
                st.session_state[last_do_key] = selected_do
                st.session_state[state_key] = None
            linked_customer = st.selectbox(
                "Customer *",
                options=choices,
                format_func=lambda cid: customer_labels.get(cid, "-- Select customer --"),
                key=state_key,
            )
        status_value = status_input_widget("service_new", DEFAULT_SERVICE_STATUS)
        status_choice = get_status_choice("service_new")
        today = datetime.now().date()
        if status_choice == "Completed":
            service_period_value = st.date_input(
                "Service period",
                value=(today, today),
                help="Select the start and end dates for the completed service.",
                key="service_new_period_completed",
            )
        elif status_choice == "In progress":
            service_period_value = st.date_input(
                "Service start date",
                value=today,
                help="Choose when this service work began.",
                key="service_new_period_start",
            )
        else:
            service_period_value = st.date_input(
                "Planned start date",
                value=today,
                help="Select when this service is scheduled to begin.",
                key="service_new_period_planned",
            )
        description = st.text_area("Service description")
        remarks = st.text_area("Remarks / updates")
        cond_cols = st.columns(2)
        with cond_cols[0]:
            condition_option = st.selectbox(
                "Generator condition after work",
                ["Not recorded"] + GENERATOR_CONDITION_OPTIONS,
                index=0,
                key="service_new_condition",
                help="Capture the condition of the generator once the work is completed.",
            )
        with cond_cols[1]:
            bill_amount_input = st.number_input(
                "Service amount",
                min_value=0.0,
                step=100.0,
                format="%.2f",
                key="service_new_bill_amount",
                help="Track the amount charged for this service.",
            )
        condition_notes = st.text_area(
            "Condition remarks",
            key="service_new_condition_notes",
            help="Add any notes about the generator condition once the job is done.",
        )
        st.markdown("**Products sold during service**")
        service_product_rows = st.session_state.get(
            "service_product_rows",
            [
                {
                    "name": "",
                    "model": "",
                    "serial": "",
                    "quantity": 1,
                }
            ],
        )
        service_product_editor = st.data_editor(
            pd.DataFrame(service_product_rows),
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "name": st.column_config.TextColumn("Product", help="Name / description"),
                "model": st.column_config.TextColumn("Model"),
                "serial": st.column_config.TextColumn("Serial"),
                "quantity": st.column_config.NumberColumn(
                    "Qty",
                    min_value=1,
                    step=1,
                    format="%d",
                ),
            },
            key="service_product_table",
        )
        service_product_entries = (
            service_product_editor.to_dict("records")
            if isinstance(service_product_editor, pd.DataFrame)
            else []
        )
        st.session_state["service_product_rows"] = service_product_entries
        service_files = st.file_uploader(
            "Attach service documents (PDF or image)",
            type=["pdf", "png", "jpg", "jpeg", "webp", "gif"],
            accept_multiple_files=True,
            key="service_new_docs",
        )
        if service_files:
            for idx, service_file in enumerate(service_files, start=1):
                _render_upload_ocr_preview(
                    service_file,
                    key_prefix=f"service_new_docs_{idx}",
                    label=f"Service document {idx} OCR",
                )
        submit = st.form_submit_button("Log service", type="primary")

    if submit:
        selected_customer = (
            linked_customer if linked_customer is not None else do_customer_map.get(selected_do)
        )
        selected_customer = int(selected_customer) if selected_customer is not None else None
        cur = conn.cursor()
        (
            service_date_str,
            service_start_str,
            service_end_str,
        ) = determine_period_strings(status_choice, service_period_value)
        valid_entry = True
        if selected_customer is None:
            st.error("Select a customer to log this service entry.")
            valid_entry = False
        if status_choice == "Completed" and (
            not service_start_str or not service_end_str
        ):
            st.error("Start and end dates are required for completed services.")
            valid_entry = False
        if status_choice != "Completed" and not service_start_str:
            st.error("Select a start date for this service entry.")
            valid_entry = False
        if valid_entry:
            _cleaned_service_products, service_product_labels = normalize_product_entries(
                service_product_entries
            )
            service_product_label = (
                "\n".join(service_product_labels) if service_product_labels else None
            )
            condition_value = (
                condition_option if condition_option in GENERATOR_CONDITION_OPTIONS else None
            )
            condition_notes_value = clean_text(condition_notes)
            bill_amount_value = None
            try:
                if bill_amount_input is not None and float(bill_amount_input) > 0:
                    bill_amount_value = round(float(bill_amount_input), 2)
            except Exception:
                bill_amount_value = None
            cur.execute(
                """
                INSERT INTO services (
                    do_number,
                    customer_id,
                    service_date,
                    service_start_date,
                    service_end_date,
                    description,
                    status,
                    remarks,
                    service_product_info,
                    condition_status,
                    condition_remarks,
                    bill_amount,
                    bill_document_path,
                    updated_at,
                    created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    selected_do,
                    selected_customer,
                    service_date_str,
                    service_start_str,
                    service_end_str,
                    clean_text(description),
                    status_value,
                    clean_text(remarks),
                    service_product_label,
                    condition_value,
                    condition_notes_value,
                    bill_amount_value,
                    None,
                    datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                    current_user_id(),
                ),
            )
            service_id = cur.lastrowid
            if selected_do and selected_customer is not None:
                link_delivery_order_to_customer(conn, selected_do, selected_customer)
                saved_docs = attach_documents(
                    conn,
                    "service_documents",
                    "service_id",
                    service_id,
                    service_files,
                    SERVICE_DOCS_DIR,
                    f"service_{service_id}",
                    allowed_extensions=DOCUMENT_UPLOAD_EXTENSIONS,
                )
                conn.commit()
                service_label = do_labels.get(selected_do) if selected_do else None
                if not service_label:
                    service_label = f"Service #{service_id}"
                customer_name = None
                if selected_customer is not None:
                    customer_name = (
                        label_by_id.get(int(selected_customer))
                        or customer_label_map.get(int(selected_customer))
                    )
                summary_parts = [service_label]
                if customer_name:
                    summary_parts.append(customer_name)
                status_label = clean_text(status_value) or DEFAULT_SERVICE_STATUS
                summary_parts.append(f"status {status_label}")
                log_activity(
                    conn,
                    event_type="service_created",
                    description=" – ".join(summary_parts),
                    entity_type="service",
                    entity_id=int(service_id),
                )
                message = "Service record saved."
                if saved_docs:
                    message = f"{message} Attached {saved_docs} document(s)."
                if bill_amount_value is not None:
                    message = f"{message} Recorded service amount {format_money(bill_amount_value)}."
                st.success(message)
                _safe_rerun()

    service_df = df_query(
        conn,
        """
        SELECT s.service_id,
               s.customer_id,
               d.customer_id AS do_customer_id,
               s.do_number,
               s.service_date,
               s.service_start_date,
               s.service_end_date,
               s.service_product_info,
               s.description,
               s.status,
               s.remarks,
               s.condition_status,
               s.condition_remarks,
               s.bill_amount,
                s.payment_receipt_path,
               s.updated_at,
               s.created_by,
               COALESCE(c.name, cdo.name, '(unknown)') AS customer,
               COUNT(sd.document_id) AS doc_count
        FROM services s
        LEFT JOIN customers c ON c.customer_id = s.customer_id
        LEFT JOIN delivery_orders d ON d.do_number = s.do_number
        LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
        LEFT JOIN service_documents sd ON sd.service_id = s.service_id
        WHERE s.deleted_at IS NULL
        GROUP BY s.service_id
        ORDER BY datetime(COALESCE(s.service_start_date, s.service_date)) DESC, s.service_id DESC
        """,
    )
    if allowed_customers is not None:
        def _service_row_allowed(row):
            service_cust = row.get("customer_id")
            do_cust = row.get("do_customer_id")
            creator_id = row.get("created_by")
            candidates = []
            if pd.notna(service_cust):
                candidates.append(int(service_cust))
            if pd.notna(do_cust):
                candidates.append(int(do_cust))
            try:
                if viewer_id is not None and pd.notna(creator_id) and int(creator_id) == int(viewer_id):
                    return True
            except Exception:
                pass
            return any(cid in allowed_customers for cid in candidates)

        service_df = service_df[service_df.apply(_service_row_allowed, axis=1)]
    if not service_df.empty:
        service_df = fmt_dates(service_df, ["service_date", "service_start_date", "service_end_date"])
        service_df["service_period"] = service_df.apply(
            lambda row: format_period_span(
                row.get("service_start_date"), row.get("service_end_date")
            ),
            axis=1,
        )
        service_records = service_df.to_dict("records")
        display_df = service_df.drop(
            columns=["customer_id", "do_customer_id", "created_by"],
            errors="ignore",
        )
        display_df["Last update"] = pd.to_datetime(display_df.get("updated_at"), errors="coerce").dt.strftime("%d-%m-%Y %H:%M")
        display_df.loc[display_df["Last update"].isna(), "Last update"] = None
        if "status" in display_df.columns:
            display_df["status"] = display_df["status"].apply(lambda x: clean_text(x) or DEFAULT_SERVICE_STATUS)
        if "condition_status" in display_df.columns:
            display_df["condition_status"] = display_df["condition_status"].apply(
                lambda x: clean_text(x) or "Not recorded"
            )
        if "bill_amount" in display_df.columns:
            display_df["service_amount_display"] = display_df["bill_amount"].apply(format_money)
        if "payment_receipt_path" in display_df.columns:
            display_df["payment_receipt_display"] = display_df["payment_receipt_path"].apply(
                lambda x: "📎" if clean_text(x) else ""
            )
        display = display_df.rename(
            columns={
                "do_number": "DO Serial",
                "service_date": "Service date",
                "service_start_date": "Service start date",
                "service_end_date": "Service end date",
                "service_period": "Service period",
                "service_product_info": "Products sold",
                "description": "Description",
                "status": "Status",
                "remarks": "Remarks",
                "condition_status": "Condition",
                "condition_remarks": "Condition notes",
                "service_amount_display": "Service amount",
                "payment_receipt_display": "Receipt",
                "customer": "Customer",
                "doc_count": "Documents",
            }
        )
        display = display.drop(columns=["payment_receipt_path"], errors="ignore")
        st.markdown("### Service history")
        st.dataframe(
            display.drop(columns=["updated_at", "service_id"], errors="ignore"),
            use_container_width=True,
        )

        records = service_records
        st.markdown("#### Update status & remarks")
        options = [int(r["service_id"]) for r in records]
        def service_label(record):
            do_ref = clean_text(record.get("do_number")) or "(no DO)"
            date_ref = clean_text(record.get("service_period")) or clean_text(
                record.get("service_date")
            )
            customer_ref = clean_text(record.get("customer"))
            parts = [do_ref]
            if date_ref:
                parts.append(f"· {date_ref}")
            if customer_ref:
                parts.append(f"· {customer_ref}")
            return " ".join(parts)

        labels = {int(r["service_id"]): service_label(r) for r in records}
        selected_service_id = st.selectbox(
            "Select service entry",
            options,
            format_func=lambda rid: labels.get(rid, str(rid)),
        )
        selected_record = next(r for r in records if int(r["service_id"]) == int(selected_service_id))
        new_status = status_input_widget(
            f"service_edit_{selected_service_id}", selected_record.get("status")
        )
        edit_status_choice = get_status_choice(f"service_edit_{selected_service_id}")
        existing_start = ensure_date(selected_record.get("service_start_date")) or ensure_date(
            selected_record.get("service_date")
        )
        existing_end = ensure_date(selected_record.get("service_end_date")) or existing_start
        today = datetime.now().date()
        default_start = existing_start or today
        default_end = existing_end or default_start
        if edit_status_choice == "Completed":
            edit_period_value = st.date_input(
                "Service period",
                value=(default_start, default_end),
                key=f"service_edit_{selected_service_id}_period_completed",
                help="Update the start and end dates for this service.",
            )
        elif edit_status_choice == "In progress":
            edit_period_value = st.date_input(
                "Service start date",
                value=default_start,
                key=f"service_edit_{selected_service_id}_period_start",
                help="Adjust when this service began.",
            )
        else:
            edit_period_value = st.date_input(
                "Planned start date",
                value=default_start,
                key=f"service_edit_{selected_service_id}_period_planned",
                help="Adjust when this service is scheduled to begin.",
            )
        new_remarks = st.text_area(
            "Remarks",
            value=clean_text(selected_record.get("remarks")) or "",
            key=f"service_edit_{selected_service_id}",
        )
        condition_cols = st.columns(2)
        existing_condition = clean_text(selected_record.get("condition_status"))
        condition_options = ["Not recorded"] + GENERATOR_CONDITION_OPTIONS
        default_condition = (
            existing_condition if existing_condition in GENERATOR_CONDITION_OPTIONS else "Not recorded"
        )
        with condition_cols[0]:
            condition_choice_edit = st.selectbox(
                "Generator condition",
                condition_options,
                index=condition_options.index(default_condition),
                key=f"service_edit_condition_{selected_service_id}",
            )
        existing_bill_amount = selected_record.get("bill_amount")
        try:
            bill_amount_default = float(existing_bill_amount) if existing_bill_amount is not None else 0.0
        except (TypeError, ValueError):
            bill_amount_default = 0.0
        with condition_cols[1]:
            bill_amount_edit = st.number_input(
                "Service amount",
                value=float(bill_amount_default),
                min_value=0.0,
                step=100.0,
                format="%.2f",
                key=f"service_edit_bill_amount_{selected_service_id}",
            )
        condition_notes_edit = st.text_area(
            "Condition remarks",
            value=clean_text(selected_record.get("condition_remarks")) or "",
            key=f"service_edit_condition_notes_{selected_service_id}",
        )
        receipt_col1, receipt_col2 = st.columns([1, 1])
        existing_receipt_path = clean_text(selected_record.get("payment_receipt_path"))
        resolved_receipt = resolve_upload_path(existing_receipt_path)
        with receipt_col1:
            receipt_upload = st.file_uploader(
                "Upload payment receipt (PDF or image)",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"service_edit_receipt_upload_{selected_service_id}",
            )
        with receipt_col2:
            clear_receipt = st.checkbox(
                "Remove receipt",
                value=False,
                key=f"service_edit_receipt_clear_{selected_service_id}",
            )
        if resolved_receipt and resolved_receipt.exists():
            st.download_button(
                "Download receipt",
                data=resolved_receipt.read_bytes(),
                file_name=resolved_receipt.name,
                key=f"service_receipt_download_{selected_service_id}",
            )
        elif existing_receipt_path:
            st.caption("Receipt file not found. Upload a new copy to replace it.")
        save_updates = st.button("Save updates", key="save_service_updates")
        if _guard_double_submit("save_service_updates", save_updates):
            (
                service_date_str,
                service_start_str,
                service_end_str,
            ) = determine_period_strings(edit_status_choice, edit_period_value)
            valid_update = True
            if edit_status_choice == "Completed" and (
                not service_start_str or not service_end_str
            ):
                st.error("Provide both start and end dates for completed services.")
                valid_update = False
            if edit_status_choice != "Completed" and not service_start_str:
                st.error("Select a start date for this service entry.")
                valid_update = False
            if valid_update:
                condition_update_value = (
                    condition_choice_edit
                    if condition_choice_edit in GENERATOR_CONDITION_OPTIONS
                    else None
                )
                condition_notes_update = clean_text(condition_notes_edit)
                bill_amount_update = None
                try:
                    if bill_amount_edit is not None and float(bill_amount_edit) > 0:
                        bill_amount_update = round(float(bill_amount_edit), 2)
                except Exception:
                    bill_amount_update = None
                receipt_path_value = clean_text(selected_record.get("payment_receipt_path"))
                replaced_receipt = False
                cleared_receipt = False
                if receipt_upload is not None:
                    receipt_path_value = store_payment_receipt(
                        receipt_upload,
                        identifier=f"service_{selected_service_id}_receipt",
                        target_dir=SERVICE_BILL_DIR,
                    )
                    replaced_receipt = bool(receipt_path_value)
                elif clear_receipt and receipt_path_value:
                    old_receipt = resolve_upload_path(receipt_path_value)
                    if old_receipt and old_receipt.exists():
                        try:
                            old_receipt.unlink()
                        except Exception:
                            pass
                    receipt_path_value = None
                    cleared_receipt = True
                conn.execute(
                    """
                    UPDATE services
                    SET status = ?,
                        remarks = ?,
                        service_date = ?,
                        service_start_date = ?,
                        service_end_date = ?,
                        condition_status = ?,
                        condition_remarks = ?,
                        bill_amount = ?,
                        payment_receipt_path = COALESCE(?, payment_receipt_path),
                        updated_at = datetime('now')
                    WHERE service_id = ?
                      AND deleted_at IS NULL
                    """,
                    (
                        new_status,
                        clean_text(new_remarks),
                        service_date_str,
                        service_start_str,
                        service_end_str,
                        condition_update_value,
                        condition_notes_update,
                        bill_amount_update,
                        receipt_path_value,
                        int(selected_service_id),
                    ),
                )
                conn.commit()
                label_text = labels.get(int(selected_service_id), f"Service #{int(selected_service_id)}")
                status_label = clean_text(new_status) or DEFAULT_SERVICE_STATUS
                message_summary = label_text
                if status_label:
                    message_summary = f"{label_text} → {status_label}"
                log_activity(
                    conn,
                    event_type="service_updated",
                    description=message_summary,
                    entity_type="service",
                    entity_id=int(selected_service_id),
                )
                message_bits = ["Service record updated."]
                if bill_amount_update is not None:
                    message_bits.append(f"Service amount {format_money(bill_amount_update)}")
                if replaced_receipt:
                    message_bits.append("Receipt uploaded")
                elif cleared_receipt:
                    message_bits.append("Receipt removed")
                st.success(". ".join(message_bits))
                _safe_rerun()

        attachments_df = df_query(
            conn,
            """
            SELECT document_id, file_path, original_name, uploaded_at
            FROM service_documents
            WHERE service_id = ?
            ORDER BY datetime(uploaded_at) DESC, document_id DESC
            """,
            (int(selected_service_id),),
        )
        st.markdown("**Attached documents**")
        if attachments_df.empty:
            st.caption("No documents attached yet.")
        else:
            for _, doc_row in attachments_df.iterrows():
                path = resolve_upload_path(doc_row.get("file_path"))
                display_name = clean_text(doc_row.get("original_name"))
                if path and path.exists():
                    label = display_name or path.name
                    st.download_button(
                        f"Download {label}",
                        data=path.read_bytes(),
                        file_name=path.name,
                        key=f"service_doc_dl_{int(doc_row['document_id'])}",
                    )
                else:
                    label = display_name or "Document"
                    st.caption(f"⚠️ Missing file: {label}")

        with st.form(f"service_doc_upload_{selected_service_id}"):
            more_docs = st.file_uploader(
                "Add more service documents (PDF or image)",
                type=["pdf", "png", "jpg", "jpeg", "webp", "gif"],
                accept_multiple_files=True,
                key=f"service_doc_files_{selected_service_id}",
            )
            upload_docs = st.form_submit_button("Upload documents")
        if upload_docs:
            if more_docs:
                saved = attach_documents(
                    conn,
                    "service_documents",
                    "service_id",
                    int(selected_service_id),
                    more_docs,
                    SERVICE_DOCS_DIR,
                    f"service_{selected_service_id}",
                    allowed_extensions=DOCUMENT_UPLOAD_EXTENSIONS,
                )
                conn.commit()
                st.success(f"Uploaded {saved} document(s).")
                _safe_rerun()
            else:
                st.info("Select at least one PDF or image to upload.")

        st.markdown("#### Delete service record")
        actor_id = current_user_id()
        is_admin = current_user_is_admin()
        deletable_df = pd.DataFrame(service_records)
        if not is_admin and actor_id is not None and not deletable_df.empty:
            deletable_df = deletable_df[
                deletable_df["created_by"].apply(
                    lambda val: int(_coerce_float(val, -1)) == actor_id
                )
            ]
        if deletable_df.empty:
            st.caption("No service records available for deletion.")
        else:
            delete_labels: dict[int, str] = {}
            for _, row in deletable_df.iterrows():
                service_id = int(row.get("service_id"))
                do_ref = clean_text(row.get("do_number")) or "Service"
                customer_ref = clean_text(row.get("customer")) or "(customer)"
                period_ref = clean_text(row.get("service_period"))
                label_parts = [do_ref, customer_ref]
                if period_ref:
                    label_parts.append(period_ref)
                delete_labels[service_id] = " • ".join(label_parts)
            delete_options = list(delete_labels.keys())
            selected_delete_id = st.selectbox(
                "Select a service entry to delete",
                delete_options,
                format_func=lambda val: delete_labels.get(val, f"Service #{val}"),
                key="service_delete_select",
            )
            confirm_delete = st.checkbox(
                "I understand this service entry will be removed from active views.",
                key="service_delete_confirm",
            )
            if st.button(
                "Delete service record",
                type="secondary",
                disabled=not confirm_delete,
                key="service_delete_button",
            ):
                conn.execute(
                    """
                    UPDATE services
                    SET deleted_at=datetime('now'),
                        deleted_by=?
                    WHERE service_id=?
                      AND deleted_at IS NULL
                    """,
                    (actor_id, selected_delete_id),
                )
                conn.commit()
                log_activity(
                    conn,
                    event_type="service_deleted",
                    description=f"Service #{selected_delete_id} deleted",
                    entity_type="service",
                    entity_id=int(selected_delete_id),
                    user_id=actor_id,
                )
                st.warning("Service record deleted.")
                _safe_rerun()
    else:
        st.info("No service records yet. Log one using the form above.")


def _build_quotation_workbook(
    *,
    metadata: dict[str, Optional[str]],
    items: list[dict[str, object]],
    totals: list[tuple[str, float]],
) -> bytes:
    buffer = io.BytesIO()
    summary_rows = [(key, metadata.get(key) or "-") for key in metadata]
    summary_df = pd.DataFrame(summary_rows, columns=["Field", "Value"])
    items_df = pd.DataFrame(items)
    totals_df = pd.DataFrame(totals, columns=["Label", "Amount"])

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Quotation", index=False)
        start_row = len(summary_df) + 2
        if not items_df.empty:
            items_df.to_excel(writer, sheet_name="Quotation", index=False, startrow=start_row)
            totals_start = start_row + len(items_df) + 2
        else:
            totals_start = start_row
        if not totals_df.empty:
            totals_df.to_excel(writer, sheet_name="Quotation", index=False, startrow=totals_start)

    buffer.seek(0)
    return buffer.read()


def _regenerate_quotation_pdf_from_workbook(file_path: Path) -> Optional[bytes]:
    """Recreate a quotation PDF from an archived Excel workbook.

    Older records may only have the Excel workbook stored; this helper rebuilds
    the PDF so dashboard downloads remain consistent.
    """

    if not file_path.exists():
        return None

    try:
        workbook = load_workbook(filename=file_path, data_only=True)
    except Exception:
        return None

    if "Quotation" not in workbook.sheetnames:
        return None

    ws = workbook["Quotation"]
    rows = list(ws.iter_rows(values_only=True))

    def _is_blank(row: tuple[object, ...]) -> bool:
        return all(cell is None or str(cell).strip() == "" for cell in row)

    metadata: dict[str, Optional[str]] = {}
    items: list[dict[str, object]] = []
    totals_rows: list[tuple[str, object]] = []

    section = "summary"
    header: list[str] = []
    totals_header: list[str] = []

    for row in rows:
        if _is_blank(row):
            if section == "summary":
                section = "items_header"
            elif section == "items":
                section = "totals_header"
            continue

        if section == "summary":
            key = clean_text(row[0]) if len(row) else None
            value = clean_text(row[1]) if len(row) > 1 else None
            if key:
                metadata[key] = value
            continue

        if section == "items_header":
            header = [clean_text(cell) or "" for cell in row if cell is not None]
            section = "items"
            continue

        if section == "items":
            if not header:
                continue
            item_values = list(row)
            if all(cell is None for cell in item_values):
                section = "totals_header"
                continue
            item: dict[str, object] = {}
            for idx, col_name in enumerate(header):
                if not col_name:
                    continue
                value = item_values[idx] if idx < len(item_values) else None
                if isinstance(value, str):
                    value = clean_text(value)
                item[col_name] = value
            if item:
                items.append(item)
            continue

        if section == "totals_header":
            totals_header = [clean_text(cell) or "" for cell in row if cell is not None]
            section = "totals"
            continue

        if section == "totals":
            if not totals_header:
                continue
            if all(cell is None for cell in row):
                break
            label = clean_text(row[0]) if len(row) else None
            value = row[1] if len(row) > 1 else None
            if label:
                totals_rows.append((label, value))

    def _total_from_label(label: str) -> Optional[float]:
        for key, value in totals_rows:
            if clean_text(key) == label:
                return _coerce_float(value, 0.0)
        return None

    for item in items:
        for tax_label in [
            "CGST (%)",
            "CGST amount",
            "SGST (%)",
            "SGST amount",
            "IGST (%)",
            "IGST amount",
        ]:
            item.pop(tax_label, None)

    totals = {
        "gross_total": _total_from_label("gross amount")
        or _total_from_label("manual total override")
        or sum(
            _coerce_float(
                item.get("Total Price, Tk.") or item.get("Gross amount"), 0.0
            )
            for item in items
        ),
        "discount_total": _total_from_label("discount total")
        or sum(_coerce_float(item.get("Discount amount"), 0.0) for item in items),
        "grand_total": _total_from_label("grand total")
        or _total_from_label("manual total override")
        or sum(
            _coerce_float(
                item.get("Total Price, Tk.") or item.get("Line total"), 0.0
            )
            for item in items
        ),
    }

    grand_total_label = format_money(totals["grand_total"]) or f"{totals['grand_total']:,.2f}"
    grand_total_words = format_amount_in_words(totals["grand_total"]) or grand_total_label

    try:
        return _build_quotation_pdf(
            metadata=metadata,
            items=items,
            totals=totals,
            grand_total_label=grand_total_label,
            grand_total_words=grand_total_words,
        )
    except Exception:
        return None


def _resolve_letterhead_path(template_choice: Optional[str] = None) -> Optional[Path]:
    template_choice = template_choice or "Default letterhead"
    base_dir = Path(__file__).resolve().parent
    sandbox_letterhead = Path("/mnt/data/eed2a8fe-ec62-4729-9ed4-aedb65953acf.png")
    default_candidates = [
        sandbox_letterhead,
        base_dir / "letterhead.png",
        base_dir / "letterhead",
        base_dir / "PS-SALES-main" / "ps_letterhead.png",
        base_dir / "ps_letterhead.png",
    ]

    preferred: list[Path] = []
    if template_choice == "PS letterhead":
        preferred = [sandbox_letterhead, base_dir / "PS-SALES-main" / "ps_letterhead.png", base_dir / "ps_letterhead.png"]
    elif template_choice == "Default letterhead":
        preferred = [sandbox_letterhead, base_dir / "letterhead.png", base_dir / "letterhead"]

    seen: set[Path] = set()
    candidates: list[Path] = []
    for path in preferred + default_candidates:
        if path in seen:
            continue
        seen.add(path)
        candidates.append(path)

    for path in candidates:
        if path.exists():
            return path
    return None




def _build_quotation_pdf(
    *,
    metadata: dict[str, Optional[str]],
    items: list[dict[str, object]],
    totals: dict[str, float],
    grand_total_label: str,
    template_choice: Optional[str] = None,
    grand_total_words: Optional[str] = None,
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="Muted",
            parent=styles["Normal"],
            textColor=colors.HexColor("#475569"),
            fontSize=10,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodySmall",
            parent=styles["Normal"],
            fontSize=11,
            leading=14,
        )
    )

    reference = metadata.get("Reference number") or metadata.get("Quotation reference") or "—"
    date_value = metadata.get("Date") or "—"
    customer_name = metadata.get("Customer company") or metadata.get("Customer / organisation") or "—"
    customer_contact = metadata.get("Customer contact name") or ""
    customer_contact_details = metadata.get("Customer contact") or ""
    customer_address = metadata.get("Customer address") or ""
    customer_district = metadata.get("Customer district") or ""
    attention_name = metadata.get("Attention name") or "—"
    prepared_by = metadata.get("Salesperson name") or "—"
    prepared_title = metadata.get("Salesperson title") or ""
    prepared_contact = metadata.get("Salesperson contact") or ""
    prepared_email = metadata.get("Salesperson email") or ""
    terms = metadata.get("Notes / terms") or ""

    story: list[object] = []
    letterhead_path = _resolve_letterhead_path(template_choice)
    letterhead_img_path = str(letterhead_path) if letterhead_path and letterhead_path.exists() else None

    def _draw_letterhead_background(canvas, _doc):
        if not letterhead_img_path:
            return
        canvas.saveState()
        try:
            reader = ImageReader(letterhead_img_path)
            img_width, img_height = reader.getSize()
            page_width, page_height = _doc.pagesize
            target_width = page_width
            target_height = target_width * (img_height / img_width)
            canvas.drawImage(
                letterhead_img_path,
                0,
                page_height - target_height,
                width=target_width,
                height=target_height,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass
        finally:
            canvas.restoreState()

    story.append(Spacer(1, 18))

    header_table = Table(
        [
            ["", Paragraph(f"<b>Date:</b> {date_value}", styles["BodySmall"])],
            ["", Paragraph(f"<b>Ref:</b> {reference}", styles["BodySmall"])],
        ],
        colWidths=[doc.width * 0.58, doc.width * 0.42],
    )
    header_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 10))

    address_lines = [customer_name, customer_contact, customer_contact_details, customer_address, customer_district]
    address_text = "<br/>".join(filter(None, address_lines)) or "—"
    story.append(Paragraph("<b>Customer</b>", styles["BodySmall"]))
    story.append(Paragraph(address_text, styles["BodySmall"]))
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            f"<b>Attention:</b> {attention_name}",
            styles["BodySmall"],
        )
    )
    story.append(Spacer(1, 8))
    story.append(
        Paragraph(
            "<b>PRICE SCHEDULE</b>",
            styles["BodySmall"],
        )
    )
    story.append(Spacer(1, 4))

    table_data = [
        [
            "Sl No.",
            "Description of Generator",
            "Qty.",
            "Unit Price, Tk.",
            "Total Price, Tk.",
        ]
    ]

    for idx, item in enumerate(items, start=1):
        description_parts = []
        description_parts.append(
            html.escape(
                clean_text(item.get("Description of Generator"))
                or clean_text(item.get("Description"))
                or clean_text(item.get("description"))
                or "Item"
            )
        )
        specs_text = clean_text(item.get("Specs")) or ""
        note_text = (
            clean_text(item.get("Notes"))
            or clean_text(item.get("Note"))
            or clean_text(item.get("note"))
            or ""
        )
        if specs_text:
            description_parts.append(
                f"<font color='#475569'>{html.escape(specs_text)}</font>"
            )
        if note_text:
            description_parts.append(
                f"<font color='#94a3b8' size='9'>{html.escape(note_text)}</font>"
            )
        description = "<br/>".join(filter(None, description_parts))
        qty_value = _coerce_float(
            item.get("Qty.") or item.get("Quantity") or item.get("quantity"), 0.0
        )
        rate_value = (
            item.get("Unit Price, Tk.")
            if "Unit Price, Tk." in item
            else item.get("Rate")
            if "Rate" in item
            else item.get("unit_price")
        )
        line_total_value = (
            item.get("Total Price, Tk.")
            if "Total Price, Tk." in item
            else item.get("Line total")
            if "Line total" in item
            else item.get("line_total")
        )
        table_data.append(
            [
                str(idx),
                Paragraph(description, styles["BodySmall"]),
                Paragraph(f"{qty_value:,.0f}", styles["BodySmall"]),
                Paragraph(format_money(rate_value) or f"{_coerce_float(rate_value, 0.0):,.2f}", styles["BodySmall"]),
                Paragraph(
                    format_money(line_total_value) or f"{_coerce_float(line_total_value, 0.0):,.2f}",
                    styles["BodySmall"],
                ),
            ]
        )

    table_data.append(
        [
            "",
            "",
            "",
            Paragraph("<b>Total Amount (Tk.)</b>", styles["BodySmall"]),
            Paragraph(f"<b>{grand_total_label}</b>", styles["BodySmall"]),
        ]
    )

    col_widths = [doc.width * 0.08, doc.width * 0.48, doc.width * 0.1, doc.width * 0.17, doc.width * 0.17]
    pricing_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    pricing_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -2), 0.5, colors.grey),
                ("BOX", (0, 0), (-1, -2), 0.5, colors.grey),
                ("SPAN", (0, -1), (2, -1)),
                ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, -1), (-1, -1), 0.5, colors.grey),
                ("BOX", (0, -1), (-1, -1), 0.5, colors.grey),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(pricing_table)
    story.append(Spacer(1, 10))

    discount_value = _coerce_float(totals.get("discount_total") if totals else None, 0.0)
    if not grand_total_words:
        amount_in_words = format_amount_in_words(totals.get("grand_total"))
        grand_total_words = amount_in_words or grand_total_label

    if discount_value:
        discount_label = format_money(discount_value) or f"{discount_value:,.2f}"
        story.append(
            Paragraph(
                f"Discount applied: {discount_label} (reflected in totals)",
                styles["Muted"],
            )
        )
        story.append(Spacer(1, 6))

    story.append(Paragraph(f"In Words: {grand_total_words}", styles["BodySmall"]))
    if terms:
        story.append(Spacer(1, 10))
        story.append(Paragraph("<b>Terms & Conditions</b>", styles["BodySmall"]))
        story.append(Spacer(1, 4))
        story.append(Paragraph(html.escape(terms).replace("\n", "<br/>"), styles["BodySmall"]))
    story.append(Spacer(1, 24))
    story.append(Paragraph(prepared_by, styles["BodySmall"]))
    if prepared_title:
        story.append(Paragraph(prepared_title, styles["BodySmall"]))
    if prepared_contact:
        story.append(Paragraph(prepared_contact, styles["BodySmall"]))
    if prepared_email:
        story.append(Paragraph(prepared_email, styles["BodySmall"]))

    doc.build(
        story,
        onFirstPage=_draw_letterhead_background,
        onLaterPages=_draw_letterhead_background,
    )

    return buffer.getvalue()


def _load_letterhead_data_uri(template_choice: Optional[str] = None) -> Optional[str]:
    """Load the configured letterhead into a data URI for the live preview."""

    template_path = _resolve_letterhead_path(template_choice)
    if not template_path:
        st.warning(
            "Letterhead template missing. Upload ps_letterhead.png to see the preview.",
            icon="⚠️",
        )
        return None

    try:
        mime = "image/png" if template_path.suffix.lower() == ".png" else "application/pdf"
        encoded = base64.b64encode(template_path.read_bytes()).decode("utf-8")
        suffix = "#page=1" if mime == "application/pdf" else ""
        return f"data:{mime};base64,{encoded}{suffix}"
    except OSError:
        st.warning(
            "Could not read the letterhead file. Please re-upload it and try again.",
            icon="⚠️",
        )
        return None


def _render_letterhead_preview(
    metadata: dict[str, Optional[str]],
    grand_total: str,
    template_choice: Optional[str] = None,
    items: Optional[list[dict[str, object]]] = None,
    totals: Optional[dict[str, float]] = None,
) -> None:
    metadata = metadata or {}

    resolved_render_id = st.session_state.get("_render_id", 0)
    overlay_id = f"letterhead-preview-{resolved_render_id}"

    template_path = _resolve_letterhead_path(template_choice)
    if not template_path:
        st.warning(
            "Letterhead template missing. Upload ps_letterhead.png to see the preview.",
            icon="⚠️",
        )
        return

    data_uri = _load_letterhead_data_uri(template_choice)
    if not data_uri:
        return

    def _format_currency(value: object) -> str:
        return format_money(value) or f"{_coerce_float(value, 0.0):,.2f}"

    items = items or []
    totals = totals or {}

    grand_total_value = _coerce_float(totals.get("grand_total"), 0.0)
    if grand_total_value <= 0:
        parsed_grand_total = parse_amount(totals.get("grand_total"))
        if parsed_grand_total is not None:
            grand_total_value = parsed_grand_total
    if grand_total_value <= 0:
        parsed_from_label = parse_amount(grand_total)
        if parsed_from_label is not None:
            grand_total_value = parsed_from_label

    if grand_total_value <= 0 and items:
        item_total = 0.0
        for item in items:
            line_total = parse_amount(
                item.get("Total Price, Tk.")
                or item.get("Line total")
                or item.get("line_total")
                or item.get("Amount")
            )
            if line_total is not None:
                item_total += line_total
        if item_total > 0:
            grand_total_value = item_total

    grand_total_label = _format_currency(grand_total_value) if grand_total_value > 0 else "—"
    grand_total_words = (
        format_amount_in_words(grand_total_value)
        if grand_total_value > 0
        else "Add item pricing to calculate totals"
    )

    customer = html.escape(
        str(
            metadata.get("Customer company")
            or metadata.get("Customer / organisation")
            or metadata.get("Hospital / customer")
            or metadata.get("Customer")
            or ""
        )
    )
    raw_contact = (
        metadata.get("Customer contact name")
        or metadata.get("Customer contact")
        or ""
    )
    contact = html.escape(str(raw_contact))
    reference = html.escape(
        str(
            metadata.get("Reference number")
            or metadata.get("Quotation reference")
            or ""
        )
    )
    address = html.escape(str(metadata.get("Customer address") or ""))
    district = html.escape(str(metadata.get("Customer district") or ""))
    date_value = html.escape(str(metadata.get("Date") or ""))
    prepared_by = html.escape(str(metadata.get("Salesperson name") or ""))
    prepared_title = html.escape(str(metadata.get("Salesperson title") or ""))
    prepared_contact = html.escape(str(metadata.get("Salesperson contact") or ""))
    attention_name = html.escape(str(metadata.get("Attention name") or ""))

    line_items = items[:8]
    discount_value = _coerce_float(totals.get("discount_total") if totals else None, 0.0)
    discount_label = (
        format_money(discount_value) or f"{discount_value:,.2f}"
        if discount_value
        else ""
    )
    rows_markup = []
    for idx, item in enumerate(line_items, start=1):
        title = html.escape(
            clean_text(item.get("Description of Generator"))
            or clean_text(item.get("Description"))
            or "Item"
        )
        specs = html.escape(clean_text(item.get("Specs")) or "")
        note = html.escape(clean_text(item.get("Notes")) or clean_text(item.get("note")) or "")
        qty = _coerce_float(item.get("Qty.") or item.get("Quantity"), 0.0)
        qty_display = f"{int(qty)}" if math.isclose(qty, round(qty)) else f"{qty:,.2f}"
        rate_display = _format_currency(
            item.get("Unit Price, Tk.")
            or item.get("Rate")
            or item.get("unit_price")
        )
        total_display = _format_currency(
            item.get("Total Price, Tk.")
            or item.get("Line total")
            or item.get("line_total")
        )
        detail_lines = [title]
        if specs:
            detail_lines.append(f"<span style='color:#475569;'>{specs}</span>")
        if note:
            detail_lines.append(
                f"<span style='color:#94a3b8; font-size:12px;'>{note}</span>"
            )
        detail_block = "<br/>".join(detail_lines)
        rows_markup.append(
            f"<tr>"
            f"<td style='padding:6px; text-align:center; border:1px solid #cbd5e1;'>{idx}</td>"
            f"<td style='padding:6px; border:1px solid #cbd5e1;'>{detail_block}</td>"
            f"<td style='padding:6px; text-align:center; border:1px solid #cbd5e1;'>{qty_display}</td>"
            f"<td style='padding:6px; text-align:right; border:1px solid #cbd5e1;'>{rate_display}</td>"
            f"<td style='padding:6px; text-align:right; border:1px solid #cbd5e1;'>{total_display}</td>"
            f"</tr>"
        )

    total_row = (
        f"<tr style='background:#f8fafc;'>"
        f"<td colspan='3' style='padding:6px; border:1px solid #cbd5e1;'></td>"
        f"<td style='padding:6px; text-align:right; border:1px solid #cbd5e1; font-weight:700;'>Total Amount (Tk.)</td>"
        f"<td style='padding:6px; text-align:right; border:1px solid #cbd5e1; font-weight:700;'>{grand_total_label}</td>"
        f"</tr>"
    )

    address_block = "<br/>".join(filter(None, [customer, contact, address, district]))

    letterhead_style = (
        f"position: relative; width: 940px; min-height: 1100px; border: 1px solid #e5e7eb; border-radius: 12px;"
        f" overflow: hidden; box-shadow: 0 18px 48px rgba(15, 23, 42, 0.14); background: #f8fafc;"
        f" background-image: url('{template_path.as_posix()}'), url('{data_uri}');"
        " background-size: contain; background-repeat: no-repeat; background-position: top center;"
    )

    preview_html = dedent(
        f"""
        <style>
          .letterhead-wrapper[data-overlay-id='{overlay_id}'] {{
            {letterhead_style}
          }}
          .letterhead-wrapper[data-overlay-id='{overlay_id}'] .letterhead-content {{
            position: relative;
            padding: 130px 72px 90px 72px;
            color: #0f172a;
            font-family: 'Arial', sans-serif;
          }}
        </style>
        <div style="margin-top: 1rem; display: flex; justify-content: center;">
          <div class="letterhead-wrapper" data-overlay-id="{overlay_id}">
            <div class="letterhead-content">
              <div style="text-align: right; font-size: 13px; line-height: 1.5;">
                <div><strong>Date:</strong> {date_value or '—'}</div>
                <div><strong>Ref:</strong> {reference or '—'}</div>
              </div>
              <div style="margin-top: 12px; font-size: 13px; line-height: 1.6;">
                <div><strong>Customer</strong></div>
                <div>{address_block or '—'}</div>
              </div>
              <div style="margin-top: 8px; font-size: 13px;"><strong>Attention:</strong> {attention_name or '—'}</div>
              <div style="margin-top: 14px; font-size: 13px; font-weight: 700;">PRICE SCHEDULE</div>
              <table style="width: 100%; border-collapse: collapse; margin-top: 6px; font-size: 12.5px;">
                <thead>
                  <tr style="background: #e2e8f0;">
                    <th style="padding: 6px; border: 1px solid #cbd5e1; text-align: center;">Sl No.</th>
                    <th style="padding: 6px; border: 1px solid #cbd5e1; text-align: left;">Description of Generator</th>
                    <th style="padding: 6px; border: 1px solid #cbd5e1; text-align: center;">Qty.</th>
                    <th style="padding: 6px; border: 1px solid #cbd5e1; text-align: right;">Unit Price, Tk.</th>
                    <th style="padding: 6px; border: 1px solid #cbd5e1; text-align: right;">Total Price, Tk.</th>
                  </tr>
                </thead>
                <tbody>
                  {''.join(rows_markup) or '<tr><td colspan="5" style="padding:8px; text-align:center; border:1px solid #cbd5e1; color:#64748b;">Add items to see the price schedule.</td></tr>'}
                  {total_row}
                </tbody>
              </table>
              {f"<div style='margin-top: 8px; font-size: 12px; color:#475569;'>Discount applied: {discount_label} (reflected above)</div>" if discount_label else ''}
              <div style="margin-top: 10px; font-size: 13px;">In Words: {grand_total_words}</div>
              <div style="margin-top: 40px; font-size: 13px; line-height: 1.4;">
                <div>{prepared_by or ''}</div>
                <div>{prepared_title or ''}</div>
                <div>{prepared_contact or ''}</div>
              </div>
            </div>
          </div>
        </div>
        """
    )

    st.markdown(preview_html, unsafe_allow_html=True)
    return preview_html

def _quotation_scope_filter() -> tuple[str, tuple[object, ...]]:
    if current_user_is_admin():
        return "WHERE deleted_at IS NULL", ()
    uid = current_user_id()
    if uid is None:
        return "WHERE 1=0", ()
    return "WHERE created_by = ? AND deleted_at IS NULL", (uid,)


def _save_quotation_record(conn, payload: dict) -> Optional[int]:
    columns = [
        "reference",
        "quote_date",
        "customer_name",
        "customer_company",
        "customer_address",
        "customer_district",
        "customer_contact",
        "attention_name",
        "attention_title",
        "subject",
        "salutation",
        "introduction",
        "closing",
        "quote_type",
        "total_amount",
        "discount_pct",
        "status",
        "payment_receipt_path",
        "follow_up_status",
        "follow_up_notes",
        "follow_up_date",
        "reminder_label",
        "letter_template",
        "salesperson_name",
        "salesperson_title",
        "salesperson_contact",
        "salesperson_email",
        "document_path",
        "items_payload",
        "remarks_internal",
        "created_by",
    ]
    placeholders = ",".join("?" for _ in columns)
    values = [payload.get(col) for col in columns]
    try:
        cur = conn.execute(
            f"INSERT INTO quotations ({','.join(columns)}) VALUES ({placeholders})",
            tuple(values),
        )
        conn.execute(
            "UPDATE quotations SET updated_at=datetime('now') WHERE quotation_id=? AND deleted_at IS NULL",
            (cur.lastrowid,),
        )
        conn.commit()
    except sqlite3.Error:
        return None
    return int(cur.lastrowid)


def _upsert_customer_from_manual_quotation(
    conn,
    *,
    name: Optional[str],
    company: Optional[str],
    phone: Optional[str],
    address: Optional[str],
    district: Optional[str],
    reference: Optional[str] = None,
    created_by: Optional[int] = None,
    lead_status: Optional[str] = None,
) -> Optional[int]:
    """Insert or backfill a customer captured from a manual quotation entry."""

    customer_name = clean_text(name)
    company_name = clean_text(company)
    phone_number = clean_text(phone)
    street_address = clean_text(address)
    district_label = clean_text(district)
    reference_label = clean_text(reference)

    if not any([customer_name, company_name, phone_number, street_address, district_label]):
        return None

    cursor = conn.cursor()

    def _fetch_existing(query: str, params: tuple[object, ...]):
        return cursor.execute(query, params).fetchone()

    existing = None
    if phone_number:
        existing = _fetch_existing(
            """
            SELECT customer_id, name, company_name, phone, address, delivery_address, remarks
            FROM customers
            WHERE TRIM(IFNULL(phone, '')) = ?
            ORDER BY customer_id DESC
            LIMIT 1
            """,
            (phone_number,),
        )
    if existing is None and company_name:
        existing = _fetch_existing(
            """
            SELECT customer_id, name, company_name, phone, address, delivery_address, remarks
            FROM customers
            WHERE LOWER(IFNULL(company_name, '')) = LOWER(?)
            ORDER BY customer_id DESC
            LIMIT 1
            """,
            (company_name,),
        )
    if existing is None and customer_name:
        existing = _fetch_existing(
            """
            SELECT customer_id, name, company_name, phone, address, delivery_address, remarks
            FROM customers
            WHERE LOWER(IFNULL(name, '')) = LOWER(?)
            ORDER BY customer_id DESC
            LIMIT 1
            """,
            (customer_name,),
        )

    if existing:
        (
            customer_id,
            existing_name,
            existing_company,
            existing_phone,
            existing_address,
            existing_delivery,
            existing_remarks,
        ) = existing

        updates: dict[str, object] = {}
        if not existing_name and customer_name:
            updates["name"] = customer_name
        if not existing_company and company_name:
            updates["company_name"] = company_name
        if not existing_phone and phone_number:
            updates["phone"] = phone_number
        if not existing_address and street_address:
            updates["address"] = street_address
        if not existing_delivery and street_address:
            updates["delivery_address"] = street_address
        if not existing_remarks and district_label:
            updates["remarks"] = f"District: {district_label}"

        if updates:
            set_clause = ", ".join(f"{col}=?" for col in updates)
            cursor.execute(
                f"UPDATE customers SET {set_clause} WHERE customer_id=?",
                (*updates.values(), customer_id),
            )
            conn.commit()
        return customer_id

    remark_parts = []
    lead_tag = clean_text(lead_status)
    if lead_tag:
        remark_parts.append(lead_tag)
    if district_label:
        remark_parts.append(f"District: {district_label}")
    if reference_label:
        remark_parts.append(f"Quotation ref: {reference_label}")
    remarks_value = " | ".join(remark_parts) if remark_parts else None

    cursor.execute(
        """
        INSERT INTO customers (name, company_name, phone, address, delivery_address, remarks, created_by, dup_flag)
        VALUES (?, ?, ?, ?, ?, ?, ?, 0)
        """,
        (
            customer_name or company_name,
            company_name,
            phone_number,
            street_address,
            street_address,
            remarks_value,
            created_by,
        ),
    )
    conn.commit()
    return cursor.lastrowid

def _persist_quotation_pdf(
    record_id: int, pdf_bytes: bytes, reference: Optional[str]
) -> Optional[str]:
    if not pdf_bytes or record_id is None:
        return None
    ensure_upload_dirs()
    safe_ref = re.sub(r"[^a-zA-Z0-9_-]+", "-", clean_text(reference) or "").strip("-")
    safe_ref = safe_ref or "quotation"
    dest = QUOTATION_RECEIPT_DIR / f"quotation_{record_id}_{safe_ref}.pdf"
    counter = 1
    while dest.exists():
        dest = QUOTATION_RECEIPT_DIR / f"quotation_{record_id}_{safe_ref}_{counter}.pdf"
        counter += 1
    try:
        with open(dest, "wb") as fh:
            fh.write(pdf_bytes)
        return str(dest.relative_to(BASE_DIR))
    except (OSError, ValueError):
        return None


def _update_quotation_records(
    conn,
    updates: Iterable[dict[str, object]],
    *,
    allow_locked: bool = False,
) -> dict[str, list[int]]:
    updated: list[int] = []
    locked: list[int] = []
    for entry in updates:
        try:
            quotation_id = int(entry.get("quotation_id"))
        except Exception:
            continue
        cur = conn.execute(
            """
            SELECT status, follow_up_status, follow_up_notes, follow_up_date, reminder_label,
                   payment_receipt_path, reference, customer_name, customer_company, customer_contact
            FROM quotations
            WHERE quotation_id=? AND deleted_at IS NULL
            """,
            (quotation_id,),
        )
        row = cur.fetchone()
        if not row:
            continue
        (
            current_status,
            current_follow_up_status,
            current_follow_up_notes,
            current_follow_up_date,
            current_reminder_label,
            current_receipt_path,
            current_reference,
            customer_name,
            customer_company,
            customer_contact,
        ) = row
        current_status = clean_text(current_status) or "pending"
        if current_status in {"paid", "rejected"} and not allow_locked:
            locked.append(quotation_id)
            continue
        status_value = clean_text(entry.get("status")) or current_status
        follow_up_status = clean_text(entry.get("follow_up_status")) or clean_text(
            current_follow_up_status
        )
        follow_up_notes = clean_text(entry.get("follow_up_notes")) or clean_text(
            current_follow_up_notes
        )
        follow_up_date = None
        if status_value != "paid":
            follow_up_date = to_iso_date(entry.get("follow_up_date") or current_follow_up_date)
        reminder_label = (
            "Payment marked as received; follow-up reminders disabled."
            if status_value == "paid"
            else clean_text(entry.get("reminder_label"))
            or clean_text(current_reminder_label)
        )
        receipt_path = clean_text(entry.get("payment_receipt_path")) or clean_text(
            current_receipt_path
        )
        if status_value == "paid" and not receipt_path and not allow_locked:
            locked.append(quotation_id)
            continue
        conn.execute(
            """
            UPDATE quotations
            SET status=?,
                follow_up_status=?,
                follow_up_notes=?,
                follow_up_date=?,
                reminder_label=?,
                payment_receipt_path=COALESCE(?, payment_receipt_path),
                updated_at=datetime('now')
            WHERE quotation_id=? AND deleted_at IS NULL
            """,
            (
                status_value,
                follow_up_status,
                follow_up_notes,
                follow_up_date,
                reminder_label,
                receipt_path,
                quotation_id,
            ),
        )
        updated.append(quotation_id)
        reference_label = clean_text(current_reference) or f"Quotation #{quotation_id}"
        if status_value != current_status:
            receipt_note = " with receipt" if receipt_path and status_value == "paid" else ""
            log_activity(
                conn,
                event_type="quotation_updated",
                description=f"{reference_label} marked as {status_value}{receipt_note}",
                entity_type="quotation",
                entity_id=quotation_id,
            )
        else:
            log_activity(
                conn,
                event_type="quotation_updated",
                description=f"{reference_label} updated",
                entity_type="quotation",
                entity_id=quotation_id,
            )
        if status_value == "paid" and status_value != current_status:
            _promote_lead_customer(
                conn,
                name=customer_name,
                company=customer_company,
                phone=customer_contact,
            )
    conn.commit()
    return {"updated": updated, "locked": locked}


def _render_quotation_section(conn, *, render_id: Optional[int] = None):
    default_date = datetime.now().date()
    result_key = "quotation_result"
    feedback = st.session_state.pop("quotation_feedback", None)
    if feedback:
        level, message = feedback
        if level == "success":
            st.success(message)
        elif level == "info":
            st.info(message)
        elif level == "warning":
            st.warning(message)
        else:
            st.write(message)

    st.session_state.setdefault("quotation_item_rows", _default_quotation_items())

    user = get_current_user()
    salesperson_profile = {
        "name": clean_text(user.get("username")) or "",
        "title": clean_text(user.get("title")) or "",
        "phone": clean_text(user.get("phone")) or "",
        "email": clean_text(user.get("email")) or "",
    }

    uid = current_user_id()
    if uid is not None:
        try:
            profile_df = df_query(
                conn,
                "SELECT username, phone, email, title FROM users WHERE user_id=?",
                (uid,),
            )
            if not profile_df.empty:
                row = profile_df.iloc[0]
                salesperson_profile = {
                    "name": clean_text(row.get("username")) or salesperson_profile["name"],
                    "title": clean_text(row.get("title")) or salesperson_profile["title"],
                    "phone": clean_text(row.get("phone")) or salesperson_profile["phone"],
                    "email": clean_text(row.get("email")) or salesperson_profile["email"],
                }
        except Exception:
            pass

    salesperson_seed = salesperson_profile["name"]
    customer_df = df_query(
        conn,
        """
        SELECT customer_id, name, company_name, address, delivery_address, phone, COALESCE(delivery_address, address) AS district
        FROM customers
        ORDER BY LOWER(COALESCE(name, company_name, phone, 'customer'))
        LIMIT 200
        """,
    )
    autofill_options = [None]
    autofill_labels = {None: "Manual entry"}
    autofill_records: dict[int, dict[str, object]] = {}
    if not customer_df.empty:
        for _, row in customer_df.iterrows():
            try:
                cid = int(row.get("customer_id"))
            except Exception:
                continue
            label_parts = [clean_text(row.get("name")) or clean_text(row.get("company_name"))]
            phone_val = clean_text(row.get("phone"))
            if phone_val:
                label_parts.append(phone_val)
            autofill_options.append(cid)
            autofill_labels[cid] = " • ".join(part for part in label_parts if part)
            autofill_records[cid] = row.to_dict()

    follow_up_presets = {
        "In 3 days": 3,
        "In 1 week": 7,
        "In 2 weeks": 14,
        "Custom date": None,
    }

    st.session_state["quotation_autofill_customer"] = None
    st.markdown("### Upload to auto-fill")
    st.caption(
        "Upload a quotation (PDF, DOCX, or TXT) to detect customer info, contact details, and line items automatically."
    )
    prefill_upload = st.file_uploader(
        "Quotation file",
        type=["pdf", "doc", "docx", "txt"],
        key="quotation_prefill_upload",
    )

    if prefill_upload:
        prefill_token = f"{prefill_upload.name}:{prefill_upload.size}"
        if st.session_state.get("quotation_prefill_token") != prefill_token:
            text, warnings = _extract_text_from_quotation_upload(prefill_upload)
            for warning in warnings:
                st.warning(warning)
            saved_prefill = save_uploaded_file(
                prefill_upload,
                QUOTATION_DOCS_DIR,
                filename=prefill_upload.name or "quotation_upload",
                allowed_extensions={".pdf", ".doc", ".docx", ".txt"},
                default_extension=".pdf",
            )
            if saved_prefill:
                try:
                    st.session_state["quotation_document_path"] = str(
                        saved_prefill.relative_to(BASE_DIR)
                    )
                except ValueError:
                    st.session_state["quotation_document_path"] = str(saved_prefill)
            updates = _extract_quotation_metadata(text)
            detected_items = updates.pop("_detected_items", None)
            applied_updates = False
            if updates:
                for key, value in updates.items():
                    if _is_blank_field(st.session_state.get(key)):
                        st.session_state[key] = value
                        applied_updates = True
            if detected_items and not _has_quotation_items(
                st.session_state.get("quotation_item_rows")
            ):
                st.session_state["quotation_item_rows"] = detected_items
                applied_updates = True
            if applied_updates:
                st.success("Quotation fields auto-filled from the uploaded file.")
            st.session_state["quotation_prefill_token"] = prefill_token

    with st.form("quotation_form"):
        st.markdown("### Quotation details")
        basic_cols = st.columns((1, 1))
        with basic_cols[0]:
            quotation_date = st.date_input(
                "Quotation date",
                value=st.session_state.get("quotation_date") or default_date,
                key="quotation_date",
            )
            customer_company = st.text_input(
                "Customer name",
                value=st.session_state.get("quotation_company_name", ""),
                key="quotation_company_name",
            )
            customer_contact_name = st.text_input(
                "Attention / contact person",
                value=st.session_state.get("quotation_customer_contact_name", ""),
                key="quotation_customer_contact_name",
            )
        with basic_cols[1]:
            reference_value = st.text_input(
                "Reference / Quotation #",
                value=st.session_state.get("quotation_reference", ""),
                key="quotation_reference",
            )
            subject_line = st.text_input(
                "Product or project subject",
                value=st.session_state.get("quotation_subject", ""),
                key="quotation_subject",
            )
            customer_contact = st.text_input(
                "Customer contact details (phone / email)",
                value=st.session_state.get("quotation_customer_contact", ""),
                key="quotation_customer_contact",
            )

        customer_address = st.text_area(
            "Delivery location / address",
            value=st.session_state.get("quotation_customer_address", ""),
            key="quotation_customer_address",
        )

        quote_type = st.selectbox(
            "Quotation type",
            ["Retail", "Wholesale"],
            index=0
            if clean_text(st.session_state.get("quotation_quote_type")) != "Wholesale"
            else 1,
            key="quotation_quote_type",
        )

        st.markdown("### Product / service details")
        item_rows = st.session_state.get("quotation_item_rows") or _default_quotation_items()
        items_df = pd.DataFrame(item_rows)
        for column in ["description", "quantity", "rate", "total_price"]:
            if column not in items_df.columns:
                default_value = 0.0 if column != "description" else ""
                items_df[column] = default_value
        items_df = items_df[["description", "quantity", "rate", "total_price"]]
        edited_df = st.data_editor(
            items_df.drop(columns=["total_price"], errors="ignore"),
            hide_index=True,
            num_rows="dynamic",
            use_container_width=True,
            key="quotation_items_editor",
            column_config={
                "description": st.column_config.TextColumn(
                    "Description of Generator", required=True
                ),
                "quantity": st.column_config.NumberColumn(
                    "Qty.", min_value=0.0, step=1.0, format="%d"
                ),
                "rate": st.column_config.NumberColumn(
                    "Unit Price, Tk.", min_value=0.0, step=100.0, format="%.2f"
                ),
            },
        )
        if isinstance(edited_df, pd.DataFrame):
            st.session_state["quotation_item_rows"] = (
                edited_df.fillna("").to_dict("records")
            )

        manual_total_value = st.number_input(
            "Manual total amount (Tk.)",
            min_value=0.0,
            step=1000.0,
            format="%.2f",
            value=_coerce_float(st.session_state.get("quotation_manual_total"), 0.0),
            key="quotation_manual_total",
            help="Optional override if the quoted amount differs from the calculated line totals.",
        )

        terms_notes = st.text_area(
            "Special notes / terms & conditions",
            value=st.session_state.get("quotation_terms", ""),
            key="quotation_terms",
        )

        st.markdown("#### Admin follow-up")
        follow_cols = st.columns((1, 1))
        follow_statuses = ["Pending", "Hot", "Possible", "Closed"]
        saved_follow_status = clean_text(st.session_state.get("quotation_follow_up_status"))
        follow_status_default = saved_follow_status.title() if saved_follow_status else "Pending"
        follow_status_default = (
            follow_status_default if follow_status_default in follow_statuses else "Pending"
        )
        with follow_cols[0]:
            follow_up_status = st.selectbox(
                "Follow-up status",
                follow_statuses,
                index=follow_statuses.index(follow_status_default),
                key="quotation_follow_up_status",
                help="Visible to admins for tracking next steps.",
            )
        with follow_cols[1]:
            follow_up_choice = st.selectbox(
                "Reminder preset",
                list(follow_up_presets.keys()),
                index=1,
                key="quotation_follow_up_choice",
            )
            custom_follow_up = follow_up_choice == "Custom date"
            if custom_follow_up and not st.session_state.get(
                "quotation_follow_up_date_toggle"
            ):
                st.session_state["quotation_follow_up_date_toggle"] = True
            enable_follow_date = st.checkbox(
                "Set follow-up date",
                value=bool(st.session_state.get("quotation_follow_up_date"))
                or custom_follow_up,
                key="quotation_follow_up_date_toggle",
                disabled=custom_follow_up,
            )
            follow_up_date_value = None
            if enable_follow_date:
                follow_up_date_value = st.date_input(
                    "Next follow-up date",
                    value=st.session_state.get("quotation_follow_up_date")
                    or datetime.now().date(),
                    key="quotation_follow_up_date",
                )
        follow_up_notes = st.text_area(
            "Follow-up remarks for admins",
            value=st.session_state.get("quotation_follow_up_notes", ""),
            key="quotation_follow_up_notes",
            help="Internal notes that help admins continue the conversation.",
        )

        form_actions = st.columns((1, 1))
        submit = form_actions[0].form_submit_button("Save quotation", type="primary")
        reset = form_actions[1].form_submit_button("Reset form")

    template_choice = "Default letterhead"
    quote_type = st.session_state.get("quotation_quote_type", "Retail")
    default_discount = _coerce_float(
        st.session_state.get("quotation_discount_default"), 0.0
    )
    customer_district = st.session_state.get("quotation_customer_district", "")
    attention_title = st.session_state.get("quotation_attention_title", "")
    admin_notes = terms_notes
    follow_up_choice = st.session_state.get("quotation_follow_up_choice")
    salesperson_title = salesperson_profile.get("title", "")
    salesperson_contact = salesperson_profile.get("phone", "")
    salesperson_email = salesperson_profile.get("email", "")
    prepared_by = (
        st.session_state.get("quotation_prepared_by") or salesperson_profile.get("name", "")
    )
    salutation = st.session_state.get("quotation_salutation", "")
    intro_text = st.session_state.get("quotation_introduction", "")
    closing_text = st.session_state.get("quotation_closing", "")
    attention_name = clean_text(customer_contact_name)
    status_value = "pending"
    if reset:
        _reset_quotation_form_state()
        st.session_state["quotation_feedback"] = (
            "info",
            "Quotation form reset to defaults.",
        )
        _safe_rerun()

    if _guard_double_submit("quotation_form_save", submit):
        prepared_items = [dict(item) for item in st.session_state.get("quotation_item_rows", [])]
        for item in prepared_items:
            if item.get("discount") in (None, ""):
                item["discount"] = default_discount

        items_clean, totals_data = normalize_quotation_items(prepared_items)

        reminder_days = follow_up_presets.get(follow_up_choice)
        follow_up_date = follow_up_date_value
        if reminder_days is not None:
            follow_up_date = quotation_date + timedelta(days=reminder_days)
        follow_up_iso = to_iso_date(follow_up_date) if follow_up_date else None
        follow_up_label = format_period_range(follow_up_iso, follow_up_iso) if follow_up_iso else ""
        reminder_label = None
        if follow_up_label and reminder_days is not None:
            reminder_label = f"Reminder scheduled in {reminder_days} days on {follow_up_label}."
        manual_total_override = max(
            _coerce_float(manual_total_value, 0.0),
            0.0,
        )
        grand_total_value = (
            manual_total_override if manual_total_override > 0 else totals_data["grand_total"]
        )

        customer_contact_combined = dedupe_join(
            [st.session_state.get("quotation_customer_contact"), attention_name], " / "
        )

        metadata = OrderedDict()
        metadata["Reference number"] = reference_value
        metadata["Date"] = quotation_date.strftime(DATE_FMT)
        metadata["Customer contact name"] = customer_contact_name
        metadata["Customer company"] = customer_company
        metadata["Customer address"] = customer_address
        metadata["Customer contact"] = customer_contact_combined
        metadata["Attention name"] = attention_name
        metadata["Quote type"] = quote_type
        metadata["Notes / terms"] = terms_notes
        metadata["Salesperson name"] = prepared_by
        metadata["Salesperson title"] = salesperson_title
        metadata["Salesperson contact"] = salesperson_contact
        metadata["Salesperson email"] = salesperson_email
        metadata["Total amount (BDT)"] = grand_total_value
        if manual_total_override > 0:
            metadata["Manual total override (BDT)"] = manual_total_override

        totals_rows = [("Gross amount", totals_data["gross_total"])]
        if totals_data["discount_total"]:
            totals_rows.append(("Discount total", totals_data["discount_total"]))
        if manual_total_override > 0 and not math.isclose(
            manual_total_override, totals_data["grand_total"]
        ):
            totals_rows.append(("Manual total override", manual_total_override))
        totals_rows.append(("Grand total", grand_total_value))

        workbook_items = [item.copy() for item in items_clean]
        workbook_bytes = _build_quotation_workbook(
            metadata=metadata,
            items=workbook_items,
            totals=totals_rows,
        )

        display_df = pd.DataFrame(workbook_items)

        column_order = [
            col
            for col in [
                "Sl No.",
                "Description of Generator",
                "Qty.",
                "Unit Price, Tk.",
                "Total Price, Tk.",
                "Notes",
            ]
            if col in display_df.columns
        ]
        if column_order:
            display_df = display_df[column_order]

        def _format_quantity_display(value: object) -> str:
            amount = _coerce_float(value, 0.0)
            if math.isclose(amount, round(amount)):
                return f"{int(round(amount))}"
            return f"{amount:,.2f}"

        money_columns = ["Unit Price, Tk.", "Total Price, Tk."]
        for col in money_columns:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(
                    lambda value: format_money(value) or f"{_coerce_float(value, 0.0):,.2f}"
                )
        for qty_col in ["Qty.", "Quantity"]:
            if qty_col in display_df.columns:
                display_df[qty_col] = display_df[qty_col].apply(_format_quantity_display)
        display_df = display_df.fillna("")

        base_filename = clean_text(reference_value) or f"quotation_{quotation_date.strftime('%Y%m%d')}"
        safe_name = _sanitize_path_component(base_filename)
        if not safe_name:
            safe_name = f"quotation_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        filename = f"{safe_name}.xlsx"

        receipt_path = None
        payload = {
            "reference": reference_value,
            "quote_date": quotation_date.isoformat(),
            "customer_name": customer_contact_name,
            "customer_company": customer_company,
            "customer_address": customer_address,
            "customer_district": customer_district,
            "customer_contact": customer_contact_combined,
            "attention_name": attention_name,
            "attention_title": attention_title,
            "subject": subject_line,
            "salutation": salutation,
            "introduction": intro_text,
            "closing": closing_text,
            "quote_type": quote_type,
            "total_amount": grand_total_value,
            "discount_pct": default_discount,
            "status": status_value,
            "payment_receipt_path": receipt_path,
            "follow_up_status": follow_up_status,
            "follow_up_notes": follow_up_notes,
            "follow_up_date": follow_up_iso,
            "reminder_label": reminder_label,
            "letter_template": template_choice,
            "salesperson_name": prepared_by,
            "salesperson_title": salesperson_title,
            "salesperson_contact": salesperson_contact,
            "salesperson_email": salesperson_email,
            "document_path": st.session_state.get("quotation_document_path"),
            "items_payload": json.dumps(items_clean, ensure_ascii=False),
            "remarks_internal": terms_notes,
            "created_by": current_user_id(),
        }
        record_id = _save_quotation_record(conn, payload)
        if clean_text(customer_company):
            lead_status = LEAD_REMARK_TAG if status_value != "paid" else None
            _upsert_customer_from_manual_quotation(
                conn,
                name=customer_contact_name,
                company=customer_company,
                phone=customer_contact,
                address=customer_address,
                district=customer_district,
                reference=reference_value,
                created_by=current_user_id(),
                lead_status=lead_status,
            )
            if status_value == "paid":
                _promote_lead_customer(
                    conn,
                    name=customer_contact_name,
                    company=customer_company,
                    phone=customer_contact,
                )

        st.session_state[result_key] = {
            "display": display_df,
            "metadata_items": list(metadata.items()),
            "totals_rows": totals_rows,
            "grand_total": grand_total_value,
            "metadata": metadata,
            "excel_bytes": workbook_bytes,
            "filename": filename,
            "record_id": record_id,
            "reminder_label": reminder_label,
            "letter_template": template_choice,
        }

    result = st.session_state.get(result_key)
    if result:
        st.success("Quotation ready. Review the details below or download the Excel file.")
        metadata_df = pd.DataFrame(result["metadata_items"], columns=["Field", "Value"])
        st.table(metadata_df)

        st.dataframe(result["display"], use_container_width=True)

        totals_rows = result.get("totals_rows", [])
        if totals_rows:
            totals_df = pd.DataFrame(totals_rows, columns=["Label", "Amount"])
            totals_df["Amount"] = totals_df["Amount"].apply(
                lambda value: format_money(value) or f"{_coerce_float(value, 0.0):,.2f}"
            )
            st.table(totals_df)

        grand_total_label = format_money(result["grand_total"]) or f"{result['grand_total']:,.2f}"
        st.markdown(f"**Grand total:** {grand_total_label}")


def _render_quotation_management(conn):
    st.markdown("### Quotation tracker")
    is_admin = current_user_is_admin()
    scope_clause, scope_params = _quotation_scope_filter()
    quotes_df = df_query(
        conn,
        dedent(
            f"""
            SELECT q.quotation_id, q.reference, q.quote_date, q.customer_company, q.customer_name, q.customer_contact,
                   q.total_amount, q.status, q.follow_up_status, q.follow_up_notes, q.follow_up_date,
                   q.reminder_label, q.payment_receipt_path, q.items_payload, q.document_path,
                   q.created_by, COALESCE(u.username, '(user)') AS created_by_name
            FROM quotations q
            LEFT JOIN users u ON u.user_id = q.created_by
            {scope_clause}
            ORDER BY datetime(q.quote_date) DESC, q.quotation_id DESC
            LIMIT 50
            """
        ),
        scope_params,
    )

    if quotes_df.empty:
        st.info("No quotations recorded yet. Create a quotation above to start tracking.")
        return

    upload_records = []
    for _, row in quotes_df.iterrows():
        document_path = clean_text(row.get("document_path"))
        if not document_path:
            continue
        upload_records.append(
            {
                "quotation_id": row.get("quotation_id"),
                "reference": clean_text(row.get("reference")) or "Quotation",
                "customer": clean_text(row.get("customer_company"))
                or clean_text(row.get("customer_name"))
                or clean_text(row.get("customer_contact")),
                "document_path": document_path,
                "quote_date": row.get("quote_date"),
                "created_by_name": clean_text(row.get("created_by_name")) or "(user)",
            }
        )
    if upload_records:
        st.markdown("#### Uploaded quotation files")
        for record in upload_records:
            path = resolve_upload_path(record.get("document_path"))
            label_parts = [
                record.get("reference"),
                record.get("customer"),
            ]
            created_by_label = record.get("created_by_name")
            if created_by_label:
                label_parts.append(created_by_label)
            label = " • ".join(part for part in label_parts if part)
            if path and path.exists():
                st.download_button(
                    label,
                    data=path.read_bytes(),
                    file_name=path.name,
                    key=f"quotation_upload_{record.get('quotation_id')}",
                )
            else:
                st.caption(f"{label} (file missing)")
        st.markdown("---")

    def _extract_product_summary(items_payload: object) -> str:
        if items_payload in (None, "", "nan", "NaT"):
            return ""
        payload_text = items_payload
        if not isinstance(payload_text, str):
            try:
                payload_text = json.dumps(payload_text)
            except (TypeError, ValueError):
                return ""
        try:
            items = json.loads(payload_text)
        except json.JSONDecodeError:
            return ""
        if not isinstance(items, list):
            return ""
        names: list[str] = []
        for item in items:
            if not isinstance(item, dict):
                continue
            description = clean_text(
                item.get("Description of Generator")
                or item.get("Description")
                or item.get("Product")
                or item.get("Item")
                or item.get("Name")
            )
            if description:
                names.append(description)
        if not names:
            return ""
        seen: set[str] = set()
        deduped: list[str] = []
        for name in names:
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            deduped.append(name)
        return ", ".join(deduped)

    def _as_editable_date(value: object) -> Optional[date]:
        if value in (None, "", "nan", "NaT"):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
        except Exception:
            return None
        if isinstance(parsed, pd.DatetimeIndex):
            parsed = parsed[0] if len(parsed) else None
        if parsed is None or pd.isna(parsed):
            return None
        if isinstance(parsed, pd.Timestamp):
            parsed = parsed.to_pydatetime()
        if isinstance(parsed, datetime):
            return parsed.date()
        return None

    quotes_df = quotes_df.copy()
    quotes_df["follow_up_date"] = quotes_df.get("follow_up_date", pd.Series(dtype=object)).apply(
        _as_editable_date
    )
    quotes_df["products"] = quotes_df.get("items_payload", pd.Series(dtype=object)).apply(
        _extract_product_summary
    )
    quotes_df = fmt_dates(quotes_df, ["quote_date"])
    quotes_df["total_amount"] = quotes_df["total_amount"].apply(
        lambda value: format_money(value) or f"{_coerce_float(value, 0.0):,.2f}"
    )

    tracker_source = quotes_df.drop(
        columns=["items_payload", "follow_up_status", "status", "payment_receipt_path"],
        errors="ignore",
    )
    if "products" in tracker_source.columns and "customer_company" in tracker_source.columns:
        columns = list(tracker_source.columns)
        columns.remove("products")
        insert_at = columns.index("customer_company") + 1
        columns.insert(insert_at, "products")
        tracker_source = tracker_source[columns]
    editable_df = tracker_source.copy()

    tracker_state_key = "quotation_tracker_rows"

    def _normalize_tracker_rows(rows: Iterable[dict[str, object]]):
        normalized: list[dict[str, object]] = []
        for row in rows:
            normalized_row: dict[str, object] = {}
            for col in editable_df.columns:
                value = row.get(col)
                if col == "follow_up_date":
                    value = _as_editable_date(value)
                normalized_row[col] = value
            normalized.append(normalized_row)
        return normalized

    current_records = _normalize_tracker_rows(editable_df.to_dict("records"))
    tracker_state: list[dict[str, object]] = st.session_state.get(tracker_state_key, [])
    current_ids = {row.get("quotation_id") for row in current_records}
    state_ids = {row.get("quotation_id") for row in tracker_state}
    if (
        not tracker_state
        or current_ids != state_ids
        or len(tracker_state) != len(current_records)
    ):
        tracker_state = current_records
        st.session_state[tracker_state_key] = tracker_state

    tracker_df = pd.DataFrame(tracker_state)
    tracker_df = tracker_df[[col for col in editable_df.columns if col in tracker_df.columns]]
    tracker_df = tracker_df.reindex(columns=editable_df.columns)

    edit_config = {
        "follow_up_notes": st.column_config.TextColumn("Follow-up notes"),
        "follow_up_date": st.column_config.DateColumn(
            "Follow-up date", format="DD-MM-YYYY"
        ),
        "reminder_label": st.column_config.TextColumn("Reminder"),
        "reference": st.column_config.TextColumn("Reference"),
        "customer_company": st.column_config.TextColumn("Customer"),
        "products": st.column_config.TextColumn("Products", disabled=True),
        "customer_name": st.column_config.TextColumn("Contact name"),
        "customer_contact": st.column_config.TextColumn("Contact"),
        "total_amount": st.column_config.TextColumn("Total amount (BDT)"),
        "quote_date": st.column_config.TextColumn("Quote date"),
        "created_by_name": st.column_config.TextColumn("Created by", disabled=True),
    }

    st.caption(
        "Edit follow-up notes and dates directly in the table, then press **Save quotation updates** to persist your changes."
    )

    edited_records: list[dict[str, object]] = []
    if not editable_df.empty:
        edited = st.data_editor(
            tracker_df,
            hide_index=True,
            use_container_width=True,
            column_config=edit_config,
            key="quotation_tracker_editor",
        )
        if isinstance(edited, pd.DataFrame):
            edited_records = _normalize_tracker_rows(edited.to_dict("records"))
            st.session_state[tracker_state_key] = edited_records
    else:
        st.info("No pending quotations to update.")

    save_disabled = editable_df.empty or not edited_records
    save_tracker = st.button(
        "Save quotation updates",
        key="quotation_tracker_save",
        disabled=save_disabled,
    )
    if _guard_double_submit("quotation_tracker_save", save_tracker) and edited_records:
        sanitized_records: list[dict[str, object]] = []
        for record in edited_records:
            sanitized_records.append(dict(record))
        result = _update_quotation_records(
            conn,
            sanitized_records,
            allow_locked=is_admin,
        )
        updated_count = len(result.get("updated", []))
        locked_count = len(result.get("locked", []))
        if updated_count:
            st.success(f"Updated {updated_count} quotation(s).")
            st.toast("Quotation tracker updated", icon="✅")
        if locked_count:
            st.info("Some quotations are locked because they are marked as paid or rejected.")

    st.markdown("#### Update saved quotation details")
    detail_labels: dict[int, str] = {}
    for _, row in quotes_df.iterrows():
        try:
            detail_id = int(row.get("quotation_id"))
        except Exception:
            continue
        reference = clean_text(row.get("reference")) or f"Quotation #{detail_id}"
        customer = clean_text(row.get("customer_company")) or clean_text(
            row.get("customer_name")
        ) or clean_text(row.get("customer_contact"))
        detail_labels[detail_id] = " • ".join(
            part for part in [reference, customer] if part
        )

    detail_choices = list(detail_labels.keys())
    if not detail_choices:
        st.info("No quotations available to edit.")
        return

    selected_detail_id = st.selectbox(
        "Select a quotation",
        detail_choices,
        format_func=lambda val: detail_labels.get(val, f"Quotation #{val}"),
        key="quotation_detail_select",
    )

    selected_row = quotes_df[
        quotes_df["quotation_id"] == selected_detail_id
    ].iloc[0]
    selected_status = clean_text(selected_row.get("status")).lower() or "due"
    existing_receipt = clean_text(selected_row.get("payment_receipt_path"))
    follow_up_status_value = clean_text(selected_row.get("follow_up_status")) or ""
    follow_up_notes_value = clean_text(selected_row.get("follow_up_notes")) or ""
    follow_up_date_seed = selected_row.get("follow_up_date")
    if isinstance(follow_up_date_seed, pd.Timestamp):
        follow_up_date_seed = follow_up_date_seed.date()

    items_payload = clean_text(selected_row.get("items_payload"))
    quotation_items = []
    if items_payload:
        try:
            quotation_items = json.loads(items_payload)
        except json.JSONDecodeError:
            quotation_items = []
    if quotation_items:
        st.markdown("#### Product details")
        items_df = pd.DataFrame(quotation_items)
        display_columns = [
            col
            for col in [
                "Description of Generator",
                "Qty.",
                "Unit Price, Tk.",
                "Total Price, Tk.",
            ]
            if col in items_df.columns
        ]
        if display_columns:
            items_df = items_df[display_columns]
        st.dataframe(items_df, use_container_width=True, hide_index=True)

    col_left, col_right = st.columns(2)
    with col_left:
        follow_up_status_input = st.text_input(
            "Follow-up status",
            value=follow_up_status_value,
            key="quotation_detail_follow_up_status",
        )
        clear_follow_up_date = st.checkbox(
            "No follow-up date",
            value=follow_up_date_seed is None,
            key="quotation_detail_clear_date",
        )
        follow_up_date_input: Optional[date] = None
        if not clear_follow_up_date:
            follow_up_date_input = st.date_input(
                "Follow-up date",
                value=follow_up_date_seed or date.today(),
                key="quotation_detail_follow_up_date",
                format="DD-MM-YYYY",
            )
    with col_right:
        follow_up_notes_input = st.text_area(
            "Follow-up notes",
            value=follow_up_notes_value,
            key="quotation_detail_follow_up_notes",
        )
        receipt_upload = None
        if selected_status == "paid":
            receipt_upload = st.file_uploader(
                "Attach receipt for this paid quotation",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"quotation_detail_receipt_{selected_detail_id}",
            )
            if not (receipt_upload or existing_receipt):
                st.caption("Upload a receipt to keep this paid quotation locked.")

    if st.button("Update quotation details", type="primary", key="quotation_detail_save"):
        receipt_path = existing_receipt
        if selected_status == "paid":
            if receipt_upload:
                safe_ref = _sanitize_path_component(
                    clean_text(selected_row.get("reference"))
                    or f"quotation_{selected_detail_id}"
                )
                receipt_path = store_payment_receipt(
                    receipt_upload, identifier=f"{safe_ref}_receipt"
                )
            if not receipt_path:
                st.error("Upload a receipt before saving a paid quotation.")
                return

        follow_up_iso = to_iso_date(follow_up_date_input) if follow_up_date_input else None
        reminder_label = (
            format_period_range(follow_up_iso, follow_up_iso)
            if follow_up_iso
            else clean_text(selected_row.get("reminder_label"))
        )

        conn.execute(
            """
             UPDATE quotations
               SET follow_up_status=?,
                   follow_up_notes=?,
                   follow_up_date=?,
                   reminder_label=?,
                   payment_receipt_path=COALESCE(?, payment_receipt_path),
                   updated_at=datetime('now')
             WHERE quotation_id=? AND deleted_at IS NULL
            """,
            (
                clean_text(follow_up_status_input) or None,
                clean_text(follow_up_notes_input) or None,
                follow_up_iso,
                reminder_label,
                receipt_path,
                selected_detail_id,
            ),
        )
        conn.commit()
        log_activity(
            conn,
            event_type="quotation_updated",
            description=f"Quotation #{selected_detail_id} details updated",
            entity_type="quotation",
            entity_id=int(selected_detail_id),
            user_id=current_user_id(),
        )
        st.success("Quotation details updated.")
        _safe_rerun()

    st.markdown("#### Delete quotation")
    current_actor = current_user_id()
    deletable_df = quotes_df.copy()
    if not is_admin and current_actor is not None:
        deletable_df = deletable_df[
            deletable_df["created_by"].apply(lambda val: int(_coerce_float(val, -1)) == current_actor)
        ]
    if deletable_df.empty:
        st.caption("No quotations available for deletion.")
        return
    delete_labels: dict[int, str] = {}
    for _, row in deletable_df.iterrows():
        try:
            quote_id = int(row.get("quotation_id"))
        except Exception:
            continue
        reference = clean_text(row.get("reference")) or f"Quotation #{quote_id}"
        customer = clean_text(row.get("customer_company")) or clean_text(
            row.get("customer_name")
        ) or clean_text(row.get("customer_contact"))
        delete_labels[quote_id] = " • ".join(part for part in [reference, customer] if part)
    delete_options = list(delete_labels.keys())
    selected_delete_id = st.selectbox(
        "Select a quotation to delete",
        delete_options,
        format_func=lambda val: delete_labels.get(val, f"Quotation #{val}"),
        key="quotation_delete_select",
    )
    confirm_delete = st.checkbox(
        "I understand this will remove the quotation from active views.",
        key="quotation_delete_confirm",
    )
    if st.button(
        "Delete quotation",
        type="secondary",
        disabled=not confirm_delete,
        key="quotation_delete_button",
    ):
        conn.execute(
            "UPDATE quotations SET deleted_at=datetime('now'), deleted_by=? WHERE quotation_id=?",
            (current_actor, selected_delete_id),
        )
        conn.commit()
        description = f"Quotation {delete_labels.get(selected_delete_id, selected_delete_id)} deleted"
        log_activity(
            conn,
            event_type="quotation_deleted",
            description=description,
            entity_type="quotation",
            entity_id=int(selected_delete_id),
            user_id=current_actor,
        )
        st.warning("Quotation deleted.")
        _safe_rerun()

    if is_admin:
        with st.expander("Quotation activity log", expanded=False):
            activity_df = _fetch_entity_activity(conn, ["quotation"], limit=50)
            if activity_df.empty:
                st.caption("No quotation activity recorded yet.")
            else:
                activity_df = fmt_dates(activity_df, ["created_at"])
                activity_df = activity_df.rename(
                    columns={
                        "created_at": "When",
                        "actor": "Staff",
                        "event_type": "Event",
                        "description": "Details",
                    }
                )
                st.dataframe(
                    activity_df[["When", "Staff", "Event", "Details"]],
                    use_container_width=True,
                    hide_index=True,
                )


def advanced_search_page(conn):
    st.subheader("🔎 Advanced Search")
    if not current_user_is_admin():
        st.warning("Advanced filters are available to admins only.")
        return

    search_text = st.text_input(
        "Keyword search",
        key="advanced_search_keyword",
        help="Search across quotations, customers, delivery orders, services, and maintenance logs.",
    )
    date_window = st.date_input(
        "Date window",
        value=(date.today() - timedelta(days=30), date.today()),
        help="Filter results by creation or activity date.",
        key="advanced_search_dates",
    )
    min_amount = st.number_input(
        "Minimum amount (for quotations)",
        min_value=0.0,
        step=100.0,
        format="%.2f",
        key="advanced_search_min_amount",
    )
    record_types = [
        "Quotations",
        "Services",
        "Maintenance",
        "Delivery orders",
        "Customers",
    ]
    selected_types = st.multiselect(
        "Record types",
        record_types,
        default=record_types,
        key="advanced_search_types",
    )

    staff_df = df_query(conn, "SELECT user_id, username FROM users ORDER BY LOWER(username)")
    staff_map = {
        int(row["user_id"]): clean_text(row.get("username")) or f"User #{int(row['user_id'])}"
        for _, row in staff_df.iterrows()
    }
    staff_choices = list(staff_map.keys())
    staff_filter = st.multiselect(
        "Staff filter",
        staff_choices,
        format_func=lambda uid: staff_map.get(uid, f"User #{uid}"),
        key="advanced_search_staff",
    )

    start_iso = end_iso = None
    if isinstance(date_window, (list, tuple)) and len(date_window) == 2:
        start_iso = to_iso_date(date_window[0])
        end_iso = to_iso_date(date_window[1])

    results: list[dict[str, object]] = []

    def _apply_date_filter(df: pd.DataFrame, column: str) -> pd.DataFrame:
        if df.empty or column not in df.columns or not start_iso or not end_iso:
            return df
        df[column] = pd.to_datetime(df[column], errors="coerce")
        start_dt = pd.to_datetime(start_iso)
        end_dt = pd.to_datetime(end_iso) + pd.Timedelta(days=1)
        return df[(df[column] >= start_dt) & (df[column] < end_dt)]

    def _append_results(df: pd.DataFrame, type_label: str, date_col: str, build_details):
        if df.empty:
            return
        filtered = _apply_date_filter(df, date_col)
        for row in filtered.to_dict("records"):
            staff_raw = row.get("created_by")
            staff_id = (
                int(staff_raw)
                if staff_raw is not None and not pd.isna(staff_raw)
                else None
            )
            if staff_filter and staff_id not in staff_filter:
                continue
            details = build_details(row)
            search_blob = " ".join(str(val) for val in details.values() if val)
            results.append(
                {
                    "Type": type_label,
                    "Title": details.get("title"),
                    "Details": details.get("details"),
                    "Date": details.get("date"),
                    "Staff": staff_map.get(staff_id),
                    "Amount": details.get("amount"),
                    "Status": details.get("status"),
                    "Attachment": details.get("attachment"),
                    "_search": search_blob.lower(),
                }
            )

    if "Quotations" in selected_types:
        quotes_df = df_query(
            conn,
            dedent(
                """
                SELECT q.quotation_id,
                       q.reference,
                       q.customer_company,
                       q.quote_date,
                       q.total_amount,
                       q.status,
                       q.document_path,
                       q.created_by,
                       q.salesperson_name,
                       u.username
                FROM quotations q
                LEFT JOIN users u ON u.user_id = q.created_by
                WHERE q.deleted_at IS NULL
                ORDER BY datetime(q.quote_date) DESC, q.quotation_id DESC
                LIMIT 200
                """
            ),
        )
        if not quotes_df.empty:
            if min_amount > 0:
                quotes_df = quotes_df[quotes_df["total_amount"].apply(lambda v: _coerce_float(v, 0.0) >= min_amount)]
            _append_results(
                quotes_df,
                "Quotation",
                "quote_date",
                lambda row: {
                    "title": clean_text(row.get("reference")) or clean_text(row.get("customer_company")),
                    "details": f"{clean_text(row.get('customer_company')) or '(customer)'} • {clean_text(row.get('salesperson_name')) or 'Sales'}",
                    "date": row.get("quote_date"),
                    "amount": _coerce_float(row.get("total_amount"), 0.0),
                    "status": clean_text(row.get("status")),
                    "attachment": clean_text(row.get("document_path")),
                },
            )

    if "Services" in selected_types:
        service_df = df_query(
            conn,
            """
            SELECT service_id, description, service_start_date, service_end_date, service_product_info, status
            FROM services
            WHERE deleted_at IS NULL
            ORDER BY datetime(COALESCE(service_start_date, service_end_date)) DESC, service_id DESC
            LIMIT 200
            """,
        )
        _append_results(
            service_df,
            "Service",
            "service_start_date",
            lambda row: {
                "title": clean_text(row.get("description")) or f"Service #{row.get('service_id')}",
                "details": clean_text(row.get("service_product_info")),
                "date": row.get("service_start_date") or row.get("service_end_date"),
                "status": clean_text(row.get("status")),
                "amount": None,
                "attachment": None,
            },
        )

    if "Maintenance" in selected_types:
        maintenance_df = df_query(
            conn,
            """
            SELECT maintenance_id, description, maintenance_start_date, maintenance_end_date, maintenance_product_info, status
            FROM maintenance_records
            WHERE deleted_at IS NULL
            ORDER BY datetime(COALESCE(maintenance_start_date, maintenance_end_date)) DESC, maintenance_id DESC
            LIMIT 200
            """,
        )
        _append_results(
            maintenance_df,
            "Maintenance",
            "maintenance_start_date",
            lambda row: {
                "title": clean_text(row.get("description")) or f"Maintenance #{row.get('maintenance_id')}",
                "details": clean_text(row.get("maintenance_product_info")),
                "date": row.get("maintenance_start_date") or row.get("maintenance_end_date"),
                "status": clean_text(row.get("status")),
                "amount": None,
                "attachment": None,
            },
        )

    if "Delivery orders" in selected_types:
        do_df = df_query(
            conn,
            """
            SELECT do_number, description, sales_person, remarks, created_at, created_by, total_amount
            FROM delivery_orders
            WHERE COALESCE(record_type, 'delivery_order') = 'delivery_order'
              AND deleted_at IS NULL
            ORDER BY datetime(created_at) DESC
            LIMIT 200
            """,
        )
        _append_results(
            do_df,
            "Delivery order",
            "created_at",
            lambda row: {
                "title": clean_text(row.get("do_number")),
                "details": clean_text(row.get("description")) or clean_text(row.get("remarks")),
                "date": row.get("created_at"),
                "status": clean_text(row.get("sales_person")),
                "amount": format_money(row.get("total_amount")) or row.get("total_amount"),
                "attachment": None,
            },
        )

    if "Customers" in selected_types:
        customer_df = df_query(
            conn,
            """
            SELECT name, company_name, phone, address, created_at, created_by
            FROM customers
            ORDER BY datetime(created_at) DESC
            LIMIT 200
            """,
        )
        _append_results(
            customer_df,
            "Customer",
            "created_at",
            lambda row: {
                "title": clean_text(row.get("name")) or clean_text(row.get("company_name")),
                "details": clean_text(row.get("phone")) or clean_text(row.get("address")),
                "date": row.get("created_at"),
                "status": None,
                "amount": None,
                "attachment": None,
            },
        )

    if not results:
        st.info("No records found for the selected filters.")
    else:
        results_df = pd.DataFrame(results)
        if search_text:
            needle = re.escape(search_text.lower())
            results_df = results_df[results_df["_search"].str.contains(needle, regex=True, na=False)]

        display_cols = ["Type", "Title", "Details", "Date", "Staff", "Amount", "Status", "Attachment"]
        results_df = results_df.fillna("")
        st.dataframe(results_df[display_cols], use_container_width=True, hide_index=True)

        csv_bytes = results_df.drop(columns=["_search"], errors="ignore").to_csv(index=False).encode("utf-8")
        st.download_button(
            "Download results",
            data=csv_bytes,
            file_name="advanced_search.csv",
            mime="text/csv",
            key="advanced_search_download",
        )

    st.markdown("### Staff activity history")
    if not staff_choices:
        st.caption("No staff accounts available for activity history.")
        return
    history_staff = st.selectbox(
        "Team member",
        staff_choices,
        format_func=lambda uid: staff_map.get(uid, f"User #{uid}"),
        key="advanced_search_history_staff",
    )
    history_range = st.date_input(
        "History range",
        value=date_window,
        key="advanced_search_history_range",
    )
    history_start = history_end = None
    if isinstance(history_range, (list, tuple)) and len(history_range) == 2:
        history_start = to_iso_date(history_range[0])
        history_end = to_iso_date(history_range[1])
    elif history_range:
        history_start = to_iso_date(history_range)
        history_end = history_start
    history_filters = ["a.user_id = ?"]
    history_params: list[object] = [int(history_staff)]
    if history_start:
        history_filters.append("date(a.created_at) >= date(?)")
        history_params.append(history_start)
    if history_end:
        history_filters.append("date(a.created_at) <= date(?)")
        history_params.append(history_end)
    history_clause = " AND ".join(history_filters)
    history_df = df_query(
        conn,
        dedent(
            f"""
            SELECT a.created_at,
                   a.event_type,
                   a.entity_type,
                   a.description
            FROM activity_log a
            WHERE {history_clause}
            ORDER BY datetime(a.created_at) DESC, a.activity_id DESC
            LIMIT 250
            """
        ),
        tuple(history_params),
    )
    if history_df.empty:
        st.caption("No activity history found for the selected staff member.")
    else:
        history_df = fmt_dates(history_df, ["created_at"])
        history_df = history_df.rename(
            columns={
                "created_at": "When",
                "event_type": "Activity",
                "entity_type": "Type",
                "description": "Details",
            }
        )
        st.dataframe(
            history_df[["When", "Activity", "Type", "Details"]],
            use_container_width=True,
            hide_index=True,
        )


def _render_maintenance_section(conn, *, show_heading: bool = True):
    if show_heading:
        st.subheader("🔧 Maintenance Records")
    _, customer_label_map = build_customer_groups(conn, only_complete=False)
    customer_options, customer_labels, _, label_by_id = fetch_customer_choices(conn)
    viewer_id = current_user_id()
    do_df = df_query(
        conn,
        """
        SELECT d.do_number, d.customer_id, d.created_by, COALESCE(c.name, '(unknown)') AS customer_name, d.description, d.remarks, d.record_type
        FROM delivery_orders d
        LEFT JOIN customers c ON c.customer_id = d.customer_id
        WHERE COALESCE(d.record_type, 'delivery_order') = 'delivery_order'
          AND d.deleted_at IS NULL
        ORDER BY datetime(d.created_at) DESC
        """,
    )
    allowed_customers = accessible_customer_ids(conn)
    do_df = filter_delivery_orders_for_view(
        do_df, allowed_customers, record_types={"delivery_order"}
    )
    do_options = [None]
    do_labels = {None: "No delivery order (manual entry)"}
    do_customer_map = {}
    do_customer_name_map = {}
    for _, row in do_df.iterrows():
        do_num = clean_text(row.get("do_number"))
        if not do_num:
            continue
        cust_id = int(row["customer_id"]) if not pd.isna(row.get("customer_id")) else None
        summary = clean_text(row.get("description"))
        cust_name = customer_label_map.get(cust_id) if cust_id else clean_text(row.get("customer_name"))
        label_parts = [do_num]
        if cust_name:
            label_parts.append(f"({cust_name})")
        if summary:
            snippet = summary[:40]
            if len(summary) > 40:
                snippet += "…"
            label_parts.append(f"– {snippet}")
        label = " ".join(part for part in label_parts if part)
        do_options.append(do_num)
        do_labels[do_num] = label
        do_customer_map[do_num] = cust_id
        do_customer_name_map[do_num] = cust_name or "(not linked)"

    with st.form("maintenance_form"):
        selected_do = st.selectbox(
            "Delivery order",
            options=do_options,
            format_func=lambda do: do_labels.get(do, str(do)),
        )
        default_customer = do_customer_map.get(selected_do)
        state_key = "maintenance_customer_link"
        last_do_key = "maintenance_customer_last_do"
        linked_customer = default_customer
        if default_customer is not None:
            st.session_state[last_do_key] = selected_do
            st.session_state[state_key] = default_customer
            customer_label = (
                customer_labels.get(default_customer)
                or customer_label_map.get(default_customer)
                or label_by_id.get(default_customer)
                or do_customer_name_map.get(selected_do)
                or f"Customer #{default_customer}"
            )
            st.text_input("Customer", value=customer_label, disabled=True)
        else:
            choices = list(customer_options)
            if st.session_state.get(last_do_key) != selected_do:
                st.session_state[last_do_key] = selected_do
                st.session_state[state_key] = None
            linked_customer = st.selectbox(
                "Customer *",
                options=choices,
                format_func=lambda cid: customer_labels.get(cid, "-- Select customer --"),
                key=state_key,
            )
        status_value = status_input_widget("maintenance_new", DEFAULT_SERVICE_STATUS)
        maintenance_status_choice = get_status_choice("maintenance_new")
        today = datetime.now().date()
        if maintenance_status_choice == "Completed":
            maintenance_period_value = st.date_input(
                "Maintenance period",
                value=(today, today),
                help="Select the start and end dates for the maintenance work.",
                key="maintenance_new_period_completed",
            )
        elif maintenance_status_choice == "In progress":
            maintenance_period_value = st.date_input(
                "Maintenance start date",
                value=today,
                help="Choose when this maintenance began.",
                key="maintenance_new_period_start",
            )
        else:
            maintenance_period_value = st.date_input(
                "Planned start date",
                value=today,
                help="Select when this maintenance is scheduled to begin.",
                key="maintenance_new_period_planned",
            )
        description = st.text_area("Maintenance description")
        remarks = st.text_area("Remarks / updates")
        maintenance_amount_input = st.number_input(
            "Maintenance amount",
            min_value=0.0,
            step=100.0,
            format="%.2f",
            key="maintenance_new_amount",
            help="Track the amount charged for this maintenance.",
        )
        st.markdown("**Products sold during maintenance**")
        maintenance_rows = st.session_state.get(
            "maintenance_product_rows",
            [
                {
                    "name": "",
                    "model": "",
                    "serial": "",
                    "quantity": 1,
                }
            ],
        )
        maintenance_editor = st.data_editor(
            pd.DataFrame(maintenance_rows),
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "name": st.column_config.TextColumn("Product", help="Name / description"),
                "model": st.column_config.TextColumn("Model"),
                "serial": st.column_config.TextColumn("Serial"),
                "quantity": st.column_config.NumberColumn(
                    "Qty",
                    min_value=1,
                    step=1,
                    format="%d",
                ),
            },
            key="maintenance_product_table",
        )
        maintenance_product_entries = (
            maintenance_editor.to_dict("records")
            if isinstance(maintenance_editor, pd.DataFrame)
            else []
        )
        st.session_state["maintenance_product_rows"] = maintenance_product_entries
        maintenance_files = st.file_uploader(
            "Attach maintenance documents (PDF or image)",
            type=["pdf", "png", "jpg", "jpeg", "webp", "gif"],
            accept_multiple_files=True,
            key="maintenance_new_docs",
        )
        if maintenance_files:
            for idx, maintenance_file in enumerate(maintenance_files, start=1):
                _render_upload_ocr_preview(
                    maintenance_file,
                    key_prefix=f"maintenance_new_docs_{idx}",
                    label=f"Maintenance document {idx} OCR",
                )
        submit = st.form_submit_button("Log maintenance", type="primary")

    if submit:
        selected_customer = (
            linked_customer if linked_customer is not None else do_customer_map.get(selected_do)
        )
        selected_customer = int(selected_customer) if selected_customer is not None else None
        cur = conn.cursor()
        (
            maintenance_date_str,
            maintenance_start_str,
            maintenance_end_str,
        ) = determine_period_strings(
            maintenance_status_choice, maintenance_period_value
        )
        valid_entry = True
        if selected_customer is None:
            st.error("Select a customer to log this maintenance entry.")
            valid_entry = False
        if maintenance_status_choice == "Completed" and (
            not maintenance_start_str or not maintenance_end_str
        ):
            st.error("Start and end dates are required for completed maintenance work.")
            valid_entry = False
        if maintenance_status_choice != "Completed" and not maintenance_start_str:
            st.error("Select a start date for this maintenance entry.")
            valid_entry = False
        if valid_entry:
            _cleaned_maintenance_products, maintenance_product_labels = normalize_product_entries(
                maintenance_product_entries
            )
            maintenance_product_label = (
                "\n".join(maintenance_product_labels)
                if maintenance_product_labels
                else None
            )
            maintenance_amount_value = None
            try:
                if maintenance_amount_input is not None and float(maintenance_amount_input) > 0:
                    maintenance_amount_value = round(float(maintenance_amount_input), 2)
            except Exception:
                maintenance_amount_value = None
            cur.execute(
                """
                INSERT INTO maintenance_records (
                    do_number,
                    customer_id,
                    maintenance_date,
                    maintenance_start_date,
                    maintenance_end_date,
                    description,
                    status,
                    remarks,
                    maintenance_product_info,
                    total_amount,
                    updated_at,
                    created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    selected_do,
                    selected_customer,
                    maintenance_date_str,
                    maintenance_start_str,
                    maintenance_end_str,
                    clean_text(description),
                    status_value,
                    clean_text(remarks),
                    maintenance_product_label,
                    maintenance_amount_value,
                    datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                    current_user_id(),
                ),
            )
            maintenance_id = cur.lastrowid
            if selected_do and selected_customer is not None:
                link_delivery_order_to_customer(conn, selected_do, selected_customer)
            saved_docs = attach_documents(
                conn,
                "maintenance_documents",
                "maintenance_id",
                maintenance_id,
                maintenance_files,
                MAINTENANCE_DOCS_DIR,
                f"maintenance_{maintenance_id}",
                allowed_extensions=DOCUMENT_UPLOAD_EXTENSIONS,
            )
            conn.commit()
            maintenance_label = do_labels.get(selected_do) if selected_do else None
            if not maintenance_label:
                maintenance_label = f"Maintenance #{maintenance_id}"
                customer_name = None
                if selected_customer is not None:
                    customer_name = (
                        label_by_id.get(int(selected_customer))
                        or customer_label_map.get(int(selected_customer))
                    )
                summary_parts = [maintenance_label]
                if customer_name:
                    summary_parts.append(customer_name)
                status_label = clean_text(status_value) or DEFAULT_SERVICE_STATUS
                summary_parts.append(f"status {status_label}")
                log_activity(
                    conn,
                    event_type="maintenance_created",
                    description=" – ".join(summary_parts),
                    entity_type="maintenance",
                    entity_id=int(maintenance_id),
                )
                message = "Maintenance record saved."
                if saved_docs:
                    message = f"{message} Attached {saved_docs} document(s)."
                st.success(message)
                _safe_rerun()

    maintenance_df = df_query(
        conn,
        """
        SELECT m.maintenance_id,
               m.customer_id,
               d.customer_id AS do_customer_id,
               m.do_number,
               m.maintenance_date,
               m.maintenance_start_date,
               m.maintenance_end_date,
               m.maintenance_product_info,
               m.description,
               m.status,
               m.remarks,
               m.total_amount,
               m.payment_receipt_path,
               m.updated_at,
               m.created_by,
               COALESCE(c.name, cdo.name, '(unknown)') AS customer,
               COUNT(md.document_id) AS doc_count
        FROM maintenance_records m
        LEFT JOIN customers c ON c.customer_id = m.customer_id
        LEFT JOIN delivery_orders d ON d.do_number = m.do_number
        LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
        LEFT JOIN maintenance_documents md ON md.maintenance_id = m.maintenance_id
        WHERE m.deleted_at IS NULL
        GROUP BY m.maintenance_id
        ORDER BY datetime(COALESCE(m.maintenance_start_date, m.maintenance_date)) DESC, m.maintenance_id DESC
        """,
    )
    if allowed_customers is not None:
        def _maintenance_row_allowed(row):
            maint_cust = row.get("customer_id")
            do_cust = row.get("do_customer_id")
            creator_id = row.get("created_by")
            candidates = []
            if pd.notna(maint_cust):
                candidates.append(int(maint_cust))
            if pd.notna(do_cust):
                candidates.append(int(do_cust))
            try:
                if viewer_id is not None and pd.notna(creator_id) and int(creator_id) == int(viewer_id):
                    return True
            except Exception:
                pass
            return any(cid in allowed_customers for cid in candidates)

        maintenance_df = maintenance_df[maintenance_df.apply(_maintenance_row_allowed, axis=1)]
    if not maintenance_df.empty:
        maintenance_df = fmt_dates(
            maintenance_df,
            ["maintenance_date", "maintenance_start_date", "maintenance_end_date"],
        )
        maintenance_records = maintenance_df.to_dict("records")
        display_df = maintenance_df.drop(
            columns=["customer_id", "do_customer_id", "created_by"],
            errors="ignore",
        )
        display_df["maintenance_period"] = display_df.apply(
            lambda row: format_period_span(
                row.get("maintenance_start_date"), row.get("maintenance_end_date")
            ),
            axis=1,
        )
        display_df["Last update"] = pd.to_datetime(display_df.get("updated_at"), errors="coerce").dt.strftime("%d-%m-%Y %H:%M")
        display_df.loc[display_df["Last update"].isna(), "Last update"] = None
        if "status" in display_df.columns:
            display_df["status"] = display_df["status"].apply(
                lambda x: clean_text(x) or DEFAULT_SERVICE_STATUS
            )
        if "total_amount" in display_df.columns:
            display_df["maintenance_amount_display"] = display_df["total_amount"].apply(
                format_money
            )
        if "payment_receipt_path" in display_df.columns:
            display_df["payment_receipt_display"] = display_df[
                "payment_receipt_path"
            ].apply(lambda x: "📎" if clean_text(x) else "")
        display = display_df.rename(
            columns={
                "do_number": "DO Serial",
                "maintenance_date": "Maintenance date",
                "maintenance_start_date": "Maintenance start date",
                "maintenance_end_date": "Maintenance end date",
                "maintenance_period": "Maintenance period",
                "maintenance_product_info": "Products sold",
                "description": "Description",
                "status": "Status",
                "remarks": "Remarks",
                "maintenance_amount_display": "Maintenance amount",
                "payment_receipt_display": "Receipt",
                "customer": "Customer",
                "doc_count": "Documents",
            }
        )
        st.markdown("### Maintenance history")
        st.dataframe(
            display.drop(columns=["updated_at", "maintenance_id"], errors="ignore"),
            use_container_width=True,
        )

        records = maintenance_records
        st.markdown("#### Update status & remarks")
        options = [int(r["maintenance_id"]) for r in records]
        def maintenance_label(record):
            do_ref = clean_text(record.get("do_number")) or "(no DO)"
            date_ref = clean_text(record.get("maintenance_period")) or clean_text(
                record.get("maintenance_date")
            )
            customer_ref = clean_text(record.get("customer"))
            parts = [do_ref]
            if date_ref:
                parts.append(f"· {date_ref}")
            if customer_ref:
                parts.append(f"· {customer_ref}")
            return " ".join(parts)

        labels = {int(r["maintenance_id"]): maintenance_label(r) for r in records}
        selected_maintenance_id = st.selectbox(
            "Select maintenance entry",
            options,
            format_func=lambda rid: labels.get(rid, str(rid)),
        )
        selected_record = next(r for r in records if int(r["maintenance_id"]) == int(selected_maintenance_id))
        new_status = status_input_widget(
            f"maintenance_edit_{selected_maintenance_id}",
            selected_record.get("status"),
        )
        maintenance_edit_choice = get_status_choice(
            f"maintenance_edit_{selected_maintenance_id}"
        )
        existing_start = ensure_date(selected_record.get("maintenance_start_date")) or ensure_date(
            selected_record.get("maintenance_date")
        )
        existing_end = ensure_date(selected_record.get("maintenance_end_date")) or existing_start
        today = datetime.now().date()
        default_start = existing_start or today
        default_end = existing_end or default_start
        if maintenance_edit_choice == "Completed":
            maintenance_period_edit = st.date_input(
                "Maintenance period",
                value=(default_start, default_end),
                key=f"maintenance_edit_{selected_maintenance_id}_period_completed",
                help="Update the start and end dates for this maintenance record.",
            )
        elif maintenance_edit_choice == "In progress":
            maintenance_period_edit = st.date_input(
                "Maintenance start date",
                value=default_start,
                key=f"maintenance_edit_{selected_maintenance_id}_period_start",
                help="Adjust when this maintenance began.",
            )
        else:
            maintenance_period_edit = st.date_input(
                "Planned start date",
                value=default_start,
                key=f"maintenance_edit_{selected_maintenance_id}_period_planned",
                help="Adjust when this maintenance is scheduled to begin.",
            )
        new_remarks = st.text_area(
            "Remarks",
            value=clean_text(selected_record.get("remarks")) or "",
            key=f"maintenance_edit_{selected_maintenance_id}",
        )
        existing_amount = selected_record.get("total_amount")
        try:
            maintenance_amount_default = (
                float(existing_amount) if existing_amount is not None else 0.0
            )
        except (TypeError, ValueError):
            maintenance_amount_default = 0.0
        maintenance_amount_edit = st.number_input(
            "Maintenance amount",
            value=float(maintenance_amount_default),
            min_value=0.0,
            step=100.0,
            format="%.2f",
            key=f"maintenance_edit_amount_{selected_maintenance_id}",
        )
        existing_receipt_path = clean_text(selected_record.get("payment_receipt_path"))
        resolved_receipt = resolve_upload_path(existing_receipt_path)
        receipt_cols = st.columns([1, 1])
        with receipt_cols[0]:
            maintenance_receipt_upload = st.file_uploader(
                "Upload payment receipt (PDF or image)",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"maintenance_edit_receipt_{selected_maintenance_id}",
            )
        with receipt_cols[1]:
            clear_maintenance_receipt = st.checkbox(
                "Remove receipt",
                value=False,
                key=f"maintenance_clear_receipt_{selected_maintenance_id}",
            )
        if resolved_receipt and resolved_receipt.exists():
            st.download_button(
                "Download receipt",
                data=resolved_receipt.read_bytes(),
                file_name=resolved_receipt.name,
                key=f"maintenance_receipt_download_{selected_maintenance_id}",
            )
        elif existing_receipt_path:
            st.caption("Receipt file not found. Upload a new copy to replace it.")
        save_maintenance = st.button(
            "Save maintenance updates",
            key="save_maintenance_updates",
        )
        if _guard_double_submit("save_maintenance_updates", save_maintenance):
            (
                maintenance_date_str,
                maintenance_start_str,
                maintenance_end_str,
            ) = determine_period_strings(
                maintenance_edit_choice, maintenance_period_edit
            )
            valid_update = True
            if maintenance_edit_choice == "Completed" and (
                not maintenance_start_str or not maintenance_end_str
            ):
                st.error(
                    "Provide both start and end dates for completed maintenance records."
                )
                valid_update = False
            if maintenance_edit_choice != "Completed" and not maintenance_start_str:
                st.error("Select a start date for this maintenance entry.")
                valid_update = False
            if valid_update:
                try:
                    maintenance_amount_update = round(float(maintenance_amount_edit or 0.0), 2)
                except Exception:
                    maintenance_amount_update = 0.0
                receipt_path_value = existing_receipt_path
                replaced_receipt = False
                cleared_receipt = False
                if maintenance_receipt_upload is not None:
                    receipt_path_value = store_payment_receipt(
                        maintenance_receipt_upload,
                        identifier=f"maintenance_{selected_maintenance_id}_receipt",
                        target_dir=MAINTENANCE_DOCS_DIR,
                    )
                    replaced_receipt = bool(receipt_path_value)
                elif clear_maintenance_receipt and receipt_path_value:
                    old_receipt = resolve_upload_path(receipt_path_value)
                    if old_receipt and old_receipt.exists():
                        try:
                            old_receipt.unlink()
                        except Exception:
                            pass
                    receipt_path_value = None
                    cleared_receipt = True
                conn.execute(
                    """
                    UPDATE maintenance_records
                    SET status = ?,
                        remarks = ?,
                        maintenance_date = ?,
                        maintenance_start_date = ?,
                        maintenance_end_date = ?,
                        total_amount = ?,
                        payment_receipt_path = COALESCE(?, payment_receipt_path),
                        updated_at = datetime('now')
                    WHERE maintenance_id = ?
                      AND deleted_at IS NULL
                    """,
                    (
                        new_status,
                        clean_text(new_remarks),
                        maintenance_date_str,
                        maintenance_start_str,
                        maintenance_end_str,
                        maintenance_amount_update,
                        receipt_path_value,
                        int(selected_maintenance_id),
                    ),
                )
                conn.commit()
                label_text = labels.get(
                    int(selected_maintenance_id),
                    f"Maintenance #{int(selected_maintenance_id)}",
                )
                status_label = clean_text(new_status) or DEFAULT_SERVICE_STATUS
                summary = f"{label_text} → {status_label}" if status_label else label_text
                log_activity(
                    conn,
                    event_type="maintenance_updated",
                    description=summary,
                    entity_type="maintenance",
                    entity_id=int(selected_maintenance_id),
                )
                message_bits = ["Maintenance record updated."]
                if maintenance_amount_update:
                    message_bits.append(f"Maintenance amount {format_money(maintenance_amount_update)}")
                if replaced_receipt:
                    message_bits.append("Receipt uploaded")
                elif cleared_receipt:
                    message_bits.append("Receipt removed")
                st.success(" ".join(message_bits))
                _safe_rerun()
                st.success("Maintenance record updated.")
                _safe_rerun()

        attachments_df = df_query(
            conn,
            """
            SELECT document_id, file_path, original_name, uploaded_at
            FROM maintenance_documents
            WHERE maintenance_id = ?
            ORDER BY datetime(uploaded_at) DESC, document_id DESC
            """,
            (int(selected_maintenance_id),),
        )
        st.markdown("**Attached documents**")
        if attachments_df.empty:
            st.caption("No documents attached yet.")
        else:
            for _, doc_row in attachments_df.iterrows():
                path = resolve_upload_path(doc_row.get("file_path"))
                display_name = clean_text(doc_row.get("original_name"))
                if path and path.exists():
                    label = display_name or path.name
                    st.download_button(
                        f"Download {label}",
                        data=path.read_bytes(),
                        file_name=path.name,
                        key=f"maintenance_doc_dl_{int(doc_row['document_id'])}",
                    )
                else:
                    label = display_name or "Document"
                    st.caption(f"⚠️ Missing file: {label}")

        with st.form(f"maintenance_doc_upload_{selected_maintenance_id}"):
            more_docs = st.file_uploader(
                "Add more maintenance documents (PDF or image)",
                type=["pdf", "png", "jpg", "jpeg", "webp", "gif"],
                accept_multiple_files=True,
                key=f"maintenance_doc_files_{selected_maintenance_id}",
            )
            upload_docs = st.form_submit_button("Upload documents")
        if upload_docs:
            if more_docs:
                saved = attach_documents(
                    conn,
                    "maintenance_documents",
                    "maintenance_id",
                    int(selected_maintenance_id),
                    more_docs,
                    MAINTENANCE_DOCS_DIR,
                    f"maintenance_{selected_maintenance_id}",
                    allowed_extensions=DOCUMENT_UPLOAD_EXTENSIONS,
                )
                conn.commit()
                st.success(f"Uploaded {saved} document(s).")
                _safe_rerun()
            else:
                st.info("Select at least one PDF or image to upload.")

        st.markdown("#### Delete maintenance record")
        actor_id = current_user_id()
        is_admin = current_user_is_admin()
        deletable_df = pd.DataFrame(maintenance_records)
        if not is_admin and actor_id is not None and not deletable_df.empty:
            deletable_df = deletable_df[
                deletable_df["created_by"].apply(
                    lambda val: int(_coerce_float(val, -1)) == actor_id
                )
            ]
        if deletable_df.empty:
            st.caption("No maintenance records available for deletion.")
        else:
            delete_labels: dict[int, str] = {}
            for _, row in deletable_df.iterrows():
                maintenance_id = int(row.get("maintenance_id"))
                do_ref = clean_text(row.get("do_number")) or "Maintenance"
                customer_ref = clean_text(row.get("customer")) or "(customer)"
                period_ref = clean_text(row.get("maintenance_period"))
                label_parts = [do_ref, customer_ref]
                if period_ref:
                    label_parts.append(period_ref)
                delete_labels[maintenance_id] = " • ".join(label_parts)
            delete_options = list(delete_labels.keys())
            selected_delete_id = st.selectbox(
                "Select a maintenance entry to delete",
                delete_options,
                format_func=lambda val: delete_labels.get(val, f"Maintenance #{val}"),
                key="maintenance_delete_select",
            )
            confirm_delete = st.checkbox(
                "I understand this maintenance entry will be removed from active views.",
                key="maintenance_delete_confirm",
            )
            if st.button(
                "Delete maintenance record",
                type="secondary",
                disabled=not confirm_delete,
                key="maintenance_delete_button",
            ):
                conn.execute(
                    """
                    UPDATE maintenance_records
                    SET deleted_at=datetime('now'),
                        deleted_by=?
                    WHERE maintenance_id=?
                      AND deleted_at IS NULL
                    """,
                    (actor_id, selected_delete_id),
                )
                conn.commit()
                log_activity(
                    conn,
                    event_type="maintenance_deleted",
                    description=f"Maintenance #{selected_delete_id} deleted",
                    entity_type="maintenance",
                    entity_id=int(selected_delete_id),
                    user_id=actor_id,
                )
                st.warning("Maintenance record deleted.")
                _safe_rerun()
    else:
        st.info("No maintenance records yet. Log one using the form above.")


def _reset_delivery_order_form_state(record_type_key: str) -> None:
    key_prefix = f"{record_type_key}_do"
    st.session_state[f"{key_prefix}_items_rows"] = _default_delivery_items()
    st.session_state[f"{key_prefix}_number"] = ""
    st.session_state[f"{key_prefix}_customer"] = None
    st.session_state[f"{key_prefix}_description"] = ""
    st.session_state[f"{key_prefix}_remarks"] = ""
    st.session_state[f"{key_prefix}_status"] = "due"
    for key in (f"{key_prefix}_form_loader", f"{key_prefix}_items_editor"):
        st.session_state.pop(key, None)
    for record_key in ("delivery_order", "work_done"):
        st.session_state.pop(f"{record_key}_receipt_upload", None)
        st.session_state.pop(f"{record_key}_document_upload", None)


DELIVERY_STATUS_OPTIONS = ["due", "advanced", "paid"]
DELIVERY_STATUS_LABELS = {
    "due": "Due",
    "advanced": "Advanced",
    "paid": "Paid",
}


def normalize_delivery_status(value: Optional[str]) -> str:
    normalized_raw = clean_text(value)
    if not normalized_raw:
        return "due"
    normalized = normalized_raw.lower()
    if normalized in DELIVERY_STATUS_OPTIONS:
        return normalized
    if normalized in {"pending", "rejected", "overdue"}:
        return "due"
    if normalized in {"advance", "advanced_payment"}:
        return "advanced"
    return "due"


def delivery_orders_page(
    conn,
    *,
    show_heading: bool = True,
    record_type_label: str = "Delivery order",
    record_type_key: str = "delivery_order",
):
    if show_heading:
        st.subheader("🚚 Delivery orders")

    is_admin = current_user_is_admin()
    record_label_input = clean_text(record_type_label)
    record_label = record_label_input or "Delivery order"
    record_label_lower = record_label.lower()
    key_prefix = f"{record_type_key}_do"
    reset_pending_key = f"{key_prefix}_form_reset_pending"
    feedback_key = f"{key_prefix}_form_feedback"
    items_rows_key = f"{key_prefix}_items_rows"
    number_key = f"{key_prefix}_number"
    customer_key = f"{key_prefix}_customer"
    description_key = f"{key_prefix}_description"
    remarks_key = f"{key_prefix}_remarks"
    status_key = f"{key_prefix}_status"
    form_loader_key = f"{key_prefix}_form_loader"
    form_key = f"{key_prefix}_form"
    items_editor_key = f"{key_prefix}_items_editor"
    filter_text_key = f"{key_prefix}_filter_text"
    filter_customer_key = f"{key_prefix}_filter_customer"
    filter_date_toggle_key = f"{key_prefix}_filter_date_toggle"
    filter_date_range_key = f"{key_prefix}_filter_date_range"

    if st.session_state.pop(reset_pending_key, False):
        _reset_delivery_order_form_state(record_type_key)
    feedback_message = st.session_state.pop(feedback_key, None)
    if feedback_message:
        st.success(feedback_message)
    st.session_state.setdefault(items_rows_key, _default_delivery_items())
    st.session_state.setdefault(number_key, "")
    st.session_state.setdefault(customer_key, None)
    st.session_state.setdefault(description_key, "")
    st.session_state.setdefault(remarks_key, "")
    st.session_state.setdefault(status_key, "due")
    autofill_customer_key = f"{record_type_key}_autofill_customer"
    st.session_state.setdefault(autofill_customer_key, None)

    customer_options, customer_labels, _, _ = fetch_customer_choices(conn)
    scope_clause, scope_params = customer_scope_filter("c")
    where_sql = f"WHERE {scope_clause}" if scope_clause else ""
    do_rows = df_query(
        conn,
        f"SELECT c.customer_id, c.delivery_order_code FROM customers c {where_sql}",
        tuple(scope_params),
    )
    customer_do_map: dict[int, str] = {}
    if not do_rows.empty:
        for _, row in do_rows.iterrows():
            customer_id = int_or_none(row.get("customer_id"))
            code = clean_text(row.get("delivery_order_code"))
            if customer_id is None or not code:
                continue
            customer_do_map[customer_id] = code
    existing_dos = df_query(
        conn,
        dedent(
            """
            SELECT d.do_number,
                   d.customer_id,
                   COALESCE(c.name, c.company_name, '(unknown)') AS customer,
                   d.description,
                   d.sales_person,
                   d.remarks,
                   d.file_path,
                   d.items_payload,
                   d.total_amount,
                   d.created_by,
                   d.created_at,
                   d.status,
                   d.payment_receipt_path,
                   d.updated_at,
                   d.record_type,
                   u.username
            FROM delivery_orders d
            LEFT JOIN customers c ON c.customer_id = d.customer_id
            LEFT JOIN users u ON u.user_id = d.created_by
            WHERE COALESCE(d.record_type, 'delivery_order') = ?
              AND d.deleted_at IS NULL
            ORDER BY datetime(d.created_at) DESC
            LIMIT 200
            """
        ),
        (record_type_key,),
    )
    allowed_customers = accessible_customer_ids(conn)
    existing_dos = filter_delivery_orders_for_view(
        existing_dos, allowed_customers, record_types={record_type_key}
    )
    load_labels: dict[Optional[str], str] = {None: f"-- New {record_label_lower} --"}
    load_choices = [None]
    if not existing_dos.empty:
        for _, row in existing_dos.iterrows():
            do_num = clean_text(row.get("do_number"))
            if not do_num:
                continue
            customer_name = clean_text(row.get("customer")) or "(customer)"
            load_choices.append(do_num)
            load_labels[do_num] = f"{do_num} • {customer_name}"

    closed_statuses = {"paid"}
    current_receipt_path: Optional[str] = None

    st.markdown(f"### Create or update a {record_label_lower}")
    selected_existing = st.selectbox(
        f"Load existing {record_label_lower}",
        load_choices,
        format_func=lambda val: load_labels.get(val, "-- New delivery order --"),
        key=form_loader_key,
    )

    if selected_existing:
        match = existing_dos[existing_dos["do_number"] == selected_existing]
        if not match.empty:
            row = match.iloc[0]
            st.session_state[number_key] = clean_text(row.get("do_number")) or ""
            cust_id = row.get("customer_id")
            st.session_state[customer_key] = int(cust_id) if pd.notna(cust_id) else None
            st.session_state[description_key] = clean_text(row.get("description")) or ""
            st.session_state[remarks_key] = clean_text(row.get("remarks")) or ""
            st.session_state[status_key] = normalize_delivery_status(row.get("status"))
            current_receipt_path = clean_text(row.get("payment_receipt_path"))
            loaded_items = parse_delivery_items_payload(row.get("items_payload"))
            st.session_state[items_rows_key] = loaded_items or _default_delivery_items()
        st.session_state[autofill_customer_key] = None
    else:
        selected_customer_state = int_or_none(st.session_state.get(customer_key))
        last_autofill_customer = int_or_none(st.session_state.get(autofill_customer_key))
        current_number = clean_text(st.session_state.get(number_key))
        suggested_code = None
        if selected_customer_state:
            suggested_code = customer_do_map.get(selected_customer_state)
            if not suggested_code:
                candidate = existing_dos[existing_dos["customer_id"] == selected_customer_state]
                if not candidate.empty:
                    suggested_code = clean_text(candidate.iloc[0].get("do_number"))
        should_autofill = (
            selected_customer_state
            and (selected_customer_state != last_autofill_customer or not current_number)
        )
        if should_autofill and suggested_code:
            st.session_state[number_key] = suggested_code
            current_number = suggested_code
            st.session_state[autofill_customer_key] = selected_customer_state
        elif selected_customer_state is None:
            st.session_state[autofill_customer_key] = None

    receipt_download = None
    receipt_download_name = None
    receipt_download_key = None

    with st.form(form_key, clear_on_submit=True):
        do_number = st.text_input(
            f"{record_label} number *",
            key=number_key,
        )
        selected_customer = st.selectbox(
            "Customer",
            customer_options,
            format_func=lambda cid: customer_labels.get(cid, "-- Select customer --"),
            key=customer_key,
        )
        description = st.text_area(
            "Description / items",
            key=description_key,
        )
        remarks = st.text_area(
            "Remarks",
            key=remarks_key,
        )
        status_value = st.selectbox(
            "Status",
            DELIVERY_STATUS_OPTIONS,
            index=DELIVERY_STATUS_OPTIONS.index(
                normalize_delivery_status(st.session_state.get(status_key))
            ),
            format_func=lambda option: DELIVERY_STATUS_LABELS.get(option, option.title()),
            help="Paid delivery orders are locked against further edits.",
            disabled=(clean_text(st.session_state.get(status_key)) or "").lower()
            in closed_statuses,
            key=status_key,
        )
        receipt_upload = None
        if status_value in {"advanced", "paid"}:
            receipt_label = (
                f"{record_label} advance receipt (PDF or image)"
                if status_value == "advanced"
                else f"{record_label} full payment receipt (PDF or image)"
            )
            receipt_help = (
                "Highly recommended: attach proof of advance payment."
                if status_value == "advanced"
                else "Highly recommended: attach proof of payment for paid records."
            )
            receipt_upload = st.file_uploader(
                receipt_label,
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                key=f"{record_type_key}_receipt_upload",
                help=receipt_help,
            )
            _render_upload_ocr_preview(
                receipt_upload,
                key_prefix=f"{record_type_key}_receipt",
                label=f"{record_label} receipt OCR",
            )
            if current_receipt_path:
                receipt_file = resolve_upload_path(current_receipt_path)
                if receipt_file and receipt_file.exists():
                    receipt_download = receipt_file.read_bytes()
                    receipt_download_name = receipt_file.name
                    receipt_download_key = (
                        f"{record_label.lower().replace(' ', '_')}_receipt_view"
                    )
                st.caption("Uploading a new file will replace the current receipt.")
        st.markdown("**Products / items**")
        items_df_seed = pd.DataFrame(
            st.session_state.get(items_rows_key, _default_delivery_items())
        )
        if "line_total" in items_df_seed.columns:
            items_df_seed = items_df_seed.drop(columns=["line_total"], errors="ignore")
        items_editor = st.data_editor(
            items_df_seed,
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            key=items_editor_key,
            column_config={
                "description": st.column_config.TextColumn(
                    "Description", help="What is being delivered / sold"
                ),
                "quantity": st.column_config.NumberColumn(
                    "Qty", min_value=0.0, step=1.0, format="%.2f"
                ),
                "unit_price": st.column_config.NumberColumn(
                    "Unit price", min_value=0.0, step=50.0, format="%.2f"
                ),
                "discount": st.column_config.NumberColumn(
                    "Discount (%)", min_value=0.0, max_value=100.0, step=0.5, format="%.2f"
                ),
            },
        )
        st.session_state[items_rows_key] = (
            items_editor.to_dict("records") if isinstance(items_editor, pd.DataFrame) else []
        )
        items_clean, estimated_total = normalize_delivery_items(
            st.session_state.get(items_rows_key, [])
        )
        st.caption(
            f"Estimated total: {format_money(estimated_total) or f'{estimated_total:,.2f}'}"
        )
        submit = st.form_submit_button(f"Save {record_label_lower}", type="primary")

    if receipt_download and receipt_download_name:
        st.download_button(
            "View current receipt",
            data=receipt_download,
            file_name=receipt_download_name,
            key=receipt_download_key,
        )

    if _guard_double_submit(f"{record_type_key}_save_form", submit):
        sales_person = clean_text(get_current_user().get("username"))
        cleaned_number = clean_text(do_number)
        if not cleaned_number:
            st.error(f"{record_label} number is required.")
        else:
            cur = conn.cursor()
            conflicting_type = df_query(
                conn,
                "SELECT COALESCE(record_type, 'delivery_order') AS record_type FROM delivery_orders WHERE do_number = ? AND COALESCE(record_type, 'delivery_order') <> ? AND deleted_at IS NULL",
                (cleaned_number, record_type_key),
            )
            if not conflicting_type.empty:
                conflict_label = clean_text(conflicting_type.iloc[0].get("record_type")) or "delivery order"
                st.error(
                    f"This number is already used for a {conflict_label.replace('_', ' ')}. Choose a different {record_label_lower} number."
                )
                return
            existing = df_query(
                conn,
                "SELECT file_path, items_payload, total_amount, created_by, status, payment_receipt_path, deleted_at FROM delivery_orders WHERE do_number = ? AND COALESCE(record_type, 'delivery_order') = ?",
                (cleaned_number, record_type_key),
            )
            stored_path = None
            existing_status = "due"
            existing_receipt = None
            existing_deleted_at = None
            if not existing.empty:
                stored_path = clean_text(existing.iloc[0].get("file_path"))
                existing_status = normalize_delivery_status(existing.iloc[0].get("status"))
                existing_receipt = clean_text(existing.iloc[0].get("payment_receipt_path"))
                existing_deleted_at = clean_text(existing.iloc[0].get("deleted_at"))
            locked_record = existing_status.lower() in closed_statuses
            receipt_only_update = (
                locked_record
                and status_value == existing_status
                and existing_status.lower() == "paid"
                and receipt_upload is not None
            )
            if locked_record and not receipt_only_update:
                st.warning(
                    f"{record_label} {cleaned_number} is locked because it is marked as {existing_status}."
                )
                return
            auto_mark_paid = False
            if (
                status_value == "advanced"
                and existing_status == "advanced"
                and existing_receipt
                and receipt_upload is not None
            ):
                auto_mark_paid = True
                status_value = "paid"
            receipt_path = existing_receipt
            if status_value in {"advanced", "paid"}:
                if receipt_upload:
                    receipt_identifier = _sanitize_path_component(cleaned_number) or "do_receipt"
                    receipt_path = store_payment_receipt(
                        receipt_upload,
                        identifier=f"{receipt_identifier}_receipt",
                        target_dir=DELIVERY_RECEIPT_DIR,
                    )
                if not receipt_path:
                    missing_label = "advance" if status_value == "advanced" else "full payment"
                    st.warning(
                        f"No {missing_label} receipt uploaded. It is highly recommended to attach one."
                    )
            if auto_mark_paid and existing_receipt and selected_customer:
                advance_receipt_path = clean_text(existing_receipt)
                resolved_advance = resolve_upload_path(advance_receipt_path)
                if resolved_advance and resolved_advance.exists():
                    conn.execute(
                        """
                        INSERT INTO customer_documents (
                            customer_id, doc_type, file_path, original_name, uploaded_by
                        ) VALUES (?, ?, ?, ?, ?)
                        """,
                        (
                            int(selected_customer),
                            f"{record_label} advance receipt",
                            advance_receipt_path,
                            resolved_advance.name,
                            current_user_id(),
                        ),
                    )
            if receipt_only_update:
                conn.execute(
                    """
                    UPDATE delivery_orders
                       SET payment_receipt_path=COALESCE(?, payment_receipt_path),
                           updated_at=datetime('now')
                     WHERE do_number=? AND COALESCE(record_type, 'delivery_order') = ? AND deleted_at IS NULL
                    """,
                    (
                        receipt_path,
                        cleaned_number,
                        record_type_key,
                    ),
                )
                conn.commit()
                st.success("Receipt added to locked record.")
                return
            items_clean, total_amount_value = normalize_delivery_items(
                st.session_state.get(items_rows_key, [])
            )
            if not items_clean:
                st.error("Add at least one product line with pricing to save the delivery order.")
                return
            items_payload = json.dumps(items_clean, ensure_ascii=False)
            creator_id = current_user_id()
            if existing.empty:
                cur.execute(
                    """
                    INSERT INTO delivery_orders (do_number, customer_id, description, sales_person, remarks, file_path, items_payload, total_amount, created_by, status, payment_receipt_path, updated_at, record_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), ?)
                    """,
                    (
                        cleaned_number,
                        int(selected_customer) if selected_customer else None,
                        clean_text(description),
                        clean_text(sales_person),
                        clean_text(remarks),
                        stored_path,
                        items_payload,
                        total_amount_value,
                        creator_id,
                        normalize_delivery_status(status_value),
                        receipt_path,
                        record_type_key,
                    ),
                )
            else:
                existing_creator = existing.iloc[0].get("created_by")
                cur.execute(
                    """
                    UPDATE delivery_orders
                       SET customer_id=?, description=?, sales_person=?, remarks=?, file_path=?, items_payload=?, total_amount=?, status=?, payment_receipt_path=COALESCE(?, payment_receipt_path), updated_at=datetime('now'), created_by=COALESCE(created_by, ?), record_type=?, deleted_at=NULL, deleted_by=NULL
                     WHERE do_number=? AND COALESCE(record_type, 'delivery_order') = ?
                    """,
                    (
                        int(selected_customer) if selected_customer else None,
                        clean_text(description),
                        clean_text(sales_person),
                        clean_text(remarks),
                        stored_path,
                        items_payload,
                        total_amount_value,
                        normalize_delivery_status(status_value) or existing_status,
                        receipt_path,
                        creator_id,
                        record_type_key,
                        cleaned_number,
                        record_type_key,
                    ),
                )
            if selected_customer:
                link_delivery_order_to_customer(conn, cleaned_number, int(selected_customer))
            conn.commit()
            if existing.empty:
                action_label = "created"
            elif existing_deleted_at:
                action_label = "restored"
            else:
                action_label = "updated"
            customer_display = customer_labels.get(
                int(selected_customer) if selected_customer else None, "(customer)"
            )
            total_display = format_money(total_amount_value) or f"{_coerce_float(total_amount_value, 0.0):,.2f}"
            event_prefix = record_label.lower().replace(" ", "_") or "delivery_order"
            event_type = f"{event_prefix}_{'updated' if not existing.empty else 'created'}"
            entity_type = clean_text(record_type_key) or (
                "work_done" if event_prefix == "work_done" else "delivery_order"
            )
            status_label = clean_text(status_value) or existing_status
            status_note = f" [{status_label.title()}]" if status_label else ""

            log_activity(
                conn,
                event_type=event_type,
                description=(
                    f"{record_label} {cleaned_number} {action_label} for {customer_display}"
                    f" ({total_display}){status_note}"
                ),
                entity_type=entity_type,
                entity_id=None,
            )

            st.session_state[reset_pending_key] = True
            st.session_state[feedback_key] = (
                f"{record_label} {cleaned_number} saved successfully."
            )
            _safe_rerun()
            return

    number_label = f"{record_label} number"
    st.markdown(f"### {record_label} search")
    filter_cols = st.columns((1.4, 1.0, 1.0))
    with filter_cols[0]:
        query_text = st.text_input(
            f"Search by {record_label.lower()} number, description or remarks",
            key=filter_text_key,
        )
    with filter_cols[1]:
        customer_filter = st.selectbox(
            "Filter by customer",
            options=[None] + [opt for opt in customer_options if opt is not None],
            format_func=lambda cid: customer_labels.get(cid, "(any)"),
            key=filter_customer_key,
        )
    with filter_cols[2]:
        use_date_filter = st.checkbox(
            "Filter by created date",
            key=filter_date_toggle_key,
        )
    date_range = None
    if use_date_filter:
        date_range = st.date_input(
            "Created between",
            value=(datetime.now().date() - timedelta(days=30), datetime.now().date()),
            key=filter_date_range_key,
        )

    do_df = df_query(
        conn,
        """
        SELECT d.do_number,
               d.customer_id,
               COALESCE(c.name, '(unknown)') AS customer,
               d.description,
               d.sales_person,
               d.remarks,
               d.created_at,
               d.file_path,
               d.total_amount,
               d.items_payload,
               d.created_by,
               d.status,
               d.payment_receipt_path,
               d.updated_at,
               COALESCE(u.username, '(user)') AS created_by_name
         FROM delivery_orders d
          LEFT JOIN customers c ON c.customer_id = d.customer_id
          LEFT JOIN users u ON u.user_id = d.created_by
         WHERE COALESCE(d.record_type, 'delivery_order') = ?
           AND d.deleted_at IS NULL
         ORDER BY datetime(d.created_at) DESC
        """,
        (record_type_key,),
    )
    allowed_customers = accessible_customer_ids(conn)
    do_df = filter_delivery_orders_for_view(do_df, allowed_customers)
    if not do_df.empty:
        do_df = fmt_dates(do_df, ["created_at", "updated_at"])
        if query_text:
            needle = query_text.lower()
            do_df = do_df[
                do_df.apply(
                    lambda row: any(
                        needle in str(row.get(col, "")).lower()
                        for col in ["do_number", "description", "remarks"]
                    ),
                    axis=1,
                )
            ]
        if customer_filter:
            do_df = do_df[do_df["customer_id"] == int(customer_filter)]
        if use_date_filter and isinstance(date_range, (list, tuple)) and len(date_range) == 2:
            start_date, end_date = date_range
            start_iso = to_iso_date(start_date)
            end_iso = to_iso_date(end_date)
            if start_iso and end_iso:
                do_df = do_df[
                    do_df["created_at"].apply(
                        lambda value: start_iso <= to_iso_date(value) <= end_iso if to_iso_date(value) else False
                    )
                ]
        do_df["Document"] = do_df["file_path"].apply(lambda fp: "📎" if clean_text(fp) else "")
        do_df["Receipt"] = do_df["payment_receipt_path"].apply(lambda fp: "📎" if clean_text(fp) else "")
        do_df["status"] = do_df["status"].apply(
            lambda s: DELIVERY_STATUS_LABELS.get(normalize_delivery_status(s), "Due")
        )
        if "total_amount" in do_df.columns:
            def _format_total_value(value: object) -> str:
                if value is None:
                    return ""
                try:
                    if pd.isna(value):
                        return ""
                except Exception:
                    pass
                return format_money(value) or f"{_coerce_float(value, 0.0):,.2f}"

            do_df["total_amount"] = do_df["total_amount"].apply(_format_total_value)
        st.markdown(f"#### {record_label} records")
        header_cols = st.columns((1.2, 1.6, 2.6, 1.0, 0.9, 0.7, 0.7, 1.2))
        header_cols[0].write(f"**{number_label}**")
        header_cols[1].write("**Customer**")
        header_cols[2].write("**Description**")
        header_cols[3].write("**Total**")
        header_cols[4].write("**Status**")
        header_cols[5].write("**Attachment**")
        header_cols[6].write("**Receipt**")
        header_cols[7].write("**Upload receipt**")

        for _, row in do_df.iterrows():
            do_number = clean_text(row.get("do_number"))
            if not do_number:
                continue
            row_key = f"{record_type_key}_{do_number}"
            row_cols = st.columns((1.2, 1.6, 2.6, 1.0, 0.9, 0.7, 0.7, 1.2))
            row_cols[0].write(do_number)
            row_cols[1].write(clean_text(row.get("customer")) or "(unknown)")
            row_cols[2].write(clean_text(row.get("description")) or "")
            row_cols[3].write(clean_text(row.get("total_amount")) or "")
            row_cols[4].write(clean_text(row.get("status")) or "")
            row_cols[5].write(clean_text(row.get("Document")) or "")
            receipt_value = clean_text(row.get("payment_receipt_path"))
            receipt_file = resolve_upload_path(receipt_value) if receipt_value else None
            if receipt_file and receipt_file.exists():
                row_cols[6].download_button(
                    "View",
                    data=receipt_file.read_bytes(),
                    file_name=receipt_file.name,
                    key=f"do_receipt_view_{row_key}",
                )
            else:
                row_cols[6].write(clean_text(row.get("Receipt")) or "")
            upload_receipt = row_cols[7].file_uploader(
                "Upload receipt",
                type=["pdf", "png", "jpg", "jpeg", "webp"],
                label_visibility="collapsed",
                key=f"do_row_receipt_upload_{row_key}",
            )
            _render_upload_ocr_preview(
                upload_receipt,
                key_prefix=f"do_row_receipt_upload_{row_key}",
                label=f"{record_label} row receipt OCR",
            )
            save_label = f"💾 Save {record_label}"
            save_doc = row_cols[7].button(
                save_label, type="secondary", key=f"do_row_save_{row_key}"
            )
            if _guard_double_submit(f"do_row_save_{row_key}", save_doc):
                if upload_receipt is None:
                    st.warning("Select a receipt to upload.")
                else:
                    updates = {}
                    if upload_receipt is not None:
                        receipt_identifier = _sanitize_path_component(do_number) or "do_receipt"
                        receipt_path = store_payment_receipt(
                            upload_receipt,
                            identifier=f"{receipt_identifier}_receipt",
                            target_dir=DELIVERY_RECEIPT_DIR,
                        )
                        updates["payment_receipt_path"] = receipt_path
                    if updates:
                        set_clause = ", ".join(f"{col}=?" for col in updates)
                        params = list(updates.values())
                        params.extend([do_number, record_type_key])
                        conn.execute(
                            f"""
                            UPDATE delivery_orders
                               SET {set_clause},
                                   updated_at=datetime('now')
                             WHERE do_number=?
                               AND COALESCE(record_type, 'delivery_order') = ?
                               AND deleted_at IS NULL
                            """,
                            tuple(params),
                        )
                        conn.commit()
                        if "file_path" in updates:
                            log_activity(
                                conn,
                                event_type=f"{record_type_key}_document_uploaded",
                                description=f"{record_label} {do_number} document uploaded",
                                entity_type=record_type_key,
                                entity_id=None,
                            )
                        if "payment_receipt_path" in updates:
                            log_activity(
                                conn,
                                event_type=f"{record_type_key}_receipt_uploaded",
                                description=f"{record_label} {do_number} receipt uploaded",
                                entity_type=record_type_key,
                                entity_id=None,
                            )
                        st.success("Upload saved.")
                        _safe_rerun()

    downloads = {}
    if not do_df.empty:
        downloads = {
            clean_text(row["do_number"]): clean_text(row["file_path"])
            for _, row in do_df.iterrows()
            if clean_text(row.get("file_path"))
        }
    if downloads:
        st.markdown(f"#### Download {record_label.lower()}")
        selected_download = st.selectbox(
            f"Pick a {record_label_lower}",
            list(downloads.keys()),
            key=f"{key_prefix}_download_select",
        )
        path_value = downloads.get(selected_download)
        file_path = resolve_upload_path(path_value)
        if file_path and file_path.exists():
            st.download_button(
                f"Download {selected_download}",
                data=file_path.read_bytes(),
                file_name=file_path.name,
                key=f"{key_prefix}_download_button",
            )
        else:
            st.info("The selected delivery order file could not be found.")
    elif st.session_state.get(filter_text_key) or query_text:
        st.caption(
            f"No matching {record_label_lower} records found for the applied filters."
        )

    st.markdown(f"#### Delete {record_label.lower()}")
    actor_id = current_user_id()
    deletable_df = do_df.copy()
    if not is_admin and actor_id is not None:
        deletable_df = deletable_df[
            deletable_df["created_by"].apply(lambda val: int(_coerce_float(val, -1)) == actor_id)
        ]
    if deletable_df.empty:
        st.caption(f"No {record_label_lower} records available for deletion.")
        return
    delete_labels: dict[str, str] = {}
    for _, row in deletable_df.iterrows():
        number = clean_text(row.get("do_number"))
        if not number:
            continue
        customer = clean_text(row.get("customer")) or "(customer)"
        delete_labels[number] = f"{number} • {customer}"
    delete_options = list(delete_labels.keys())
    selected_delete = st.selectbox(
        f"Select a {record_label_lower} to delete",
        delete_options,
        format_func=lambda val: delete_labels.get(val, val),
        key=f"{key_prefix}_delete_select",
    )
    confirm_delete = st.checkbox(
        f"I understand this {record_label_lower} will be removed from active views.",
        key=f"{key_prefix}_delete_confirm",
    )
    if st.button(
        f"Delete {record_label_lower}",
        type="secondary",
        disabled=not confirm_delete,
        key=f"{record_label_lower}_delete_button",
    ):
        delete_row = deletable_df[
            deletable_df["do_number"].apply(lambda val: clean_text(val) == selected_delete)
        ]
        delete_info = {}
        if not delete_row.empty:
            delete_info = delete_row.iloc[0].to_dict()
        attachment_path = clean_text(delete_info.get("file_path"))
        receipt_path = clean_text(delete_info.get("payment_receipt_path"))
        description_text = clean_text(delete_info.get("description"))
        total_amount = clean_text(delete_info.get("total_amount"))
        deleted_by_label = clean_text(delete_info.get("created_by_name")) or "(staff)"
        conn.execute(
            """
            UPDATE delivery_orders
               SET deleted_at=datetime('now'),
                   deleted_by=?
             WHERE do_number=?
               AND COALESCE(record_type, 'delivery_order') = ?
            """,
            (actor_id, selected_delete, record_type_key),
        )
        conn.commit()
        detail_parts = [
            f"{record_label} {selected_delete} deleted",
            f"by {deleted_by_label}",
        ]
        if description_text:
            detail_parts.append(f"desc: {description_text}")
        if total_amount:
            detail_parts.append(f"total: {total_amount}")
        if attachment_path:
            detail_parts.append(f"doc: {attachment_path}")
        if receipt_path:
            detail_parts.append(f"receipt: {receipt_path}")
        description = " | ".join(detail_parts)
        log_activity(
            conn,
            event_type=f"{record_type_key}_deleted",
            description=description,
            entity_type=record_type_key,
            entity_id=None,
            user_id=actor_id,
        )
        st.warning(f"{record_label} deleted.")
        _safe_rerun()


def quotation_page(conn, *, render_id: Optional[int] = None):
    st.subheader("🧾 Quotation")
    _render_quotation_section(conn, render_id=render_id)
    st.markdown("---")
    _render_quotation_management(conn)


def work_done_page(conn):
    st.subheader("✅ Work done")
    st.caption("Create, update, and download work completion slips just like delivery orders.")
    delivery_orders_page(
        conn,
        show_heading=False,
        record_type_label="Work done",
        record_type_key="work_done",
    )


def service_maintenance_page(conn):
    st.subheader("🛠️ Maintenance and Service")
    tabs = st.tabs(["Service", "Maintenance"])
    with tabs[0]:
        st.markdown("### Service records")
        _render_service_section(conn, show_heading=False)
    with tabs[1]:
        st.markdown("### Maintenance records")
        _render_maintenance_section(conn, show_heading=False)
    st.markdown("---")
    st.info("Create new quotations from the dedicated 'Quotation' page in the sidebar.")


def customer_summary_page(conn):
    st.subheader("📒 Customer Summary")
    blank_label = "(blank)"
    show_complete_only = st.checkbox(
        "Only show customers with phone + address",
        value=False,
        help="Enable this to hide incomplete customer records from the summary.",
    )
    complete_clause = customer_complete_clause()
    name_clause = "TRIM(COALESCE(name, '')) <> ''"
    scope_clause, scope_params = customer_scope_filter()
    where_parts = [complete_clause if show_complete_only else name_clause]
    params: list[object] = []
    if scope_clause:
        where_parts.append(scope_clause)
        params.extend(scope_params)
    where_sql = " AND ".join(where_parts)
    customers = df_query(
        conn,
        f"""
        SELECT TRIM(name) AS name, GROUP_CONCAT(customer_id) AS ids, COUNT(*) AS cnt
        FROM customers
        WHERE {where_sql}
        GROUP BY TRIM(name)
        ORDER BY TRIM(name) ASC
        """,
        tuple(params),
    )
    if customers.empty:
        if show_complete_only:
            st.info(
                "No complete customers available for your account. Check the Scraps page for records that need details."
            )
        else:
            st.info("No customers available for your account yet.")
        return

    names = customers["name"].tolist()
    name_map = {
        row["name"]: f"{row['name']} ({int(row['cnt'])} records)" if int(row["cnt"]) > 1 else row["name"]
        for _, row in customers.iterrows()
    }
    sel_name = st.selectbox("Select customer", names, format_func=lambda n: name_map.get(n, n))
    row = customers[customers["name"] == sel_name].iloc[0]
    ids = [int(i) for i in str(row["ids"]).split(",") if i]
    cnt = int(row["cnt"])

    placeholder_block = ','.join('?' * len(ids))
    info = df_query(
        conn,
        f"""
        SELECT
            MAX(name) AS name,
            GROUP_CONCAT(DISTINCT phone) AS phone,
            GROUP_CONCAT(DISTINCT address) AS address,
            GROUP_CONCAT(DISTINCT purchase_date) AS purchase_dates,
            GROUP_CONCAT(DISTINCT product_info) AS products,
            GROUP_CONCAT(DISTINCT delivery_order_code) AS do_codes
        FROM customers
        WHERE customer_id IN ({placeholder_block})
        """,
        ids,
    ).iloc[0].to_dict()

    st.write("**Name:**", info.get("name") or blank_label)
    st.write("**Phone:**", info.get("phone"))
    st.write("**Address:**", info.get("address"))
    st.write("**Product:**", info.get("products"))
    st.write("**Delivery order:**", info.get("do_codes"))
    if cnt > 1:
        st.caption(f"Merged from {cnt} duplicates")

    st.markdown("---")
    placeholders = ",".join("?" * len(ids))

    warr = df_query(
        conn,
        f"""
        SELECT w.warranty_id as id, c.name as customer, p.name as product, p.model, w.serial, w.issue_date, w.expiry_date, w.status, w.remarks, w.dup_flag
        FROM warranties w
        LEFT JOIN customers c ON c.customer_id = w.customer_id
        LEFT JOIN products p ON p.product_id = w.product_id
        WHERE w.customer_id IN ({placeholders})
          AND p.name IS NOT NULL
          AND TRIM(p.name) != ''
        ORDER BY date(w.expiry_date) DESC
        """,
        ids,
    )
    warr = fmt_dates(warr, ["issue_date", "expiry_date"])
    if "dup_flag" in warr.columns:
        warr = warr.assign(duplicate=warr["dup_flag"].apply(lambda x: "🔁 duplicate" if int(x) == 1 else ""))
    warr_display = format_warranty_table(warr)

    service_df = df_query(
        conn,
        f"""
        SELECT s.service_id,
               s.do_number,
               s.service_date,
               s.service_start_date,
               s.service_end_date,
               s.service_product_info,
               s.description,
               s.remarks,
               s.bill_document_path,
               s.payment_receipt_path,
               COALESCE(c.name, cdo.name, '(unknown)') AS customer,
               COUNT(sd.document_id) AS doc_count
        FROM services s
        LEFT JOIN customers c ON c.customer_id = s.customer_id
        LEFT JOIN delivery_orders d ON d.do_number = s.do_number
        LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
        LEFT JOIN service_documents sd ON sd.service_id = s.service_id
        WHERE s.deleted_at IS NULL
          AND COALESCE(s.customer_id, d.customer_id) IN ({placeholders})
        GROUP BY s.service_id
        ORDER BY datetime(COALESCE(s.service_start_date, s.service_date)) DESC, s.service_id DESC
        """,
        ids,
    )
    service_df = fmt_dates(service_df, ["service_date", "service_start_date", "service_end_date"])
    if not service_df.empty:
        service_df["service_period"] = service_df.apply(
            lambda row: format_period_span(
                row.get("service_start_date"), row.get("service_end_date")
            ),
            axis=1,
        )

    maintenance_df = df_query(
        conn,
        f"""
        SELECT m.maintenance_id,
               m.do_number,
               m.maintenance_date,
               m.maintenance_start_date,
               m.maintenance_end_date,
               m.maintenance_product_info,
               m.description,
               m.remarks,
               m.payment_receipt_path,
               COALESCE(c.name, cdo.name, '(unknown)') AS customer,
               COUNT(md.document_id) AS doc_count
        FROM maintenance_records m
        LEFT JOIN customers c ON c.customer_id = m.customer_id
        LEFT JOIN delivery_orders d ON d.do_number = m.do_number
        LEFT JOIN customers cdo ON cdo.customer_id = d.customer_id
        LEFT JOIN maintenance_documents md ON md.maintenance_id = m.maintenance_id
        WHERE m.deleted_at IS NULL
          AND COALESCE(m.customer_id, d.customer_id) IN ({placeholders})
        GROUP BY m.maintenance_id
        ORDER BY datetime(COALESCE(m.maintenance_start_date, m.maintenance_date)) DESC, m.maintenance_id DESC
        """,
        ids,
    )
    maintenance_df = fmt_dates(
        maintenance_df,
        ["maintenance_date", "maintenance_start_date", "maintenance_end_date"],
    )
    if not maintenance_df.empty:
        maintenance_df["maintenance_period"] = maintenance_df.apply(
            lambda row: format_period_span(
                row.get("maintenance_start_date"), row.get("maintenance_end_date")
            ),
            axis=1,
        )

    do_df = df_query(
        conn,
        f"""
        SELECT d.do_number,
               COALESCE(c.name, '(unknown)') AS customer,
               d.description,
               d.sales_person,
               d.remarks,
               d.created_at,
               d.file_path,
               d.payment_receipt_path
        FROM delivery_orders d
        LEFT JOIN customers c ON c.customer_id = d.customer_id
        WHERE d.customer_id IN ({placeholders})
          AND d.deleted_at IS NULL
        ORDER BY datetime(d.created_at) DESC
        """,
        ids,
    )
    if not do_df.empty:
        do_df = fmt_dates(do_df, ["created_at"])
        do_df["do_number"] = do_df["do_number"].apply(clean_text)
        do_df["Document"] = do_df["file_path"].apply(lambda fp: "📎" if clean_text(fp) else "")
        do_df["Receipt"] = do_df["payment_receipt_path"].apply(lambda fp: "📎" if clean_text(fp) else "")

    do_numbers = set()
    if not do_df.empty and "do_number" in do_df.columns:
        do_numbers.update(val for val in do_df["do_number"].tolist() if val)
    if not service_df.empty and "do_number" in service_df.columns:
        do_numbers.update(clean_text(val) for val in service_df["do_number"].tolist() if clean_text(val))
    if not maintenance_df.empty and "do_number" in maintenance_df.columns:
        do_numbers.update(clean_text(val) for val in maintenance_df["do_number"].tolist() if clean_text(val))
    do_numbers = {val for val in do_numbers if val}

    present_dos = set()
    if not do_df.empty and "do_number" in do_df.columns:
        present_dos.update(val for val in do_df["do_number"].tolist() if val)
    missing_dos = sorted(do for do in do_numbers if do not in present_dos)
    if missing_dos:
        extra_df = df_query(
            conn,
            f"""
            SELECT d.do_number,
                   COALESCE(c.name, '(unknown)') AS customer,
                   d.description,
                   d.sales_person,
                   d.remarks,
                   d.created_at,
                   d.file_path,
                   d.payment_receipt_path
            FROM delivery_orders d
            LEFT JOIN customers c ON c.customer_id = d.customer_id
            WHERE d.do_number IN ({','.join('?' * len(missing_dos))})
              AND d.deleted_at IS NULL
            """,
            missing_dos,
        )
        if not extra_df.empty:
            extra_df = fmt_dates(extra_df, ["created_at"])
            extra_df["do_number"] = extra_df["do_number"].apply(clean_text)
            do_df = pd.concat([do_df, extra_df], ignore_index=True) if not do_df.empty else extra_df
            present_dos.update(val for val in extra_df["do_number"].tolist() if val)
    orphan_dos = sorted(do for do in do_numbers if do not in present_dos)

    if do_df is not None and not do_df.empty:
        do_df["Document"] = do_df["file_path"].apply(lambda fp: "📎" if clean_text(fp) else "")
        do_df["Receipt"] = do_df["payment_receipt_path"].apply(lambda fp: "📎" if clean_text(fp) else "")

    st.markdown("**Delivery orders**")
    if (do_df is None or do_df.empty) and not orphan_dos:
        st.info("No delivery orders found for this customer.")
    else:
        if do_df is not None and not do_df.empty:
            st.dataframe(
                do_df.rename(
                    columns={
                        "do_number": "DO Serial",
                        "customer": "Customer",
                        "description": "Description",
                        "sales_person": "Sales person",
                        "remarks": "Remarks",
                        "created_at": "Created",
                        "Document": "Document",
                        "Receipt": "Receipt",
                    }
                ).drop(columns=["file_path", "payment_receipt_path"], errors="ignore"),
                use_container_width=True,
            )
        if orphan_dos:
            st.caption("Referenced DO codes without a recorded delivery order: " + ", ".join(orphan_dos))

    st.markdown("**Warranties**")
    if warr_display is None or warr_display.empty:
        st.info("No warranties recorded for this customer.")
    else:
        st.dataframe(warr_display)

    st.markdown("**Service records**")
    if service_df.empty:
        st.info("No service records found for this customer.")
    else:
        service_display = service_df.rename(
            columns={
                "do_number": "DO Serial",
                "service_date": "Service date",
                "service_start_date": "Service start date",
                "service_end_date": "Service end date",
                "service_period": "Service period",
                "service_product_info": "Products sold",
                "description": "Description",
                "remarks": "Remarks",
                "customer": "Customer",
                "doc_count": "Documents",
            }
        )
        st.dataframe(
            service_display.drop(
                columns=["service_id", "bill_document_path", "payment_receipt_path"],
                errors="ignore",
            ),
            use_container_width=True,
        )

    st.markdown("**Maintenance records**")
    if maintenance_df.empty:
        st.info("No maintenance records found for this customer.")
    else:
        maintenance_display = maintenance_df.rename(
            columns={
                "do_number": "DO Serial",
                "maintenance_date": "Maintenance date",
                "maintenance_start_date": "Maintenance start date",
                "maintenance_end_date": "Maintenance end date",
                "maintenance_period": "Maintenance period",
                "maintenance_product_info": "Products sold",
                "description": "Description",
                "remarks": "Remarks",
                "customer": "Customer",
                "doc_count": "Documents",
            }
        )
        st.dataframe(
            maintenance_display.drop(
                columns=["maintenance_id", "payment_receipt_path"],
                errors="ignore",
            ),
            use_container_width=True,
        )

    documents = []
    customer_attachments = df_query(
        conn,
        f"""
        SELECT customer_id, attachment_path
        FROM customers
        WHERE customer_id IN ({placeholders})
          AND attachment_path IS NOT NULL
          AND attachment_path != ''
        """,
        ids,
    )
    if not customer_attachments.empty:
        for _, row in customer_attachments.iterrows():
            path = resolve_upload_path(row.get("attachment_path"))
            if not path or not path.exists():
                continue
            customer_id = int(row.get("customer_id"))
            display_name = path.name
            archive_name = "/".join(
                [
                    _sanitize_path_component("customer"),
                    f"{_sanitize_path_component(str(customer_id))}_{_sanitize_path_component(display_name)}",
                ]
            )
            documents.append(
                {
                    "source": "Customer",
                    "reference": f"Customer #{customer_id}",
                    "display": display_name,
                    "path": path,
                    "archive_name": archive_name,
                    "key": f"customer_{customer_id}",
                }
            )
    customer_docs = df_query(
        conn,
        f"""
        SELECT document_id, customer_id, doc_type, file_path, original_name, uploaded_at
        FROM customer_documents
        WHERE customer_id IN ({placeholders})
          AND deleted_at IS NULL
        ORDER BY datetime(uploaded_at) DESC, document_id DESC
        """,
        ids,
    )
    if not customer_docs.empty:
        customer_docs = customer_docs.drop_duplicates(
            subset=["customer_id", "doc_type", "file_path", "original_name"],
            keep="first",
        )
    if not customer_docs.empty:
        for _, row in customer_docs.iterrows():
            path = resolve_upload_path(row.get("file_path"))
            if not path or not path.exists():
                continue
            customer_id = int(row.get("customer_id"))
            doc_type = clean_text(row.get("doc_type")) or "Customer document"
            display_name = clean_text(row.get("original_name")) or path.name
            uploaded = pd.to_datetime(row.get("uploaded_at"), errors="coerce")
            documents.append(
                {
                    "source": doc_type,
                    "reference": f"Customer #{customer_id}",
                    "display": display_name,
                    "path": path,
                    "archive_name": "/".join(
                        [
                            _sanitize_path_component(doc_type),
                            f"{_sanitize_path_component(str(customer_id))}_{_sanitize_path_component(display_name)}",
                        ]
                    ),
                    "uploaded": uploaded if pd.notna(uploaded) else None,
                    "key": f"customer_doc_{customer_id}_{int(row.get('document_id'))}",
                }
            )
    if do_df is not None and not do_df.empty:
        for _, row in do_df.iterrows():
            path = resolve_upload_path(row.get("file_path"))
            if not path or not path.exists():
                continue
            do_ref = clean_text(row.get("do_number")) or "delivery_order"
            display_name = path.name
            archive_name = "/".join(
                [
                    _sanitize_path_component("delivery_orders"),
                    f"{_sanitize_path_component(do_ref)}_{_sanitize_path_component(display_name)}",
                ]
            )
            documents.append(
                {
                    "source": "Delivery order",
                    "reference": do_ref,
                    "display": display_name,
                    "path": path,
                    "archive_name": archive_name,
                    "key": f"do_{do_ref}",
                }
            )
            receipt_path = resolve_upload_path(row.get("payment_receipt_path"))
            if receipt_path and receipt_path.exists():
                receipt_name = receipt_path.name
                receipt_archive = "/".join(
                    [
                        _sanitize_path_component("delivery_order_receipts"),
                        f"{_sanitize_path_component(do_ref)}_{_sanitize_path_component(receipt_name)}",
                    ]
                )
                documents.append(
                    {
                        "source": "Delivery order receipt",
                        "reference": do_ref,
                        "display": receipt_name,
                        "path": receipt_path,
                        "archive_name": receipt_archive,
                        "key": f"do_receipt_{do_ref}",
                    }
                )

    service_docs = pd.DataFrame()
    if "service_id" in service_df.columns and not service_df.empty:
        service_ids = [int(val) for val in service_df["service_id"].dropna().astype(int).tolist()]
        if service_ids:
            service_docs = df_query(
                conn,
                f"""
                SELECT document_id, service_id, file_path, original_name, uploaded_at
                FROM service_documents
                WHERE service_id IN ({','.join('?' * len(service_ids))})
                ORDER BY datetime(uploaded_at) DESC, document_id DESC
                """,
                service_ids,
            )
    service_lookup = {}
    if "service_id" in service_df.columns and not service_df.empty:
        for _, row in service_df.iterrows():
            if pd.isna(row.get("service_id")):
                continue
            service_lookup[int(row["service_id"])] = row
    if not service_docs.empty:
        for _, doc_row in service_docs.iterrows():
            path = resolve_upload_path(doc_row.get("file_path"))
            if not path or not path.exists():
                continue
            service_id = int(doc_row.get("service_id"))
            record = service_lookup.get(service_id, {})
            reference = clean_text(record.get("do_number")) or f"Service #{service_id}"
            display_name = clean_text(doc_row.get("original_name")) or path.name
            uploaded = pd.to_datetime(doc_row.get("uploaded_at"), errors="coerce")
            uploaded_fmt = uploaded.strftime("%d-%m-%Y %H:%M") if pd.notna(uploaded) else None
            archive_name = "/".join(
                [
                    _sanitize_path_component("service"),
                    f"{_sanitize_path_component(reference)}_{_sanitize_path_component(display_name)}",
                ]
            )
            documents.append(
                {
                    "source": "Service",
                    "reference": reference,
                    "display": display_name,
                    "uploaded": uploaded_fmt,
                    "path": path,
                    "archive_name": archive_name,
                    "key": f"service_{service_id}_{int(doc_row['document_id'])}",
                }
            )
    if not service_df.empty:
        for _, row in service_df.iterrows():
            if pd.isna(row.get("service_id")):
                continue
            service_id = int(row.get("service_id"))
            reference = clean_text(row.get("do_number")) or f"Service #{service_id}"
            bill_path = resolve_upload_path(row.get("bill_document_path"))
            if bill_path and bill_path.exists():
                bill_name = bill_path.name
                bill_archive = "/".join(
                    [
                        _sanitize_path_component("service_bills"),
                        f"{_sanitize_path_component(reference)}_{_sanitize_path_component(bill_name)}",
                    ]
                )
                documents.append(
                    {
                        "source": "Service bill",
                        "reference": reference,
                        "display": bill_name,
                        "path": bill_path,
                        "archive_name": bill_archive,
                        "key": f"service_bill_{service_id}",
                    }
                )
            receipt_path = resolve_upload_path(row.get("payment_receipt_path"))
            if receipt_path and receipt_path.exists():
                receipt_name = receipt_path.name
                receipt_archive = "/".join(
                    [
                        _sanitize_path_component("service_receipts"),
                        f"{_sanitize_path_component(reference)}_{_sanitize_path_component(receipt_name)}",
                    ]
                )
                documents.append(
                    {
                        "source": "Service receipt",
                        "reference": reference,
                        "display": receipt_name,
                        "path": receipt_path,
                        "archive_name": receipt_archive,
                        "key": f"service_receipt_{service_id}",
                    }
                )

    maintenance_docs = pd.DataFrame()
    if "maintenance_id" in maintenance_df.columns and not maintenance_df.empty:
        maintenance_ids = [int(val) for val in maintenance_df["maintenance_id"].dropna().astype(int).tolist()]
        if maintenance_ids:
            maintenance_docs = df_query(
                conn,
                f"""
                SELECT document_id, maintenance_id, file_path, original_name, uploaded_at
                FROM maintenance_documents
                WHERE maintenance_id IN ({','.join('?' * len(maintenance_ids))})
                ORDER BY datetime(uploaded_at) DESC, document_id DESC
                """,
                maintenance_ids,
            )
    maintenance_lookup = {}
    if "maintenance_id" in maintenance_df.columns and not maintenance_df.empty:
        for _, row in maintenance_df.iterrows():
            if pd.isna(row.get("maintenance_id")):
                continue
            maintenance_lookup[int(row["maintenance_id"])] = row
    if not maintenance_docs.empty:
        for _, doc_row in maintenance_docs.iterrows():
            path = resolve_upload_path(doc_row.get("file_path"))
            if not path or not path.exists():
                continue
            maintenance_id = int(doc_row.get("maintenance_id"))
            record = maintenance_lookup.get(maintenance_id, {})
            reference = clean_text(record.get("do_number")) or f"Maintenance #{maintenance_id}"
            display_name = clean_text(doc_row.get("original_name")) or path.name
            uploaded = pd.to_datetime(doc_row.get("uploaded_at"), errors="coerce")
            uploaded_fmt = uploaded.strftime("%d-%m-%Y %H:%M") if pd.notna(uploaded) else None
            archive_name = "/".join(
                [
                    _sanitize_path_component("maintenance"),
                    f"{_sanitize_path_component(reference)}_{_sanitize_path_component(display_name)}",
                ]
            )
            documents.append(
                {
                    "source": "Maintenance",
                    "reference": reference,
                    "display": display_name,
                    "uploaded": uploaded_fmt,
                    "path": path,
                    "archive_name": archive_name,
                    "key": f"maintenance_{maintenance_id}_{int(doc_row['document_id'])}",
                }
            )
    if not maintenance_df.empty:
        for _, row in maintenance_df.iterrows():
            if pd.isna(row.get("maintenance_id")):
                continue
            maintenance_id = int(row.get("maintenance_id"))
            reference = clean_text(row.get("do_number")) or f"Maintenance #{maintenance_id}"
            receipt_path = resolve_upload_path(row.get("payment_receipt_path"))
            if receipt_path and receipt_path.exists():
                receipt_name = receipt_path.name
                receipt_archive = "/".join(
                    [
                        _sanitize_path_component("maintenance_receipts"),
                        f"{_sanitize_path_component(reference)}_{_sanitize_path_component(receipt_name)}",
                    ]
                )
                documents.append(
                    {
                        "source": "Maintenance receipt",
                        "reference": reference,
                        "display": receipt_name,
                        "path": receipt_path,
                        "archive_name": receipt_archive,
                        "key": f"maintenance_receipt_{maintenance_id}",
                    }
                )

    customer_name = clean_text(info.get("name"))
    raw_phone = clean_text(info.get("phone"))
    phone_tokens = []
    if raw_phone:
        phone_tokens = [
            token
            for token in (clean_text(val) for val in re.split(r"[,\n]+", raw_phone))
            if token
        ]
    quote_filters: list[str] = []
    quote_params: list[object] = []
    if customer_name:
        quote_filters.append("LOWER(COALESCE(customer_name, '')) = LOWER(?)")
        quote_filters.append("LOWER(COALESCE(customer_company, '')) = LOWER(?)")
        quote_params.extend([customer_name, customer_name])
    for phone in phone_tokens:
        quote_filters.append("customer_contact LIKE ?")
        quote_params.append(f"%{phone}%")
    if quote_filters:
        quote_clause = " OR ".join(quote_filters)
        quotation_docs = df_query(
            conn,
            dedent(
                f"""
                SELECT quotation_id,
                       reference,
                       document_path,
                       payment_receipt_path
                FROM quotations
                WHERE deleted_at IS NULL
                  AND ({quote_clause})
                ORDER BY datetime(quote_date) DESC, quotation_id DESC
                """
            ),
            tuple(quote_params),
        )
        if not quotation_docs.empty:
            for _, row in quotation_docs.iterrows():
                quote_id = int(row.get("quotation_id"))
                reference = clean_text(row.get("reference")) or f"Quotation #{quote_id}"
                doc_path = resolve_upload_path(row.get("document_path"))
                if doc_path and doc_path.exists():
                    doc_name = doc_path.name
                    doc_archive = "/".join(
                        [
                            _sanitize_path_component("quotations"),
                            f"{_sanitize_path_component(reference)}_{_sanitize_path_component(doc_name)}",
                        ]
                    )
                    documents.append(
                        {
                            "source": "Quotation",
                            "reference": reference,
                            "display": doc_name,
                            "path": doc_path,
                            "archive_name": doc_archive,
                            "key": f"quotation_doc_{quote_id}",
                        }
                    )
                receipt_path = resolve_upload_path(row.get("payment_receipt_path"))
                if receipt_path and receipt_path.exists():
                    receipt_name = receipt_path.name
                    receipt_archive = "/".join(
                        [
                            _sanitize_path_component("quotation_receipts"),
                            f"{_sanitize_path_component(reference)}_{_sanitize_path_component(receipt_name)}",
                        ]
                    )
                    documents.append(
                        {
                            "source": "Quotation receipt",
                            "reference": reference,
                            "display": receipt_name,
                            "path": receipt_path,
                            "archive_name": receipt_archive,
                            "key": f"quotation_receipt_{quote_id}",
                        }
                    )

    deduped_documents: dict[tuple[str, str, str, str], dict[str, object]] = {}
    for entry in documents:
        path_value = entry.get("path")
        path_key = ""
        if isinstance(path_value, Path):
            try:
                path_key = str(path_value.resolve())
            except OSError:
                path_key = str(path_value)
        elif path_value:
            path_key = str(path_value)
        dedupe_key = (
            clean_text(entry.get("source")) or "",
            clean_text(entry.get("reference")) or "",
            clean_text(entry.get("display")) or "",
            path_key,
        )
        existing = deduped_documents.get(dedupe_key)
        if existing is None:
            deduped_documents[dedupe_key] = entry
            continue
        existing_uploaded = existing.get("uploaded")
        new_uploaded = entry.get("uploaded")
        if not existing_uploaded and new_uploaded:
            deduped_documents[dedupe_key] = entry
    documents = list(deduped_documents.values())

    documents.sort(key=lambda d: (d["source"], d.get("reference") or "", d.get("display") or ""))

    st.markdown("**Documents**")
    if not documents:
        st.info("No documents attached for this customer.")
    else:
        for idx, doc in enumerate(documents, start=1):
            path = doc.get("path")
            if not path or not path.exists():
                continue
            label = f"{doc['source']}: {doc['reference']} – {doc['display']}"
            if doc.get("uploaded"):
                label = f"{label} (uploaded {doc['uploaded']})"
            st.download_button(
                f"Download {label}",
                data=path.read_bytes(),
                file_name=path.name,
                key=f"cust_doc_{doc['key']}_{idx}",
            )
        zip_buffer = bundle_documents_zip(documents)
        if zip_buffer is not None:
            archive_title = _sanitize_path_component(info.get("name") or blank_label)
            st.download_button(
                "⬇️ Download all documents (.rar)",
                data=zip_buffer.getvalue(),
                file_name=f"{archive_title}_documents.rar",
                mime="application/x-rar-compressed",
                key="cust_docs_zip",
            )

    pdf_bytes = generate_customer_summary_pdf(
        info.get("name") or blank_label,
        info,
        warr_display,
        service_df,
        maintenance_df,
    )
    package_buffer = bundle_customer_package(
        documents,
        pdf_bytes,
        info.get("name") or blank_label,
    )
    download_cols = st.columns(2)
    download_cols[0].download_button(
        "⬇️ Download summary (PDF)",
        data=pdf_bytes,
        file_name=f"customer_summary_{clean_text(info.get('name')) or 'customer'}.pdf",
        mime="application/pdf",
    )
    if package_buffer is not None:
        archive_name = _sanitize_path_component(info.get("name") or blank_label)
        download_cols[1].download_button(
            "⬇️ Download full customer package (.rar)",
            data=package_buffer.getvalue(),
            file_name=f"{archive_name}_full_package.rar",
            mime="application/x-rar-compressed",
            key="cust_full_package",
        )


def customers_hub_page(conn):
    tabs = st.tabs(["Customers", "Customer Summary", "Import", "Scraps", "Duplicates"])
    with tabs[0]:
        customers_page(conn)
    with tabs[1]:
        customer_summary_page(conn)
    with tabs[2]:
        import_page(conn)
    with tabs[3]:
        scraps_page(conn)
    with tabs[4]:
        duplicates_page(conn)


def scraps_page(conn):
    st.subheader("🗂️ Scraps (Incomplete Records)")
    st.caption(
        "Rows listed here are missing key details (name, phone, or address). They stay hidden from summaries until completed."
    )
    scope_clause, scope_params = customer_scope_filter()
    where_parts = [customer_incomplete_clause()]
    params: list[object] = []
    if scope_clause:
        where_parts.append(scope_clause)
        params.extend(scope_params)
    where_sql = " AND ".join(where_parts)
    scraps = df_query(
        conn,
        f"""
        SELECT customer_id as id, name, phone, address, remarks, purchase_date, product_info, delivery_order_code, created_at
        FROM customers
        WHERE {where_sql}
        ORDER BY datetime(created_at) DESC
        """,
        tuple(params),
    )
    scraps = fmt_dates(scraps, ["created_at", "purchase_date"])
    if scraps.empty:
        st.success("No scraps! All customer rows have the required details.")
    else:
        def missing_fields(row):
            missing = []
            for col, label in REQUIRED_CUSTOMER_FIELDS.items():
                val = row.get(col)
                if pd.isna(val) or str(val).strip() == "":
                    missing.append(label)
            return ", ".join(missing)

        scraps = scraps.assign(missing=scraps.apply(missing_fields, axis=1))
        display_cols = ["name", "phone", "address", "remarks", "purchase_date", "product_info", "delivery_order_code", "missing", "created_at"]
        st.dataframe(scraps[display_cols])

        st.markdown("### Update scrap record")
        records = scraps.to_dict("records")
        option_keys = [int(r["id"]) for r in records]
        option_labels = {}
        for r in records:
            rid = int(r["id"])
            name_label = clean_text(r.get("name")) or "(no name)"
            missing_label = clean_text(r.get("missing")) or "—"
            details = missing_label or "complete"
            created = clean_text(r.get("created_at"))
            created_fmt = f" – added {created}" if created else ""
            option_labels[rid] = f"{name_label or '(no name)'} (missing: {details}){created_fmt}"
        selected_id = st.selectbox(
            "Choose a record to fix",
            option_keys,
            format_func=lambda k: option_labels[k],
        )
        selected = next(r for r in records if int(r["id"]) == selected_id)

        def existing_value(key):
            return clean_text(selected.get(key)) or ""

        with st.form("scrap_update_form"):
            name = st.text_input("Name", existing_value("name"))
            phone = st.text_input("Phone", existing_value("phone"))
            address = st.text_area("Address", existing_value("address"))
            purchase = st.text_input("Purchase date (DD-MM-YYYY)", existing_value("purchase_date"))
            product = st.text_input("Product", existing_value("product_info"))
            do_code = st.text_input("Delivery order code", existing_value("delivery_order_code"))
            remarks_text = st.text_area("Remarks", existing_value("remarks"))
            col1, col2 = st.columns(2)
            save = col1.form_submit_button("Save changes", type="primary")
            delete = col2.form_submit_button("Delete scrap")

        if save:
            new_name = clean_text(name)
            new_phone = clean_text(phone)
            new_address = clean_text(address)
            new_remarks = clean_text(remarks_text)
            purchase_str, _ = date_strings_from_input(purchase)
            new_product = clean_text(product)
            new_do = clean_text(do_code)
            old_phone = clean_text(selected.get("phone"))
            conn.execute(
                "UPDATE customers SET name=?, phone=?, address=?, remarks=?, purchase_date=?, product_info=?, delivery_order_code=?, dup_flag=0 WHERE customer_id=?",
                (
                    new_name,
                    new_phone,
                    new_address,
                    new_remarks,
                    purchase_str,
                    new_product,
                    new_do,
                    int(selected_id),
                ),
            )
            old_do = clean_text(selected.get("delivery_order_code"))
            if new_do:
                conn.execute(
                    """
                    INSERT INTO delivery_orders (do_number, customer_id, order_id, description, sales_person, remarks, file_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(do_number) DO UPDATE SET
                        customer_id=excluded.customer_id,
                        description=excluded.description,
                        remarks=excluded.remarks,
                        deleted_at=NULL,
                        deleted_by=NULL
                    """,
                    (
                        new_do,
                        int(selected_id),
                        None,
                        new_product,
                        None,
                        new_remarks,
                        None,
                    ),
                )
            if old_do and old_do != new_do:
                conn.execute(
                    "DELETE FROM delivery_orders WHERE do_number=? AND (customer_id IS NULL OR customer_id=?)",
                    (old_do, int(selected_id)),
                )
            if old_phone and old_phone != new_phone:
                recalc_customer_duplicate_flag(conn, old_phone)
            if new_phone:
                recalc_customer_duplicate_flag(conn, new_phone)
            conn.commit()
            conn.execute(
                "UPDATE import_history SET customer_name=?, phone=?, address=?, delivery_address=?, product_label=?, do_number=?, original_date=? WHERE customer_id=? AND deleted_at IS NULL",
                (
                    new_name,
                    new_phone,
                    new_address,
                    new_address,
                    new_product,
                    new_do,
                    purchase_str,
                    int(selected_id),
                ),
            )
            conn.commit()
            if new_name and new_phone and new_address:
                st.success("Details saved. This record is now complete and will appear in other pages.")
            else:
                st.info("Details saved, but the record is still incomplete and will remain in Scraps until all required fields are filled.")
            _safe_rerun()

        if delete:
            conn.execute("DELETE FROM customers WHERE customer_id=?", (int(selected_id),))
            conn.commit()
            st.warning("Scrap record deleted.")
            _safe_rerun()

    if current_user_is_admin():
        st.markdown("### Deleted staff submissions")
        deleted_log = df_query(
            conn,
            dedent(
                """
                SELECT a.created_at,
                       a.event_type,
                       a.description,
                       COALESCE(u.username, 'User #' || a.user_id) AS actor
                FROM activity_log a
                LEFT JOIN users u ON u.user_id = a.user_id
                WHERE a.event_type LIKE '%_deleted'
                ORDER BY datetime(a.created_at) DESC
                LIMIT 200
                """
            ),
        )
        if deleted_log.empty:
            st.caption("No deleted submissions recorded yet.")
        else:
            deleted_log = fmt_dates(deleted_log, ["created_at"])
            deleted_log = deleted_log.rename(
                columns={
                    "created_at": "When",
                    "event_type": "Event",
                    "description": "Details",
                    "actor": "Staff",
                }
            )
            st.dataframe(
                deleted_log[["When", "Staff", "Event", "Details"]],
                use_container_width=True,
                hide_index=True,
            )

# ---------- Import helpers ----------
def refine_multiline(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    def _as_list(value: object) -> list[object]:
        if isinstance(value, str):
            items = [s.strip() for s in value.splitlines() if s.strip()]
            return items or [None]
        if pd.isna(value):
            return [None]
        return [value]

    listified = {col: df[col].apply(_as_list) for col in df.columns}

    normalized_rows: list[dict[str, object]] = []
    for _, row in pd.DataFrame(listified).iterrows():
        lengths = [len(values) for values in row]
        max_len = max(lengths) if lengths else 0
        if max_len == 0:
            normalized_rows.append({col: None for col in df.columns})
            continue
        expanded = {
            col: values + [None] * (max_len - len(values))
            for col, values in row.items()
        }
        for idx in range(max_len):
            normalized_rows.append({col: expanded[col][idx] for col in df.columns})

    return pd.DataFrame(normalized_rows, columns=df.columns)


_TRAILING_ZERO_NUMBER = re.compile(r"^-?\d+\.0+$")


def _normalize_sort_value(value: object) -> str:
    text = clean_text(value)
    if text is None:
        return ""
    if _TRAILING_ZERO_NUMBER.match(text):
        return text.split(".", 1)[0]
    return text


def _sort_dataframe_safe(df: pd.DataFrame, sort_columns: Iterable[str]) -> pd.DataFrame:
    columns = [col for col in sort_columns if col in df.columns]
    if not columns:
        return df

    def _sort_key(series: pd.Series) -> pd.Series:
        if pd.api.types.is_datetime64_any_dtype(series):
            return series
        name = str(series.name or "").lower()
        if "date" in name:
            converted = pd.to_datetime(series, errors="coerce", dayfirst=True)
            if not converted.isna().all():
                return converted
        return series.map(_normalize_sort_value)

    return df.sort_values(by=columns, key=_sort_key, na_position="last")


def normalize_headers(cols):
    norm = []
    for c in cols:
        s = str(c).strip().lower().replace(" ", "_")
        norm.append(s)
    return norm

HEADER_MAP = {
    "date": {"date", "delivery_date", "issue_date", "order_date", "dt", "d_o", "d", "sale_date"},
    "purchase_date": {"purchase_date", "purchase", "purchase_dt", "buy_date", "purchase_on"},
    "follow_up_date": {"follow_up_date", "followup_date", "follow_up", "followup", "next_follow_up", "next_followup"},
    "customer_name": {"customer_name", "customer", "company", "company_name", "client", "party", "name"},
    "address": {"address", "addr", "street", "location"},
    "delivery_address": {"delivery_address", "delivery_addr", "shipping_address", "ship_to", "delivery"},
    "phone": {"phone", "mobile", "contact", "contact_no", "phone_no", "phone_number", "cell", "whatsapp"},
    "product": {"product", "item", "generator", "model", "description"},
    "do_code": {"do_code", "delivery_order", "delivery_order_code", "delivery_order_no", "do", "d_o_code", "do_number"},
    "work_done_code": {"work_done_code", "work_done", "work_done_no", "work_done_number", "workdone", "work_done_ref"},
    "service_code": {"service_code", "service", "service_no", "service_number", "service_ref"},
    "maintenance_code": {"maintenance_code", "maintenance", "maintenance_no", "maintenance_number", "maintenance_ref"},
    "quantity": {"quantity", "qty", "count", "units", "pcs", "pieces"},
    "remarks": {"remarks", "remark", "notes", "note", "comments", "comment"},
    "amount_spent": {"amount", "amount_spent", "value", "price", "invoice_amount", "total", "total_amount", "amt"},
}

def map_headers_guess(cols):
    cols_norm = normalize_headers(cols)
    mapping = {k: None for k in HEADER_MAP.keys()}
    for i, cn in enumerate(cols_norm):
        for target, aliases in HEADER_MAP.items():
            if cn in aliases and mapping[target] is None:
                mapping[target] = i
                break
    default_order = [
        "date",
        "purchase_date",
        "customer_name",
        "address",
        "delivery_address",
        "phone",
        "product",
        "do_code",
        "work_done_code",
        "service_code",
        "maintenance_code",
        "quantity",
    ]
    if cols_norm[: len(default_order)] == default_order:
        mapping = {field: idx for idx, field in enumerate(default_order)}
    return mapping


def split_product_label(label: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    if label is None:
        return None, None
    text = clean_text(label)
    if not text:
        return None, None
    if "-" in text:
        left, right = text.split("-", 1)
        return clean_text(left), clean_text(right)
    return text, None


def parse_date_value(value) -> Optional[pd.Timestamp]:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return None
    if isinstance(dt, pd.DatetimeIndex):
        dt = dt[0]
    return dt.normalize()


def date_strings_from_input(value) -> tuple[Optional[str], Optional[str]]:
    dt = parse_date_value(value)
    if dt is None:
        return None, None
    expiry = dt + pd.Timedelta(days=365)
    return dt.strftime("%Y-%m-%d"), expiry.strftime("%Y-%m-%d")


def int_or_none(value) -> Optional[int]:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        return int(value)
    except (TypeError, ValueError):
        return None

def coerce_excel_date(series):
    s = pd.to_datetime(series, errors="coerce", dayfirst=True)
    if s.isna().mean() > 0.5:
        try:
            num = pd.to_numeric(series, errors="coerce")
            if num.notna().sum() > 0 and (num.dropna().median() > 20000):
                s = pd.to_datetime(num, unit="d", origin="1899-12-30", errors="coerce")
        except Exception:
            pass
    return s

def import_page(conn):
    st.subheader("⬆️ Import from Excel/CSV (append)")
    st.caption("We’ll auto-detect columns; you can override mapping. Dates accept DD-MM-YYYY or Excel serials.")
    f = st.file_uploader("Upload .xlsx or .csv", type=["xlsx","csv"])
    if f is None:
        st.markdown("---")
        manage_import_history(conn)
        return
    # Streamlit reruns the script whenever widgets change state. This means the
    # uploaded file object is reused across runs and its pointer sits at the end
    # after the first read. Attempting to read again (e.g. after a selectbox
    # change) would therefore raise an "Excel file format cannot be determined"
    # error or return empty data, effectively restarting the app view. Reset the
    # pointer before every read so interactive mapping works reliably.
    f.seek(0)
    if f.name.endswith(".csv"):
        df = pd.read_csv(f)
    else:
        df = pd.read_excel(f)
    st.write("Preview:", df.head())

    guess = map_headers_guess(list(df.columns))
    cols = list(df.columns)
    opts = ["(blank)"] + cols
    col1, col2, col3 = st.columns(3)
    col4, col5, col6 = st.columns(3)
    col7, col8, col9 = st.columns(3)
    col10, col11, col12 = st.columns(3)
    col13, col14, col15 = st.columns(3)
    sel_date = col1.selectbox(
        "Date (fallback)",
        options=opts,
        index=(guess["date"] + 1) if guess.get("date") is not None else 0,
    )
    sel_purchase_date = col2.selectbox(
        "Purchase date",
        options=opts,
        index=(guess.get("purchase_date", None) + 1) if guess.get("purchase_date") is not None else 0,
    )
    sel_name = col3.selectbox(
        "Customer name", options=opts, index=(guess["customer_name"] + 1) if guess.get("customer_name") is not None else 0
    )
    sel_addr = col4.selectbox(
        "Address", options=opts, index=(guess["address"] + 1) if guess.get("address") is not None else 0
    )
    sel_do = col5.selectbox(
        "Delivery address", options=opts, index=(guess["delivery_address"] + 1) if guess.get("delivery_address") is not None else 0
    )
    sel_phone = col6.selectbox(
        "Phone", options=opts, index=(guess["phone"] + 1) if guess.get("phone") is not None else 0
    )
    sel_prod = col7.selectbox(
        "Product", options=opts, index=(guess["product"] + 1) if guess.get("product") is not None else 0
    )
    sel_do_code = col8.selectbox(
        "Delivery order code (optional)", options=opts, index=(guess["do_code"] + 1) if guess.get("do_code") is not None else 0
    )
    sel_work_done = col9.selectbox(
        "Work done code (optional)",
        options=opts,
        index=(guess.get("work_done_code", None) + 1) if guess.get("work_done_code") is not None else 0,
    )
    sel_service_code = col10.selectbox(
        "Service code (optional)",
        options=opts,
        index=(guess.get("service_code", None) + 1) if guess.get("service_code") is not None else 0,
    )
    sel_maintenance_code = col11.selectbox(
        "Maintenance code (optional)",
        options=opts,
        index=(guess.get("maintenance_code", None) + 1) if guess.get("maintenance_code") is not None else 0,
    )
    sel_follow_up = col12.selectbox(
        "Follow-up date (optional)",
        options=opts,
        index=(guess.get("follow_up_date", None) + 1) if guess.get("follow_up_date") is not None else 0,
    )
    sel_remarks = col13.selectbox(
        "Remarks", options=opts, index=(guess.get("remarks", None) + 1) if guess.get("remarks") is not None else 0
    )
    sel_amount = col14.selectbox(
        "Amount spent", options=opts, index=(guess.get("amount_spent", None) + 1) if guess.get("amount_spent") is not None else 0
    )
    sel_quantity = col15.selectbox(
        "Quantity", options=opts, index=(guess.get("quantity", None) + 1) if guess.get("quantity") is not None else 0
    )

    def pick(col_name):
        return df[col_name] if col_name != "(blank)" else pd.Series([None] * len(df))

    df_norm = pd.DataFrame(
        {
            "date": pick(sel_date),
            "purchase_date": pick(sel_purchase_date),
            "customer_name": pick(sel_name),
            "address": pick(sel_addr),
            "delivery_address": pick(sel_do),
            "phone": pick(sel_phone),
            "product": pick(sel_prod),
            "do_code": pick(sel_do_code),
            "work_done_code": pick(sel_work_done),
            "service_code": pick(sel_service_code),
            "maintenance_code": pick(sel_maintenance_code),
            "follow_up_date": pick(sel_follow_up),
            "remarks": pick(sel_remarks),
            "amount_spent": pick(sel_amount),
            "quantity": pick(sel_quantity),
        }
    )
    text_columns = [
        "customer_name",
        "address",
        "delivery_address",
        "phone",
        "product",
        "do_code",
        "work_done_code",
        "service_code",
        "maintenance_code",
        "remarks",
    ]
    for column in text_columns:
        if column in df_norm.columns:
            df_norm[column] = df_norm[column].apply(
                lambda value: "" if pd.isna(value) else str(value)
            )
    df_norm["quantity"] = df_norm["quantity"].apply(parse_quantity)
    skip_blanks = st.checkbox("Skip blank rows", value=True)
    df_norm = refine_multiline(df_norm)
    df_norm["date"] = coerce_excel_date(df_norm["date"])
    df_norm["purchase_date"] = coerce_excel_date(df_norm["purchase_date"])
    df_norm["follow_up_date"] = coerce_excel_date(df_norm["follow_up_date"])
    df_norm = df_norm.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    if skip_blanks:
        df_norm = df_norm.dropna(how="all")
    df_norm = df_norm.drop_duplicates()
    df_norm = _sort_dataframe_safe(
        df_norm, ["purchase_date", "date", "customer_name", "phone", "do_code"]
    ).reset_index(drop=True)
    st.markdown("#### Review & edit rows before importing")
    preview = df_norm.copy()
    preview["Action"] = "Import"
    preview_text_columns = [
        "customer_name",
        "address",
        "delivery_address",
        "phone",
        "product",
        "do_code",
        "work_done_code",
        "service_code",
        "maintenance_code",
        "remarks",
    ]
    for column in preview_text_columns:
        if column in preview.columns:
            preview[column] = preview[column].fillna("").astype(str)
    editor = st.data_editor(
        preview,
        key="import_editor",
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "date": st.column_config.DateColumn("Date", format="DD-MM-YYYY", required=False),
            "purchase_date": st.column_config.DateColumn(
                "Purchase date", format="DD-MM-YYYY", required=False
            ),
            "follow_up_date": st.column_config.DateColumn(
                "Follow-up date", format="DD-MM-YYYY", required=False
            ),
            "work_done_code": st.column_config.TextColumn(
                "Work done code", required=False
            ),
            "service_code": st.column_config.TextColumn(
                "Service code", required=False
            ),
            "maintenance_code": st.column_config.TextColumn(
                "Maintenance code", required=False
            ),
            "Action": st.column_config.SelectboxColumn("Action", options=["Import", "Skip"], required=True),
            "remarks": st.column_config.TextColumn("Remarks", required=False),
            "amount_spent": st.column_config.NumberColumn(
                "Amount spent", min_value=0.0, step=0.01, format="%.2f", required=False
            ),
            "quantity": st.column_config.NumberColumn(
                "Quantity", min_value=1, step=1, format="%d", required=False
            ),
        },
    )

    if st.button("Append into database"):
        editor = editor if isinstance(editor, pd.DataFrame) else pd.DataFrame(editor)
        ready = editor[editor["Action"].fillna("Import").str.lower() == "import"].copy()
        ready.drop(columns=["Action"], inplace=True, errors="ignore")
        seeded, d_c, d_p = _import_clean6(conn, ready, tag="Manual import (mapped)")
        if seeded == 0:
            st.warning("No rows added (rows empty/invalid). Check mapping or file.")
        else:
            st.success(f"Imported {seeded} rows. Duplicates flagged — customers: {d_c}, products: {d_p}.")

    st.markdown("---")
    manage_import_history(conn)

def manual_merge_section(conn, customers_df: pd.DataFrame) -> None:
    if customers_df is None or customers_df.empty:
        return

    if "id" not in customers_df.columns:
        return

    work_df = customers_df.copy()
    work_df["id"] = work_df["id"].apply(int_or_none)
    work_df = work_df[work_df["id"].notna()]
    if work_df.empty:
        return

    work_df["id"] = work_df["id"].astype(int)

    def build_label(row):
        name_val = clean_text(row.get("name")) or "(no name)"
        phone_val = clean_text(row.get("phone")) or "(no phone)"
        address_val = clean_text(row.get("address")) or "-"
        product_val = clean_text(row.get("product_info")) or "-"
        do_val = clean_text(row.get("delivery_order_code")) or "-"
        date_dt = parse_date_value(row.get("purchase_date"))
        if date_dt is not None:
            date_label = date_dt.strftime(DATE_FMT)
        else:
            date_label = clean_text(row.get("purchase_date")) or "-"
        return f"#{row['id']} – {name_val} | Phone: {phone_val} | Date: {date_label} | Product: {product_val} | DO: {do_val}"

    work_df["_label"] = work_df.apply(build_label, axis=1)
    def _compose_search_blob(row) -> str:
        fields = [
            row.get("name"),
            row.get("company_name"),
            row.get("phone"),
            row.get("address"),
            row.get("delivery_address"),
            row.get("product_info"),
            row.get("delivery_order_code"),
            row.get("remarks"),
            row.get("sales_person"),
            row.get("purchase_date"),
        ]
        cleaned: list[str] = []
        for value in fields:
            text = clean_text(value)
            if not text:
                continue
            cleaned.append(" ".join(text.split()))
        return " ".join(cleaned).lower()

    work_df["_search_blob"] = work_df.apply(_compose_search_blob, axis=1)
    work_df["_search_blob"] = work_df["_search_blob"].fillna("")

    label_map = {row["id"]: row["_label"] for row in work_df.to_dict("records")}

    st.divider()
    st.markdown("#### Manual customer merge")
    st.caption(
        "Select multiple customer records that refer to the same person even if the phone number or purchase date differs. "
        "The earliest record will be kept and enriched with the combined details."
    )

    filter_value = st.text_input(
        "Filter customers by name, phone, address, product, or DO (optional)",
        key="manual_merge_filter",
    ).strip()
    filter_normalized = " ".join(filter_value.lower().split())

    filtered_df = work_df
    if filter_normalized:
        mask = filtered_df["_search_blob"].str.contains(
            filter_normalized, regex=False, na=False
        )
        filtered_df = filtered_df[mask]

    options = filtered_df["id"].tolist()
    if not options:
        st.info("No customer records match the current filter.")
        return

    with st.form("manual_merge_form"):
        selected_ids = st.multiselect(
            "Select customer records to merge",
            options=options,
            format_func=lambda cid: label_map.get(cid, f"#{cid}"),
        )

        preview_df = work_df[work_df["id"].isin(selected_ids)]
        if not preview_df.empty:
            preview_df = preview_df.copy()
            preview_df["purchase_date"] = pd.to_datetime(preview_df["purchase_date"], errors="coerce")
            preview_df["purchase_date"] = preview_df["purchase_date"].dt.strftime(DATE_FMT)
            preview_df["purchase_date"] = preview_df["purchase_date"].fillna("-")
            preview_cols = [
                col
                for col in [
                    "id",
                    "name",
                    "phone",
                    "address",
                    "purchase_date",
                    "product_info",
                    "delivery_order_code",
                    "created_at",
                ]
                if col in preview_df.columns
            ]
            st.dataframe(
                preview_df[preview_cols]
                .rename(
                    columns={
                        "id": "ID",
                        "name": "Name",
                        "phone": "Phone",
                        "address": "Address",
                        "purchase_date": "Purchase date",
                        "product_info": "Product",
                        "delivery_order_code": "DO code",
                        "created_at": "Created",
                    }
                )
                .sort_values("ID"),
                use_container_width=True,
                hide_index=True,
            )

        submitted = st.form_submit_button("Merge selected customers", type="primary")

    if submitted:
        if len(selected_ids) < 2:
            st.warning("Select at least two customers to merge.")
            return
        if merge_customer_records(conn, selected_ids):
            st.success(f"Merged {len(selected_ids)} customer records.")
            _safe_rerun()
        else:
            st.error("Could not merge the selected customers. Please try again.")


def duplicates_page(conn):
    st.subheader("⚠️ Possible Duplicates")
    if auto_merge_matching_customers(conn):
        st.info(
            "Automatically merged customers sharing the same name and address.",
            icon="✅",
        )
        _safe_rerun()
        return
    scope_clause, scope_params = customer_scope_filter("c")
    where_sql = f"WHERE {scope_clause}" if scope_clause else ""
    cust_raw = df_query(
        conn,
        f"""
        SELECT
            c.customer_id as id,
            c.name,
            c.company_name,
            c.phone,
            c.address,
            c.delivery_address,
            c.remarks,
            c.purchase_date,
            c.product_info,
            c.delivery_order_code,
            c.sales_person,
            c.amount_spent,
            c.dup_flag,
            c.created_at
        FROM customers c
        {where_sql}
        ORDER BY datetime(c.created_at) DESC
        """,
        scope_params if scope_clause else (),
    )
    warr = df_query(
        conn,
        "SELECT w.warranty_id as id, c.name as customer, p.name as product, p.model, w.serial, w.issue_date, w.expiry_date, w.remarks, w.dup_flag FROM warranties w LEFT JOIN customers c ON c.customer_id = w.customer_id LEFT JOIN products p ON p.product_id = w.product_id ORDER BY date(w.issue_date) DESC",
    )
    duplicate_customers = pd.DataFrame()
    if not cust_raw.empty:
        duplicate_customers = cust_raw[cust_raw["dup_flag"] == 1].copy()
    if duplicate_customers.empty:
        st.success("No customer duplicates detected at the moment.")
    else:
        editor_df = duplicate_customers.copy()
        editor_df["__group_key"] = [
            " | ".join(
                [
                    clean_text(row.get("phone")) or "(no phone)",
                    (
                        parse_date_value(row.get("purchase_date")).strftime(DATE_FMT)
                        if parse_date_value(row.get("purchase_date")) is not None
                        else "-"
                    ),
                    clean_text(row.get("product_info")) or "-",
                ]
            )
            for _, row in editor_df.iterrows()
        ]
        preview_df = editor_df.assign(
            duplicate="🔁 duplicate phone",
            purchase_date_fmt=pd.to_datetime(editor_df["purchase_date"], errors="coerce").dt.strftime(DATE_FMT),
            created_at_fmt=pd.to_datetime(editor_df["created_at"], errors="coerce").dt.strftime("%d-%m-%Y %H:%M"),
        )
        if "amount_spent" in preview_df.columns:
            def _format_amount_cell(val: object) -> str:
                if val in (None, ""):
                    return ""
                try:
                    if pd.isna(val):
                        return ""
                except Exception:
                    pass
                return format_money(val) or f"{_coerce_float(val, 0.0):,.2f}"

            preview_df["amount_spent_fmt"] = preview_df["amount_spent"].apply(
                _format_amount_cell
            )
        preview_cols = [
            col
            for col in [
                "__group_key",
                "name",
                "company_name",
                "phone",
                "address",
                "delivery_address",
                "remarks",
                "purchase_date_fmt",
                "product_info",
                "sales_person",
                "created_at_fmt",
            ]
            if col in preview_df.columns
        ]
        if preview_cols:
            display_df = (
                preview_df[preview_cols]
                .rename(
                    columns={
                        "__group_key": "Duplicate set",
                        "company_name": "Company",
                        "purchase_date_fmt": "Purchase date",
                        "product_info": "Product",
                        "delivery_address": "Delivery address",
                        "remarks": "Remarks",
                        "sales_person": "Sales person",
                        "created_at_fmt": "Created",
                    }
                )
                .sort_values(by=["Duplicate set", "Created"], na_position="last")
            )
            display_df["Purchase date"] = display_df["Purchase date"].fillna("-")
            display_df["Created"] = display_df["Created"].fillna("-")
            st.markdown("#### Duplicate rows")
            st.caption(
                "Each duplicate set groups rows sharing the same phone, purchase date, and product so you can double-check real multi-unit sales."
            )
            st.dataframe(display_df, use_container_width=True, hide_index=True)
            combined_preview = (
                preview_df.groupby(["name", "phone"], dropna=False)
                .apply(
                    lambda g: pd.Series(
                        {
                            "Address": dedupe_join(
                                [
                                    clean_text(val)
                                    for val in g.get("address", pd.Series(dtype=object)).tolist()
                                    if clean_text(val)
                                ]
                            ),
                            "Products": dedupe_join(
                                [
                                    clean_text(val)
                                    for val in g.get("product_info", pd.Series(dtype=object)).tolist()
                                    if clean_text(val)
                                ]
                            ),
                            "Purchase dates": dedupe_join(
                                [
                                    val
                                    for val in g.get("purchase_date_fmt", pd.Series(dtype=object)).tolist()
                                    if val
                                ]
                            ),
                            "Created": dedupe_join(
                                [
                                    val
                                    for val in g.get("created_at_fmt", pd.Series(dtype=object)).tolist()
                                    if val
                                ]
                            ),
                        }
                    )
                )
                .reset_index()
            )
            if not combined_preview.empty:
                combined_preview = combined_preview.rename(
                    columns={"name": "Name", "phone": "Phone"}
                )
                st.caption("Condensed view (grouped by customer and phone)")
                st.dataframe(
                    combined_preview,
                    use_container_width=True,
                    hide_index=True,
                )
        group_counts = editor_df.groupby("__group_key").size().to_dict()
        selection_options = [(None, "All duplicate rows")] + [
            (label, f"{label} ({group_counts.get(label, 0)} row(s))") for label in sorted(editor_df["__group_key"].unique())
        ]
        selected_group, _ = st.selectbox(
            "Focus on a duplicate set (optional)",
            options=selection_options,
            index=0,
            format_func=lambda opt: opt[1],
        )
        if selected_group:
            editor_df = editor_df[editor_df["__group_key"] == selected_group]
        if editor_df.empty:
            st.info("No rows match the selected duplicate set.")
        else:
            editor_df["duplicate"] = "🔁 duplicate phone"
            editor_df["purchase_date"] = pd.to_datetime(editor_df["purchase_date"], errors="coerce")
            editor_df["created_at"] = pd.to_datetime(editor_df["created_at"], errors="coerce")
            editor_df["Action"] = "Keep"
            st.markdown("#### Edit duplicate entries")
            header_cols = st.columns(
                (
                    1.4,
                    1.2,
                    1.1,
                    1.6,
                    1.6,
                    1.4,
                    1.1,
                    1.1,
                    1.6,
                    0.9,
                )
            )
            header_cols[0].write("**Name**")
            header_cols[1].write("**Company**")
            header_cols[2].write("**Phone**")
            header_cols[3].write("**Address**")
            header_cols[4].write("**Delivery address**")
            header_cols[5].write("**Remarks**")
            header_cols[6].write("**Sales person**")
            header_cols[7].write("**Purchase date**")
            header_cols[8].write("**Product**")
            header_cols[9].write("**Action**")
            editor_rows = []
            for _, row in editor_df.iterrows():
                cid = int_or_none(row.get("id"))
                if cid is None:
                    continue
                row_cols = st.columns(
                    (
                        1.4,
                        1.2,
                        1.1,
                        1.6,
                        1.6,
                        1.4,
                        1.1,
                        1.1,
                        1.6,
                        0.9,
                    )
                )
                name_key = f"dup_name_{cid}"
                company_key = f"dup_company_{cid}"
                phone_key = f"dup_phone_{cid}"
                address_key = f"dup_address_{cid}"
                delivery_address_key = f"dup_delivery_address_{cid}"
                remarks_key = f"dup_remarks_{cid}"
                sales_key = f"dup_sales_{cid}"
                purchase_key = f"dup_purchase_{cid}"
                product_key = f"dup_product_{cid}"
                action_key = f"dup_action_{cid}"
                purchase_date_value = row.get("purchase_date")
                purchase_date_label = ""
                if isinstance(purchase_date_value, pd.Timestamp) and not pd.isna(purchase_date_value):
                    purchase_date_label = purchase_date_value.strftime(DATE_FMT)
                row_cols[0].text_input(
                    "Name",
                    value=clean_text(row.get("name")) or "",
                    key=name_key,
                    label_visibility="collapsed",
                )
                row_cols[1].text_input(
                    "Company",
                    value=clean_text(row.get("company_name")) or "",
                    key=company_key,
                    label_visibility="collapsed",
                )
                row_cols[2].text_input(
                    "Phone",
                    value=clean_text(row.get("phone")) or "",
                    key=phone_key,
                    label_visibility="collapsed",
                )
                row_cols[3].text_input(
                    "Address",
                    value=clean_text(row.get("address")) or "",
                    key=address_key,
                    label_visibility="collapsed",
                )
                row_cols[4].text_input(
                    "Delivery address",
                    value=clean_text(row.get("delivery_address")) or "",
                    key=delivery_address_key,
                    label_visibility="collapsed",
                )
                row_cols[5].text_input(
                    "Remarks",
                    value=clean_text(row.get("remarks")) or "",
                    key=remarks_key,
                    label_visibility="collapsed",
                )
                row_cols[6].text_input(
                    "Sales person",
                    value=clean_text(row.get("sales_person")) or "",
                    key=sales_key,
                    label_visibility="collapsed",
                )
                row_cols[7].text_input(
                    "Purchase date",
                    value=purchase_date_label,
                    key=purchase_key,
                    label_visibility="collapsed",
                )
                row_cols[8].text_input(
                    "Product",
                    value=clean_text(row.get("product_info")) or "",
                    key=product_key,
                    label_visibility="collapsed",
                )
                row_cols[9].selectbox(
                    "Action",
                    options=["Keep", "Delete"],
                    key=action_key,
                    label_visibility="collapsed",
                )
                editor_rows.append(
                    {
                        "id": cid,
                        "name": st.session_state.get(name_key),
                        "company_name": st.session_state.get(company_key),
                        "phone": st.session_state.get(phone_key),
                        "address": st.session_state.get(address_key),
                        "delivery_address": st.session_state.get(delivery_address_key),
                        "remarks": st.session_state.get(remarks_key),
                        "sales_person": st.session_state.get(sales_key),
                        "purchase_date": st.session_state.get(purchase_key),
                        "product_info": st.session_state.get(product_key),
                        "Action": st.session_state.get(action_key),
                    }
                )
            user = st.session_state.user or {}
            is_admin = user.get("role") == "admin"
            if not is_admin:
                st.caption("Deleting rows requires admin privileges; non-admin delete actions will be ignored.")
            raw_map = {int(row["id"]): row for row in duplicate_customers.to_dict("records") if int_or_none(row.get("id")) is not None}
            if st.button("Apply duplicate table updates", type="primary"):
                editor_result = pd.DataFrame(editor_rows)
                if editor_result.empty:
                    st.info("No rows to update.")
                else:
                    phones_to_recalc: set[str] = set()
                    updates = deletes = 0
                    errors: list[str] = []
                    made_updates = False
                    for row in editor_result.to_dict("records"):
                        cid = int_or_none(row.get("id"))
                        if cid is None or cid not in raw_map:
                            continue
                        action = str(row.get("Action") or "Keep").strip().lower()
                        if action == "delete":
                            if is_admin:
                                delete_customer_record(conn, cid)
                                deletes += 1
                            else:
                                errors.append(f"Only admins can delete customers (ID #{cid}).")
                            continue
                        new_name = clean_text(row.get("name"))
                        new_company = clean_text(row.get("company_name"))
                        new_phone = clean_text(row.get("phone"))
                        new_address = clean_text(row.get("address"))
                        new_delivery_address = clean_text(row.get("delivery_address"))
                        new_remarks = clean_text(row.get("remarks"))
                        new_sales_person = clean_text(row.get("sales_person"))
                        purchase_str, _ = date_strings_from_input(row.get("purchase_date"))
                        product_label = clean_text(row.get("product_info"))
                        original_row = raw_map[cid]
                        old_name = clean_text(original_row.get("name"))
                        old_company = clean_text(original_row.get("company_name"))
                        old_phone = clean_text(original_row.get("phone"))
                        old_address = clean_text(original_row.get("address"))
                        old_delivery_address = clean_text(original_row.get("delivery_address"))
                        old_remarks = clean_text(original_row.get("remarks"))
                        old_sales_person = clean_text(original_row.get("sales_person"))
                        old_purchase = clean_text(original_row.get("purchase_date"))
                        old_product = clean_text(original_row.get("product_info"))
                        old_amount = parse_amount(original_row.get("amount_spent"))
                        old_do = clean_text(original_row.get("delivery_order_code"))
                        new_amount = old_amount
                        new_do = old_do
                        if (
                            new_name == old_name
                            and new_company == old_company
                            and new_phone == old_phone
                            and new_address == old_address
                            and new_delivery_address == old_delivery_address
                            and new_remarks == old_remarks
                            and new_sales_person == old_sales_person
                            and purchase_str == old_purchase
                            and product_label == old_product
                        ):
                            continue
                        conn.execute(
                            """
                            UPDATE customers
                               SET name=?,
                                   company_name=?,
                                   phone=?,
                                   address=?,
                                   delivery_address=?,
                                   remarks=?,
                                   purchase_date=?,
                                   product_info=?,
                                   delivery_order_code=?,
                                   sales_person=?,
                                   amount_spent=?,
                                   dup_flag=0
                             WHERE customer_id=?
                            """,
                            (
                                new_name,
                                new_company,
                                new_phone,
                                new_address,
                                new_delivery_address,
                                new_remarks,
                                purchase_str,
                                product_label,
                                new_do,
                                new_sales_person,
                                new_amount,
                                cid,
                            ),
                        )
                        if new_do:
                            conn.execute(
                                """
                                INSERT INTO delivery_orders (do_number, customer_id, order_id, description, sales_person, remarks, file_path)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                                ON CONFLICT(do_number) DO UPDATE SET
                                    customer_id=excluded.customer_id,
                                    description=excluded.description,
                                    remarks=excluded.remarks
                                """,
                                (
                                    new_do,
                                    cid,
                                    None,
                                    product_label,
                                    new_sales_person,
                                    new_remarks,
                                    None,
                                ),
                            )
                        if old_do and old_do != new_do:
                            conn.execute(
                                "DELETE FROM delivery_orders WHERE do_number=? AND (customer_id IS NULL OR customer_id=?)",
                                (old_do, cid),
                            )
                        conn.execute(
                            """
                            UPDATE import_history
                               SET customer_name=?,
                                   phone=?,
                                   address=?,
                                   delivery_address=?,
                                   product_label=?,
                                   notes=?,
                                   amount_spent=?,
                                   do_number=?,
                                   original_date=?
                             WHERE customer_id=? AND deleted_at IS NULL
                            """,
                            (
                                new_name,
                                new_phone,
                                new_address,
                                new_delivery_address,
                                product_label,
                                new_remarks,
                                new_amount,
                                new_do,
                                purchase_str,
                                cid,
                            ),
                        )
                        if old_phone and old_phone != new_phone:
                            phones_to_recalc.add(old_phone)
                        if new_phone:
                            phones_to_recalc.add(new_phone)
                        updates += 1
                        made_updates = True
                    if made_updates:
                        conn.commit()
                    if phones_to_recalc:
                        for phone_value in phones_to_recalc:
                            recalc_customer_duplicate_flag(conn, phone_value)
                        conn.commit()
                    if errors:
                        for err in errors:
                            st.error(err)
                    if updates or deletes:
                        st.success(f"Updated {updates} row(s) and deleted {deletes} row(s).")
                        if not errors:
                            _safe_rerun()
                    elif not errors:
                        st.info("No changes detected.")
    manual_merge_section(conn, cust_raw)

    if not warr.empty:
        warr = fmt_dates(warr, ["issue_date", "expiry_date"])
        warr = warr.assign(duplicate=warr["dup_flag"].apply(lambda x: "🔁 duplicate serial" if int(x)==1 else ""))
        st.markdown("**Warranties (duplicate serial)**")
        st.dataframe(
            warr[warr["dup_flag"] == 1].drop(columns=["id", "dup_flag"], errors="ignore"),
            use_container_width=True,
        )


def scraps_duplicates_page(conn):
    tabs = st.tabs(["Scraps", "Duplicates"])
    with tabs[0]:
        scraps_page(conn)
    with tabs[1]:
        duplicates_page(conn)


def users_admin_page(conn):
    ensure_auth(role="admin")
    st.subheader("👤 Users (Admin)")
    users = df_query(
        conn,
        """
        SELECT user_id as id, username, phone, email, title, role, created_at
        FROM users
        ORDER BY datetime(created_at) DESC
        """,
    )
    users = users.assign(created_at=pd.to_datetime(users["created_at"], errors="coerce").dt.strftime(DATE_FMT))
    st.dataframe(users.drop(columns=["id"], errors="ignore"))

    with st.expander("Add user"):
        with st.form("add_user"):
            u = st.text_input("Username")
            p = st.text_input("Password", type="password")
            phone = st.text_input("Phone number (required)")
            email = st.text_input("Email (optional)")
            title = st.text_input("Title / role", help="Shown on quotations by default")
            role = st.selectbox("Role", ["staff", "admin"])
            ok = st.form_submit_button("Create")
            if ok and u.strip() and p.strip():
                if not phone.strip():
                    st.error("Phone number is required for staff accounts.")
                    return
                h = hashlib.sha256(p.encode("utf-8")).hexdigest()
                try:
                    conn.execute(
                        "INSERT INTO users (username, pass_hash, phone, email, title, role) VALUES (?, ?, ?, ?, ?, ?)",
                        (
                            u.strip(),
                            h,
                            clean_text(phone),
                            clean_text(email),
                            clean_text(title),
                            role,
                        ),
                    )
                    conn.commit()
                    st.success("User added")
                except sqlite3.IntegrityError:
                    st.error("Username already exists")

    with st.expander("Reset password / delete"):
        uid = st.number_input("User ID", min_value=1, step=1)
        newp = st.text_input("New password", type="password")
        col1, col2 = st.columns(2)
        if col1.button("Set new password"):
            h = hashlib.sha256(newp.encode("utf-8")).hexdigest()
            conn.execute("UPDATE users SET pass_hash=? WHERE user_id=?", (h, int(uid)))
            conn.commit()
            st.success("Password updated")
        if col2.button("Delete user"):
            role_row = conn.execute(
                "SELECT role FROM users WHERE user_id=?",
                (int(uid),),
            ).fetchone()
            role_value = clean_text(role_row[0]) if role_row else ""
            if role_value.lower() == "admin":
                st.error("Admin users cannot be deleted.")
            else:
                conn.execute("DELETE FROM users WHERE user_id=?", (int(uid),))
                conn.commit()
                st.warning("User deleted")

# ---------- Import engine ----------
def _import_clean6(conn, df, tag="Import"):
    """Import cleaned dataframe into database.

    The function is resilient to messy input: it normalizes and sorts the
    dataframe internally so callers can pass raw data without pre-processing.
    """
    # ensure dataframe is normalized even if caller didn't pre-clean
    df = df.copy()
    df = refine_multiline(df)
    if "date" in df.columns:
        df["date"] = coerce_excel_date(df["date"])
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
    sort_cols = [
        col
        for col in ["purchase_date", "date", "customer_name", "phone", "do_code"]
        if col in df.columns
    ]
    if not sort_cols:
        sort_cols = df.columns.tolist()
    df = df.dropna(how="all").drop_duplicates()
    df = _sort_dataframe_safe(df, sort_cols).reset_index(drop=True)

    cur = conn.cursor()
    seeded = 0
    d_c = d_p = 0
    phones_to_recalc: set[str] = set()
    created_by = current_user_id()
    for _, r in df.iterrows():
        purchase_input = r.get("purchase_date") if "purchase_date" in df.columns else None
        try:
            if pd.isna(purchase_input):
                purchase_input = None
        except Exception:
            pass
        if purchase_input is None:
            purchase_input = r.get("date", pd.NaT)
        d = purchase_input
        cust = clean_text(r.get("customer_name"))
        addr = clean_text(r.get("address"))
        delivery_addr = clean_text(r.get("delivery_address"))
        phone = clean_text(r.get("phone"))
        product_label = clean_text(r.get("product"))
        do_serial = clean_text(r.get("do_code"))
        work_done_code = clean_text(r.get("work_done_code"))
        service_code = clean_text(r.get("service_code"))
        maintenance_code = clean_text(r.get("maintenance_code"))
        follow_up_input = r.get("follow_up_date")
        remarks_val = clean_text(r.get("remarks"))
        amount_value = parse_amount(r.get("amount_spent"))
        quantity_value = parse_quantity(r.get("quantity"), default=1)
        if cust is None and phone is None and product_label is None:
            continue
        purchase_dt = parse_date_value(d)
        purchase_str = purchase_dt.strftime("%Y-%m-%d") if isinstance(purchase_dt, pd.Timestamp) else None
        # dup checks
        def exists_phone(phone_value, purchase_value, do_value, product_value):
            normalized_phone = clean_text(phone_value)
            if not normalized_phone:
                return False
            clauses = ["phone = ?"]
            params: list[object] = [normalized_phone]
            if purchase_value:
                clauses.append("IFNULL(purchase_date, '') = ?")
                params.append(purchase_value)
            else:
                clauses.append("(purchase_date IS NULL OR purchase_date = '')")
            if do_value:
                clauses.append("LOWER(IFNULL(delivery_order_code, '')) = LOWER(?)")
                params.append(do_value)
            elif product_value:
                clauses.append("LOWER(IFNULL(product_info, '')) = LOWER(?)")
                params.append(product_value)
            query = f"SELECT 1 FROM customers WHERE {' AND '.join(clauses)} LIMIT 1"
            cur.execute(query, tuple(params))
            return cur.fetchone() is not None

        dupc = 1 if exists_phone(phone, purchase_str, do_serial, product_label) else 0
        cur.execute(
            "INSERT INTO customers (name, phone, address, delivery_address, remarks, amount_spent, created_by, dup_flag) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                cust,
                phone,
                addr,
                delivery_addr,
                remarks_val,
                amount_value,
                created_by,
                dupc,
            ),
        )
        cid = cur.lastrowid
        if dupc:
            d_c += 1
        if phone:
            normalized_phone = clean_text(phone)
            if normalized_phone:
                phones_to_recalc.add(normalized_phone)

        base_dt = purchase_dt or pd.Timestamp.now().normalize()
        pid = None
        oid = None
        order_item_id = None
        warranty_id = None
        name, model = split_product_label(product_label)
        if name:
            def exists_prod(name, model):
                cur.execute(
                    "SELECT 1 FROM products WHERE name = ? AND IFNULL(model,'') = IFNULL(?, '') LIMIT 1",
                    (name, model),
                )
                return cur.fetchone() is not None

            dupp = 1 if exists_prod(name, model) else 0
            cur.execute(
                "INSERT INTO products (name, model, dup_flag) VALUES (?, ?, ?)",
                (name, model, dupp),
            )
            pid = cur.lastrowid
            if dupp:
                d_p += 1

            # we still record orders (hidden) to keep a timeline if needed
            order_date = base_dt
            delivery_date = base_dt
            cur.execute(
                "INSERT INTO orders (customer_id, order_date, delivery_date, notes) VALUES (?, ?, ?, ?)",
                (
                    cid,
                    order_date.strftime("%Y-%m-%d") if order_date is not None else None,
                    delivery_date.strftime("%Y-%m-%d") if delivery_date is not None else None,
                    f"Imported ({tag})",
                ),
            )
            oid = cur.lastrowid
            cur.execute(
                "INSERT INTO order_items (order_id, product_id, quantity) VALUES (?, ?, ?)",
                (oid, pid, quantity_value),
            )
            order_item_id = cur.lastrowid

            base = base_dt
            expiry = base + pd.Timedelta(days=365)
            cur.execute(
                "INSERT INTO warranties (customer_id, product_id, serial, issue_date, expiry_date, status, remarks, dup_flag) VALUES (?, ?, ?, ?, ?, 'active', ?, 0)",
                (
                    cid,
                    pid,
                    None,
                    base.strftime("%Y-%m-%d"),
                    expiry.strftime("%Y-%m-%d"),
                    remarks_val,
                ),
            )
            warranty_id = cur.lastrowid

        if do_serial and oid is not None:
            description = product_label
            cur.execute(
                "INSERT OR IGNORE INTO delivery_orders (do_number, customer_id, order_id, description, sales_person, remarks, file_path) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    do_serial,
                    cid,
                    oid,
                    description,
                    None,
                    remarks_val,
                    None,
                ),
            )
        purchase_date = purchase_str or (
            base_dt.strftime("%Y-%m-%d") if isinstance(base_dt, pd.Timestamp) else None
        )
        cur.execute(
            "UPDATE customers SET purchase_date=?, product_info=?, delivery_order_code=?, remarks=?, amount_spent=?, delivery_address=? WHERE customer_id=?",
            (
                purchase_date,
                product_label,
                do_serial,
                remarks_val,
                amount_value,
                delivery_addr,
                cid,
            ),
        )
        follow_up_date = parse_date_value(follow_up_input)
        follow_up_str = (
            follow_up_date.strftime("%Y-%m-%d")
            if isinstance(follow_up_date, pd.Timestamp)
            else None
        )
        if follow_up_str:
            cur.execute(
                "INSERT INTO customer_notes (customer_id, note, remind_on) VALUES (?, ?, ?)",
                (
                    cid,
                    "Follow-up",
                    follow_up_str,
                ),
            )
        if work_done_code:
            cur.execute(
                """
                INSERT OR IGNORE INTO delivery_orders (
                    do_number, customer_id, order_id, description, sales_person, remarks,
                    record_type, status, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, 'work_done', 'pending', ?)
                """,
                (
                    work_done_code,
                    cid,
                    oid,
                    product_label,
                    None,
                    remarks_val,
                    created_by,
                ),
            )
        if service_code:
            cur.execute(
                """
                INSERT INTO services (
                    do_number, customer_id, service_date, description, remarks, service_product_info, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    service_code,
                    cid,
                    purchase_date,
                    product_label or remarks_val,
                    remarks_val,
                    product_label,
                    created_by,
                ),
            )
        if maintenance_code:
            cur.execute(
                """
                INSERT INTO maintenance_records (
                    do_number, customer_id, maintenance_date, description, remarks, maintenance_product_info,
                    total_amount, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    maintenance_code,
                    cid,
                    purchase_date,
                    product_label or remarks_val,
                    remarks_val,
                    product_label,
                    None,
                    created_by,
                ),
            )
        cur.execute(
            "INSERT INTO import_history (customer_id, product_id, order_id, order_item_id, warranty_id, do_number, import_tag, original_date, customer_name, address, phone, product_label, notes, amount_spent, imported_by, delivery_address, quantity) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                cid,
                pid,
                oid,
                order_item_id,
                warranty_id,
                do_serial,
                tag,
                purchase_date,
                cust,
                addr,
                phone,
                product_label,
                remarks_val,
                amount_value,
                created_by,
                delivery_addr,
                quantity_value,
            ),
        )
        seeded += 1
    conn.commit()
    for p in phones_to_recalc:
        recalc_customer_duplicate_flag(conn, p)
    conn.commit()
    return seeded, d_c, d_p


def update_import_entry(conn, record: dict, updates: dict) -> None:
    cur = conn.cursor()
    import_id = int_or_none(record.get("import_id"))
    if import_id is None:
        return

    customer_id = int_or_none(record.get("customer_id"))
    product_id = int_or_none(record.get("product_id"))
    order_id = int_or_none(record.get("order_id"))
    order_item_id = int_or_none(record.get("order_item_id"))
    warranty_id = int_or_none(record.get("warranty_id"))

    old_phone = clean_text(record.get("live_phone")) or clean_text(record.get("phone"))
    old_do = clean_text(record.get("do_number"))

    new_name = clean_text(updates.get("customer_name"))
    new_phone = clean_text(updates.get("phone"))
    new_address = clean_text(updates.get("address"))
    new_delivery_address = clean_text(updates.get("delivery_address"))
    purchase_date_str, expiry_str = date_strings_from_input(updates.get("purchase_date"))
    product_label = clean_text(updates.get("product_label"))
    new_do = clean_text(updates.get("do_number"))
    new_remarks = clean_text(updates.get("remarks"))
    new_amount = parse_amount(updates.get("amount_spent"))
    quantity_value = parse_quantity(updates.get("quantity"), default=1)
    product_name, product_model = split_product_label(product_label)

    if customer_id is not None:
        cur.execute(
            "UPDATE customers SET name=?, phone=?, address=?, delivery_address=?, purchase_date=?, product_info=?, delivery_order_code=?, remarks=?, amount_spent=?, dup_flag=0 WHERE customer_id=?",
            (
                new_name,
                new_phone,
                new_address,
                new_delivery_address,
                purchase_date_str,
                product_label,
                new_do,
                new_remarks,
                new_amount,
                customer_id,
            ),
        )

    if order_id is not None:
        cur.execute(
            "UPDATE orders SET order_date=?, delivery_date=? WHERE order_id=?",
            (purchase_date_str, purchase_date_str, order_id),
        )

    if order_item_id is not None:
        cur.execute(
            "UPDATE order_items SET quantity=? WHERE order_item_id=?",
            (quantity_value, order_item_id),
        )

    if product_id is not None:
        cur.execute(
            "UPDATE products SET name=?, model=? WHERE product_id=?",
            (product_name, product_model, product_id),
        )

    if warranty_id is not None:
        cur.execute(
            "UPDATE warranties SET issue_date=?, expiry_date=?, status='active', remarks=? WHERE warranty_id=?",
            (purchase_date_str, expiry_str, new_remarks, warranty_id),
        )

    if new_do:
        cur.execute(
            """
            INSERT INTO delivery_orders (do_number, customer_id, order_id, description, sales_person, remarks, file_path)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(do_number) DO UPDATE SET
                customer_id=excluded.customer_id,
                order_id=excluded.order_id,
                description=excluded.description,
                remarks=excluded.remarks
            """,
            (
                new_do,
                customer_id,
                order_id,
                product_label,
                None,
                new_remarks,
                None,
            ),
        )
    if old_do and old_do != new_do:
        params = [old_do]
        query = "DELETE FROM delivery_orders WHERE do_number=?"
        if order_id is not None:
            query += " AND (order_id IS NULL OR order_id=?)"
            params.append(order_id)
        cur.execute(query, tuple(params))

    cur.execute(
        "UPDATE import_history SET original_date=?, customer_name=?, address=?, delivery_address=?, phone=?, product_label=?, do_number=?, notes=?, amount_spent=?, quantity=? WHERE import_id=?",
        (
            purchase_date_str,
            new_name,
            new_address,
            new_delivery_address,
            new_phone,
            product_label,
            new_do,
            new_remarks,
            new_amount,
            quantity_value,
            import_id,
        ),
    )
    conn.commit()

    if old_phone and old_phone != new_phone:
        recalc_customer_duplicate_flag(conn, old_phone)
    if new_phone:
        recalc_customer_duplicate_flag(conn, new_phone)
    conn.commit()


def delete_import_entry(conn, record: dict) -> None:
    cur = conn.cursor()
    import_id = int_or_none(record.get("import_id"))
    if import_id is None:
        return
    deleted_by = current_user_id()

    customer_id = int_or_none(record.get("customer_id"))
    product_id = int_or_none(record.get("product_id"))
    order_id = int_or_none(record.get("order_id"))
    order_item_id = int_or_none(record.get("order_item_id"))
    warranty_id = int_or_none(record.get("warranty_id"))
    do_number = clean_text(record.get("do_number"))
    attachment_path = record.get("live_attachment_path")

    old_phone = clean_text(record.get("live_phone")) or clean_text(record.get("phone"))

    if do_number:
        params = [do_number]
        query = "DELETE FROM delivery_orders WHERE do_number=?"
        if order_id is not None:
            query += " AND (order_id IS NULL OR order_id=?)"
            params.append(order_id)
        cur.execute(query, tuple(params))

    if warranty_id is not None:
        cur.execute("DELETE FROM warranties WHERE warranty_id=?", (warranty_id,))
    if order_item_id is not None:
        cur.execute("DELETE FROM order_items WHERE order_item_id=?", (order_item_id,))
    if order_id is not None:
        cur.execute("DELETE FROM orders WHERE order_id=?", (order_id,))
    if product_id is not None:
        cur.execute("DELETE FROM products WHERE product_id=?", (product_id,))
    if customer_id is not None:
        cur.execute("DELETE FROM customers WHERE customer_id=?", (customer_id,))

    cur.execute(
        "UPDATE import_history SET deleted_at = datetime('now'), deleted_by=? WHERE import_id=?",
        (deleted_by, import_id),
    )
    conn.commit()

    if attachment_path:
        path = resolve_upload_path(attachment_path)
        if path and path.exists():
            try:
                path.unlink()
            except Exception:
                pass

    if old_phone:
        recalc_customer_duplicate_flag(conn, old_phone)
        conn.commit()


def delete_work_report(
    conn,
    *,
    report_id: int,
    owner_id: Optional[int],
    owner_label: Optional[str],
    period_type: Optional[str],
    period_start: Optional[str],
    period_end: Optional[str],
    report_template: Optional[str],
) -> None:
    conn.execute("DELETE FROM work_reports WHERE report_id=?", (int(report_id),))
    conn.commit()
    template_key = _normalize_report_template(report_template)
    template_label = REPORT_TEMPLATE_LABELS.get(template_key, "Service report")
    cadence_label = format_period_label(period_type)
    period_label = format_period_range(period_start, period_end)
    owner_text = owner_label or (
        f"User #{int(owner_id)}" if owner_id is not None else "team member"
    )
    description = (
        f"Deleted {template_label.lower()} {cadence_label.lower()} report for "
        f"{owner_text} ({period_label})"
    )
    log_activity(
        conn,
        event_type="report_deleted",
        description=description,
        entity_type="report",
        entity_id=int(report_id),
        user_id=current_user_id(),
    )


def _normalize_report_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    cleaned = clean_text(value)
    return cleaned


def _lookup_customer_id_by_name(conn, customer_name: Optional[str]) -> Optional[int]:
    name_value = clean_text(customer_name)
    if not name_value:
        return None
    row = conn.execute(
        """
        SELECT customer_id
        FROM customers
        WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))
           OR LOWER(TRIM(company_name)) = LOWER(TRIM(?))
        ORDER BY customer_id ASC
        LIMIT 1
        """,
        (name_value, name_value),
    ).fetchone()
    if row:
        return int(row[0])
    return None


def _sync_report_payment_records(
    conn,
    *,
    report_id: int,
    report_owner_id: int,
    template_key: str,
    grid_rows: Iterable[dict],
    period_end: str,
) -> None:
    if template_key not in {"service", "follow_up"}:
        return

    normalized_rows = _normalize_grid_rows(grid_rows, template_key=template_key)
    if not normalized_rows:
        return

    for idx, row in enumerate(normalized_rows):
        if template_key == "service":
            status_value = clean_text(row.get("payment_status")) or "pending"
            customer_name = clean_text(row.get("customer_name"))
            description = clean_text(row.get("reported_complaints")) or clean_text(
                row.get("details_remarks")
            )
            remarks = clean_text(row.get("details_remarks"))
            product_info = clean_text(row.get("product_details"))
            service_status = clean_text(row.get("status")) or DEFAULT_SERVICE_STATUS
            service_date = row.get("work_done_date") or period_end
            bill_amount = _coerce_float(row.get("bill_tk"), 0.0)
        else:
            status_value = clean_text(row.get("status")) or "pending"
            customer_name = clean_text(row.get("client_name"))
            description = clean_text(row.get("notes")) or clean_text(row.get("product_detail"))
            remarks = clean_text(row.get("notes"))
            product_info = clean_text(row.get("product_detail"))
            service_status = DEFAULT_SERVICE_STATUS
            service_date = row.get("follow_up_date") or period_end
            bill_amount = None

        normalized_status = status_value.strip().lower()
        if normalized_status not in {"paid", "pending"}:
            normalized_status = "pending"

        customer_id = _lookup_customer_id_by_name(conn, customer_name)
        existing = conn.execute(
            """
            SELECT service_id
            FROM services
            WHERE report_id=? AND report_row_index=?
              AND deleted_at IS NULL
            LIMIT 1
            """,
            (int(report_id), int(idx)),
        ).fetchone()

        if existing:
            conn.execute(
                """
                UPDATE services
                SET customer_id=COALESCE(?, customer_id),
                    service_date=?,
                    service_start_date=?,
                    service_end_date=?,
                    description=?,
                    status=?,
                    remarks=?,
                    service_product_info=?,
                    payment_status=?,
                    bill_amount=COALESCE(?, bill_amount),
                    updated_at=datetime('now')
                WHERE service_id=?
                  AND deleted_at IS NULL
                """,
                (
                    customer_id,
                    service_date,
                    service_date,
                    service_date,
                    description,
                    service_status,
                    remarks,
                    product_info,
                    normalized_status,
                    bill_amount,
                    int(existing[0]),
                ),
            )
            continue

        if normalized_status != "paid":
            continue

        conn.execute(
            """
            INSERT INTO services (
                do_number,
                customer_id,
                service_date,
                service_start_date,
                service_end_date,
                description,
                status,
                remarks,
                service_product_info,
                payment_status,
                bill_amount,
                report_id,
                report_row_index,
                updated_at,
                created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), ?)
            """,
            (
                None,
                customer_id,
                service_date,
                service_date,
                service_date,
                description,
                service_status,
                remarks,
                product_info,
                normalized_status,
                bill_amount,
                int(report_id),
                int(idx),
                report_owner_id,
            ),
        )

    conn.commit()


def upsert_work_report(
    conn,
    *,
    report_id: Optional[int],
    user_id: int,
    period_type: str,
    period_start,
    period_end,
    tasks: Optional[str],
    remarks: Optional[str],
    research: Optional[str],
    report_template: Optional[str] = None,
    grid_rows: Optional[Iterable[dict]] = None,
    attachment_path=_ATTACHMENT_UNCHANGED,
    current_attachment: Optional[str] = None,
    import_file_path=_ATTACHMENT_UNCHANGED,
    current_import_file: Optional[str] = None,
) -> int:
    if user_id is None:
        raise ValueError("User is required to save a report.")

    key, normalized_start, normalized_end = normalize_report_window(
        period_type, period_start, period_end
    )
    start_iso = normalized_start.isoformat()
    end_iso = normalized_end.isoformat()

    tasks_val = _normalize_report_text(tasks)
    remarks_val = _normalize_report_text(remarks)
    research_val = _normalize_report_text(research)
    template_key = _normalize_report_template(report_template)
    grid_payload_val = prepare_report_grid_payload(
        grid_rows or [], template_key=template_key
    )

    cur = conn.cursor()
    effective_id = report_id
    cur.execute(
        """
        SELECT report_id, attachment_path, import_file_path
        FROM work_reports
        WHERE user_id=? AND period_type=? AND period_start=?
        LIMIT 1
        """,
        (user_id, key, start_iso),
    )
    row = cur.fetchone()
    if row:
        existing_id = int(row[0])
        if effective_id is None:
            effective_id = existing_id
            if current_attachment is None:
                current_attachment = row[1]
            if current_import_file is None:
                current_import_file = row[2]
        elif existing_id != effective_id:
            raise ValueError(
                "Another report already exists for this period. Select it from the dropdown to edit."
            )

    if effective_id is not None and current_attachment is None:
        cur.execute(
            "SELECT attachment_path, import_file_path FROM work_reports WHERE report_id=?",
            (effective_id,),
        )
        match = cur.fetchone()
        if match:
            current_attachment = match[0]
            if current_import_file is None:
                current_import_file = match[1]

    if attachment_path is _ATTACHMENT_UNCHANGED:
        attachment_value = current_attachment
    else:
        attachment_value = attachment_path

    if import_file_path is _ATTACHMENT_UNCHANGED:
        import_value = current_import_file
    else:
        import_value = import_file_path

    created_new = False
    if effective_id is None:
        try:
            cur.execute(
                """
                INSERT INTO work_reports (user_id, period_type, period_start, period_end, tasks, remarks, research, grid_payload, attachment_path, import_file_path, report_template)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    user_id,
                    key,
                    start_iso,
                    end_iso,
                    tasks_val,
                    remarks_val,
                    research_val,
                    grid_payload_val,
                    attachment_value,
                    import_value,
                    template_key,
                ),
            )
        except sqlite3.IntegrityError as exc:
            raise ValueError(
                "Another report already exists for this period. Select it from the dropdown to edit."
            ) from exc
        effective_id = int(cur.lastrowid)
        created_new = True
    else:
        try:
            cur.execute(
                """
                UPDATE work_reports
                SET period_type=?, period_start=?, period_end=?, tasks=?, remarks=?, research=?, grid_payload=?, attachment_path=?, import_file_path=?, report_template=?, updated_at=datetime('now')
                WHERE report_id=?
                """,
                (
                    key,
                    start_iso,
                    end_iso,
                    tasks_val,
                    remarks_val,
                    research_val,
                    grid_payload_val,
                    attachment_value,
                    import_value,
                    template_key,
                    effective_id,
                ),
            )
        except sqlite3.IntegrityError as exc:
            raise ValueError(
                "Another report already exists for this period. Select it from the dropdown to edit."
            ) from exc

    conn.commit()
    _sync_report_payment_records(
        conn,
        report_id=int(effective_id),
        report_owner_id=user_id,
        template_key=template_key,
        grid_rows=grid_rows or [],
        period_end=end_iso,
    )
    cadence_label = REPORT_PERIOD_OPTIONS.get(key, key.title())
    period_label = format_period_range(start_iso, end_iso)
    owner_label = None
    try:
        owner_row = conn.execute(
            "SELECT username FROM users WHERE user_id=?",
            (user_id,),
        ).fetchone()
        if owner_row:
            owner_label = clean_text(owner_row[0])
    except sqlite3.Error:
        owner_label = None
    actor = owner_label or f"User #{user_id}"
    event_type = "report_submitted" if created_new else "report_updated"
    verb = "submitted" if created_new else "updated"
    description = f"{actor} {verb} {cadence_label.lower()} report ({period_label})"
    log_activity(
        conn,
        event_type=event_type,
        description=description,
        entity_type="report",
        entity_id=int(effective_id),
        user_id=user_id,
    )
    return effective_id


def manage_import_history(conn):
    st.subheader("🗃️ Manage imported rows")
    user = get_current_user()
    is_admin = user.get("role") == "admin"
    where_parts = ["ih.deleted_at IS NULL"]
    params: list[object] = []
    if not is_admin:
        user_id = current_user_id()
        if user_id is None:
            where_parts.append("1=0")
        else:
            where_parts.append(
                "(ih.imported_by = ? OR (ih.imported_by IS NULL AND c.created_by = ?))"
            )
            params.extend([user_id, user_id])
    where_clause = " AND ".join(where_parts)
    hist = df_query(
        conn,
        f"""
        SELECT ih.*, c.name AS live_customer_name, c.address AS live_address, c.phone AS live_phone,
               c.purchase_date AS live_purchase_date, c.product_info AS live_product_info,
               c.delivery_order_code AS live_do_code, c.delivery_address AS live_delivery_address,
               c.attachment_path AS live_attachment_path, c.created_by AS live_created_by
        FROM import_history ih
        LEFT JOIN customers c ON c.customer_id = ih.customer_id
        WHERE {where_clause}
        ORDER BY ih.import_id DESC
        LIMIT 200
        """,
        tuple(params),
    )
    if hist.empty:
        st.info("No imported rows yet. Upload a file to get started.")
        return

    display_cols = [
        "import_id",
        "customer_name",
        "phone",
        "delivery_address",
        "product_label",
        "do_number",
        "quantity",
        "amount_spent",
    ]
    display = hist.copy()
    display = display[display_cols]
    display.rename(
        columns={
            "import_id": "ID",
            "customer_name": "Customer",
            "phone": "Phone",
            "delivery_address": "Delivery address",
            "product_label": "Product",
            "do_number": "DO code",
            "quantity": "Quantity",
            "amount_spent": "Amount spent",
        },
        inplace=True,
    )
    st.dataframe(display, use_container_width=True)

    ids = hist["import_id"].astype(int).tolist()
    label_map = {}
    for _, row in hist.iterrows():
        name = clean_text(row.get("customer_name")) or clean_text(row.get("live_customer_name")) or "(no name)"
        tag = clean_text(row.get("import_tag")) or "import"
        label_map[int(row["import_id"])] = f"#{int(row['import_id'])} – {name} ({tag})"

    selected_id = st.selectbox(
        "Select an import entry",
        ids,
        format_func=lambda x: label_map.get(int(x), str(x)),
    )
    selected = hist[hist["import_id"] == selected_id].iloc[0].to_dict()
    current_name = clean_text(selected.get("live_customer_name")) or clean_text(selected.get("customer_name")) or ""
    current_phone = clean_text(selected.get("live_phone")) or clean_text(selected.get("phone")) or ""
    current_address = clean_text(selected.get("live_address")) or clean_text(selected.get("address")) or ""
    current_delivery_address = (
        clean_text(selected.get("live_delivery_address"))
        or clean_text(selected.get("delivery_address"))
        or ""
    )
    current_product = clean_text(selected.get("live_product_info")) or clean_text(selected.get("product_label")) or ""
    current_do = clean_text(selected.get("live_do_code")) or clean_text(selected.get("do_number")) or ""
    purchase_seed = selected.get("live_purchase_date") or selected.get("original_date")
    purchase_str = clean_text(purchase_seed) or ""
    amount_seed = selected.get("amount_spent")
    amount_value = parse_amount(amount_seed)
    amount_display = ""
    if amount_value is not None:
        amount_display = format_money(amount_value) or f"{amount_value:,.2f}"
    current_quantity = parse_quantity(selected.get("quantity"), default=1)

    user = st.session_state.user or {}
    is_admin = user.get("role") == "admin"
    viewer_id = current_user_id()
    imported_by = int_or_none(selected.get("imported_by"))
    created_by = int_or_none(selected.get("live_created_by"))
    can_delete = bool(
        is_admin
        or (viewer_id is not None and imported_by is not None and viewer_id == imported_by)
        or (viewer_id is not None and imported_by is None and created_by is not None and viewer_id == created_by)
    )

    with st.form(f"manage_import_{selected_id}"):
        name_input = st.text_input("Customer name", value=current_name)
        phone_input = st.text_input("Phone", value=current_phone)
        address_input = st.text_area("Address", value=current_address)
        delivery_address_input = st.text_area(
            "Delivery address", value=current_delivery_address
        )
        purchase_input = st.text_input("Purchase date (DD-MM-YYYY)", value=purchase_str)
        product_input = st.text_input("Product", value=current_product)
        do_input = st.text_input("Delivery order code", value=current_do)
        remarks_input = st.text_area(
            "Remarks",
            value=clean_text(selected.get("notes")) or "",
            help="Optional remarks stored with this import entry.",
        )
        amount_input = st.text_input(
            "Amount spent",
            value=amount_display,
            help="Track how much was spent for this imported row.",
        )
        quantity_input = st.number_input(
            "Quantity",
            min_value=1,
            value=current_quantity,
            step=1,
            help="How many units were recorded on this import line.",
        )
        col1, col2 = st.columns(2)
        save_btn = col1.form_submit_button("Save changes", type="primary")
        delete_btn = col2.form_submit_button("Delete import", disabled=not can_delete)

    if save_btn:
        try:
            update_import_entry(
                conn,
                selected,
                {
                    "customer_name": name_input,
                    "phone": phone_input,
                    "address": address_input,
                    "delivery_address": delivery_address_input,
                    "purchase_date": purchase_input,
                    "product_label": product_input,
                    "do_number": do_input,
                    "remarks": remarks_input,
                    "amount_spent": amount_input,
                    "quantity": quantity_input,
                },
            )
            conn.execute(
                "UPDATE import_history SET notes=?, amount_spent=? WHERE import_id=?",
                (
                    clean_text(remarks_input),
                    parse_amount(amount_input),
                    int(selected_id),
                ),
            )
            conn.commit()
        except sqlite3.Error as exc:
            st.error(f"Unable to amend this import entry. {exc}")
        else:
            st.success("Import entry updated.")
            _safe_rerun()

    if delete_btn and can_delete:
        delete_import_entry(conn, selected)
        st.warning("Import entry deleted.")
        _safe_rerun()
    elif delete_btn and not can_delete:
        st.error("Only admins can delete import rows.")

# ---------- Reports ----------
def reports_page(conn):
    st.subheader("📈 Work reports")
    user = get_current_user()
    if not user:
        st.info("Log in to capture and review team reports.")
        return

    viewer_id = current_user_id()
    if viewer_id is None:
        st.warning("Unable to determine your account. Please log in again.")
        return

    is_admin = user.get("role") == "admin"
    today = datetime.now().date()
    current_week_start = today - timedelta(days=today.weekday())
    current_week_end = current_week_start + timedelta(days=6)
    current_month_start = date(today.year, today.month, 1)
    current_month_end = date(
        today.year, today.month, monthrange(today.year, today.month)[1]
    )
    st.caption(
        "Staff can see only their own entries. Admins can review every team member's submissions."
    )

    directory = df_query(
        conn,
        "SELECT user_id, username, role FROM users ORDER BY LOWER(username)",
    )
    user_labels: dict[int, str] = {}
    if not directory.empty:
        for _, row in directory.iterrows():
            try:
                uid = int(row["user_id"])
            except Exception:
                continue
            username = clean_text(row.get("username")) or f"User #{uid}"
            role_label = clean_text(row.get("role"))
            if role_label == "admin":
                username = f"{username} (admin)"
            user_labels[uid] = username
    if viewer_id not in user_labels:
        user_labels[viewer_id] = clean_text(user.get("username")) or f"User #{viewer_id}"

    sorted_users = sorted(user_labels.items(), key=lambda item: item[1].lower())
    user_ids = [uid for uid, _ in sorted_users]
    label_map = {uid: label for uid, label in sorted_users}
    if not user_ids:
        user_ids = [viewer_id]
        label_map[viewer_id] = clean_text(user.get("username")) or f"User #{viewer_id}"

    report_owner_id = viewer_id
    if not is_admin:
        st.info(
            f"Recording progress for **{label_map.get(viewer_id, 'you')}**.",
            icon="📝",
        )
        st.caption(
            "Daily entries are limited to today. Weekly and monthly reports can be logged for any selected window."
        )

    def _date_or(value, fallback: date) -> date:
        if value is None:
            return fallback
        parsed_iso = to_iso_date(value)
        if parsed_iso:
            try:
                return datetime.strptime(parsed_iso, "%Y-%m-%d").date()
            except ValueError:
                pass
        try:
            parsed = pd.to_datetime(value, errors="coerce")
        except Exception:
            return fallback
        if pd.isna(parsed):
            return fallback
        if isinstance(parsed, pd.DatetimeIndex):
            if len(parsed) == 0:
                return fallback
            parsed = parsed[0]
        return pd.Timestamp(parsed).date()

    def _staff_report_window_allows_edit(
        row,
        *,
        today: date,
        week_start: date,
        week_end: date,
        month_start: date,
        month_end: date,
    ) -> bool:
        return True


    owner_reports = df_query(
        conn,
        dedent(
            """
            SELECT report_id, user_id, period_type, period_start, period_end, tasks, remarks, research, grid_payload, attachment_path, import_file_path, report_template, created_at, updated_at
            FROM work_reports
            WHERE user_id=?
            ORDER BY date(period_start) DESC, report_id DESC
            LIMIT 50
            """
        ),
        (report_owner_id,),
    )
    record_labels: dict[int, str] = {}
    selectable_reports = owner_reports.copy()
    if not owner_reports.empty:
        selectable_reports["report_id"] = selectable_reports["report_id"].apply(
            lambda val: int(float(val))
        )
        for _, row in selectable_reports.iterrows():
            rid = int(row["report_id"])
            record_labels[rid] = (
                f"{format_period_label(row.get('period_type'))} – "
                f"{format_period_range(row.get('period_start'), row.get('period_end'))}"
            )
        if not is_admin:
            selectable_reports["__staff_can_edit__"] = selectable_reports.apply(
                lambda row: _staff_report_window_allows_edit(
                    row,
                    today=today,
                    week_start=current_week_start,
                    week_end=current_week_end,
                    month_start=current_month_start,
                    month_end=current_month_end,
                ),
                axis=1,
            )
            selectable_reports = selectable_reports[
                selectable_reports["__staff_can_edit__"] == True  # noqa: E712
            ].copy()
            selectable_reports.drop(
                columns=["__staff_can_edit__"],
                inplace=True,
                errors="ignore",
            )

    selectable_ids: list[int] = []
    if not selectable_reports.empty and "report_id" in selectable_reports.columns:
        selectable_ids = [
            int(val) for val in selectable_reports["report_id"].tolist() if not pd.isna(val)
        ]

    def _format_report_choice(value):
        try:
            return record_labels.get(int(value), f"Report #{int(value)}")
        except Exception:
            return "Report"

    if "report_edit_select_pending" in st.session_state:
        pending_selection = st.session_state.pop("report_edit_select_pending")
        if pending_selection is not None:
            try:
                st.session_state["report_edit_select"] = int(pending_selection)
            except Exception:
                st.session_state["report_edit_select"] = None

    current_selection = st.session_state.get("report_edit_select")
    try:
        if current_selection is not None and int(current_selection) not in selectable_ids:
            st.session_state["report_edit_select"] = None
    except Exception:
        st.session_state["report_edit_select"] = None

    selected_report_id = None
    if selectable_ids:
        selected_report_id = st.selectbox(
            "Load an existing report",
            selectable_ids,
            format_func=_format_report_choice,
            key="report_edit_select",
            index=None,
            placeholder="Select a report to edit",
        )

    # Preserve spreadsheet-style edits while working on the same report, but
    # reset when switching to another record.
    if st.session_state.get("report_grid_current_id") != selected_report_id:
        st.session_state["report_grid_current_id"] = selected_report_id
        st.session_state.pop("report_grid_editor_state", None)
        st.session_state.pop("report_grid_import_rows", None)
        st.session_state.pop("report_grid_import_payload", None)
        st.session_state.pop("report_grid_mapping_choices", None)
        st.session_state.pop("report_grid_mapping_saved", None)

    editing_record: Optional[dict] = None
    if selected_report_id is not None and not selectable_reports.empty:
        match = selectable_reports[
            selectable_reports["report_id"] == int(selected_report_id)
        ]
        if not match.empty:
            editing_record = match.iloc[0].to_dict()
    if selected_report_id is not None and editing_record:
        owner_seed = editing_record.get("user_id")
        owner_id = None
        if owner_seed is not None:
            owner_id = int(_coerce_float(owner_seed, -1))
            if owner_id < 0:
                owner_id = None
        owner_label = label_map.get(owner_id, f"User #{owner_id}") if owner_id else None
        can_delete = is_admin
        if can_delete:
            with st.expander("Delete report", expanded=False):
                st.warning(
                    "Deleting removes the report from the system. Uploaded files stay in storage for admin review."
                )
                confirm_delete = st.checkbox(
                    "I understand, delete this report",
                    key=f"report_delete_confirm_{selected_report_id}",
                )
                if st.button(
                    "Delete report",
                    type="secondary",
                    disabled=not confirm_delete,
                    key=f"report_delete_button_{selected_report_id}",
                ):
                    delete_work_report(
                        conn,
                        report_id=int(selected_report_id),
                        owner_id=owner_id,
                        owner_label=owner_label,
                        period_type=clean_text(editing_record.get("period_type")),
                        period_start=clean_text(editing_record.get("period_start")),
                        period_end=clean_text(editing_record.get("period_end")),
                        report_template=clean_text(editing_record.get("report_template")),
                    )
                    st.warning("Report deleted.")
                    _safe_rerun()
    template_key = _normalize_report_template(
        editing_record.get("report_template") if editing_record else None
    )
    if not editing_record:
        template_key = _normalize_report_template(
            st.session_state.get("report_template_select")
        )
        template_key = template_key or "service"
        template_options = list(REPORT_TEMPLATE_LABELS.keys())
        template_index = template_options.index(template_key)
        template_key = st.selectbox(
            "Report template",
            template_options,
            index=template_index,
            format_func=lambda key: REPORT_TEMPLATE_LABELS.get(
                key, key.replace("_", " ").title()
            ),
            key="report_template_select",
        )
    else:
        st.selectbox(
            "Report template",
            list(REPORT_TEMPLATE_LABELS.keys()),
            index=list(REPORT_TEMPLATE_LABELS.keys()).index(template_key),
            format_func=lambda key: REPORT_TEMPLATE_LABELS.get(
                key, key.replace("_", " ").title()
            ),
            key="report_template_select",
            disabled=True,
        )
    if st.session_state.get("report_template_current") != template_key:
        st.session_state["report_template_current"] = template_key
        st.session_state.pop("report_grid_import_rows", None)
        st.session_state.pop("report_grid_import_payload", None)
        st.session_state.pop("report_grid_mapping_choices", None)
        st.session_state.pop("report_grid_mapping_saved", None)
        st.session_state.pop("report_grid_editor_state", None)
    default_period_key = "daily"
    if editing_record:
        seed_period = clean_text(editing_record.get("period_type"))
        if seed_period:
            seed_period = seed_period.lower()
            if seed_period in REPORT_PERIOD_OPTIONS:
                default_period_key = seed_period
    period_keys = list(REPORT_PERIOD_OPTIONS.keys())
    if not is_admin:
        allowed_periods = ["daily", "weekly", "monthly"]
        period_keys = [key for key in period_keys if key in allowed_periods]
    if not period_keys:
        period_keys = ["daily"]
    if default_period_key not in period_keys:
        default_period_key = period_keys[0]
    period_index = (
        period_keys.index(default_period_key)
        if default_period_key in period_keys
        else 0
    )

    default_start = _date_or(editing_record.get("period_start") if editing_record else None, today)
    default_end = _date_or(editing_record.get("period_end") if editing_record else None, default_start)
    legacy_tasks = clean_text(editing_record.get("tasks")) if editing_record else None
    legacy_remarks = clean_text(editing_record.get("remarks")) if editing_record else None
    legacy_research = clean_text(editing_record.get("research")) if editing_record else None

    start_date = default_start
    end_date = default_end

    existing_attachment_value: Optional[str] = (
        editing_record.get("attachment_path") if editing_record else None
    )
    existing_import_value: Optional[str] = (
        editing_record.get("import_file_path") if editing_record else None
    )

    st.markdown("##### Import report data")
    import_payload = st.session_state.get("report_grid_import_payload")
    import_payload_is_new = False
    if st.session_state.pop("report_grid_importer_reset", False):
        st.session_state.pop("report_grid_importer", None)
        st.session_state.pop("report_grid_import_payload", None)
        st.session_state.pop("report_grid_mapping_choices", None)
        st.session_state.pop("report_grid_mapping_saved", None)

    import_file = st.file_uploader(
        "Upload report grid (Excel or CSV)",
        type=["xlsx", "xls", "csv"],
        help="Populate the grid below by importing a spreadsheet with columns matching the report table.",
        key="report_grid_importer",
    )
    uploaded_df: Optional[pd.DataFrame] = None
    if import_file is not None:
        import_payload_is_new = True
        import_payload = {
            "name": import_file.name,
            "data": import_file.getvalue(),
        }
        st.session_state["report_grid_import_payload"] = import_payload
        st.session_state.pop("report_grid_mapping_choices", None)
        st.session_state["report_grid_mapping_saved"] = False

    if import_payload:
        uploaded_df = _load_report_grid_dataframe(
            import_payload.get("data", b""), import_payload.get("name", "")
        )
        if uploaded_df is not None:
            suggestions = _suggest_report_column_mapping(
                uploaded_df.columns, template_key=template_key
            )
            mapping_seed = st.session_state.get("report_grid_mapping_choices", {})
            map_options = ["(Do not import)"] + list(uploaded_df.columns)
            selected_mapping: dict[str, str] = {}
            load_clicked = False
            with st.form("report_grid_import_mapper"):
                st.caption(
                    "Align columns from the uploaded file to the report grid fields. Skipped columns will be ignored."
                )
                for key, config in _get_report_grid_fields(template_key).items():
                    default_choice = mapping_seed.get(key) or suggestions.get(key)
                    if default_choice not in map_options:
                        default_choice = "(Do not import)"
                    choice = st.selectbox(
                        config["label"],
                        options=map_options,
                        index=map_options.index(default_choice)
                        if default_choice in map_options
                        else 0,
                        key=f"report_map_{key}",
                        help=f"Select the column that represents '{config['label']}'.",
                    )
                    if choice != "(Do not import)":
                        selected_mapping[choice] = key
                load_clicked = st.form_submit_button("Load mapped rows into grid")

            if load_clicked:
                st.session_state["report_grid_mapping_choices"] = {
                    key: st.session_state.get(f"report_map_{key}")
                    for key in _get_report_grid_fields(template_key).keys()
                }
                imported_rows = _import_report_grid_from_dataframe(
                    uploaded_df, selected_mapping, template_key=template_key
                )
                if imported_rows:
                    st.session_state["report_grid_import_rows"] = imported_rows
                    st.success(
                        f"Loaded {len(imported_rows)} row(s) using the selected mapping."
                    )
                    st.session_state.pop("report_grid_import_payload", None)
                    st.session_state.pop("report_grid_mapping_choices", None)
                    st.session_state.pop("report_grid_mapping_saved", None)
                    st.session_state["report_grid_importer_reset"] = True
                    _safe_rerun()
                else:
                    st.warning(
                        "No rows were imported with that mapping. Please review the selections and try again.",
                        icon="⚠️",
                    )
        else:
            st.warning(
                "We could not read any matching columns from that file. Ensure the headers match the report grid labels.",
                icon="⚠️",
            )

    grid_seed_rows = (
        parse_report_grid_payload(
            editing_record.get("grid_payload"), template_key=template_key
        )
        if editing_record
        else []
    )
    if "report_grid_import_rows" in st.session_state:
        grid_seed_rows = st.session_state.pop("report_grid_import_rows") or grid_seed_rows
    elif st.session_state.get("report_grid_editor_state"):
        grid_seed_rows = st.session_state["report_grid_editor_state"] or grid_seed_rows
    if not grid_seed_rows:
        if template_key == "service":
            fallback_row = _default_report_grid_row(template_key)
            if legacy_tasks:
                fallback_row["reported_complaints"] = legacy_tasks
            if legacy_remarks:
                fallback_row["details_remarks"] = legacy_remarks
            if legacy_research:
                fallback_row["product_details"] = legacy_research
            if any(val not in (None, "") for val in fallback_row.values()):
                grid_seed_rows = [fallback_row]
    existing_attachment_path = (
        resolve_upload_path(existing_attachment_value)
        if existing_attachment_value
        else None
    )
    existing_attachment_bytes: Optional[bytes] = None
    existing_attachment_name: Optional[str] = None
    if existing_attachment_path and existing_attachment_path.exists():
        existing_attachment_name = existing_attachment_path.name
        try:
            existing_attachment_bytes = existing_attachment_path.read_bytes()
        except OSError:
            existing_attachment_bytes = None
    existing_import_path = (
        resolve_upload_path(existing_import_value)
        if existing_import_value
        else None
    )
    existing_import_bytes: Optional[bytes] = None
    existing_import_name: Optional[str] = None
    if existing_import_path and existing_import_path.exists():
        existing_import_name = existing_import_path.name
        try:
            existing_import_bytes = existing_import_path.read_bytes()
        except OSError:
            existing_import_bytes = None

    if existing_attachment_value:
        st.caption("Current attachment")
        if existing_attachment_bytes and existing_attachment_name:
            st.download_button(
                "Download current attachment",
                data=existing_attachment_bytes,
                file_name=existing_attachment_name,
                key="report_attachment_download",
            )
        else:
            st.warning(
                "The saved attachment could not be located on disk.",
                icon="⚠️",
            )
    if existing_import_value:
        st.caption("Imported file")
        if existing_import_bytes and existing_import_name:
            st.download_button(
                "Download imported file",
                data=existing_import_bytes,
                file_name=existing_import_name,
                key="report_import_download",
            )
        else:
            st.warning(
                "The saved import file could not be located on disk.",
                icon="⚠️",
            )

    with st.form("work_report_form"):
        period_choice = st.selectbox(
            "Report cadence",
            period_keys,
            index=period_index,
            format_func=lambda key: REPORT_PERIOD_OPTIONS.get(key, key.title()),
            key="report_period_type",
        )
        if period_choice == "daily":
            day_kwargs: dict[str, object] = {}
            if not is_admin and not editing_record:
                day_kwargs["min_value"] = today
                day_kwargs["max_value"] = today
            day_value = st.date_input(
                "Report date",
                value=default_start,
                key="report_period_daily",
                **day_kwargs,
            )
            start_date = day_value
            end_date = day_value
        elif period_choice == "weekly":
            base_start = default_start if editing_record else today - timedelta(days=today.weekday())
            base_end = default_end if editing_record else base_start + timedelta(days=6)
            week_value = st.date_input(
                "Week range",
                value=(base_start, base_end),
                key="report_period_weekly",
            )
            if isinstance(week_value, (list, tuple)) and len(week_value) == 2:
                start_date, end_date = week_value
            else:
                start_date = week_value
                end_date = week_value + timedelta(days=6)
            st.caption(
                f"Selected window: {format_period_range(to_iso_date(start_date), to_iso_date(end_date))}"
            )
        else:
            base_month = default_start if editing_record else today
            try:
                month_seed = base_month.replace(day=1)
            except Exception:
                month_seed = date(today.year, today.month, 1)
            month_value = st.date_input(
                "Month",
                value=month_seed,
                key="report_period_monthly",
            )
            if isinstance(month_value, (list, tuple)) and month_value:
                month_seed = month_value[0]
            else:
                month_seed = month_value
            if not isinstance(month_seed, date):
                month_seed = month_seed.to_pydatetime().date() if hasattr(month_seed, "to_pydatetime") else month_seed
            if not isinstance(month_seed, date):
                month_seed = date(today.year, today.month, 1)
            month_start = month_seed.replace(day=1)
            last_day = monthrange(month_start.year, month_start.month)[1]
            month_end = date(month_start.year, month_start.month, last_day)
            start_date, end_date = month_start, month_end
            st.caption(
                f"Selected window: {format_period_range(to_iso_date(start_date), to_iso_date(end_date))}"
            )

        template_label = REPORT_TEMPLATE_LABELS.get(
            template_key, template_key.replace("_", " ").title()
        )
        st.caption(
            f"Log {template_label.lower()} progress in a spreadsheet-style grid. "
            "Add rows for each customer or job completed."
        )
        seed_for_editor = grid_seed_rows or [_default_report_grid_row(template_key)]
        editor_seed = _grid_rows_for_editor(
            seed_for_editor, template_key=template_key
        )
        if not editor_seed:
            editor_seed = _grid_rows_for_editor(
                [_default_report_grid_row(template_key)],
                template_key=template_key,
            )
        fields = _get_report_grid_fields(template_key)
        grid_df_seed = pd.DataFrame(editor_seed, columns=fields.keys())
        column_config = _build_report_column_config(fields)
        disabled_columns: list[str] = []
        if "remarks_history" in fields:
            disabled_columns.append("remarks_history")
        if editing_record and not is_admin:
            if template_key in {"service", "sales"}:
                editable_columns = {"details_remarks", "progress_status"}
                helper_label = "Staff can update only the remarks and progress columns for existing reports."
            elif template_key == "follow_up":
                editable_columns = {"notes", "reminder_date", "progress_status"}
                helper_label = (
                    "Staff can update only the remarks, progress, and reminder date columns for existing follow-up reports."
                )
            else:
                editable_columns = set()
                helper_label = ""
            if editable_columns:
                disabled_columns = [
                    key for key in fields.keys() if key not in editable_columns
                ]
                if helper_label:
                    st.info(helper_label, icon="📝")
        if template_key == "service":
            customer_scope, customer_params = customer_scope_filter()
            customer_clause = f"WHERE {customer_scope}" if customer_scope else ""
            customer_df = df_query(
                conn,
                f"""
                SELECT name, company_name
                FROM customers
                {customer_clause}
                ORDER BY LOWER(COALESCE(name, company_name, ''))
                """,
                customer_params,
            )
            customer_options: list[str] = []
            if not customer_df.empty:
                for _, row in customer_df.iterrows():
                    display = clean_text(row.get("name")) or clean_text(
                        row.get("company_name")
                    )
                    if display and display not in customer_options:
                        customer_options.append(display)
            existing_customers = {
                clean_text(row.get("customer_name"))
                for row in (grid_seed_rows or [])
                if clean_text(row.get("customer_name"))
            }
            for name in sorted(existing_customers):
                if name and name not in customer_options:
                    customer_options.append(name)
            if customer_options:
                column_config["customer_name"] = st.column_config.SelectboxColumn(
                    "Customer Name",
                    options=[""] + customer_options,
                    help="Choose an existing customer for this service report.",
                )
        report_grid_df = st.data_editor(
            grid_df_seed,
            column_config=column_config,
            column_order=list(fields.keys()),
            hide_index=True,
            num_rows="dynamic",
            use_container_width=True,
            disabled=disabled_columns,
            key="report_grid_editor",
        )
        st.session_state["report_grid_editor_state"] = _grid_rows_from_editor(
            report_grid_df, template_key=template_key
        )
        remove_attachment = False
        attachment_upload = None
        if existing_attachment_value:
            remove_attachment = st.checkbox(
                "Remove current attachment",
                key="report_remove_attachment",
            )
        attachment_upload = st.file_uploader(
            "Attach supporting document (PDF, image, or Excel)",
            type=["pdf", "png", "jpg", "jpeg", "webp", "gif", "xlsx", "xls"],
            key="report_attachment_uploader",
            help="Optional proof of work, photos, or documentation. Excel uploads are supported for imported reports.",
        )
        _render_upload_ocr_preview(
            attachment_upload,
            key_prefix="report_attachment_uploader",
            label="Report attachment OCR",
        )
        submitted = st.form_submit_button("Save report", type="primary")

    if submitted:
        cleanup_path: Optional[str] = None
        cleanup_import_path: Optional[str] = None
        grid_rows_to_store = _grid_rows_from_editor(
            report_grid_df, template_key=template_key
        )
        if editing_record:
            previous_grid_rows = parse_report_grid_payload(
                clean_text(editing_record.get("grid_payload")),
                template_key=template_key,
            )
            grid_rows_to_store = _append_report_remarks_history(
                grid_rows_to_store,
                previous_grid_rows,
                template_key=template_key,
            )
        st.session_state.pop("report_grid_editor_state", None)
        summary_fields = REPORT_TEMPLATE_SUMMARY_FIELDS.get(
            template_key, REPORT_TEMPLATE_SUMMARY_FIELDS["service"]
        )
        tasks_summary = _summarize_grid_column(
            grid_rows_to_store, summary_fields["tasks"]
        )
        remarks_summary = _summarize_grid_column(
            grid_rows_to_store, summary_fields["remarks"]
        )
        research_summary = _summarize_grid_column(
            grid_rows_to_store, summary_fields["research"]
        )
        try:
            normalized_key, normalized_start, normalized_end = normalize_report_window(
                period_choice, start_date, end_date
            )
        except ValueError as err:
            st.error(str(err))
        else:
            attachment_to_store = _ATTACHMENT_UNCHANGED
            import_file_to_store = _ATTACHMENT_UNCHANGED
            attachment_save_failed = False
            save_allowed = True

            if not is_admin:
                validation_error: Optional[str] = None
                allow_relaxed_edit = bool(
                    editing_record and template_key in {"service", "sales", "follow_up"}
                )
                if not allow_relaxed_edit:
                    if normalized_key == "daily":
                        if normalized_start != today or normalized_end != today:
                            validation_error = "Daily reports can only be submitted for today."
                    elif normalized_key == "weekly":
                        if today.weekday() != 5:
                            validation_error = "Weekly reports can only be submitted on Saturdays."
                        elif not (
                            normalized_start == current_week_start
                            and normalized_end == current_week_end
                        ):
                            validation_error = (
                                "Weekly reports must cover the current week (Monday to Sunday)."
                            )
                    elif normalized_key == "monthly":
                        if not (
                            normalized_start == current_month_start
                            and normalized_end == current_month_end
                        ):
                            validation_error = "Monthly reports must cover the current month."
                if validation_error:
                    st.error(validation_error)
                    save_allowed = False

            if save_allowed and attachment_upload is not None:
                identifier = (
                    f"user{report_owner_id}_{normalized_key}_{normalized_start.isoformat()}"
                )
                stored_path = store_report_attachment(
                    attachment_upload,
                    identifier=identifier,
                )
                if stored_path:
                    attachment_to_store = stored_path
                    if existing_attachment_value:
                        cleanup_path = existing_attachment_value
                else:
                    st.error("Attachment could not be saved. Please try again.")
                    attachment_save_failed = True
            elif save_allowed and remove_attachment and existing_attachment_value:
                attachment_to_store = None
                cleanup_path = existing_attachment_value

            if save_allowed and import_payload_is_new and import_payload:
                identifier = (
                    f"user{report_owner_id}_{normalized_key}_{normalized_start.isoformat()}"
                )
                stored_import = store_report_import_file(
                    import_payload.get("name") or "report_import",
                    import_payload.get("data") or b"",
                    identifier=identifier,
                )
                if stored_import:
                    import_file_to_store = stored_import
                    if existing_import_value:
                        cleanup_import_path = existing_import_value
                else:
                    st.error("Imported file could not be saved. Please try again.")
                    attachment_save_failed = True

            if save_allowed and not attachment_save_failed:
                if not grid_rows_to_store:
                    st.error("Add at least one row to the report grid before saving.")
                else:
                    try:
                        saved_id = upsert_work_report(
                            conn,
                            report_id=int(selected_report_id) if selected_report_id is not None else None,
                            user_id=int(report_owner_id),
                            period_type=normalized_key,
                            period_start=normalized_start,
                            period_end=normalized_end,
                            tasks=tasks_summary,
                            remarks=remarks_summary,
                            research=research_summary,
                            report_template=template_key,
                            grid_rows=grid_rows_to_store,
                            attachment_path=attachment_to_store,
                            current_attachment=existing_attachment_value,
                            import_file_path=import_file_to_store,
                            current_import_file=existing_import_value,
                        )
                    except ValueError as err:
                        st.error(str(err))
                    else:
                        st.success("Report saved successfully.")
                        if cleanup_path:
                            old_path = resolve_upload_path(cleanup_path)
                            if old_path and old_path.exists():
                                with contextlib.suppress(OSError):
                                    old_path.unlink()
                        if cleanup_import_path:
                            old_import = resolve_upload_path(cleanup_import_path)
                            if old_import and old_import.exists():
                                with contextlib.suppress(OSError):
                                    old_import.unlink()
                        st.session_state["report_edit_select_pending"] = saved_id
                        st.session_state.pop("report_attachment_uploader", None)
                        st.session_state.pop("report_grid_import_payload", None)
                        _safe_rerun()

    st.markdown("---")
    st.markdown("#### Report history")

    def _display_text(value: Optional[object]) -> str:
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        return str(value).strip()

    history_user = viewer_id
    if is_admin:
        history_options: list[Optional[int]] = [None] + user_ids

        def _history_label(uid: Optional[int]) -> str:
            if uid is None:
                return "All team members"
            return label_map.get(uid, f"User #{uid}")

        default_history_index = (
            history_options.index(report_owner_id)
            if report_owner_id in history_options
            else 0
        )
        history_user = st.selectbox(
            "Team member",
            history_options,
            index=default_history_index,
            format_func=_history_label,
            key="report_history_user",
        )

    period_keys = list(REPORT_PERIOD_OPTIONS.keys())
    history_periods = st.multiselect(
        "Cadence",
        period_keys,
        default=period_keys,
        format_func=lambda key: REPORT_PERIOD_OPTIONS.get(key, key.title()),
        key="report_history_periods",
    )

    default_history_start = today - timedelta(days=30)
    history_range = st.date_input(
        "Period range",
        value=(default_history_start, today),
        key="report_history_range",
    )
    range_start = range_end = None
    if isinstance(history_range, (list, tuple)) and len(history_range) == 2:
        range_start, range_end = history_range
    elif history_range:
        range_start = history_range
        range_end = history_range

    search_term = st.text_input(
        "Search notes",
        key="report_history_search",
        placeholder="Keyword in tasks, remarks, or research",
    )

    filters: list[str] = []
    params: list[object] = []
    if not is_admin or history_user is not None:
        target = history_user if history_user is not None else viewer_id
        filters.append("wr.user_id = ?")
        params.append(int(target))
    if history_periods and len(history_periods) != len(period_keys):
        placeholders = ",".join("?" for _ in history_periods)
        filters.append(f"wr.period_type IN ({placeholders})")
        params.extend(history_periods)
    if range_start:
        filters.append("date(wr.period_start) >= date(?)")
        params.append(to_iso_date(range_start))
    if range_end:
        filters.append("date(wr.period_end) <= date(?)")
        params.append(to_iso_date(range_end))
    if search_term:
        keyword = search_term.strip()
        if keyword:
            filters.append(
                "(wr.tasks LIKE '%'||?||'%' OR wr.remarks LIKE '%'||?||'%' OR wr.research LIKE '%'||?||'%')"
            )
            params.extend([keyword, keyword, keyword])

    where_clause = " AND ".join(filters) if filters else "1=1"
    history_df = df_query(
        conn,
        dedent(
            f"""
            SELECT wr.report_id, wr.user_id, wr.period_type, wr.period_start, wr.period_end,
                   wr.tasks, wr.remarks, wr.research, wr.grid_payload, wr.attachment_path, wr.import_file_path, wr.report_template, wr.created_at, wr.updated_at,
                   u.username
            FROM work_reports wr
            JOIN users u ON u.user_id = wr.user_id
            WHERE {where_clause}
            ORDER BY date(wr.period_start) DESC, wr.report_id DESC
            """
        ),
        tuple(params),
    )

    if history_df.empty:
        st.info("No reports found for the selected filters.")
        return

    history_df["report_id"] = history_df["report_id"].apply(lambda val: int(float(val)))
    history_df["username"] = history_df.apply(
        lambda row: clean_text(row.get("username")) or f"User #{int(row['user_id'])}",
        axis=1,
    )

    history_df["template_key"] = history_df["report_template"].apply(
        _normalize_report_template
    )
    history_df["grid_rows"] = history_df.apply(
        lambda row: parse_report_grid_payload(
            row.get("grid_payload"), template_key=row.get("template_key")
        ),
        axis=1,
    )

    def _legacy_rows(row: pd.Series) -> list[dict[str, object]]:
        fallback = _default_report_grid_row()
        legacy_flag = False
        if clean_text(row.get("tasks")):
            fallback["reported_complaints"] = clean_text(row.get("tasks"))
            legacy_flag = True
        if clean_text(row.get("remarks")):
            fallback["details_remarks"] = clean_text(row.get("remarks"))
            legacy_flag = True
        if clean_text(row.get("research")):
            fallback["product_details"] = clean_text(row.get("research"))
            legacy_flag = True
        return [fallback] if legacy_flag else []

    history_df["grid_rows"] = history_df.apply(
        lambda row: row.get("grid_rows") or _legacy_rows(row),
        axis=1,
    )

    entry_records: list[dict[str, object]] = []
    download_records: list[dict[str, object]] = []
    for _, record in history_df.iterrows():
        owner = record.get("username") or f"User #{int(record.get('user_id'))}"
        cadence_label = format_period_label(record.get("period_type"))
        period_label = format_period_range(
            record.get("period_start"), record.get("period_end")
        )
        template_key = record.get("template_key")
        template_label = REPORT_TEMPLATE_LABELS.get(
            template_key, str(template_key).replace("_", " ").title()
        )
        grid_rows = record.get("grid_rows") or []
        display_df = format_report_grid_rows_for_display(
            grid_rows, empty_ok=True, template_key=template_key
        )
        if display_df.empty:
            continue
        for entry in display_df.to_dict("records"):
            entry_record = {"Template": template_label}
            entry_record.update(entry)
            entry_records.append(entry_record)
            download_entry = {
                "Team member": owner,
                "Template": template_label,
                "Cadence": cadence_label,
                "Period": period_label,
            }
            download_entry.update(entry)
            download_records.append(download_entry)

    entry_table = pd.DataFrame(entry_records)
    if not entry_table.empty:
        for fields in REPORT_TEMPLATE_FIELDS.values():
            for key, config in fields.items():
                label = config["label"]
                if label not in entry_table.columns:
                    entry_table[label] = pd.NA
                if config["type"] == "number":
                    entry_table[label] = pd.to_numeric(
                        entry_table[label], errors="coerce"
                    )
                else:
                    entry_table[label] = entry_table[label].fillna("")
        if "Template" not in entry_table.columns:
            entry_table["Template"] = ""
        entry_table = entry_table.reindex(
            columns=["Template", *ALL_REPORT_DISPLAY_COLUMNS]
        )
        st.dataframe(entry_table, use_container_width=True)
        progress_label = "Progress"
        if progress_label in entry_table.columns:
            progress_series = (
                entry_table[progress_label]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.lower()
            )
            progress_series = progress_series.replace({"pending": "ongoing"})
            total_rows = int(progress_series.shape[0])
            done_rows = int((progress_series == "done").sum())
            rejected_rows = int((progress_series == "rejected").sum())
            ongoing_rows = int((progress_series == "ongoing").sum())
            progress_cols = st.columns(4)
            progress_cols[0].metric("Report rows", total_rows)
            progress_cols[1].metric("Ongoing", ongoing_rows)
            progress_cols[2].metric("Done", done_rows)
            progress_cols[3].metric("Rejected", rejected_rows)
    else:
        st.info(
            "No structured report entries are available for the selected filters."
        )

    download_df = pd.DataFrame(download_records)
    if not download_df.empty:
        desired_columns = [
            "Team member",
            "Template",
            "Cadence",
            "Period",
            *ALL_REPORT_DISPLAY_COLUMNS,
        ]
        download_df = download_df.reindex(columns=desired_columns, fill_value="")
    elif not entry_table.empty:
        download_df = entry_table.reindex(
            columns=["Template", *ALL_REPORT_DISPLAY_COLUMNS], fill_value=""
        )
    if not download_df.empty:
        csv_data = download_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Download filtered reports",
            data=csv_data,
            file_name="work_reports.csv",
            mime="text/csv",
            key="reports_download",
        )

    cadence_summary = (
        history_df.assign(label=history_df["period_type"].apply(format_period_label))
        .groupby("label")["report_id"]
        .count()
        .sort_index()
    )
    if not cadence_summary.empty:
        st.markdown("##### Cadence summary")
        cols = st.columns(len(cadence_summary))
        for col, (label, count) in zip(cols, cadence_summary.items()):
            col.metric(label, int(count))

    if is_admin and history_user is None:
        coverage = (
            history_df.assign(
                label=history_df["period_type"].apply(format_period_label),
                owner=history_df["username"],
            )
            .pivot_table(
                index="owner",
                columns="label",
                values="report_id",
                aggfunc="count",
                fill_value=0,
            )
            .astype(int)
        )
        coverage.columns.name = None
        coverage = coverage.reset_index().rename(columns={"owner": "Team member"})
        st.markdown("##### Reports by team member")
        st.dataframe(coverage, use_container_width=True)

    detail_limit = min(len(history_df), 20)
    if detail_limit:
        st.markdown("##### Quick read")
        for _, row in history_df.head(detail_limit).iterrows():
            header = (
                f"{row['username']} – {format_period_label(row['period_type'])} "
                f"({format_period_range(row.get('period_start'), row.get('period_end'))})"
            )
            with st.expander(header, expanded=False):
                grid_df = format_report_grid_rows_for_display(
                    row.get("grid_rows"),
                    empty_ok=True,
                    template_key=row.get("template_key"),
                )
                if not grid_df.empty:
                    st.dataframe(grid_df, use_container_width=True)
                else:
                    st.write("No structured entries recorded for this report.")
                    legacy_blocks = [
                        ("Tasks completed", _display_text(row.get("tasks"))),
                        ("Remarks / blockers", _display_text(row.get("remarks"))),
                        ("Research / learnings", _display_text(row.get("research"))),
                    ]
                    for title, text in legacy_blocks:
                        if text:
                            st.markdown(f"**{title}**")
                            st.write(text)
                created_label = format_period_range(
                    row.get("created_at"), row.get("created_at")
                )
                updated_label = format_period_range(
                    row.get("updated_at"), row.get("updated_at")
                )
                st.caption(f"Logged on {created_label} • Last updated {updated_label}")

# ---------- Main ----------
def main():
    st.session_state["_render_id"] = st.session_state.get("_render_id", 0) + 1
    render_id = st.session_state["_render_id"]
    init_ui()
    apply_theme_css()
    _ensure_quotation_editor_server()
    conn = get_conn()
    init_schema(conn)
    _purge_expired_sessions(conn)
    if st.session_state.pop("logout_requested", False):
        _clear_session_for_logout(conn)
        st.rerun()
    _, backup_error = ensure_monthly_backup(
        BACKUP_DIR,
        "ps_crm_backup",
        lambda: export_full_archive(conn),
        BACKUP_RETENTION_COUNT,
        BACKUP_MIRROR_PATH,
    )
    st.session_state["auto_backup_error"] = backup_error
    _restore_user_session(conn)
    login_box(conn, render_id=render_id)
    # Auth gate: stop rendering any dashboard/navigation without a user session.
    if not st.session_state.get("user"):
        st.stop()
    _touch_session(conn, st.session_state.get("session_token"))

    if "page" not in st.session_state:
        st.session_state.page = "Dashboard"

    user = st.session_state.user or {}
    role = user.get("role")
    if role == "admin":
        pages = [
            "Dashboard",
            "Customers",
            "Quotation",
            "Operations",
            "Warranties",
            "Advanced Search",
            "Reports",
            "Users (Admin)",
        ]
    else:
        pages = [
            "Dashboard",
            "Customers",
            "Quotation",
            "Operations",
            "Warranties",
            "Reports",
        ]

    if "nav_page" not in st.session_state:
        st.session_state["nav_page"] = st.session_state.get("page", pages[0])
    if st.session_state.get("nav_page") not in pages:
        st.session_state["nav_page"] = pages[0]
    current_page = st.session_state.get("nav_page", pages[0])

    def _sync_nav_choice(key: str) -> None:
        selection = st.session_state.get(key, pages[0])
        if selection not in pages:
            selection = pages[0]
        st.session_state["nav_page"] = selection
        st.session_state["page"] = selection

    if "nav_selection_sidebar" not in st.session_state:
        st.session_state["nav_selection_sidebar"] = current_page
    elif st.session_state.get("nav_selection_sidebar") not in pages:
        st.session_state["nav_selection_sidebar"] = current_page

    def _render_mobile_nav() -> None:
        if "nav_selection_mobile" not in st.session_state:
            st.session_state["nav_selection_mobile"] = current_page
        elif st.session_state.get("nav_selection_mobile") not in pages:
            st.session_state["nav_selection_mobile"] = current_page
        st.markdown('<div class="ps-mobile-nav">', unsafe_allow_html=True)
        if hasattr(st, "popover"):
            with st.popover("☰"):
                st.radio(
                    "Navigate",
                    pages,
                    key="nav_selection_mobile",
                    on_change=lambda: _sync_nav_choice("nav_selection_mobile"),
                )
                if st.button("Logout", key="mobile_logout", use_container_width=True):
                    _request_logout()
                    st.rerun()
        else:
            with st.expander("☰ Menu", expanded=False):
                st.radio(
                    "Navigate",
                    pages,
                    key="nav_selection_mobile",
                    on_change=lambda: _sync_nav_choice("nav_selection_mobile"),
                )
                if st.button(
                    "Logout", key="mobile_logout_expander", use_container_width=True
                ):
                    _request_logout()
                    st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    _render_mobile_nav()

    with st.sidebar:
        sidebar_dark = st.toggle(
            "Mode",
            value=get_theme() == "dark",
            key="sidebar_theme_toggle",
            help="Toggle between light and dark.",
        )
        set_theme(sidebar_dark)
        apply_theme_css()
        st.markdown("### Navigation")
        st.radio(
            "Navigate",
            pages,
            key="nav_selection_sidebar",
            on_change=lambda: _sync_nav_choice("nav_selection_sidebar"),
        )
        st.divider()
        if st.button("Logout", key="sidebar_logout_main", use_container_width=True):
            _request_logout()
            st.rerun()

    page = st.session_state.get("nav_page", pages[0])
    st.session_state.page = page
    show_expiry_notifications(conn)

    if page == "Dashboard":
        dashboard(conn)
    elif page == "Quotation":
        quotation_page(conn, render_id=render_id)
    elif page == "Customers":
        customers_hub_page(conn)
    elif page == "Operations":
        operations_page(conn)
    elif page == "Warranties":
        warranties_page(conn)
    elif page == "Advanced Search":
        advanced_search_page(conn)
    elif page == "Reports":
        reports_page(conn)
    elif page == "Users (Admin)":
        users_admin_page(conn)

if __name__ == "__main__":
    if _streamlit_runtime_active():
        main()
    else:
        _bootstrap_streamlit_app()
